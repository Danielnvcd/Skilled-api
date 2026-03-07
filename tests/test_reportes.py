import pytest
from io import BytesIO
import openpyxl

class TestReportesIndividuales:
    """Pruebas para las descargas de Excel individuales por módulo."""

    def test_exportar_acceso_denegado_si_no_logueado(self, client):
        """Verificar que un usuario no logueado redirige a login."""
        # Probamos con una ruta genérica de ejemplo
        response = client.get('/reportes/excel_prenomina/2026-01-01')
        assert response.status_code == 302
        assert '/login' in response.headers.get('Location', '')

    def test_exportar_acceso_denegado_coordinador(self, logged_in_coordinador):
        """Coordinadores no deberían poder descargar los reportes financieros."""
        response = logged_in_coordinador.get('/reportes/excel_prenomina/2026-01-01')
        assert response.status_code in (302, 403)

    def test_exportacion_prenomina_exitosa_admin(self, logged_in_admin, db, proyecto, trabajador):
        """Prueba de exportación de prenómina (vista previa o guardada)."""
        from app.models import ReporteSemanal, RegistroDiarioHoras
        from datetime import date, timedelta
        
        # Usar una fecha fija para evitar problemas de desfase
        fecha_fix = date(2026, 3, 2) # Un lunes de ejemplo
        
        # Crear un reporte cerrado para que la prenormina tenga qué calcular
        rep = ReporteSemanal(proyecto_id=proyecto.id, fecha_inicio_semana=fecha_fix, 
                             fecha_fin_semana=fecha_fix+timedelta(days=6), estado='PRENOMINA_CERRADA')
        db.session.add(rep)
        db.session.flush()
        
        # El trabajador debe tener registros en este reporte
        reg = RegistroDiarioHoras(reporte_id=rep.id, trabajador_id=trabajador.id, 
                                  fecha=fecha_fix, horas_productivas=8)
        db.session.add(reg)
        db.session.commit()
        
        response = logged_in_admin.get(f'/reportes/excel_prenomina/2026-03-02')
        assert response.status_code == 200
        assert 'Reporte_Prenomina' in response.headers.get('Content-Disposition', '')
        
        excel_data = BytesIO(response.data)
        workbook = openpyxl.load_workbook(excel_data)
        assert 'Prenómina' in workbook.sheetnames

    def test_exportacion_historico_exitosa_admin(self, logged_in_admin, db, proyecto, trabajador):
        """Prueba de exportación de histórico."""
        from app.models import Prenomina, ReporteSemanal
        from datetime import date, timedelta
        
        # Crear un reporte cerrado y una prenomina
        rep = ReporteSemanal(proyecto_id=proyecto.id, fecha_inicio_semana=date.today(), 
                             fecha_fin_semana=date.today()+timedelta(days=6), estado='CERRADO')
        db.session.add(rep)
        db.session.flush()
        
        pre = Prenomina(fecha_inicio=rep.fecha_inicio_semana, fecha_fin=rep.fecha_fin_semana, 
                        trabajador_id=trabajador.id, estado='APROBADO', total_a_pagar=1000)
        db.session.add(pre)
        db.session.commit()
        
        fecha_str = pre.fecha_inicio.strftime('%Y-%m-%d')
        response = logged_in_admin.get(f'/reportes/excel_historico/{fecha_str}')
        assert response.status_code == 200
        assert 'Reporte_Historico' in response.headers.get('Content-Disposition', '')
        
        excel_data = BytesIO(response.data)
        workbook = openpyxl.load_workbook(excel_data)
        assert 'Histórico' in workbook.sheetnames

    def test_exportacion_proyecto_exitosa_admin(self, logged_in_admin, db, proyecto, trabajador):
        """Prueba de exportación de proyecto total."""
        from app.models import ReporteSemanal, Prenomina, RegistroDiarioHoras
        from datetime import date, timedelta
        
        # 1. Crear un reporte cerrado
        rep = ReporteSemanal(proyecto_id=proyecto.id, fecha_inicio_semana=date.today(), 
                             fecha_fin_semana=date.today()+timedelta(days=6), estado='PRENOMINA_CERRADA')
        db.session.add(rep)
        db.session.flush()
        
        # 2. El trabajador debe tener registros en este reporte
        reg = RegistroDiarioHoras(reporte_id=rep.id, trabajador_id=trabajador.id, 
                                  fecha=date.today(), horas_productivas=8)
        db.session.add(reg)
        
        # 3. Debe haber una prenomina aprobada para este trabajador en esta semana
        pre = Prenomina(fecha_inicio=rep.fecha_inicio_semana, fecha_fin=rep.fecha_fin_semana, 
                        trabajador_id=trabajador.id, estado='APROBADO', total_a_pagar=1000)
        db.session.add(pre)
        db.session.commit()
        
        response = logged_in_admin.get(f'/reportes/excel_proyecto_total/{proyecto.id}')
        assert response.status_code == 200
        assert 'Reporte_ProyectoTotal' in response.headers.get('Content-Disposition', '')
        
        excel_data = BytesIO(response.data)
        workbook = openpyxl.load_workbook(excel_data)
        assert 'Proyecto Total' in workbook.sheetnames

    def test_exportacion_prestamos_exitosa_admin(self, logged_in_admin, db, trabajador):
        """Prueba de exportación de préstamos por trabajador."""
        from app.models import Prestamo
        p = Prestamo(trabajador_id=trabajador.id, monto_total=1000, monto_restante=1000, 
                     plazo_semanas=10, descuento_semanal=100)
        db.session.add(p)
        db.session.commit()
        
        response = logged_in_admin.get(f'/reportes/excel_prestamos/{trabajador.id}')
        assert response.status_code == 200
        assert 'Prestamos_' in response.headers.get('Content-Disposition', '')
        
        excel_data = BytesIO(response.data)
        workbook = openpyxl.load_workbook(excel_data)
        assert 'Préstamos' in workbook.sheetnames

    def test_exportacion_ajustes_exitosa_admin(self, logged_in_admin, db, trabajador):
        """Prueba de exportación de ajustes por periodo."""
        from app.models import AjustePeriodo, AjusteDescuento
        from datetime import date
        
        periodo = AjustePeriodo(nombre="Test Periodo", fecha_inicio=date.today(), fecha_fin=date.today())
        db.session.add(periodo)
        db.session.flush()
        
        desc = AjusteDescuento(periodo_id=periodo.id, trabajador_id=trabajador.id, monto=100, fecha_descuento=date.today())
        db.session.add(desc)
        db.session.commit()
        
        response = logged_in_admin.get(f'/reportes/excel_ajustes/{periodo.id}')
        assert response.status_code == 200
        assert 'Ajuste_Inbursa' in response.headers.get('Content-Disposition', '')
        
        excel_data = BytesIO(response.data)
        workbook = openpyxl.load_workbook(excel_data)
        assert 'Ajuste Inbursa' in workbook.sheetnames

