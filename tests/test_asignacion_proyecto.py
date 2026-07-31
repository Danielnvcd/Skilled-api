"""Tests de la asignación de material por proyecto.

Cubre `previsualizar`, `asignar` y `devolver`, que son la base de la sección
«Material por proyecto».

Lo que hace segura la decisión de «asignar lo disponible cuando no alcanza» es
que la previsualización lo muestre ANTES de aplicar. Por eso el test central de
este archivo es que previsualizar y aplicar coincidan: si divergen, la vista
previa deja de servir para lo único que sirve, que es confiar en ella.
"""
import pytest
from werkzeug.security import generate_password_hash

from app.models import (
    Almacen, MovimientoInventario, Producto, Proyecto,
    StockAlmacenProyecto, StockPorAlmacen, User,
)
from app.routes.api_auth import _encode_access_token


def _login(client, user):
    client.environ_base['HTTP_AUTHORIZATION'] = f'Bearer {_encode_access_token(user)}'


@pytest.fixture
def admin(db):
    u = User(username='as_admin', password_hash=generate_password_hash('Pass123!'), role='admin')
    db.session.add(u); db.session.commit()
    return u


@pytest.fixture
def almacen(db):
    a = Almacen(nombre='Central', qr_code='QR-AS-1', activo=True)
    db.session.add(a); db.session.commit()
    return a


@pytest.fixture
def proyecto(db):
    p = Proyecto(numero_proyecto='AS-001', nombre='Obra Demo', activo=True)
    db.session.add(p); db.session.commit()
    return p


@pytest.fixture
def producto(db, almacen):
    """100 unidades, todas en el bucket GENERAL."""
    p = Producto(codigo='CEM-01', descripcion='Cemento', categoria='Obra', unidad='kg',
                 stock_actual=100, stock_minimo=0, precio_unitario=50)
    db.session.add(p); db.session.flush()
    db.session.add(StockAlmacenProyecto(producto_id=p.id, almacen_id=almacen.id,
                                        proyecto_id=None, cantidad=100))
    db.session.add(StockPorAlmacen(producto_id=p.id, almacen_id=almacen.id, cantidad=100))
    db.session.commit()
    return p


def _bucket(db, producto, almacen, proyecto_id):
    f = (db.session.query(StockAlmacenProyecto)
         .filter_by(producto_id=producto.id, almacen_id=almacen.id,
                    proyecto_id=proyecto_id).first())
    return float(f.cantidad) if f else 0.0


def _url(proyecto, sufijo=''):
    return f'/api/v1/proyectos-materiales/{proyecto.id}/{sufijo}'


# ─── Previsualización ─────────────────────────────────────────────────────────

class TestPrevisualizar:

    def test_no_escribe_nada(self, client, db, admin, almacen, proyecto, producto):
        _login(client, admin)
        r = client.post(_url(proyecto, 'asignar/previsualizar'),
                        json={'lineas': [{'sku': 'CEM-01', 'cantidad': 30}]})
        assert r.status_code == 200, r.get_json()
        assert r.get_json()['resumen']['ok'] == 1

        db.session.expire_all()
        assert _bucket(db, producto, almacen, None) == 100, 'la previsualización escribió'
        assert _bucket(db, producto, almacen, proyecto.id) == 0

    def test_muestra_el_resultado_no_el_incremento(self, client, admin, almacen, proyecto, producto):
        """«0 → 30» dice más que «+30»."""
        _login(client, admin)
        r = client.post(_url(proyecto, 'asignar/previsualizar'),
                        json={'lineas': [{'sku': 'CEM-01', 'cantidad': 30}]})
        linea = r.get_json()['lineas'][0]
        assert linea['actual'] == 0
        assert linea['resultado'] == 30

    def test_sku_inexistente_no_rompe_el_lote(self, client, admin, almacen, proyecto, producto):
        _login(client, admin)
        r = client.post(_url(proyecto, 'asignar/previsualizar'),
                        json={'lineas': [
                            {'sku': 'CEM-01', 'cantidad': 10},
                            {'sku': 'NO-EXISTE', 'cantidad': 5},
                        ]})
        res = r.get_json()
        assert res['resumen']['ok'] == 1
        assert res['resumen']['errores'] == 1
        malo = [f for f in res['lineas'] if f['sku'] == 'NO-EXISTE'][0]
        assert 'no existe en el catálogo' in malo['motivo']

    def test_bodega_mal_escrita_sugiere_la_correcta(self, client, admin, almacen, proyecto, producto):
        """Un typo en el nombre de la bodega es el error más común del Excel."""
        _login(client, admin)
        r = client.post(_url(proyecto, 'asignar/previsualizar'),
                        json={'lineas': [{'sku': 'CEM-01', 'cantidad': 10, 'almacen': 'Cent'}]})
        motivo = r.get_json()['lineas'][0]['motivo']
        assert 'no existe' in motivo
        assert 'Central' in motivo, f'debería sugerir la bodega parecida: {motivo}'

    def test_sin_suficiente_avisa_y_ajusta(self, client, admin, almacen, proyecto, producto):
        """Decisión: se asigna lo disponible con aviso, no se bloquea el lote."""
        _login(client, admin)
        r = client.post(_url(proyecto, 'asignar/previsualizar'),
                        json={'lineas': [{'sku': 'CEM-01', 'cantidad': 500}]})
        linea = r.get_json()['lineas'][0]
        assert linea['estado'] == 'aviso'
        assert linea['cantidad_aplicada'] == 100
        assert '100' in linea['motivo']

    def test_cantidad_cero_es_error(self, client, admin, almacen, proyecto, producto):
        _login(client, admin)
        r = client.post(_url(proyecto, 'asignar/previsualizar'),
                        json={'lineas': [{'sku': 'CEM-01', 'cantidad': 0}]})
        assert r.get_json()['lineas'][0]['estado'] == 'error'


# ─── El invariante central ────────────────────────────────────────────────────

class TestPrevisualizarYAplicarCoinciden:

    def test_con_ajuste_por_faltante(self, client, db, admin, almacen, proyecto, producto):
        """Si la previsualización promete un resultado y aplicar hace otro, la
        vista previa no sirve para nada."""
        _login(client, admin)
        cuerpo = {'lineas': [{'sku': 'CEM-01', 'cantidad': 500}]}   # dispara el ajuste

        prev = client.post(_url(proyecto, 'asignar/previsualizar'), json=cuerpo).get_json()
        aplic = client.post(_url(proyecto, 'asignar'), json=cuerpo).get_json()

        assert prev['lineas'][0]['cantidad_aplicada'] == aplic['lineas'][0]['cantidad_aplicada']
        assert prev['lineas'][0]['resultado'] == aplic['lineas'][0]['resultado']
        db.session.expire_all()
        assert _bucket(db, producto, almacen, proyecto.id) == prev['lineas'][0]['resultado']

    def test_con_lote_mixto(self, client, db, admin, almacen, proyecto, producto):
        _login(client, admin)
        cuerpo = {'lineas': [
            {'sku': 'CEM-01', 'cantidad': 30},
            {'sku': 'NO-EXISTE', 'cantidad': 5},
        ]}
        prev = client.post(_url(proyecto, 'asignar/previsualizar'), json=cuerpo).get_json()
        aplic = client.post(_url(proyecto, 'asignar'), json=cuerpo).get_json()
        assert prev['resumen'] == aplic['resumen']


# ─── Aplicación ───────────────────────────────────────────────────────────────

class TestAplicar:

    def test_mueve_de_general_al_proyecto(self, client, db, admin, almacen, proyecto, producto):
        _login(client, admin)
        r = client.post(_url(proyecto, 'asignar'),
                        json={'lineas': [{'sku': 'CEM-01', 'cantidad': 30}]})
        assert r.status_code == 200, r.get_json()
        db.session.expire_all()
        assert _bucket(db, producto, almacen, proyecto.id) == 30
        assert _bucket(db, producto, almacen, None) == 70, 'debe salir de General'
        # El total NO cambia: el material solo cambió de etiqueta.
        assert float(Producto.query.get(producto.id).stock_actual) == 100

    def test_entrada_nueva_si_aumenta_el_total(self, client, db, admin, almacen, proyecto, producto):
        """origen='entrada' es material que llega de fuera: crea stock."""
        _login(client, admin)
        r = client.post(_url(proyecto, 'asignar'),
                        json={'origen': 'entrada', 'lineas': [{'sku': 'CEM-01', 'cantidad': 40}]})
        assert r.status_code == 200, r.get_json()
        db.session.expire_all()
        assert _bucket(db, producto, almacen, proyecto.id) == 40
        assert _bucket(db, producto, almacen, None) == 100, 'General no se toca'
        assert float(Producto.query.get(producto.id).stock_actual) == 140

    def test_deja_rastro_en_el_kardex(self, client, db, admin, almacen, proyecto, producto):
        """Si algo mueve stock sin quedar registrado, el inventario deja de ser
        auditable — y esa es la razón de reutilizar los helpers de movimientos."""
        _login(client, admin)
        antes = MovimientoInventario.query.count()
        client.post(_url(proyecto, 'asignar'),
                    json={'lineas': [{'sku': 'CEM-01', 'cantidad': 30}]})
        movs = MovimientoInventario.query.all()
        assert len(movs) == antes + 1
        m = movs[-1]
        assert m.tipo == 'REASIGNACION'
        assert m.proyecto_destino_id == proyecto.id
        assert m.proyecto_origen_id is None

    def test_sumar_es_el_modo_por_defecto(self, client, db, admin, almacen, proyecto, producto):
        _login(client, admin)
        for _ in range(2):
            client.post(_url(proyecto, 'asignar'),
                        json={'lineas': [{'sku': 'CEM-01', 'cantidad': 20}]})
        db.session.expire_all()
        assert _bucket(db, producto, almacen, proyecto.id) == 40, 'sumó, no reemplazó'

    def test_modo_reemplazar_ajusta_al_objetivo(self, client, db, admin, almacen, proyecto, producto):
        _login(client, admin)
        client.post(_url(proyecto, 'asignar'),
                    json={'lineas': [{'sku': 'CEM-01', 'cantidad': 40}]})
        # Reemplazar a 25: sobran 15 que vuelven a General.
        r = client.post(_url(proyecto, 'asignar'),
                        json={'modo': 'reemplazar', 'lineas': [{'sku': 'CEM-01', 'cantidad': 25}]})
        assert r.status_code == 200, r.get_json()
        db.session.expire_all()
        assert _bucket(db, producto, almacen, proyecto.id) == 25
        assert _bucket(db, producto, almacen, None) == 75

    def test_las_lineas_con_error_se_omiten_sin_abortar(
        self, client, db, admin, almacen, proyecto, producto,
    ):
        _login(client, admin)
        r = client.post(_url(proyecto, 'asignar'),
                        json={'lineas': [
                            {'sku': 'CEM-01', 'cantidad': 30},
                            {'sku': 'NO-EXISTE', 'cantidad': 5},
                        ]})
        assert r.status_code == 200
        assert r.get_json()['aplicadas'] == 1
        db.session.expire_all()
        assert _bucket(db, producto, almacen, proyecto.id) == 30


# ─── Devolver y mover ─────────────────────────────────────────────────────────

class TestDevolver:

    def _asignar(self, client, proyecto, cantidad):
        client.post(_url(proyecto, 'asignar'),
                    json={'lineas': [{'sku': 'CEM-01', 'cantidad': cantidad}]})

    def test_devolver_a_general(self, client, db, admin, almacen, proyecto, producto):
        _login(client, admin)
        self._asignar(client, proyecto, 60)
        r = client.post(_url(proyecto, 'devolver'),
                        json={'lineas': [{'producto_id': producto.id,
                                          'almacen_id': almacen.id, 'cantidad': 25}]})
        assert r.status_code == 200, r.get_json()
        assert r.get_json()['destino'] == 'General'
        db.session.expire_all()
        assert _bucket(db, producto, almacen, proyecto.id) == 35
        assert _bucket(db, producto, almacen, None) == 65

    def test_mover_a_otro_proyecto(self, client, db, admin, almacen, proyecto, producto):
        otro = Proyecto(numero_proyecto='AS-002', nombre='Otra obra', activo=True)
        db.session.add(otro); db.session.commit()
        _login(client, admin)
        self._asignar(client, proyecto, 50)
        r = client.post(_url(proyecto, 'devolver'),
                        json={'destino_proyecto_id': otro.id,
                              'lineas': [{'producto_id': producto.id,
                                          'almacen_id': almacen.id, 'cantidad': 20}]})
        assert r.status_code == 200, r.get_json()
        db.session.expire_all()
        assert _bucket(db, producto, almacen, proyecto.id) == 30
        assert _bucket(db, producto, almacen, otro.id) == 20

    def test_no_devolver_mas_de_lo_que_hay(self, client, db, admin, almacen, proyecto, producto):
        _login(client, admin)
        self._asignar(client, proyecto, 10)
        r = client.post(_url(proyecto, 'devolver'),
                        json={'lineas': [{'producto_id': producto.id,
                                          'almacen_id': almacen.id, 'cantidad': 999}]})
        assert r.get_json()['aplicadas'] == 0
        assert r.get_json()['problemas']
        db.session.expire_all()
        assert _bucket(db, producto, almacen, proyecto.id) == 10, 'no debió tocarse'

    def test_destino_igual_al_origen_422(self, client, admin, almacen, proyecto, producto):
        _login(client, admin)
        r = client.post(_url(proyecto, 'devolver'),
                        json={'destino_proyecto_id': proyecto.id,
                              'lineas': [{'producto_id': producto.id,
                                          'almacen_id': almacen.id, 'cantidad': 1}]})
        assert r.status_code == 422


# ─── Validaciones y permisos ──────────────────────────────────────────────────

class TestValidaciones:

    def test_sin_lineas_422(self, client, admin, proyecto):
        _login(client, admin)
        assert client.post(_url(proyecto, 'asignar'), json={'lineas': []}).status_code == 422

    def test_origen_invalido_422(self, client, admin, proyecto):
        _login(client, admin)
        r = client.post(_url(proyecto, 'asignar'),
                        json={'origen': 'inventado', 'lineas': [{'sku': 'X', 'cantidad': 1}]})
        assert r.status_code == 422

    def test_modo_invalido_422(self, client, admin, proyecto):
        _login(client, admin)
        r = client.post(_url(proyecto, 'asignar'),
                        json={'modo': 'inventado', 'lineas': [{'sku': 'X', 'cantidad': 1}]})
        assert r.status_code == 422

    def test_proyecto_inexistente_404(self, client, admin):
        _login(client, admin)
        r = client.post('/api/v1/proyectos-materiales/999999/asignar',
                        json={'lineas': [{'sku': 'X', 'cantidad': 1}]})
        assert r.status_code == 404

    def test_tope_de_lineas(self, client, admin, proyecto):
        """Tope anti-DoS: un Excel enorme no debe poder tumbar la transacción."""
        _login(client, admin)
        r = client.post(_url(proyecto, 'asignar'),
                        json={'lineas': [{'sku': 'X', 'cantidad': 1}] * 2001})
        assert r.status_code == 422

    def test_coordinador_no_puede_asignar(self, client, db, proyecto):
        u = User(username='as_coord', password_hash=generate_password_hash('Pass123!'),
                 role='coordinador')
        db.session.add(u); db.session.commit()
        _login(client, u)
        r = client.post(_url(proyecto, 'asignar'),
                        json={'lineas': [{'sku': 'X', 'cantidad': 1}]})
        assert r.status_code == 403


# ─── Resumen y plantilla ──────────────────────────────────────────────────────

class TestResumenYPlantilla:

    def test_general_va_primero(self, client, admin, almacen, proyecto, producto):
        """General no es un proyecto más: es el stock libre del que sale casi
        toda asignación, y el punto de referencia contra el que se leen los
        demás. Va primero siempre."""
        _login(client, admin)
        r = client.get('/api/v1/proyectos-materiales/resumen-asignacion')
        assert r.status_code == 200, r.get_json()
        primera = r.get_json()['tarjetas'][0]
        assert primera['es_general'] is True
        assert primera['proyecto_id'] is None

    def test_cuenta_lo_apartado_por_proyecto(self, client, admin, almacen, proyecto, producto):
        _login(client, admin)
        client.post(_url(proyecto, 'asignar'),
                    json={'lineas': [{'sku': 'CEM-01', 'cantidad': 30}]})

        tarjetas = client.get('/api/v1/proyectos-materiales/resumen-asignacion').get_json()
        general = tarjetas['tarjetas'][0]
        obra = [t for t in tarjetas['tarjetas'] if t['proyecto_id'] == proyecto.id][0]

        assert obra['unidades'] == 30
        assert obra['materiales'] == 1
        assert obra['valor'] == 1500          # 30 × 50
        assert general['unidades'] == 70      # lo que quedó libre

    def test_incluye_proyectos_sin_material(self, client, db, admin, proyecto):
        """Ver que una obra no tiene nada apartado ES información, y además es
        el punto de partida para asignarle."""
        _login(client, admin)
        tarjetas = client.get('/api/v1/proyectos-materiales/resumen-asignacion').get_json()
        ids = {t['proyecto_id'] for t in tarjetas['tarjetas']}
        assert proyecto.id in ids
        obra = [t for t in tarjetas['tarjetas'] if t['proyecto_id'] == proyecto.id][0]
        assert obra['unidades'] == 0

    def test_plantilla_tiene_solo_tres_columnas(self, client, admin, almacen, proyecto, producto):
        """La de catálogo pide trece porque sirve para dar de alta materiales.
        Aquí el material ya existe: solo varía cuánto y dónde."""
        import io
        import openpyxl
        _login(client, admin)
        r = client.get(_url(proyecto, 'plantilla-asignacion'))
        assert r.status_code == 200
        ws = openpyxl.load_workbook(io.BytesIO(r.data)).active
        encabezados = [c.value for c in ws[3] if c.value]
        assert encabezados == ['SKU', 'Cantidad', 'Bodega']
        assert 'Proyecto' not in encabezados, 'el proyecto es contexto, no columna'

    def test_plantilla_prellena_sku_pero_no_cantidad(
        self, client, admin, almacen, proyecto, producto,
    ):
        """Prellenar la cantidad invitaría a subir el archivo sin tocarlo y
        duplicar todo el material del proyecto. Lo tedioso se da hecho; lo que
        decide el resultado se escribe a conciencia."""
        import io
        import openpyxl
        _login(client, admin)
        client.post(_url(proyecto, 'asignar'),
                    json={'lineas': [{'sku': 'CEM-01', 'cantidad': 40}]})

        r = client.get(_url(proyecto, 'plantilla-asignacion'))
        ws = openpyxl.load_workbook(io.BytesIO(r.data)).active
        assert ws.cell(row=4, column=1).value == 'CEM-01'    # SKU sí
        assert ws.cell(row=4, column=2).value is None        # cantidad NO
        assert ws.cell(row=4, column=3).value == 'Central'   # bodega sí

    def test_plantilla_de_proyecto_inexistente_404(self, client, admin):
        _login(client, admin)
        r = client.get('/api/v1/proyectos-materiales/999999/plantilla-asignacion')
        assert r.status_code == 404


# ─── Importación del Excel ────────────────────────────────────────────────────

def _excel(filas, encabezados=('SKU', 'Cantidad', 'Bodega'), fila_encabezado=3):
    """Arma un .xlsx en memoria como el que subiría el usuario."""
    import io
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, h in enumerate(encabezados, 1):
        ws.cell(row=fila_encabezado, column=i, value=h)
    for r, fila in enumerate(filas, fila_encabezado + 1):
        for c, v in enumerate(fila, 1):
            ws.cell(row=r, column=c, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class TestImportarExcel:

    def _subir(self, client, proyecto, buf, **form):
        return client.post(
            _url(proyecto, 'asignar/importar'),
            data={'archivo': (buf, 'asignacion.xlsx'), **form},
            content_type='multipart/form-data',
        )

    def test_no_escribe_nada(self, client, db, admin, almacen, proyecto, producto):
        """Subir el archivo es previsualizar, no aplicar. Si escribiera, el
        usuario perdería el único momento en que puede echarse atrás."""
        _login(client, admin)
        r = self._subir(client, proyecto, _excel([('CEM-01', 25, 'Central')]))
        assert r.status_code == 200, r.get_json()
        assert r.get_json()['resumen']['ok'] == 1

        db.session.expire_all()
        assert _bucket(db, producto, almacen, proyecto.id) == 0
        assert _bucket(db, producto, almacen, None) == 100

    def test_el_archivo_y_la_captura_dan_lo_mismo(
        self, client, admin, almacen, proyecto, producto,
    ):
        """Subir un Excel no es un camino distinto al de capturar a mano: es otra
        forma de llenar las mismas líneas. Si divergieran, habría dos reglas de
        negocio que mantener sincronizadas a mano."""
        _login(client, admin)
        del_archivo = self._subir(
            client, proyecto, _excel([('CEM-01', 25, 'Central')]),
        ).get_json()
        a_mano = client.post(
            _url(proyecto, 'asignar/previsualizar'),
            json={'lineas': [{'sku': 'CEM-01', 'cantidad': 25, 'almacen': 'Central'}]},
        ).get_json()

        assert del_archivo['lineas'] == a_mano['lineas']
        assert del_archivo['resumen'] == a_mano['resumen']

    def test_las_filas_en_blanco_no_son_errores(
        self, client, admin, almacen, proyecto, producto,
    ):
        """La plantilla trae 200 filas vacías a propósito, y los SKU del proyecto
        ya escritos. Reportar ambas cosas como errores llenaría la vista previa
        de ruido y escondería los problemas de verdad.

        (Las vacías del FINAL ni siquiera llegan: pandas las recorta al leer. Se
        prueban las de en medio, que son las que sí hay que saltar.)"""
        _login(client, admin)
        r = self._subir(client, proyecto, _excel([
            ('CEM-01', 25, 'Central'),
            (None, None, None),            # fila en blanco intercalada
            ('CEM-01', None, 'Central'),   # prellenada, sin cantidad: no se toca
        ]))
        res = r.get_json()
        assert res['resumen']['errores'] == 0
        assert len(res['lineas']) == 1
        assert res['filas_ignoradas'] == 2

    def test_sku_prellenado_sin_cantidad_se_ignora(
        self, client, admin, almacen, proyecto, producto,
    ):
        """La plantilla trae los SKU del proyecto ya escritos. Dejar uno sin
        cantidad significa «este no lo toco», no «error»."""
        _login(client, admin)
        r = self._subir(client, proyecto, _excel([('CEM-01', None, 'Central')]))
        assert r.status_code == 422
        assert 'Cantidad' in r.get_json()['detail']

    def test_encuentra_los_encabezados_aunque_se_muevan(
        self, client, admin, almacen, proyecto, producto,
    ):
        """Se buscan por contenido, no por número de fila: si alguien inserta o
        borra una fila arriba, el archivo sigue sirviendo."""
        _login(client, admin)
        r = self._subir(client, proyecto,
                        _excel([('CEM-01', 10, 'Central')], fila_encabezado=1))
        assert r.status_code == 200
        assert r.get_json()['resumen']['ok'] == 1

    def test_acepta_codigo_como_alias_de_sku(
        self, client, admin, almacen, proyecto, producto,
    ):
        _login(client, admin)
        r = self._subir(client, proyecto,
                        _excel([('CEM-01', 10, 'Central')],
                               encabezados=('Código', 'Cantidad', 'Almacén')))
        assert r.status_code == 200
        assert r.get_json()['resumen']['ok'] == 1

    def test_sin_encabezados_reconocibles_da_error_claro(
        self, client, admin, almacen, proyecto, producto,
    ):
        _login(client, admin)
        r = self._subir(client, proyecto,
                        _excel([('CEM-01', 10, 'Central')],
                               encabezados=('Cosa', 'Otra', 'Más')))
        assert r.status_code == 400
        assert 'plantilla' in r.get_json()['detail'].lower()

    def test_rechaza_lo_que_no_es_excel(self, client, admin, proyecto):
        import io
        _login(client, admin)
        r = self._subir(client, proyecto, io.BytesIO(b'x' * 200))
        assert r.status_code == 400

    def test_la_vuelta_completa(self, client, db, admin, almacen, proyecto, producto):
        """Bajar la plantilla, llenarla, subirla y aplicar: el recorrido real."""
        import io
        import openpyxl
        _login(client, admin)

        # 1. Se baja la plantilla del proyecto.
        plantilla = client.get(_url(proyecto, 'plantilla-asignacion'))
        wb = openpyxl.load_workbook(io.BytesIO(plantilla.data))
        ws = wb.active

        # 2. Se escribe SKU y cantidad, como haría el usuario.
        ws.cell(row=4, column=1, value='CEM-01')
        ws.cell(row=4, column=2, value=60)
        ws.cell(row=4, column=3, value='Central')
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        # 3. Se sube: previsualización.
        prev = self._subir(client, proyecto, buf).get_json()
        assert prev['resumen']['ok'] == 1
        assert prev['lineas'][0]['resultado'] == 60

        # 4. Se confirma con el mismo /asignar de la captura a mano.
        r = client.post(_url(proyecto, 'asignar'),
                        json={'lineas': [{'sku': f['sku'], 'cantidad': f['cantidad_aplicada'],
                                          'almacen_id': f['almacen_id']}
                                         for f in prev['lineas'] if f['estado'] != 'error']})
        assert r.status_code == 200, r.get_json()

        db.session.expire_all()
        assert _bucket(db, producto, almacen, proyecto.id) == 60
        assert _bucket(db, producto, almacen, None) == 40


class TestDisponibleEnGeneral:

    def test_lo_reporta_para_capturar_sin_adivinar(
        self, client, admin, almacen, proyecto, producto,
    ):
        """La interfaz muestra el disponible ANTES de que se escriba la
        cantidad. Sin ese dato se captura a ciegas."""
        _login(client, admin)
        r = client.post(_url(proyecto, 'asignar/previsualizar'),
                        json={'lineas': [{'sku': 'CEM-01', 'cantidad': 10}]})
        assert r.get_json()['lineas'][0]['disponible'] == 100

    def test_en_entrada_nueva_no_aplica(self, client, admin, almacen, proyecto, producto):
        """El material llega de fuera: no hay tope contra el que capturar, y
        mostrar el de General ahí sería un dato engañoso."""
        _login(client, admin)
        r = client.post(_url(proyecto, 'asignar/previsualizar'),
                        json={'lineas': [{'sku': 'CEM-01', 'cantidad': 10}],
                              'origen': 'entrada'})
        assert r.get_json()['lineas'][0]['disponible'] is None


class TestConfirmarLaVistaPrevia:
    """El SPA confirma reenviando `cantidad_pedida`, no `cantidad_aplicada`.

    En modo reemplazar la cantidad es el OBJETIVO, no el incremento: reenviar el
    delta haría que el backend lo tomara por objetivo y asignara de menos. Estos
    tests fijan que reenviar lo PEDIDO da el resultado que prometió la vista
    previa en los dos modos.
    """

    def _confirmar(self, client, proyecto, prev, **extra):
        return client.post(_url(proyecto, 'asignar'), json={
            'lineas': [{'producto_id': f['producto_id'], 'almacen_id': f['almacen_id'],
                        'cantidad': f['cantidad_pedida']}
                       for f in prev['lineas'] if f['estado'] != 'error'],
            **extra,
        })

    def test_en_modo_sumar(self, client, db, admin, almacen, proyecto, producto):
        _login(client, admin)
        cuerpo = {'lineas': [{'sku': 'CEM-01', 'cantidad': 30}]}
        prev = client.post(_url(proyecto, 'asignar/previsualizar'), json=cuerpo).get_json()
        self._confirmar(client, proyecto, prev)

        db.session.expire_all()
        assert _bucket(db, producto, almacen, proyecto.id) == prev['lineas'][0]['resultado']

    def test_en_modo_reemplazar(self, client, db, admin, almacen, proyecto, producto):
        _login(client, admin)
        # Punto de partida: el proyecto ya tiene 20.
        client.post(_url(proyecto, 'asignar'),
                    json={'lineas': [{'sku': 'CEM-01', 'cantidad': 20}]})

        cuerpo = {'lineas': [{'sku': 'CEM-01', 'cantidad': 60}], 'modo': 'reemplazar'}
        prev = client.post(_url(proyecto, 'asignar/previsualizar'), json=cuerpo).get_json()
        assert prev['lineas'][0]['cantidad_aplicada'] == 40   # el delta
        assert prev['lineas'][0]['resultado'] == 60           # el objetivo

        self._confirmar(client, proyecto, prev, modo='reemplazar')
        db.session.expire_all()
        # Si se hubiera reenviado el delta (40) como objetivo, aquí habría 40.
        assert _bucket(db, producto, almacen, proyecto.id) == 60

    def test_en_modo_reemplazar_bajando(self, client, db, admin, almacen, proyecto, producto):
        """Reemplazar con menos devuelve el sobrante a General."""
        _login(client, admin)
        client.post(_url(proyecto, 'asignar'),
                    json={'lineas': [{'sku': 'CEM-01', 'cantidad': 50}]})

        cuerpo = {'lineas': [{'sku': 'CEM-01', 'cantidad': 20}], 'modo': 'reemplazar'}
        prev = client.post(_url(proyecto, 'asignar/previsualizar'), json=cuerpo).get_json()
        self._confirmar(client, proyecto, prev, modo='reemplazar')

        db.session.expire_all()
        assert _bucket(db, producto, almacen, proyecto.id) == 20
        assert _bucket(db, producto, almacen, None) == 80


# ─── Stock libre (General) ────────────────────────────────────────────────────

class TestExistenciasGenerales:

    URL = '/api/v1/proyectos-materiales/general/existencias'

    def test_lista_lo_que_no_esta_apartado(self, client, admin, almacen, proyecto, producto):
        _login(client, admin)
        r = client.get(self.URL)
        assert r.status_code == 200, r.get_json()
        d = r.get_json()
        assert d['total'] == 1
        m = d['materiales'][0]
        assert m['codigo'] == 'CEM-01'
        assert m['total'] == 100
        assert m['por_almacen'] == {str(almacen.id): 100}

    def test_lo_apartado_deja_de_ser_libre(self, client, admin, almacen, proyecto, producto):
        """Es el punto entero de la pantalla: lo que ya tiene dueño no se puede
        volver a repartir."""
        _login(client, admin)
        client.post(_url(proyecto, 'asignar'),
                    json={'lineas': [{'sku': 'CEM-01', 'cantidad': 40}]})

        m = client.get(self.URL).get_json()['materiales'][0]
        assert m['total'] == 60

    def test_desaparece_al_apartarse_entero(self, client, admin, almacen, proyecto, producto):
        _login(client, admin)
        client.post(_url(proyecto, 'asignar'),
                    json={'lineas': [{'sku': 'CEM-01', 'cantidad': 100}]})

        d = client.get(self.URL).get_json()
        assert d['total'] == 0
        assert d['materiales'] == []

    def test_busca_por_codigo_y_por_descripcion(self, client, admin, almacen, proyecto, producto):
        _login(client, admin)
        assert client.get(self.URL, query_string={'q': 'CEM'}).get_json()['total'] == 1
        assert client.get(self.URL, query_string={'q': 'cemento'}).get_json()['total'] == 1
        assert client.get(self.URL, query_string={'q': 'zzz'}).get_json()['total'] == 0

    def test_pagina_por_material_no_por_bucket(self, client, db, admin, almacen, proyecto):
        """Un material repartido en dos bodegas es UNA fila de la tabla. Si se
        paginara por bucket, ese material saldría partido entre dos páginas."""
        _login(client, admin)
        otra = Almacen(nombre='Norte', qr_code='QR-AS-2', activo=True)
        db.session.add(otra); db.session.flush()
        p = Producto(codigo='VAR-01', descripcion='Varilla', categoria='Obra',
                     unidad='pz', stock_actual=30, stock_minimo=0, precio_unitario=10)
        db.session.add(p); db.session.flush()
        db.session.add(StockAlmacenProyecto(producto_id=p.id, almacen_id=almacen.id,
                                            proyecto_id=None, cantidad=20))
        db.session.add(StockAlmacenProyecto(producto_id=p.id, almacen_id=otra.id,
                                            proyecto_id=None, cantidad=10))
        db.session.commit()

        d = client.get(self.URL).get_json()
        assert d['total'] == 1, 'dos buckets del mismo material son una sola fila'
        m = d['materiales'][0]
        assert m['total'] == 30
        assert set(m['por_almacen'].values()) == {20.0, 10.0}
        assert len(d['almacenes']) == 2

    def test_ordena_por_cantidad_descendente(self, client, db, admin, almacen, proyecto, producto):
        """Lo que más sobra es lo primero que conviene repartir."""
        _login(client, admin)
        p = Producto(codigo='ARE-01', descripcion='Arena', categoria='Obra', unidad='m3',
                     stock_actual=500, stock_minimo=0, precio_unitario=1)
        db.session.add(p); db.session.flush()
        db.session.add(StockAlmacenProyecto(producto_id=p.id, almacen_id=almacen.id,
                                            proyecto_id=None, cantidad=500))
        db.session.commit()

        codigos = [m['codigo'] for m in client.get(self.URL).get_json()['materiales']]
        assert codigos == ['ARE-01', 'CEM-01']

    def test_respeta_el_tope_por_pagina(self, client, admin, almacen, proyecto, producto):
        """Sin tope, General bajaría el catálogo casi entero en cada visita."""
        _login(client, admin)
        d = client.get(self.URL, query_string={'per_page': 9999}).get_json()
        assert d['per_page'] == 200
