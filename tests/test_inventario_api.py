"""
Tests del blueprint `inventario_api` (Flask, JWT API-only).

Cobertura: auth, autorización por rol, CRUD, lógica de negocio (stock,
ajustes, traspasos), solicitudes, auditoría, edge cases.

Notas:
  - CSRF está desactivado vía conftest (WTF_CSRF_ENABLED=False).
  - Rate limiter está desactivado vía conftest (RATELIMIT_ENABLED=False).
  - Auth: JWT real en `Authorization: Bearer …` — el helper `_login` setea
    `environ_base['HTTP_AUTHORIZATION']` así que todas las requests siguientes
    del mismo client ya van firmadas.
  - Códigos esperados: 400 (regla de negocio), 422 (validación), 404 (no existe),
    403 (sin rol), 401 (sin token), 204 (DELETE OK), 200 (resto OK).
"""
import uuid
import pytest
from werkzeug.security import generate_password_hash

from app.extensions import db as flask_db
from app.models import (
    User, Almacen, Estante, Producto, StockPorAlmacen, StockAlmacenProyecto,
    MovimientoInventario, Proyecto, Trabajador,
    SolicitudMaterial, SolicitudMaterialDetalle, AuditLog,
    TomaInventario,
)
from app.routes.api_auth import _encode_access_token


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _login(client, user_id: int, role: str = None, username: str = None):
    """Firma el client con un JWT real del usuario indicado.

    Compat con el flujo antiguo de sesión: se llama igual (`_login(client, u.id, 'role')`)
    pero ahora emite un access token y lo deja en `environ_base`, de modo que TODAS
    las requests subsiguientes del mismo `client` viajan con `Authorization: Bearer …`.

    `role` y `username` se ignoran — el JWT toma el rol real del User en BD; los
    parámetros se conservan solo para no romper las llamadas existentes.
    """
    user = flask_db.session.get(User, user_id)
    assert user is not None, f"_login: user_id={user_id} no existe en BD"
    token = _encode_access_token(user)
    client.environ_base['HTTP_AUTHORIZATION'] = f'Bearer {token}'


def _logout(client):
    client.environ_base.pop('HTTP_AUTHORIZATION', None)


# ─── Fixtures locales ─────────────────────────────────────────────────────────
# Se aprovecha conftest.py: `client`, `db`, `app`. Cada test corre dentro de una
# transacción aislada (rollback automático).

@pytest.fixture
def inv_admin(db):
    u = User(username='inv_admin', password_hash=generate_password_hash('Pass123!'), role='admin')
    db.session.add(u)
    db.session.commit()
    return u


@pytest.fixture
def inv_user(db):
    u = User(username='inv_user', password_hash=generate_password_hash('Pass123!'), role='inventario')
    db.session.add(u)
    db.session.commit()
    return u


@pytest.fixture
def inv_solicitante(db):
    u = User(username='inv_sol', password_hash=generate_password_hash('Pass123!'), role='solicitante_material')
    db.session.add(u)
    db.session.commit()
    return u


@pytest.fixture
def inv_outsider(db):
    """Usuario sin rol de inventario (coordinador sí puede leer inventario
    desde 05-25 — usamos `visitor` que no figura en ninguna whitelist)."""
    u = User(username='inv_out', password_hash=generate_password_hash('Pass123!'), role='visitor')
    db.session.add(u)
    db.session.commit()
    return u


# ═══════════════════════════════════════════════════════════════════════════════
# 1. HEALTH CHECK
# ═══════════════════════════════════════════════════════════════════════════════

class TestHealth:
    def test_health_ok(self, client):
        resp = client.get('/api/v1/health')
        assert resp.status_code == 200
        assert resp.get_json() == {'status': 'ok'}


# ═══════════════════════════════════════════════════════════════════════════════
# 2. AUTENTICACIÓN Y AUTORIZACIÓN
# ═══════════════════════════════════════════════════════════════════════════════

class TestAutenticacion:

    def test_sin_sesion_retorna_401(self, client):
        resp = client.get('/api/v1/productos/')
        assert resp.status_code == 401

    def test_rol_sin_permiso_retorna_403(self, client, inv_outsider):
        _login(client, inv_outsider.id, 'coordinador')
        resp = client.get('/api/v1/productos/')
        assert resp.status_code == 403

    def test_admin_puede_acceder(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.get('/api/v1/productos/')
        assert resp.status_code == 200

    def test_inventario_puede_leer(self, client, inv_user):
        _login(client, inv_user.id, 'inventario')
        resp = client.get('/api/v1/productos/')
        assert resp.status_code == 200

    def test_solicitante_puede_leer_productos(self, client, inv_solicitante):
        _login(client, inv_solicitante.id, 'solicitante_material')
        resp = client.get('/api/v1/productos/')
        assert resp.status_code == 200

    def test_solicitante_no_puede_crear_producto(self, client, inv_solicitante):
        _login(client, inv_solicitante.id, 'solicitante_material')
        resp = client.post('/api/v1/productos/', json={
            'codigo': 'HACK-001', 'descripcion': 'X', 'categoria': 'T', 'unidad': 'pza',
            'stock_actual': 0, 'stock_minimo': 0,
        })
        assert resp.status_code == 403

    def test_solicitante_no_puede_borrar_almacen(self, client, inv_solicitante):
        _login(client, inv_solicitante.id, 'solicitante_material')
        resp = client.delete('/api/v1/almacenes/1')
        assert resp.status_code == 403


# ═══════════════════════════════════════════════════════════════════════════════
# 3. PRODUCTOS — CRUD
# ═══════════════════════════════════════════════════════════════════════════════

class TestProductos:

    def test_crear_producto_valido(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/productos/', json={
            'codigo': 'PROD-001', 'descripcion': 'Test', 'categoria': 'Herramientas',
            'unidad': 'pza', 'stock_actual': 50.0, 'stock_minimo': 5.0,
        })
        assert resp.status_code == 200
        data = resp.get_json()
        assert data['codigo'] == 'PROD-001'
        assert data['stock_actual'] == 50.0
        assert data['activo'] is True

    def test_crear_producto_codigo_duplicado(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        db.session.add(Producto(codigo='DUP-001', descripcion='Existente',
                                categoria='T', unidad='pza', stock_actual=0, stock_minimo=0))
        db.session.commit()
        resp = client.post('/api/v1/productos/', json={
            'codigo': 'DUP-001', 'descripcion': 'Duplicado', 'categoria': 'T', 'unidad': 'pza',
            'stock_actual': 0, 'stock_minimo': 0,
        })
        assert resp.status_code == 400
        assert 'ya existe' in resp.get_json()['detail'].lower()

    def test_crear_producto_codigo_invalido(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/productos/', json={
            'codigo': 'PROD TEST 001!', 'descripcion': 'X', 'categoria': 'T', 'unidad': 'pza',
            'stock_actual': 0, 'stock_minimo': 0,
        })
        assert resp.status_code == 422

    def test_crear_producto_stock_negativo(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/productos/', json={
            'codigo': 'NEG-001', 'descripcion': 'X', 'categoria': 'T', 'unidad': 'pza',
            'stock_actual': -10.0, 'stock_minimo': 0,
        })
        assert resp.status_code == 422

    def test_crear_producto_sin_campo_requerido(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/productos/', json={
            'codigo': 'MISS-001', 'categoria': 'T', 'unidad': 'pza',
        })
        assert resp.status_code == 422

    def test_crear_producto_codigo_demasiado_largo(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/productos/', json={
            'codigo': 'A' * 51, 'descripcion': 'X', 'categoria': 'T', 'unidad': 'pza',
            'stock_actual': 0, 'stock_minimo': 0,
        })
        assert resp.status_code == 422

    def test_listar_productos(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        db.session.add(Producto(codigo='LIST-001', descripcion='Test', categoria='T',
                                unidad='pza', stock_actual=0, stock_minimo=0))
        db.session.commit()
        resp = client.get('/api/v1/productos/')
        assert resp.status_code == 200
        codigos = [p['codigo'] for p in resp.get_json()]
        assert 'LIST-001' in codigos

    def test_listar_paginacion(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.get('/api/v1/productos/?skip=0&limit=1')
        assert resp.status_code == 200
        assert len(resp.get_json()) <= 1

    def test_listar_paginado_por_paginas(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        # Categoría propia para aislar el conteo de otros productos del fixture.
        for i in range(3):
            db.session.add(Producto(codigo=f'PG-{i}', descripcion=f'P{i}', categoria='PAGINADO',
                                    unidad='pza', stock_actual=0, stock_minimo=0))
        db.session.commit()

        r1 = client.get('/api/v1/productos/paginado?categoria=PAGINADO&per_page=2&page=1')
        assert r1.status_code == 200
        d1 = r1.get_json()
        assert d1['total'] == 3
        assert d1['pages'] == 2
        assert d1['page'] == 1
        assert d1['per_page'] == 2
        assert len(d1['items']) == 2

        r2 = client.get('/api/v1/productos/paginado?categoria=PAGINADO&per_page=2&page=2')
        d2 = r2.get_json()
        assert len(d2['items']) == 1
        # Sin traslape entre páginas (orden determinista por id).
        ids1 = {p['id'] for p in d1['items']}
        ids2 = {p['id'] for p in d2['items']}
        assert ids1.isdisjoint(ids2)

    # ── Filtros avanzados del catálogo ──────────────────────────────────────────

    def test_filtro_stock_bajo(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        db.session.add_all([
            Producto(codigo='SB-LOW', descripcion='Bajo', categoria='F', unidad='pza', stock_actual=2, stock_minimo=5),
            Producto(codigo='SB-OK', descripcion='Ok', categoria='F', unidad='pza', stock_actual=50, stock_minimo=5),
        ])
        db.session.commit()
        resp = client.get('/api/v1/productos/?stock=bajo')
        assert resp.status_code == 200
        codigos = {p['codigo'] for p in resp.get_json()}
        assert 'SB-LOW' in codigos
        assert 'SB-OK' not in codigos

    def test_filtro_imagen_sin(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        db.session.add_all([
            Producto(codigo='IMG-CON', descripcion='Con', categoria='F', unidad='pza',
                     stock_actual=1, stock_minimo=0, imagen_url='http://x/y.jpg'),
            Producto(codigo='IMG-SIN', descripcion='Sin', categoria='F', unidad='pza',
                     stock_actual=1, stock_minimo=0),
        ])
        db.session.commit()
        resp = client.get('/api/v1/productos/?imagen=sin')
        assert resp.status_code == 200
        codigos = {p['codigo'] for p in resp.get_json()}
        assert 'IMG-SIN' in codigos
        assert 'IMG-CON' not in codigos

    def test_filtro_unidad_y_endpoint_unidades(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        db.session.add_all([
            Producto(codigo='U-KG', descripcion='Kilo', categoria='F', unidad='kg', stock_actual=1, stock_minimo=0),
            Producto(codigo='U-PZA', descripcion='Pieza', categoria='F', unidad='pza', stock_actual=1, stock_minimo=0),
        ])
        db.session.commit()
        ru = client.get('/api/v1/productos/unidades/')
        assert ru.status_code == 200
        assert 'kg' in ru.get_json()
        resp = client.get('/api/v1/productos/?unidad=kg')
        assert resp.status_code == 200
        codigos = {p['codigo'] for p in resp.get_json()}
        assert codigos == {'U-KG'}

    def test_filtro_compra_activa(self, client, inv_admin, db):
        from app.models import SolicitudCompra, SolicitudCompraDetalle
        _login(client, inv_admin.id, 'admin')
        p_en = Producto(codigo='C-EN', descripcion='EnCompra', categoria='F', unidad='pza', stock_actual=1, stock_minimo=0)
        p_no = Producto(codigo='C-NO', descripcion='SinCompra', categoria='F', unidad='pza', stock_actual=1, stock_minimo=0)
        db.session.add_all([p_en, p_no]); db.session.flush()
        sc = SolicitudCompra(solicitado_por_id=inv_admin.id, estatus='PENDIENTE')
        db.session.add(sc); db.session.flush()
        db.session.add(SolicitudCompraDetalle(
            solicitud_compra_id=sc.id, producto_id=p_en.id, cantidad_solicitada=10,
        ))
        db.session.commit()
        resp = client.get('/api/v1/productos/?compra=activa')
        assert resp.status_code == 200
        codigos = {p['codigo'] for p in resp.get_json()}
        assert 'C-EN' in codigos
        assert 'C-NO' not in codigos

    def test_actualizar_producto(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = Producto(codigo='UPD-001', descripcion='Orig', categoria='T',
                     unidad='pza', stock_actual=0, stock_minimo=0)
        db.session.add(p); db.session.commit()
        resp = client.put(f'/api/v1/productos/{p.id}', json={'descripcion': 'Actualizada'})
        assert resp.status_code == 200
        assert resp.get_json()['descripcion'] == 'Actualizada'

    def test_actualizar_producto_inexistente(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.put('/api/v1/productos/999999', json={'descripcion': 'x'})
        assert resp.status_code == 404

    def test_actualizar_codigo_a_uno_duplicado(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        a = Producto(codigo='A-001', descripcion='A', categoria='T', unidad='pza',
                     stock_actual=0, stock_minimo=0)
        b = Producto(codigo='B-001', descripcion='B', categoria='T', unidad='pza',
                     stock_actual=0, stock_minimo=0)
        db.session.add_all([a, b]); db.session.commit()
        resp = client.put(f'/api/v1/productos/{b.id}', json={'codigo': 'A-001'})
        assert resp.status_code == 400

    def test_borrar_producto_soft_delete(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = Producto(codigo='DEL-001', descripcion='X', categoria='T', unidad='pza',
                     stock_actual=0, stock_minimo=0)
        db.session.add(p); db.session.commit()
        pid = p.id
        resp = client.delete(f'/api/v1/productos/{pid}')
        assert resp.status_code == 204
        p_db = db.session.get(Producto, pid)
        assert p_db.activo is False

    def test_borrar_producto_inexistente(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.delete('/api/v1/productos/999999')
        assert resp.status_code == 404

    def test_stock_minimo_supera_maximo(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/productos/', json={
            'codigo': 'STOCK-OVER', 'descripcion': 'X', 'categoria': 'T', 'unidad': 'pza',
            'stock_actual': 0, 'stock_minimo': 2_000_000,
        })
        assert resp.status_code == 422


# ═══════════════════════════════════════════════════════════════════════════════
# 3-bis. PRODUCTOS DE CABLE — Tipo + Tamaño obligatorios y unidad forzada a M
# ═══════════════════════════════════════════════════════════════════════════════

class TestProductosCable:

    def test_crear_cable_sin_tipo_tamano_falla(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/productos/', json={
            'codigo': 'CAB-001', 'descripcion': 'Cable THHN', 'categoria': 'Cable',
            'unidad': 'M', 'stock_actual': 100, 'stock_minimo': 10,
        })
        assert resp.status_code == 422
        assert 'cable' in resp.get_json()['detail'].lower()

    def test_crear_cable_completo_fuerza_unidad_m(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        # Aunque mandemos 'rollo', el backend debe forzar la unidad a 'M'.
        resp = client.post('/api/v1/productos/', json={
            'codigo': 'CAB-002', 'descripcion': 'Cable THHN cal 12', 'categoria': 'Cables',
            'unidad': 'rollo', 'stock_actual': 250.5, 'stock_minimo': 20,
            'cable_tipo': 'THHN', 'cable_calibre': '12',
        })
        assert resp.status_code == 200, resp.get_json()
        data = resp.get_json()
        assert data['unidad'] == 'M'
        assert data['cable_tipo'] == 'THHN'
        assert data['cable_calibre'] == '12'
        # unidad M admite decimales → el stock 250.5 se conserva.
        assert data['stock_actual'] == 250.5

    def test_crear_cable_detecta_categoria_con_acentos_y_espacios(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/productos/', json={
            'codigo': 'CAB-003', 'descripcion': 'Cable desnudo', 'categoria': 'Cablería Eléctrica',
            'unidad': 'pza', 'stock_actual': 5, 'stock_minimo': 0,
            'cable_tipo': 'Desnudo', 'cable_calibre': '2/0',
        })
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['unidad'] == 'M'

    def test_producto_normal_no_exige_cable(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/productos/', json={
            'codigo': 'NOCAB-001', 'descripcion': 'Tornillo', 'categoria': 'Tornillería',
            'unidad': 'pza', 'stock_actual': 10, 'stock_minimo': 0,
        })
        assert resp.status_code == 200
        data = resp.get_json()
        assert data['cable_tipo'] is None
        assert data['cable_calibre'] is None

    def test_producto_normal_ignora_campos_cable(self, client, inv_admin):
        """Un producto no-cable no debe arrastrar cable_tipo/calibre aunque el
        cliente los mande (evita datos huérfanos)."""
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/productos/', json={
            'codigo': 'NOCAB-002', 'descripcion': 'Tuerca', 'categoria': 'Tuercas',
            'unidad': 'pza', 'stock_actual': 10, 'stock_minimo': 0,
            'cable_tipo': 'THHN', 'cable_calibre': '12',
        })
        assert resp.status_code == 200
        assert resp.get_json()['cable_tipo'] is None

    def test_update_a_cable_exige_tipo_tamano(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = Producto(codigo='UPCAB-001', descripcion='X', categoria='General',
                     unidad='pza', stock_actual=0, stock_minimo=0)
        db.session.add(p); db.session.commit()
        # Cambiar la categoría a cable sin dar Tipo/Tamaño → 422.
        resp = client.put(f'/api/v1/productos/{p.id}', json={'categoria': 'Cable'})
        assert resp.status_code == 422
        # Con los datos completos → OK y unidad forzada a M.
        resp2 = client.put(f'/api/v1/productos/{p.id}', json={
            'categoria': 'Cable', 'cable_tipo': 'THW', 'cable_calibre': '10',
        })
        assert resp2.status_code == 200, resp2.get_json()
        assert resp2.get_json()['unidad'] == 'M'
        assert resp2.get_json()['cable_calibre'] == '10'

    def test_crear_producto_guarda_el_precio(self, client, inv_admin, db):
        """El precio capturado al ALTA debe guardarse.

        Se perdía: el SPA lo mandaba y el schema lo aceptaba, pero el alta no lo
        asignaba y el producto quedaba en 0 hasta volver a editarlo — mientras
        que la importación por Excel sí lo guardaba. De ese precio salen los
        costos por proyecto, así que un 0 silencioso desalinea los reportes.
        """
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/productos/', json={
            'codigo': 'PRECIO-1', 'descripcion': 'Tornillo', 'categoria': 'Tornillería',
            'unidad': 'pza', 'stock_actual': 0, 'stock_minimo': 0,
            'precio_unitario': 37.25,
        })
        assert resp.status_code == 200, resp.get_json()
        assert float(resp.get_json()['precio_unitario']) == pytest.approx(37.25, abs=0.001)
        p = Producto.query.filter_by(codigo='PRECIO-1').first()
        assert float(p.precio_unitario) == pytest.approx(37.25, abs=0.001)

    def test_update_parcial_de_cable_conserva_datos(self, client, inv_admin, db):
        """Editar solo el precio de un cable NO debe borrar Tipo/Tamaño ni fallar
        por 'faltan datos de cable'."""
        _login(client, inv_admin.id, 'admin')
        p = Producto(codigo='UPCAB-002', descripcion='Cable', categoria='Cable',
                     unidad='M', cable_tipo='THHN', cable_calibre='14',
                     stock_actual=0, stock_minimo=0)
        db.session.add(p); db.session.commit()
        resp = client.put(f'/api/v1/productos/{p.id}', json={'precio_unitario': 99.5})
        assert resp.status_code == 200, resp.get_json()
        data = resp.get_json()
        assert data['cable_tipo'] == 'THHN'
        assert data['cable_calibre'] == '14'

    def test_solicitud_detalle_incluye_datos_cable(self, client, inv_solicitante, db):
        """El detalle de la solicitud expone cable_tipo/calibre del producto para
        que el carrito/solicitudes/PDF puedan mostrarlos."""
        from decimal import Decimal
        p = Producto(codigo='CAB-SOL', descripcion='Cable THHN', categoria='Cable',
                     unidad='M', cable_tipo='THHN', cable_calibre='12',
                     stock_actual=Decimal('500'), stock_minimo=0)
        db.session.add(p); db.session.commit()
        _login(client, inv_solicitante.id, 'solicitante_material')
        resp = client.post('/api/v1/solicitudes/', json={
            'proyecto': 'Cableado nave 1',
            'detalles': [{'producto_id': p.id, 'cantidad_solicitada': 30.5}],
        })
        assert resp.status_code == 200, resp.get_json()
        det = resp.get_json()['detalles'][0]
        assert det['cable_tipo'] == 'THHN'
        assert det['cable_calibre'] == '12'


# ═══════════════════════════════════════════════════════════════════════════════
# 3-ter. IMPORTACIÓN EXCEL — columnas de cable a prueba de tontos
# ═══════════════════════════════════════════════════════════════════════════════

def _build_import_xlsx(rows):
    """Construye un .xlsx en memoria con la fila de encabezados oficiales + las
    filas de datos dadas (dicts con llave = nombre de columna)."""
    import io as _io
    from openpyxl import Workbook
    HEADERS = [
        'Código (SKU)', 'Descripción', 'Categoría', 'Unidad',
        'Stock Inicial', 'Stock Mínimo', 'Precio Unitario', 'URL Imagen (opcional)',
        'Tipo (cable)', 'Tamaño mm²/AWG (cable)',
        # Feature stock por proyecto: destino del stock inicial por fila.
        'Almacén', 'Proyecto',
    ]
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for r in rows:
        ws.append([r.get(h, '') for h in HEADERS])
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


class TestImportarCable:

    def _post(self, client, rows):
        import io as _io
        data = {'archivo': (_io.BytesIO(_build_import_xlsx(rows)), 'materiales.xlsx')}
        return client.post('/api/v1/productos/importar', data=data,
                           content_type='multipart/form-data')

    def test_importar_cable_nuevo_fuerza_m(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        resp = self._post(client, [{
            'Código (SKU)': 'IMP-CAB-1', 'Descripción': 'Cable THHN', 'Categoría': 'Cable',
            'Unidad': 'rollo', 'Stock Inicial': 100, 'Stock Mínimo': 10,
            'Tipo (cable)': 'THHN', 'Tamaño mm²/AWG (cable)': '12',
        }])
        assert resp.status_code == 200, resp.get_json()
        body = resp.get_json()
        assert body['exitosos'] == 1 and body['errores'] == []
        p = Producto.query.filter_by(codigo='IMP-CAB-1').first()
        assert p is not None
        assert p.unidad == 'M'  # forzado aunque el Excel dijera "rollo"
        assert p.cable_tipo == 'THHN' and p.cable_calibre == '12'

    def test_importar_cable_sin_datos_es_error(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        resp = self._post(client, [{
            'Código (SKU)': 'IMP-CAB-2', 'Descripción': 'Cable X', 'Categoría': 'Cables',
            'Unidad': 'M', 'Stock Inicial': 5, 'Stock Mínimo': 0,
            # Falta Tipo y Tamaño → debe reportar error y NO crear el producto.
        }])
        assert resp.status_code == 200, resp.get_json()
        body = resp.get_json()
        assert body['exitosos'] == 0
        assert len(body['errores']) == 1 and 'cable' in body['errores'][0].lower()
        assert Producto.query.filter_by(codigo='IMP-CAB-2').first() is None

    def test_importar_rellena_cable_existente(self, client, inv_admin, db):
        """Un cable ya existente sin Tipo/Tamaño se completa al reimportar."""
        _login(client, inv_admin.id, 'admin')
        p = Producto(codigo='IMP-CAB-3', descripcion='Cable viejo', categoria='Cable',
                     unidad='M', cable_tipo=None, cable_calibre=None,
                     stock_actual=0, stock_minimo=0)
        db.session.add(p); db.session.commit()
        resp = self._post(client, [{
            'Código (SKU)': 'IMP-CAB-3', 'Descripción': 'Cable viejo', 'Categoría': 'Cable',
            'Unidad': 'M', 'Stock Inicial': 0, 'Stock Mínimo': 0,
            'Tipo (cable)': 'THW', 'Tamaño mm²/AWG (cable)': '10',
        }])
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['actualizados'] == 1
        db.session.refresh(p)
        assert p.cable_tipo == 'THW' and p.cable_calibre == '10'

    def test_importar_no_cable_ignora_columnas_cable(self, client, inv_admin, db):
        """Llenar Tipo/Tamaño en una categoría que no es cable NO debe guardarlos."""
        _login(client, inv_admin.id, 'admin')
        resp = self._post(client, [{
            'Código (SKU)': 'IMP-NOCAB-1', 'Descripción': 'Tornillo', 'Categoría': 'Tornillería',
            'Unidad': 'pza', 'Stock Inicial': 10, 'Stock Mínimo': 0,
            'Tipo (cable)': 'THHN', 'Tamaño mm²/AWG (cable)': '12',
        }])
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['exitosos'] == 1
        p = Producto.query.filter_by(codigo='IMP-NOCAB-1').first()
        assert p.cable_tipo is None and p.cable_calibre is None
        assert p.unidad == 'pza'  # no se fuerza a M en no-cable

    def test_exportar_catalogo_trae_productos(self, client, inv_admin, db):
        """El export debe traer los productos existentes ya llenos, con los
        encabezados oficiales (incl. cable)."""
        import io as _io
        from openpyxl import load_workbook
        from decimal import Decimal
        db.session.add_all([
            Producto(codigo='EXP-1', descripcion='Tornillo', categoria='Tornillería',
                     unidad='pza', stock_actual=Decimal('10'), stock_minimo=2, precio_unitario=Decimal('3.5')),
            Producto(codigo='EXP-CAB', descripcion='Cable', categoria='Cable', unidad='M',
                     cable_tipo='THHN', cable_calibre='12', stock_actual=Decimal('100'), stock_minimo=0),
        ])
        db.session.commit()
        _login(client, inv_admin.id, 'admin')
        resp = client.get('/api/v1/productos/exportar')
        assert resp.status_code == 200
        ws = load_workbook(_io.BytesIO(resp.data)).active
        # Encabezados en fila 4; datos desde fila 5.
        headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
        assert 'Tamaño mm²/AWG (cable)' in headers
        codigos = {ws.cell(row=r, column=1).value for r in range(5, ws.max_row + 1)}
        assert {'EXP-1', 'EXP-CAB'}.issubset(codigos)

    def test_roundtrip_export_reimport_sin_cambios_falsos(self, client, inv_admin, db):
        """Exportar el catálogo y reimportar ese MISMO archivo sin editar no debe
        reportar ningún cambio (sin falsos positivos por formato de números)."""
        import io as _io
        from decimal import Decimal
        db.session.add_all([
            Producto(codigo='RT-1', descripcion='Tornillo', categoria='Tornillería',
                     unidad='pza', stock_actual=Decimal('10'), stock_minimo=2, precio_unitario=Decimal('3.5')),
            Producto(codigo='RT-2', descripcion='Cable THHN', categoria='Cable', unidad='M',
                     cable_tipo='THHN', cable_calibre='12', stock_actual=Decimal('250.5'),
                     stock_minimo=0, precio_unitario=Decimal('18.75')),
            Producto(codigo='RT-3', descripcion='Bote pintura', categoria='Pinturas',
                     unidad='Lts', stock_actual=Decimal('4'), stock_minimo=1, precio_unitario=Decimal('0')),
        ])
        db.session.commit()
        _login(client, inv_admin.id, 'admin')

        exp = client.get('/api/v1/productos/exportar')
        assert exp.status_code == 200
        data = {'archivo': (_io.BytesIO(exp.data), 'catalogo_materiales.xlsx')}
        imp = client.post('/api/v1/productos/importar', data=data,
                          content_type='multipart/form-data')
        assert imp.status_code == 200, imp.get_json()
        body = imp.get_json()
        assert body['exitosos'] == 0
        assert body['actualizados'] == 0, body['cambios_detalle']
        assert body['sin_cambios'] == 3

    def test_reimport_sin_cambios(self, client, inv_admin, db):
        """Reimportar una fila idéntica no debe actualizar: cuenta 'sin cambios'."""
        from decimal import Decimal
        p = Producto(codigo='DIF-0', descripcion='Tuerca', categoria='Tuercas',
                     unidad='pza', stock_actual=Decimal('5'), stock_minimo=1, precio_unitario=Decimal('2'))
        db.session.add(p); db.session.commit()
        _login(client, inv_admin.id, 'admin')
        resp = self._post(client, [{
            'Código (SKU)': 'DIF-0', 'Descripción': 'Tuerca', 'Categoría': 'Tuercas',
            'Unidad': 'pza', 'Stock Inicial': 5, 'Stock Mínimo': 1, 'Precio Unitario': 2,
        }])
        assert resp.status_code == 200, resp.get_json()
        body = resp.get_json()
        assert body['actualizados'] == 0
        assert body['sin_cambios'] == 1
        assert body['cambios_detalle'] == []

    def test_reimport_detecta_solo_el_cambio(self, client, inv_admin, db):
        """Cambiar un solo campo → actualizados=1 y el detalle lista ese campo.
        El stock actual NO se modifica aunque venga otro Stock Inicial."""
        from decimal import Decimal
        p = Producto(codigo='DIF-1', descripcion='Tuerca', categoria='Tuercas',
                     unidad='pza', stock_actual=Decimal('5'), stock_minimo=1, precio_unitario=Decimal('2'))
        db.session.add(p); db.session.commit()
        pid = p.id
        _login(client, inv_admin.id, 'admin')
        resp = self._post(client, [{
            'Código (SKU)': 'DIF-1', 'Descripción': 'Tuerca', 'Categoría': 'Tuercas',
            'Unidad': 'pza', 'Stock Inicial': 999, 'Stock Mínimo': 1, 'Precio Unitario': 7.5,
        }])
        assert resp.status_code == 200, resp.get_json()
        body = resp.get_json()
        assert body['actualizados'] == 1 and body['sin_cambios'] == 0
        assert body['cambios_detalle'][0]['codigo'] == 'DIF-1'
        assert any('precio' in c for c in body['cambios_detalle'][0]['cambios'])
        db.session.refresh(p)
        assert float(p.precio_unitario) == 7.5
        assert float(p.stock_actual) == 5.0  # el stock NO cambió

    def test_plantilla_incluye_columnas_cable(self, client, inv_admin):
        """La plantilla descargable debe traer las 2 columnas de cable."""
        import io as _io
        from openpyxl import load_workbook
        _login(client, inv_admin.id, 'admin')
        resp = client.get('/api/v1/productos/plantilla-importar')
        assert resp.status_code == 200
        ws = load_workbook(_io.BytesIO(resp.data)).active
        headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
        assert 'Tipo (cable)' in headers
        assert 'Tamaño mm²/AWG (cable)' in headers

    def test_categoria_ambigua_pide_confirmacion(self, client, inv_admin, db):
        """Importar 'Cable azul' cuando existe 'Cable' → pide confirmación (no crea nada)."""
        db.session.add(Producto(codigo='EXIST-CAB', descripcion='Cable base', categoria='Cable',
                                unidad='M', cable_tipo='THHN', cable_calibre='12',
                                stock_actual=0, stock_minimo=0))
        db.session.commit()
        _login(client, inv_admin.id, 'admin')
        resp = self._post(client, [{
            'Código (SKU)': 'CAB-AZ-1', 'Descripción': 'Cable azul THHN', 'Categoría': 'Cable azul',
            'Unidad': 'M', 'Stock Inicial': 10, 'Stock Mínimo': 0,
            'Tipo (cable)': 'THHN', 'Tamaño mm²/AWG (cable)': '12',
        }])
        assert resp.status_code == 200, resp.get_json()
        body = resp.get_json()
        assert body.get('necesita_confirmacion') is True
        amb = body['categorias_ambiguas']
        assert any(a['nombre'] == 'Cable azul' and a['sugerencia'] == 'Cable' for a in amb)
        assert 'Cable' in body['categorias_existentes']
        # No se creó nada todavía.
        assert Producto.query.filter_by(codigo='CAB-AZ-1').first() is None

    def test_confirmar_mapea_a_existente(self, client, inv_admin, db):
        """Con el mapeo 'Cable azul'→'Cable', el producto entra en 'Cable'."""
        import io as _io
        db.session.add(Producto(codigo='EXIST-CAB2', descripcion='Cable base', categoria='Cable',
                                unidad='M', cable_tipo='THHN', cable_calibre='12',
                                stock_actual=0, stock_minimo=0))
        db.session.commit()
        _login(client, inv_admin.id, 'admin')
        import json as _json
        data = {
            'archivo': (_io.BytesIO(_build_import_xlsx([{
                'Código (SKU)': 'CAB-AZ-2', 'Descripción': 'Cable azul THHN', 'Categoría': 'Cable azul',
                'Unidad': 'M', 'Stock Inicial': 10, 'Stock Mínimo': 0,
                'Tipo (cable)': 'THHN', 'Tamaño mm²/AWG (cable)': '12',
            }])), 'materiales.xlsx'),
            'categoria_mapeo': _json.dumps({'Cable azul': 'Cable'}),
        }
        resp = client.post('/api/v1/productos/importar', data=data,
                           content_type='multipart/form-data')
        assert resp.status_code == 200, resp.get_json()
        body = resp.get_json()
        assert body.get('necesita_confirmacion') is not True
        assert body['exitosos'] == 1
        p = Producto.query.filter_by(codigo='CAB-AZ-2').first()
        assert p is not None and p.categoria == 'Cable'  # se agregó a la existente

    def test_confirmar_crear_nueva(self, client, inv_admin, db):
        """Con el mapeo 'Cable azul'→'' (crear nueva), se crea como categoría propia."""
        import io as _io, json as _json
        db.session.add(Producto(codigo='EXIST-CAB3', descripcion='Cable base', categoria='Cable',
                                unidad='M', cable_tipo='THHN', cable_calibre='12',
                                stock_actual=0, stock_minimo=0))
        db.session.commit()
        _login(client, inv_admin.id, 'admin')
        data = {
            'archivo': (_io.BytesIO(_build_import_xlsx([{
                'Código (SKU)': 'CAB-AZ-3', 'Descripción': 'Cable azul THHN', 'Categoría': 'Cable azul',
                'Unidad': 'M', 'Stock Inicial': 10, 'Stock Mínimo': 0,
                'Tipo (cable)': 'THHN', 'Tamaño mm²/AWG (cable)': '12',
            }])), 'materiales.xlsx'),
            'categoria_mapeo': _json.dumps({'Cable azul': ''}),
        }
        resp = client.post('/api/v1/productos/importar', data=data,
                           content_type='multipart/form-data')
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['exitosos'] == 1
        p = Producto.query.filter_by(codigo='CAB-AZ-3').first()
        assert p is not None and p.categoria == 'Cable azul'  # categoría nueva propia

    def test_columnas_cable_junto_a_categoria(self, client, inv_admin):
        """Tipo y Tamaño deben ir INMEDIATAMENTE después de Categoría."""
        import io as _io
        from openpyxl import load_workbook
        _login(client, inv_admin.id, 'admin')
        resp = client.get('/api/v1/productos/plantilla-importar')
        ws = load_workbook(_io.BytesIO(resp.data)).active
        headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
        i_cat = headers.index('Categoría')
        assert headers[i_cat + 1] == 'Tipo (cable)'
        assert headers[i_cat + 2] == 'Tamaño mm²/AWG (cable)'

    def test_export_seccionado_por_categoria(self, client, inv_admin, db):
        """El export agrupa por categoría con filas de sección (marcador '#')."""
        import io as _io
        from openpyxl import load_workbook
        from decimal import Decimal
        db.session.add_all([
            Producto(codigo='SEC-A1', descripcion='A1', categoria='Alfa', unidad='pza',
                     stock_actual=Decimal('1'), stock_minimo=0),
            Producto(codigo='SEC-A2', descripcion='A2', categoria='Alfa', unidad='pza',
                     stock_actual=Decimal('1'), stock_minimo=0),
            Producto(codigo='SEC-B1', descripcion='B1', categoria='Beta', unidad='pza',
                     stock_actual=Decimal('1'), stock_minimo=0),
        ])
        db.session.commit()
        _login(client, inv_admin.id, 'admin')
        resp = client.get('/api/v1/productos/exportar')
        assert resp.status_code == 200
        ws = load_workbook(_io.BytesIO(resp.data)).active
        col_a = [ws.cell(row=r, column=1).value for r in range(5, ws.max_row + 1)]
        secciones = [v for v in col_a if isinstance(v, str) and v.startswith('#')]
        # Al menos una sección por categoría (Alfa, Beta).
        assert len(secciones) >= 2
        assert any('Alfa' in s for s in secciones)
        assert any('Beta' in s for s in secciones)

    def test_import_ignora_filas_de_seccion(self, client, inv_admin, db):
        """Una fila cuyo código empieza con '#' (sección) se salta sin error."""
        _login(client, inv_admin.id, 'admin')
        resp = self._post(client, [
            {'Código (SKU)': '#  ▸  Tornillería  (1)'},  # fila de sección
            {'Código (SKU)': 'SEC-OK', 'Descripción': 'Tornillo', 'Categoría': 'Tornillería',
             'Unidad': 'pza', 'Stock Inicial': 3, 'Stock Mínimo': 0},
        ])
        assert resp.status_code == 200, resp.get_json()
        body = resp.get_json()
        assert body['exitosos'] == 1
        assert body['errores'] == []  # la fila de sección NO es un error
        assert Producto.query.filter_by(codigo='SEC-OK').first() is not None

    def test_importar_cable_existente_sin_columnas_no_rompe(self, client, inv_admin, db):
        """Reimportar un cable con plantilla vieja (sin columnas de cable) NO debe
        fallar ni borrar el Tipo/Tamaño ya guardado."""
        _login(client, inv_admin.id, 'admin')
        p = Producto(codigo='IMP-CAB-4', descripcion='Cable', categoria='Cable',
                     unidad='M', cable_tipo='THHN', cable_calibre='8',
                     stock_actual=0, stock_minimo=0)
        db.session.add(p); db.session.commit()
        # Fila sin las columnas de cable (se envían vacías, como plantilla vieja).
        resp = self._post(client, [{
            'Código (SKU)': 'IMP-CAB-4', 'Descripción': 'Cable', 'Categoría': 'Cable',
            'Unidad': 'M', 'Stock Inicial': 0, 'Stock Mínimo': 0, 'Precio Unitario': 55,
        }])
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['actualizados'] == 1
        db.session.refresh(p)
        assert p.cable_tipo == 'THHN' and p.cable_calibre == '8'


# ═══════════════════════════════════════════════════════════════════════════════
# 4. ALMACENES — CRUD
# ═══════════════════════════════════════════════════════════════════════════════

class TestAlmacenes:

    def test_crear_almacen(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/almacenes/', json={
            'nombre': 'Almacén Central', 'ubicacion': 'Planta 1', 'activo': True,
        })
        assert resp.status_code == 200
        data = resp.get_json()
        assert data['nombre'] == 'Almacén Central'
        assert data['qr_code']

    def test_crear_almacen_sin_nombre(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/almacenes/', json={'ubicacion': 'P2'})
        assert resp.status_code == 422

    def test_crear_almacen_nombre_demasiado_largo(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/almacenes/', json={'nombre': 'A' * 101, 'activo': True})
        assert resp.status_code == 422

    def test_listar_almacenes(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        db.session.add(Almacen(nombre='X', qr_code=str(uuid.uuid4()), activo=True))
        db.session.commit()
        resp = client.get('/api/v1/almacenes/')
        assert resp.status_code == 200
        assert isinstance(resp.get_json(), list)

    def test_validar_qr_almacen(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        qr = str(uuid.uuid4())
        db.session.add(Almacen(nombre='QR Test', qr_code=qr, activo=True))
        db.session.commit()
        resp = client.get(f'/api/v1/almacenes/{qr}/validar')
        assert resp.status_code == 200

    def test_validar_qr_inexistente(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.get(f'/api/v1/almacenes/{uuid.uuid4()}/validar')
        assert resp.status_code == 404

    def test_actualizar_almacen(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        a = Almacen(nombre='Orig', qr_code=str(uuid.uuid4()), activo=True)
        db.session.add(a); db.session.commit()
        resp = client.put(f'/api/v1/almacenes/{a.id}', json={'ubicacion': 'P-X'})
        assert resp.status_code == 200
        assert resp.get_json()['ubicacion'] == 'P-X'

    def test_actualizar_almacen_inexistente(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.put('/api/v1/almacenes/999999', json={'nombre': 'X'})
        assert resp.status_code == 404

    def test_borrar_almacen_soft_delete(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        a = Almacen(nombre='Borrar', qr_code=str(uuid.uuid4()), activo=True)
        db.session.add(a); db.session.commit()
        aid = a.id
        resp = client.delete(f'/api/v1/almacenes/{aid}')
        assert resp.status_code == 204
        assert db.session.get(Almacen, aid).activo is False

    def test_borrar_almacen_inexistente(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.delete('/api/v1/almacenes/999999')
        assert resp.status_code == 404


# ═══════════════════════════════════════════════════════════════════════════════
# 5. ESTANTES — CRUD
# ═══════════════════════════════════════════════════════════════════════════════

class TestEstantes:

    def _crear_almacen(self, db):
        a = Almacen(nombre='Alm Est', qr_code=str(uuid.uuid4()), activo=True)
        db.session.add(a); db.session.commit()
        return a

    def test_crear_estante_valido(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        a = self._crear_almacen(db)
        resp = client.post('/api/v1/estantes/', json={
            'nombre': 'Estante A-1', 'descripcion': 'Primero', 'almacen_id': a.id,
        })
        assert resp.status_code == 200
        assert resp.get_json()['nombre'] == 'Estante A-1'

    def test_crear_estante_almacen_inexistente(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/estantes/', json={'nombre': 'X', 'almacen_id': 999999})
        assert resp.status_code == 404

    def test_crear_estante_sin_almacen_id(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/estantes/', json={'nombre': 'X'})
        assert resp.status_code == 422

    def test_qr_image_estante_es_png(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        a = self._crear_almacen(db)
        e = Estante(nombre='E1', almacen_id=a.id, qr_code=str(uuid.uuid4()), activo=True)
        db.session.add(e); db.session.commit()
        resp = client.get(f'/api/v1/estantes/{e.id}/qr-image')
        assert resp.status_code == 200
        assert resp.headers['Content-Type'] == 'image/png'
        assert resp.data[:4] == b'\x89PNG'

    def test_qr_image_estante_inexistente(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.get('/api/v1/estantes/999999/qr-image')
        assert resp.status_code == 404

    def test_validar_qr_estante(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        a = self._crear_almacen(db)
        qr = str(uuid.uuid4())
        e = Estante(nombre='E1', almacen_id=a.id, qr_code=qr, activo=True)
        db.session.add(e); db.session.commit()
        resp = client.get(f'/api/v1/estantes/{qr}/validar')
        assert resp.status_code == 200

    def test_validar_qr_estante_inexistente(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.get(f'/api/v1/estantes/{uuid.uuid4()}/validar')
        assert resp.status_code == 404

    def test_actualizar_estante_almacen_destino_invalido(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        a = self._crear_almacen(db)
        e = Estante(nombre='E1', almacen_id=a.id, qr_code=str(uuid.uuid4()), activo=True)
        db.session.add(e); db.session.commit()
        resp = client.put(f'/api/v1/estantes/{e.id}', json={'almacen_id': 999999})
        assert resp.status_code == 404

    def test_get_estantes_por_almacen(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        a = self._crear_almacen(db)
        resp = client.get(f'/api/v1/almacenes/{a.id}/estantes')
        assert resp.status_code == 200
        assert isinstance(resp.get_json(), list)


# ═══════════════════════════════════════════════════════════════════════════════
# 5-bis. REJILLA DE ESTANTE + STOCK POR CELDA (Pausa 11)
# ═══════════════════════════════════════════════════════════════════════════════

class TestEstanteLayout:

    def _setup(self, db, *, stock=100, filas=3, columnas=4):
        a = Almacen(nombre='Alm Grid', qr_code=str(uuid.uuid4()), activo=True)
        db.session.add(a); db.session.commit()
        e = Estante(nombre='Rack G', almacen_id=a.id, qr_code=str(uuid.uuid4()),
                    activo=True, filas=filas, columnas=columnas)
        p = Producto(codigo='GR-001', descripcion='Tubo', categoria='F', unidad='pza',
                     stock_actual=stock, stock_minimo=0)
        db.session.add_all([e, p]); db.session.commit()
        db.session.add(StockPorAlmacen(producto_id=p.id, almacen_id=a.id, cantidad=stock))
        db.session.commit()
        return a, e, p

    def test_crear_estante_con_rejilla(self, client, inv_admin, db):
        _login(client, inv_admin.id)
        a = Almacen(nombre='A', qr_code=str(uuid.uuid4()), activo=True)
        db.session.add(a); db.session.commit()
        resp = client.post('/api/v1/estantes/', json={
            'nombre': 'R1', 'almacen_id': a.id, 'filas': 2, 'columnas': 5,
        })
        assert resp.status_code == 200
        body = resp.get_json()
        assert body['filas'] == 2 and body['columnas'] == 5

    def test_guardar_y_leer_layout(self, client, inv_admin, db):
        _login(client, inv_admin.id)
        a, e, p = self._setup(db)
        resp = client.put(f'/api/v1/estantes/{e.id}/layout', json={
            'posiciones': [{'producto_id': p.id, 'fila': 3, 'columna': 2, 'cantidad': 40}],
        })
        assert resp.status_code == 200, resp.get_json()
        lay = client.get(f'/api/v1/estantes/{e.id}/layout').get_json()
        assert len(lay['celdas']) == 1
        celda = lay['celdas'][0]
        assert celda['fila'] == 3 and celda['columna'] == 2 and celda['cantidad'] == 40
        assert lay['stock_almacen'][str(p.id)] == 100

    def test_layout_rechaza_posicion_fuera_de_rejilla(self, client, inv_admin, db):
        _login(client, inv_admin.id)
        a, e, p = self._setup(db, filas=2, columnas=2)
        resp = client.put(f'/api/v1/estantes/{e.id}/layout', json={
            'posiciones': [{'producto_id': p.id, 'fila': 9, 'columna': 1, 'cantidad': 1}],
        })
        assert resp.status_code == 422

    def test_layout_rechaza_cantidad_mayor_que_stock(self, client, inv_admin, db):
        _login(client, inv_admin.id)
        a, e, p = self._setup(db, stock=10)
        resp = client.put(f'/api/v1/estantes/{e.id}/layout', json={
            'posiciones': [{'producto_id': p.id, 'fila': 1, 'columna': 1, 'cantidad': 25}],
        })
        assert resp.status_code == 422
        assert 'errores' in resp.get_json()

    def test_layout_fila_sin_columna_es_invalido(self, client, inv_admin, db):
        _login(client, inv_admin.id)
        a, e, p = self._setup(db)
        resp = client.put(f'/api/v1/estantes/{e.id}/layout', json={
            'posiciones': [{'producto_id': p.id, 'fila': 1, 'columna': None, 'cantidad': 1}],
        })
        assert resp.status_code == 422

    def test_legacy_set_productos_sigue_funcionando(self, client, inv_admin, db):
        _login(client, inv_admin.id)
        a, e, p = self._setup(db)
        resp = client.put(f'/api/v1/estantes/{e.id}/productos', json={'producto_ids': [p.id]})
        assert resp.status_code == 200
        lay = client.get(f'/api/v1/estantes/{e.id}/layout').get_json()
        # Asignado pero sin ubicar (fila/columna NULL, cantidad 0).
        assert len(lay['celdas']) == 1
        assert lay['celdas'][0]['fila'] is None and lay['celdas'][0]['cantidad'] == 0

    def test_scan_qr_devuelve_productos_del_rack_con_ubicacion(self, client, inv_admin, db):
        _login(client, inv_admin.id)
        a, e, p = self._setup(db)
        client.put(f'/api/v1/estantes/{e.id}/layout', json={
            'posiciones': [{'producto_id': p.id, 'fila': 2, 'columna': 3, 'cantidad': 12}],
        })
        resp = client.get(f'/api/v1/estantes/{e.qr_code}/inventario')
        assert resp.status_code == 200
        body = resp.get_json()
        assert len(body['productos']) == 1
        prod = body['productos'][0]
        assert prod['id'] == p.id
        assert prod['ubicacion'] == {'fila': 2, 'columna': 3, 'cantidad': 12}

    def test_reducir_rejilla_deja_sin_ubicar(self, client, inv_admin, db):
        _login(client, inv_admin.id)
        a, e, p = self._setup(db, filas=4, columnas=4)
        client.put(f'/api/v1/estantes/{e.id}/layout', json={
            'posiciones': [{'producto_id': p.id, 'fila': 4, 'columna': 4, 'cantidad': 5}],
        })
        # Reducir a 2x2 → la posición (4,4) queda fuera y pasa a "sin ubicar".
        resp = client.put(f'/api/v1/estantes/{e.id}', json={'filas': 2, 'columnas': 2})
        assert resp.status_code == 200
        lay = client.get(f'/api/v1/estantes/{e.id}/layout').get_json()
        celda = lay['celdas'][0]
        assert celda['fila'] is None and celda['columna'] is None
        # No pierde la cantidad capturada.
        assert celda['cantidad'] == 5

    def test_mover_estante_de_almacen_resetea_celdas(self, client, inv_admin, db):
        """Mover un estante a otro almacén deja sus celdas 'sin ubicar' y en 0:
        el sub-libro de celdas es POR almacén y las unidades físicas no viajan
        con el registro del estante (evita romper Σceldas ≤ stock en el destino)."""
        _login(client, inv_admin.id)
        a, e, p = self._setup(db)
        client.put(f'/api/v1/estantes/{e.id}/layout', json={
            'posiciones': [{'producto_id': p.id, 'fila': 1, 'columna': 1, 'cantidad': 30}],
        })
        destino = Almacen(nombre='Alm Destino', qr_code=str(uuid.uuid4()), activo=True)
        db.session.add(destino); db.session.commit()

        resp = client.put(f'/api/v1/estantes/{e.id}', json={'almacen_id': destino.id})
        assert resp.status_code == 200

        lay = client.get(f'/api/v1/estantes/{e.id}/layout').get_json()
        assert lay['estante']['almacen_id'] == destino.id
        # La colocación sigue (el producto continúa asignado al estante) pero
        # queda sin ubicar y en 0 — sin 'sobrante' contra el almacén destino.
        celda = lay['celdas'][0]
        assert celda['fila'] is None and celda['columna'] is None
        assert celda['cantidad'] == 0
        assert lay['sobrante_por_producto'] == {}

    def test_editar_estante_sin_cambiar_almacen_conserva_celdas(self, client, inv_admin, db):
        """Re-enviar el mismo almacen_id (o editar solo el nombre) NO debe
        resetear las celdas: el reseteo es exclusivo del cambio real de bodega."""
        _login(client, inv_admin.id)
        a, e, p = self._setup(db)
        client.put(f'/api/v1/estantes/{e.id}/layout', json={
            'posiciones': [{'producto_id': p.id, 'fila': 2, 'columna': 2, 'cantidad': 15}],
        })
        resp = client.put(f'/api/v1/estantes/{e.id}', json={
            'nombre': 'Rack G renombrado', 'almacen_id': a.id,
        })
        assert resp.status_code == 200
        lay = client.get(f'/api/v1/estantes/{e.id}/layout').get_json()
        celda = lay['celdas'][0]
        assert celda['fila'] == 2 and celda['columna'] == 2 and celda['cantidad'] == 15


# ═══════════════════════════════════════════════════════════════════════════════
# 6. MOVIMIENTOS
# ═══════════════════════════════════════════════════════════════════════════════

class TestMovimientos:

    @pytest.fixture(autouse=True)
    def _bodega_default(self, db):
        """Movimientos requieren al menos un Almacén activo (fallback
        cuando el payload no manda almacén_origen/destino). Creamos DOS
        bodegas activas: la primera es la default (origen), la segunda se
        usa como destino en los TRASPASOs. IDs expuestos vía
        `self._bodega_id` y `self._bodega_dest_id`."""
        a = Almacen(nombre='Bodega Tests', qr_code=str(uuid.uuid4()), activo=True)
        b = Almacen(nombre='Bodega Tests B', qr_code=str(uuid.uuid4()), activo=True)
        db.session.add_all([a, b]); db.session.commit()
        self._bodega_id = a.id
        self._bodega_dest_id = b.id
        return a

    def _crear_producto(self, db, stock=100):
        from decimal import Decimal
        p = Producto(
            codigo=f'MOV-{uuid.uuid4().hex[:6]}', descripcion='X', categoria='T',
            unidad='pza', stock_actual=Decimal(str(stock)), stock_minimo=10,
        )
        db.session.add(p); db.session.flush()
        # Feature 'stock por proyecto': la fuente de verdad es
        # stock_almacen_proyecto (bucket general = proyecto NULL); stock_por_almacen
        # y stock_actual son caches que se recalculan desde los buckets. Sembramos
        # ambos para que los movimientos partan del stock esperado.
        db.session.add(StockAlmacenProyecto(
            producto_id=p.id, almacen_id=self._bodega_id, proyecto_id=None,
            cantidad=Decimal(str(stock)),
        ))
        db.session.add(StockPorAlmacen(
            producto_id=p.id, almacen_id=self._bodega_id,
            cantidad=Decimal(str(stock)),
        ))
        db.session.commit()
        return p

    def test_entrada_incrementa_stock(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._crear_producto(db, stock=100)
        resp = client.post('/api/v1/movimientos/', json={
            'tipo': 'ENTRADA', 'producto_id': p.id, 'cantidad': 25.0,
        })
        assert resp.status_code == 200
        db.session.refresh(p)
        assert float(p.stock_actual) == 125.0

    def test_entrada_cantidad_cero_falla(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._crear_producto(db)
        resp = client.post('/api/v1/movimientos/', json={
            'tipo': 'ENTRADA', 'producto_id': p.id, 'cantidad': 0.0,
        })
        assert resp.status_code in (400, 422)

    def test_entrada_cantidad_negativa_falla(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._crear_producto(db)
        resp = client.post('/api/v1/movimientos/', json={
            'tipo': 'ENTRADA', 'producto_id': p.id, 'cantidad': -10.0,
        })
        assert resp.status_code in (400, 422)

    def test_salida_reduce_stock(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._crear_producto(db, stock=100)
        resp = client.post('/api/v1/movimientos/', json={
            'tipo': 'SALIDA', 'producto_id': p.id, 'cantidad': 10.0,
        })
        assert resp.status_code == 200
        db.session.refresh(p)
        assert float(p.stock_actual) == pytest.approx(90.0, abs=0.01)

    def test_salida_stock_insuficiente(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._crear_producto(db, stock=10)
        resp = client.post('/api/v1/movimientos/', json={
            'tipo': 'SALIDA', 'producto_id': p.id, 'cantidad': 9999.0,
        })
        assert resp.status_code == 400
        assert 'insuficiente' in resp.get_json()['detail'].lower()

    def test_ajuste_positivo(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._crear_producto(db, stock=100)
        resp = client.post('/api/v1/movimientos/', json={
            'tipo': 'AJUSTE', 'producto_id': p.id, 'cantidad': 5.0,
        })
        assert resp.status_code == 200
        db.session.refresh(p)
        assert float(p.stock_actual) == pytest.approx(105.0, abs=0.01)

    def test_ajuste_negativo_valido(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._crear_producto(db, stock=100)
        resp = client.post('/api/v1/movimientos/', json={
            'tipo': 'AJUSTE', 'producto_id': p.id, 'cantidad': -5.0,
        })
        assert resp.status_code == 200
        db.session.refresh(p)
        assert float(p.stock_actual) == pytest.approx(95.0, abs=0.01)

    def test_ajuste_provoca_stock_negativo(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._crear_producto(db, stock=10)
        resp = client.post('/api/v1/movimientos/', json={
            'tipo': 'AJUSTE', 'producto_id': p.id, 'cantidad': -50.0,
        })
        assert resp.status_code == 400
        assert 'negativo' in resp.get_json()['detail'].lower()

    def test_traspaso_reduce_stock(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._crear_producto(db, stock=100)
        resp = client.post('/api/v1/movimientos/', json={
            'tipo': 'TRASPASO', 'producto_id': p.id, 'cantidad': 5.0,
            'almacen_origen_id': self._bodega_id,
            'almacen_destino_id': self._bodega_dest_id,
        })
        assert resp.status_code == 200, resp.get_json()

    def test_traspaso_stock_insuficiente(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._crear_producto(db, stock=5)
        resp = client.post('/api/v1/movimientos/', json={
            'tipo': 'TRASPASO', 'producto_id': p.id, 'cantidad': 9999.0,
            'almacen_origen_id': self._bodega_id,
            'almacen_destino_id': self._bodega_dest_id,
        })
        assert resp.status_code == 400

    def test_movimiento_producto_inexistente(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/movimientos/', json={
            'tipo': 'ENTRADA', 'producto_id': 999999, 'cantidad': 10.0,
        })
        assert resp.status_code == 404

    def test_tipo_movimiento_invalido(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._crear_producto(db)
        resp = client.post('/api/v1/movimientos/', json={
            'tipo': 'TYPO_INEXISTENTE', 'producto_id': p.id, 'cantidad': 10.0,
        })
        assert resp.status_code == 422

    def test_cantidad_excede_limite_maximo(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._crear_producto(db)
        resp = client.post('/api/v1/movimientos/', json={
            'tipo': 'ENTRADA', 'producto_id': p.id, 'cantidad': 999_999.0,
        })
        assert resp.status_code == 422


# ═══════════════════════════════════════════════════════════════════════════════
# 7. SOLICITUDES DE MATERIAL
# ═══════════════════════════════════════════════════════════════════════════════

class TestSolicitudes:

    def _crear_producto(self, db):
        from decimal import Decimal
        p = Producto(
            codigo=f'SOL-{uuid.uuid4().hex[:6]}', descripcion='X', categoria='M',
            unidad='pza', stock_actual=Decimal('200'), stock_minimo=5,
        )
        db.session.add(p); db.session.commit()
        return p

    def test_solicitante_crea_solicitud(self, client, inv_solicitante, db):
        _login(client, inv_solicitante.id, 'solicitante_material')
        p = self._crear_producto(db)
        resp = client.post('/api/v1/solicitudes/', json={
            'proyecto': 'Alpha',
            'detalles': [{'producto_id': p.id, 'cantidad_solicitada': 10.0}],
        })
        assert resp.status_code == 200
        data = resp.get_json()
        assert data['estatus'] == 'PENDIENTE'
        assert len(data['detalles']) == 1

    def test_solicitud_sin_detalles(self, client, inv_solicitante):
        _login(client, inv_solicitante.id, 'solicitante_material')
        resp = client.post('/api/v1/solicitudes/', json={'proyecto': 'X', 'detalles': []})
        assert resp.status_code == 422

    def test_solicitud_sin_proyecto_rechaza(self, client, inv_solicitante, db):
        """Proyecto es obligatorio: sin él (o vacío) debe responder 422."""
        _login(client, inv_solicitante.id, 'solicitante_material')
        p = self._crear_producto(db)
        det = [{'producto_id': p.id, 'cantidad_solicitada': 5.0}]
        # Falta la clave proyecto.
        resp = client.post('/api/v1/solicitudes/', json={'detalles': det})
        assert resp.status_code == 422
        # Proyecto en blanco tampoco se acepta.
        resp2 = client.post('/api/v1/solicitudes/', json={'proyecto': '   ', 'detalles': det})
        assert resp2.status_code == 422

    def test_solicitud_cantidad_cero(self, client, inv_solicitante, db):
        _login(client, inv_solicitante.id, 'solicitante_material')
        p = self._crear_producto(db)
        resp = client.post('/api/v1/solicitudes/', json={
            'proyecto': 'Proyecto Test',
            'detalles': [{'producto_id': p.id, 'cantidad_solicitada': 0.0}],
        })
        assert resp.status_code == 422

    def test_solicitud_producto_inexistente_retorna_400(self, client, inv_solicitante):
        _login(client, inv_solicitante.id, 'solicitante_material')
        resp = client.post('/api/v1/solicitudes/', json={
            'proyecto': 'Proyecto Test',
            'detalles': [{'producto_id': 999999, 'cantidad_solicitada': 5.0}],
        })
        assert resp.status_code == 400
        # `detail` puede ser str o lista de strings (errores multilínea).
        detail = resp.get_json()['detail']
        text = ' '.join(detail) if isinstance(detail, list) else detail
        assert '999999' in text

    def test_rol_no_autorizado_no_crea_solicitud(self, client, inv_outsider, db):
        _login(client, inv_outsider.id, 'coordinador')
        p = self._crear_producto(db)
        resp = client.post('/api/v1/solicitudes/', json={
            'proyecto': 'Proyecto Test',
            'detalles': [{'producto_id': p.id, 'cantidad_solicitada': 5.0}],
        })
        # El decorador de login del blueprint inventario_api bloquea coordinador con 403.
        assert resp.status_code == 403

    def test_solicitante_solo_ve_suyas(self, client, inv_solicitante, inv_admin, db):
        # Solicitud de otro usuario
        otra = SolicitudMaterial(solicitante_id=inv_admin.id, proyecto='Otra', estatus='PENDIENTE')
        db.session.add(otra); db.session.commit()

        _login(client, inv_solicitante.id, 'solicitante_material')
        p = self._crear_producto(db)
        client.post('/api/v1/solicitudes/', json={
            'proyecto': 'Proyecto Test',
            'detalles': [{'producto_id': p.id, 'cantidad_solicitada': 1.0}],
        })
        resp = client.get('/api/v1/solicitudes/')
        assert resp.status_code == 200
        for s in resp.get_json():
            assert s['solicitante_id'] == inv_solicitante.id

    def test_admin_ve_todas(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.get('/api/v1/solicitudes/')
        assert resp.status_code == 200
        assert isinstance(resp.get_json(), list)

    def test_actualizar_estado_solicitud(self, client, inv_solicitante, inv_admin, db):
        p = self._crear_producto(db)
        _login(client, inv_solicitante.id, 'solicitante_material')
        r = client.post('/api/v1/solicitudes/', json={
            'proyecto': 'Proyecto Test',
            'detalles': [{'producto_id': p.id, 'cantidad_solicitada': 3.0}],
        })
        sol_id = r.get_json()['id']
        _logout(client)
        _login(client, inv_admin.id, 'admin')
        resp = client.patch(f'/api/v1/solicitudes/{sol_id}/estado', json={'estatus': 'APROBADA'})
        assert resp.status_code == 200
        data = resp.get_json()
        assert data['estatus'] == 'APROBADA'
        assert data['fecha_cierre'] is not None

    def test_actualizar_estado_inexistente(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.patch('/api/v1/solicitudes/999999/estado', json={'estatus': 'APROBADA'})
        assert resp.status_code == 404

    def test_actualizar_estado_typo(self, client, inv_solicitante, inv_admin, db):
        p = self._crear_producto(db)
        _login(client, inv_solicitante.id, 'solicitante_material')
        r = client.post('/api/v1/solicitudes/', json={
            'proyecto': 'Proyecto Test',
            'detalles': [{'producto_id': p.id, 'cantidad_solicitada': 1.0}],
        })
        sol_id = r.get_json()['id']
        _logout(client)
        _login(client, inv_admin.id, 'admin')
        resp = client.patch(f'/api/v1/solicitudes/{sol_id}/estado', json={'estatus': 'APROBANDO'})
        assert resp.status_code == 422

    def test_solicitante_no_aprueba_su_solicitud(self, client, inv_solicitante, db):
        p = self._crear_producto(db)
        _login(client, inv_solicitante.id, 'solicitante_material')
        r = client.post('/api/v1/solicitudes/', json={
            'proyecto': 'Proyecto Test',
            'detalles': [{'producto_id': p.id, 'cantidad_solicitada': 1.0}],
        })
        sol_id = r.get_json()['id']
        resp = client.patch(f'/api/v1/solicitudes/{sol_id}/estado', json={'estatus': 'APROBADA'})
        assert resp.status_code == 403

    def test_volver_a_pendiente_limpia_fecha_cierre(self, client, inv_solicitante, inv_admin, db):
        p = self._crear_producto(db)
        _login(client, inv_solicitante.id, 'solicitante_material')
        r = client.post('/api/v1/solicitudes/', json={
            'proyecto': 'Proyecto Test',
            'detalles': [{'producto_id': p.id, 'cantidad_solicitada': 1.0}],
        })
        sol_id = r.get_json()['id']
        _logout(client)
        _login(client, inv_admin.id, 'admin')

        ap = client.patch(f'/api/v1/solicitudes/{sol_id}/estado', json={'estatus': 'APROBADA'})
        assert ap.get_json()['fecha_cierre'] is not None

        pend = client.patch(f'/api/v1/solicitudes/{sol_id}/estado', json={'estatus': 'PENDIENTE'})
        assert pend.status_code == 200
        assert pend.get_json()['fecha_cierre'] is None


# ═══════════════════════════════════════════════════════════════════════════════
# 8. PROYECTOS
# ═══════════════════════════════════════════════════════════════════════════════

class TestProyectos:

    def test_listar_proyectos(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.get('/api/v1/proyectos/')
        assert resp.status_code == 200
        assert isinstance(resp.get_json(), list)

    def test_listar_proyectos_sin_autenticar(self, client):
        resp = client.get('/api/v1/proyectos/')
        assert resp.status_code == 401


# ═══════════════════════════════════════════════════════════════════════════════
# 9. AUDITORÍA
# ═══════════════════════════════════════════════════════════════════════════════

class TestAuditoria:

    def test_crear_producto_genera_audit(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        antes = db.session.query(AuditLog).count()
        client.post('/api/v1/productos/', json={
            'codigo': f'AUD-{uuid.uuid4().hex[:6]}', 'descripcion': 'X', 'categoria': 'T',
            'unidad': 'pza', 'stock_actual': 0, 'stock_minimo': 0,
        })
        despues = db.session.query(AuditLog).count()
        assert despues > antes

    def test_crear_almacen_genera_audit(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        antes = db.session.query(AuditLog).count()
        client.post('/api/v1/almacenes/', json={
            'nombre': f'Alm Audit {uuid.uuid4().hex[:4]}', 'activo': True,
        })
        despues = db.session.query(AuditLog).count()
        assert despues > antes


# ═══════════════════════════════════════════════════════════════════════════════
# 10. EDGE CASES
# ═══════════════════════════════════════════════════════════════════════════════

class TestEdgeCases:

    def test_skip_negativo(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.get('/api/v1/productos/?skip=-1&limit=10')
        # Validación clamp: skip<0 → 422
        assert resp.status_code in (200, 422)

    def test_limit_cero(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.get('/api/v1/productos/?skip=0&limit=0')
        assert resp.status_code == 200
        assert resp.get_json() == []

    def test_campos_extra_son_ignorados(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/almacenes/', json={
            'nombre': f'Alm Extra {uuid.uuid4().hex[:4]}', 'activo': True,
            'campo_inexistente': 'X',
        })
        assert resp.status_code == 200

    def test_id_no_numerico_rechazado(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        # Flask's <int:id> converter devuelve 404 cuando el path no parsea como int
        resp = client.put('/api/v1/productos/abc', json={})
        assert resp.status_code == 404

    def test_payload_vacio(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/productos/', data=b'',
                           content_type='application/json')
        assert resp.status_code == 422

    def test_content_type_incorrecto(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.post('/api/v1/productos/',
                           data={'codigo': 'FORM-001', 'descripcion': 'Test'})
        # request.get_json(silent=True) devuelve None → schema rechaza
        assert resp.status_code == 422


# ═══════════════════════════════════════════════════════════════════════════════
# 12. MOVIMIENTOS — PARTES (quién entrega / quién recibe) + VALE PDF
# ═══════════════════════════════════════════════════════════════════════════════

class TestMovimientoPartes:

    @pytest.fixture(autouse=True)
    def _seed(self, db):
        self._bodega = Almacen(nombre='Bodega Partes', qr_code=str(uuid.uuid4()), activo=True)
        db.session.add(self._bodega); db.session.commit()

    def _producto(self, db, stock=100):
        from decimal import Decimal
        p = Producto(codigo=f'PT-{uuid.uuid4().hex[:6]}', descripcion='X', categoria='T',
                     unidad='pza', stock_actual=Decimal(str(stock)), stock_minimo=0)
        db.session.add(p); db.session.flush()
        db.session.add(StockAlmacenProyecto(producto_id=p.id, almacen_id=self._bodega.id,
                                            proyecto_id=None, cantidad=Decimal(str(stock))))
        db.session.add(StockPorAlmacen(producto_id=p.id, almacen_id=self._bodega.id,
                                       cantidad=Decimal(str(stock))))
        db.session.commit()
        return p

    def _trabajador(self, db, activo=True):
        t = Trabajador(no_empleado=f'TP-{uuid.uuid4().hex[:5]}', nombre='Juan',
                       nombre_apellidos='Pérez López', activo=activo)
        db.session.add(t); db.session.commit()
        return t

    def test_salida_con_partes_persiste_y_serializa(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._producto(db, stock=100)
        t = self._trabajador(db)
        resp = client.post('/api/v1/movimientos/', json={
            'tipo': 'SALIDA', 'producto_id': p.id, 'cantidad': 10.0,
            'almacen_origen_id': self._bodega.id,
            'entrega_nombre': 'Almacén Central',
            'recibe_trabajador_id': t.id,
        })
        assert resp.status_code == 200, resp.get_json()
        body = resp.get_json()
        assert body['entrega_nombre'] == 'Almacén Central'
        assert body['recibe_nombre'] == t.nombre_completo
        assert body['recibe_trabajador_id'] == t.id
        mov = flask_db.session.get(MovimientoInventario, body['id'])
        assert mov.recibe_trabajador_id == t.id
        assert mov.entrega_nombre == 'Almacén Central'

    def test_vale_pdf_responde_ok(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._producto(db, stock=50)
        resp = client.post('/api/v1/movimientos/', json={
            'tipo': 'ENTRADA', 'producto_id': p.id, 'cantidad': 5.0,
            'almacen_destino_id': self._bodega.id,
            'entrega_nombre': 'Proveedor X', 'recibe_nombre': 'Bodeguero',
        })
        assert resp.status_code == 200, resp.get_json()
        mov_id = resp.get_json()['id']
        pdf = client.get(f'/api/v1/movimientos/{mov_id}/pdf')
        assert pdf.status_code == 200
        assert pdf.mimetype == 'application/pdf'
        assert pdf.data[:4] == b'%PDF'

    def test_vale_pdf_404_si_no_existe(self, client, inv_admin):
        _login(client, inv_admin.id, 'admin')
        resp = client.get('/api/v1/movimientos/999999/pdf')
        assert resp.status_code == 404

    def test_trabajador_inexistente_422(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._producto(db)
        resp = client.post('/api/v1/movimientos/', json={
            'tipo': 'ENTRADA', 'producto_id': p.id, 'cantidad': 5.0,
            'almacen_destino_id': self._bodega.id,
            'recibe_trabajador_id': 999999,
        })
        assert resp.status_code == 422

    def test_trabajador_inactivo_422(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._producto(db)
        t = self._trabajador(db, activo=False)
        resp = client.post('/api/v1/movimientos/', json={
            'tipo': 'ENTRADA', 'producto_id': p.id, 'cantidad': 5.0,
            'almacen_destino_id': self._bodega.id,
            'entrega_trabajador_id': t.id,
        })
        assert resp.status_code == 422


# ═══════════════════════════════════════════════════════════════════════════════
# 13. EDITOR DE STOCK POR BUCKET (ajustar-buckets → AJUSTES)
# ═══════════════════════════════════════════════════════════════════════════════

class TestAjustarBuckets:

    @pytest.fixture(autouse=True)
    def _seed(self, db):
        self._bodega = Almacen(nombre='Bodega Buckets', qr_code=str(uuid.uuid4()), activo=True)
        db.session.add(self._bodega); db.session.commit()

    def _producto(self, db, general=50):
        from decimal import Decimal
        p = Producto(codigo=f'BK-{uuid.uuid4().hex[:6]}', descripcion='X', categoria='T',
                     unidad='pza', stock_actual=Decimal(str(general)), stock_minimo=0)
        db.session.add(p); db.session.flush()
        db.session.add(StockAlmacenProyecto(producto_id=p.id, almacen_id=self._bodega.id,
                                            proyecto_id=None, cantidad=Decimal(str(general))))
        db.session.add(StockPorAlmacen(producto_id=p.id, almacen_id=self._bodega.id,
                                       cantidad=Decimal(str(general))))
        db.session.commit()
        return p

    def test_subir_bucket_genera_ajuste(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._producto(db, general=50)
        resp = client.post(f'/api/v1/productos/{p.id}/ajustar-buckets', json={
            'buckets': [{'almacen_id': self._bodega.id, 'proyecto_id': None, 'cantidad_objetivo': 60}],
        })
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['buckets_ajustados'] == 1
        db.session.refresh(p)
        assert float(p.stock_actual) == pytest.approx(60.0, abs=0.01)
        ajustes = MovimientoInventario.query.filter_by(producto_id=p.id, tipo='AJUSTE').all()
        assert len(ajustes) == 1 and float(ajustes[0].cantidad) == pytest.approx(10.0, abs=0.01)

    def test_agregar_bucket_proyecto(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._producto(db, general=50)
        proy = Proyecto(numero_proyecto=f'PB-{uuid.uuid4().hex[:5]}', nombre='Obra', activo=True)
        db.session.add(proy); db.session.commit()
        resp = client.post(f'/api/v1/productos/{p.id}/ajustar-buckets', json={
            'buckets': [
                {'almacen_id': self._bodega.id, 'proyecto_id': None, 'cantidad_objetivo': 50},
                {'almacen_id': self._bodega.id, 'proyecto_id': proy.id, 'cantidad_objetivo': 20},
            ],
        })
        assert resp.status_code == 200, resp.get_json()
        db.session.refresh(p)
        assert float(p.stock_actual) == pytest.approx(70.0, abs=0.01)
        bucket = StockAlmacenProyecto.query.filter_by(
            producto_id=p.id, almacen_id=self._bodega.id, proyecto_id=proy.id).first()
        assert bucket is not None and float(bucket.cantidad) == pytest.approx(20.0, abs=0.01)

    def test_ajuste_respeta_reservas(self, client, inv_admin, db):
        from decimal import Decimal
        _login(client, inv_admin.id, 'admin')
        p = self._producto(db, general=50)
        p.stock_reservado = Decimal('40')
        db.session.commit()
        resp = client.post(f'/api/v1/productos/{p.id}/ajustar-buckets', json={
            'buckets': [{'almacen_id': self._bodega.id, 'proyecto_id': None, 'cantidad_objetivo': 5}],
        })
        assert resp.status_code == 409
        db.session.refresh(p)
        assert float(p.stock_actual) == pytest.approx(50.0, abs=0.01)  # rollback

    def test_bucket_duplicado_422(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._producto(db)
        resp = client.post(f'/api/v1/productos/{p.id}/ajustar-buckets', json={
            'buckets': [
                {'almacen_id': self._bodega.id, 'proyecto_id': None, 'cantidad_objetivo': 5},
                {'almacen_id': self._bodega.id, 'proyecto_id': None, 'cantidad_objetivo': 7},
            ],
        })
        assert resp.status_code == 422

    def test_almacen_inexistente_422(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._producto(db)
        resp = client.post(f'/api/v1/productos/{p.id}/ajustar-buckets', json={
            'buckets': [{'almacen_id': 999999, 'proyecto_id': None, 'cantidad_objetivo': 5}],
        })
        assert resp.status_code == 422


# ═══════════════════════════════════════════════════════════════════════════════
# 14. IMPORTACIÓN — destino por Almacén + Proyecto
# ═══════════════════════════════════════════════════════════════════════════════

class TestImportarBuckets:

    def _post(self, client, rows):
        import io as _io
        data = {'archivo': (_io.BytesIO(_build_import_xlsx(rows)), 'materiales.xlsx')}
        return client.post('/api/v1/productos/importar', data=data,
                           content_type='multipart/form-data')

    def test_importar_deposita_en_bucket(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        cdmx = Almacen(nombre='CDMX', qr_code=str(uuid.uuid4()), activo=True)
        db.session.add(cdmx)
        proy = Proyecto(numero_proyecto='PROY-A', nombre='Nave', activo=True)
        db.session.add(proy); db.session.commit()
        resp = self._post(client, [{
            'Código (SKU)': 'IMP-BK-1', 'Descripción': 'Tornillo', 'Categoría': 'Tornillería',
            'Unidad': 'pza', 'Stock Inicial': 30, 'Stock Mínimo': 0,
            'Almacén': 'CDMX', 'Proyecto': 'PROY-A',
        }])
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['exitosos'] == 1
        p = Producto.query.filter_by(codigo='IMP-BK-1').first()
        assert p is not None
        bucket = StockAlmacenProyecto.query.filter_by(
            producto_id=p.id, almacen_id=cdmx.id, proyecto_id=proy.id).first()
        assert bucket is not None and float(bucket.cantidad) == pytest.approx(30.0, abs=0.01)
        cache = StockPorAlmacen.query.filter_by(producto_id=p.id, almacen_id=cdmx.id).first()
        assert cache is not None and float(cache.cantidad) == pytest.approx(30.0, abs=0.01)

    def test_importar_almacen_inexistente_error(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        # Debe existir al menos una bodega activa (default) para el resto del flujo.
        db.session.add(Almacen(nombre='Central', qr_code=str(uuid.uuid4()), activo=True))
        db.session.commit()
        resp = self._post(client, [{
            'Código (SKU)': 'IMP-BK-2', 'Descripción': 'Tuerca', 'Categoría': 'Tornillería',
            'Unidad': 'pza', 'Stock Inicial': 10, 'Stock Mínimo': 0,
            'Almacén': 'NoExiste',
        }])
        assert resp.status_code == 200, resp.get_json()
        body = resp.get_json()
        assert body['exitosos'] == 0
        assert len(body['errores']) == 1 and 'almac' in body['errores'][0].lower()
        assert Producto.query.filter_by(codigo='IMP-BK-2').first() is None


# ═══════════════════════════════════════════════════════════════════════════════
# 15. RESUMEN POR PROYECTO Y ALMACÉN (portada)
# ═══════════════════════════════════════════════════════════════════════════════

class TestResumenProyectos:

    def _seed(self, db):
        from decimal import Decimal
        a = Almacen(nombre='Alm A', qr_code=str(uuid.uuid4()), activo=True)
        b = Almacen(nombre='Alm B', qr_code=str(uuid.uuid4()), activo=True)
        db.session.add_all([a, b])
        proy = Proyecto(numero_proyecto=f'RP-{uuid.uuid4().hex[:5]}', nombre='Obra', activo=True)
        p = Producto(codigo=f'RP-{uuid.uuid4().hex[:6]}', descripcion='X', categoria='T',
                     unidad='pza', stock_actual=Decimal('150'), stock_minimo=0)
        db.session.add_all([proy, p]); db.session.flush()
        # General en A(100) y B(10); proyecto en A(40).
        db.session.add_all([
            StockAlmacenProyecto(producto_id=p.id, almacen_id=a.id, proyecto_id=None, cantidad=Decimal('100')),
            StockAlmacenProyecto(producto_id=p.id, almacen_id=b.id, proyecto_id=None, cantidad=Decimal('10')),
            StockAlmacenProyecto(producto_id=p.id, almacen_id=a.id, proyecto_id=proy.id, cantidad=Decimal('40')),
        ])
        db.session.commit()
        return a, b, proy

    def test_resumen_matriz(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        a, b, proy = self._seed(db)
        resp = client.get('/api/v1/almacenes/resumen-proyectos')
        assert resp.status_code == 200, resp.get_json()
        data = resp.get_json()
        # Columnas: ambos almacenes activos.
        nombres_alm = {c['nombre'] for c in data['almacenes']}
        assert {'Alm A', 'Alm B'}.issubset(nombres_alm)
        assert data['total_unidades'] == pytest.approx(150.0, abs=0.01)
        # Productos distintos con existencia en todo el inventario (1 SKU).
        assert data['total_productos'] == 1
        # General primero.
        assert data['filas'][0]['es_general'] is True
        general = data['filas'][0]
        assert general['total_unidades'] == pytest.approx(110.0, abs=0.01)
        assert general['total_productos'] == 1
        assert general['celdas'][str(a.id)]['unidades'] == pytest.approx(100.0, abs=0.01)
        assert general['celdas'][str(b.id)]['unidades'] == pytest.approx(10.0, abs=0.01)
        # Fila del proyecto.
        fila_proy = next(f for f in data['filas'] if f['proyecto_id'] == proy.id)
        assert fila_proy['total_unidades'] == pytest.approx(40.0, abs=0.01)
        assert fila_proy['celdas'][str(a.id)]['unidades'] == pytest.approx(40.0, abs=0.01)
        assert str(b.id) not in fila_proy['celdas']

    def test_resumen_requiere_login(self, client):
        resp = client.get('/api/v1/almacenes/resumen-proyectos')
        assert resp.status_code == 401


# ═══════════════════════════════════════════════════════════════════════════════
# TOMAS FÍSICAS — atomicidad del cierre
# ═══════════════════════════════════════════════════════════════════════════════

class TestCierreTomaAtomico:
    """El cierre de una toma aplica N ajustes: o todos, o ninguno.

    Regresión: `_perform_movimiento` commiteaba cada ajuste por separado, así que
    un fallo a media lista dejaba la toma ABIERTA con parte del stock ya movido.
    Como `cantidad_sistema` es un snapshot fijado al abrir la toma, al reintentar
    el cierre se recalculaba la misma diferencia y los ajustes ya aplicados se
    volvían a aplicar — duplicando el descuento.
    """

    @pytest.fixture(autouse=True)
    def _setup(self, db, inv_admin, client):
        from decimal import Decimal
        self.bodega = Almacen(
            nombre=f'B-{uuid.uuid4().hex[:5]}', qr_code=str(uuid.uuid4()), activo=True,
        )
        db.session.add(self.bodega)
        db.session.commit()
        _login(client, inv_admin.id)
        self._db = db
        self._admin = inv_admin

    def _producto(self, stock, reservado=0):
        from decimal import Decimal
        p = Producto(
            codigo=f'TA-{uuid.uuid4().hex[:6]}', descripcion='Prod toma',
            categoria='T', unidad='pza',
            stock_actual=Decimal(str(stock)), stock_minimo=0,
            stock_reservado=Decimal(str(reservado)),
        )
        self._db.session.add(p)
        self._db.session.flush()
        self._db.session.add(StockAlmacenProyecto(
            producto_id=p.id, almacen_id=self.bodega.id, proyecto_id=None,
            cantidad=Decimal(str(stock)),
        ))
        self._db.session.add(StockPorAlmacen(
            producto_id=p.id, almacen_id=self.bodega.id, cantidad=Decimal(str(stock)),
        ))
        self._db.session.commit()
        return p

    def test_ajuste_sin_autocommit_es_revertible(self, app):
        """El corazón del arreglo: con `autocommit=False` el ajuste NO se
        commitea, así que el llamador puede deshacerlo.

        Antes `_perform_movimiento` commiteaba internamente, y por eso el cierre
        de toma no podía revertir las líneas ya aplicadas cuando otra fallaba.
        Se prueba a este nivel y no vía el endpoint porque el `db` de conftest
        envuelve cada test en su propia transacción, así que un
        `session.rollback()` dentro del request no se puede observar desde fuera.
        """
        from decimal import Decimal
        from app.routes.inventario_api.movimientos import _perform_movimiento

        prod = self._producto(stock=10)
        punto = self._db.session.begin_nested()
        with app.test_request_context():
            resp = _perform_movimiento(
                {
                    'tipo': 'AJUSTE', 'producto_id': prod.id,
                    'cantidad': Decimal('-4'), 'almacen_origen_id': self.bodega.id,
                    'motivo': 'toma test', 'reconciliar': True,
                },
                self._admin,
                autocommit=False,
            )
        assert not isinstance(resp, tuple), getattr(resp, 'json', resp)

        # Dentro de la transacción el descuento ya se ve...
        assert flask_db.session.get(Producto, prod.id).stock_actual == 6

        # ...pero deshacer el savepoint lo revierte: no hubo commit interno.
        punto.rollback()
        self._db.session.expire_all()
        assert flask_db.session.get(Producto, prod.id).stock_actual == 10, (
            'el ajuste sobrevivió al rollback — _perform_movimiento commiteó por dentro'
        )
        bucket = StockAlmacenProyecto.query.filter_by(
            producto_id=prod.id, almacen_id=self.bodega.id, proyecto_id=None,
        ).first()
        assert bucket.cantidad == 10, 'el bucket quedó descuadrado tras el rollback'

    def test_cierre_con_linea_invalida_deja_la_toma_abierta(self, client):
        """`bloqueado` no se puede ajustar (invadiría stock apartado): el cierre
        responde 409, reporta el error y la toma NO pasa a CERRADA."""
        bloqueado = self._producto(stock=10, reservado=10)

        toma_id = client.post(
            '/api/v1/tomas/', json={'almacen_id': self.bodega.id},
        ).get_json()['id']
        detalles = client.get(f'/api/v1/tomas/{toma_id}').get_json()['detalles']
        det = next(d for d in detalles if d['producto_id'] == bloqueado.id)
        client.patch(
            f"/api/v1/tomas/{toma_id}/detalles/{det['id']}",
            json={'cantidad_fisica': 5},
        )

        resp = client.post(f'/api/v1/tomas/{toma_id}/cerrar', json={})
        assert resp.status_code == 409, resp.get_json()
        cuerpo = resp.get_json()
        assert cuerpo['errores'], 'debe reportar qué línea falló'
        assert cuerpo['errores'][0]['producto_id'] == bloqueado.id
        assert flask_db.session.get(TomaInventario, toma_id).estatus == 'ABIERTA'

    def test_cierre_sin_fallos_aplica_todo(self, client):
        """Camino feliz: sin errores, los ajustes SÍ se aplican y la toma cierra."""
        a = self._producto(stock=10)
        b = self._producto(stock=8)

        toma_id = client.post(
            '/api/v1/tomas/', json={'almacen_id': self.bodega.id},
        ).get_json()['id']
        detalles = client.get(f'/api/v1/tomas/{toma_id}').get_json()['detalles']
        por_prod = {d['producto_id']: d for d in detalles}

        for prod, fisico in ((a, 7), (b, 8)):
            client.patch(
                f"/api/v1/tomas/{toma_id}/detalles/{por_prod[prod.id]['id']}",
                json={'cantidad_fisica': fisico},
            )

        resp = client.post(f'/api/v1/tomas/{toma_id}/cerrar', json={})
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['ajustes_creados'] == 1  # solo `a` tenía diferencia

        self._db.session.expire_all()
        assert flask_db.session.get(Producto, a.id).stock_actual == 7
        assert flask_db.session.get(Producto, b.id).stock_actual == 8
        assert flask_db.session.get(TomaInventario, toma_id).estatus == 'CERRADA'


# ═══════════════════════════════════════════════════════════════════════════════
# CATEGORÍAS CON BARRA EN EL NOMBRE
# ═══════════════════════════════════════════════════════════════════════════════

class TestCategoriaConBarraEnElNombre:
    """Regresión: `Tubería/Accesorios` no se podía editar ni borrar.

    Las rutas usaban el convertidor `<string:nombre>`, que NO acepta barras. El
    navegador manda el nombre con `%2F`, pero la capa WSGI lo decodifica a `/`
    ANTES del enrutado, así que la ruta dejaba de coincidir y Flask respondía
    404 desde el router — sin llegar nunca al endpoint. La pista para
    diagnosticarlo fue justamente que el 404 venía con cuerpo vacío en lugar del
    JSON con `detail` que devuelve la vista.

    El arreglo es `<path:nombre>`, que sí acepta barras. Es seguro porque ese
    nombre solo se usa para consultar la base, nunca para construir rutas de
    archivos.
    """

    NOMBRE = 'Tubería/Accesorios'

    @pytest.fixture(autouse=True)
    def _setup(self, client, db, inv_admin):
        _login(client, inv_admin.id)
        self._db = db

    def _crear_config(self, db, nombre):
        from app.models.inventario import CategoriaConfig
        cfg = CategoriaConfig(nombre=nombre)
        db.session.add(cfg)
        db.session.commit()
        return cfg

    def test_borrar_categoria_con_barra(self, client, db):
        from app.models.inventario import CategoriaConfig
        self._crear_config(db, self.NOMBRE)

        # `quote(safe='')` reproduce lo que hace encodeURIComponent en el SPA.
        from urllib.parse import quote
        r = client.delete(
            f"/api/v1/categorias-config/{quote(self.NOMBRE, safe='')}?con_productos=1"
        )
        assert r.status_code == 200, r.get_json()
        assert CategoriaConfig.query.filter_by(nombre=self.NOMBRE).first() is None

    def test_borrar_categoria_con_barra_sin_productos(self, client, db):
        """El caso reportado: la categoría no tiene productos y aun así fallaba."""
        from urllib.parse import quote
        from app.models.inventario import CategoriaConfig
        self._crear_config(db, 'ZZ Sin Productos/Con Barra')

        r = client.delete(
            f"/api/v1/categorias-config/{quote('ZZ Sin Productos/Con Barra', safe='')}"
            f"?con_productos=1"
        )
        assert r.status_code == 200, r.get_json()
        assert CategoriaConfig.query.filter_by(nombre='ZZ Sin Productos/Con Barra').first() is None

    def test_editar_categoria_con_barra(self, client, db):
        from urllib.parse import quote
        from app.models.inventario import CategoriaConfig
        self._crear_config(db, self.NOMBRE)

        r = client.put(
            f"/api/v1/categorias-config/{quote(self.NOMBRE, safe='')}",
            json={'imagen_url': 'https://ejemplo.test/tuberia.webp'},
        )
        assert r.status_code == 200, r.get_json()
        cfg = CategoriaConfig.query.filter_by(nombre=self.NOMBRE).first()
        assert cfg.imagen_url == 'https://ejemplo.test/tuberia.webp'

    def test_categoria_sin_barra_sigue_funcionando(self, client, db):
        """El cambio de convertidor no debe alterar el caso normal."""
        from urllib.parse import quote
        from app.models.inventario import CategoriaConfig
        self._crear_config(db, 'ZZ Normal')

        r = client.delete(
            f"/api/v1/categorias-config/{quote('ZZ Normal', safe='')}?con_productos=1"
        )
        assert r.status_code == 200, r.get_json()
        assert CategoriaConfig.query.filter_by(nombre='ZZ Normal').first() is None

    def test_categoria_inexistente_sigue_devolviendo_404_del_endpoint(self, client):
        """Este 404 debe venir de la vista (con `detail`), no del router.

        Distinguirlos importa: un 404 sin cuerpo significa que la ruta ni
        siquiera coincidió, que es exactamente el bug que se corrigió.
        """
        r = client.delete('/api/v1/categorias-config/ZZ No Existe/Nada?con_productos=1')
        assert r.status_code == 404
        assert r.get_json() is not None, 'el 404 vino del router, no del endpoint'
        assert 'detail' in r.get_json()


# ═══════════════════════════════════════════════════════════════════════════════
# 20. EXPORT DE CATÁLOGO Y PLANTILLA CON DESTINO (almacén / proyecto)
# ═══════════════════════════════════════════════════════════════════════════════

class TestMinimosMasivos:
    """El stock mínimo se puede fijar en masa y sugerir desde el consumo real."""

    def _producto(self, db, codigo, unidad='pza', minimo=0):
        from decimal import Decimal
        p = Producto(codigo=codigo, descripcion=f'Prod {codigo}', categoria='ZZ Min',
                     unidad=unidad, stock_actual=Decimal('100'),
                     stock_minimo=Decimal(str(minimo)))
        db.session.add(p); db.session.commit()
        return p

    def test_fijar_el_mismo_minimo_a_varios(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        a = self._producto(db, 'MIN-1')
        b = self._producto(db, 'MIN-2')
        resp = client.patch('/api/v1/productos/minimos',
                            json={'producto_ids': [a.id, b.id], 'stock_minimo': 12})
        assert resp.status_code == 200, resp.get_json()
        body = resp.get_json()
        assert body['actualizados'] == 2 and body['errores'] == []
        db.session.refresh(a); db.session.refresh(b)
        assert float(a.stock_minimo) == 12.0 and float(b.stock_minimo) == 12.0

    def test_minimos_distintos_por_producto(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        a = self._producto(db, 'MIN-3')
        b = self._producto(db, 'MIN-4')
        resp = client.patch('/api/v1/productos/minimos', json={'items': [
            {'id': a.id, 'stock_minimo': 5},
            {'id': b.id, 'stock_minimo': 9},
        ]})
        assert resp.status_code == 200, resp.get_json()
        db.session.refresh(a); db.session.refresh(b)
        assert float(a.stock_minimo) == 5.0 and float(b.stock_minimo) == 9.0

    def test_no_toca_el_stock_real(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._producto(db, 'MIN-5')
        client.patch('/api/v1/productos/minimos',
                     json={'producto_ids': [p.id], 'stock_minimo': 3})
        db.session.refresh(p)
        assert float(p.stock_actual) == 100.0

    def test_rechaza_decimales_en_unidad_entera(self, client, inv_admin, db):
        """Misma regla que el alta manual y que la importación."""
        _login(client, inv_admin.id, 'admin')
        p = self._producto(db, 'MIN-6', unidad='pza', minimo=1)
        resp = client.patch('/api/v1/productos/minimos',
                            json={'producto_ids': [p.id], 'stock_minimo': 2.5})
        assert resp.status_code == 200, resp.get_json()
        body = resp.get_json()
        assert body['actualizados'] == 0 and len(body['errores']) == 1
        db.session.refresh(p)
        assert float(p.stock_minimo) == 1.0

    def test_acepta_decimales_en_unidad_continua(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._producto(db, 'MIN-7', unidad='Mts')
        resp = client.patch('/api/v1/productos/minimos',
                            json={'producto_ids': [p.id], 'stock_minimo': 2.5})
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['actualizados'] == 1

    def test_sugerencia_usa_el_consumo_real(self, client, inv_admin, db):
        """30 unidades salidas en 30 días = 1/día → 15 días de cobertura = 15."""
        import datetime
        from decimal import Decimal
        _login(client, inv_admin.id, 'admin')
        p = self._producto(db, 'MIN-8')
        alm = Almacen(nombre=f'MIN-BOD-{uuid.uuid4().hex[:4]}', qr_code=str(uuid.uuid4()), activo=True)
        db.session.add(alm); db.session.commit()
        db.session.add(MovimientoInventario(
            tipo='SALIDA', producto_id=p.id, almacen_origen_id=alm.id,
            cantidad=Decimal('30'), usuario_id=inv_admin.id,
            fecha=datetime.datetime.now() - datetime.timedelta(days=5),
        ))
        db.session.commit()

        resp = client.post('/api/v1/productos/minimos/sugerencia', json={
            'producto_ids': [p.id], 'dias_consumo': 30, 'dias_cobertura': 15,
        })
        assert resp.status_code == 200, resp.get_json()
        item = resp.get_json()['items'][0]
        assert item['consumo_diario'] == pytest.approx(1.0, abs=0.01)
        assert item['sugerido'] == pytest.approx(15.0, abs=0.01)
        assert item['sin_consumo'] is False

    def test_sugerencia_marca_los_que_no_se_mueven(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._producto(db, 'MIN-9')
        resp = client.post('/api/v1/productos/minimos/sugerencia',
                           json={'producto_ids': [p.id]})
        assert resp.status_code == 200, resp.get_json()
        item = resp.get_json()['items'][0]
        assert item['sin_consumo'] is True and item['sugerido'] == 0

    def test_sugerencia_redondea_hacia_arriba_en_piezas(self, client, inv_admin, db):
        """1.2 piezas de mínimo no existe: se pide 2."""
        import datetime
        from decimal import Decimal
        _login(client, inv_admin.id, 'admin')
        p = self._producto(db, 'MIN-10')
        db.session.add(MovimientoInventario(
            tipo='SALIDA', producto_id=p.id, cantidad=Decimal('12'),
            usuario_id=inv_admin.id, fecha=datetime.datetime.now() - datetime.timedelta(days=2),
        ))
        db.session.commit()
        resp = client.post('/api/v1/productos/minimos/sugerencia', json={
            'producto_ids': [p.id], 'dias_consumo': 30, 'dias_cobertura': 3,
        })
        item = resp.get_json()['items'][0]
        assert item['consumo_diario'] == pytest.approx(0.4, abs=0.01)
        assert item['sugerido'] == 2  # ceil(1.2)

    def test_valida_parametros(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        p = self._producto(db, 'MIN-11')
        assert client.post('/api/v1/productos/minimos/sugerencia',
                           json={'producto_ids': []}).status_code == 422
        assert client.post('/api/v1/productos/minimos/sugerencia',
                           json={'producto_ids': [p.id], 'dias_consumo': 5000}).status_code == 422
        assert client.patch('/api/v1/productos/minimos', json={}).status_code == 422

    def test_quien_no_escribe_inventario_no_puede(self, client, inv_solicitante, db):
        """Mismo guard que el resto de la escritura del catálogo: inventario y
        admin sí; un solicitante de material, no."""
        _login(client, inv_solicitante.id, 'solicitante_material')
        p = self._producto(db, 'MIN-12')
        resp = client.patch('/api/v1/productos/minimos',
                            json={'producto_ids': [p.id], 'stock_minimo': 5})
        assert resp.status_code == 403
        resp = client.post('/api/v1/productos/minimos/sugerencia',
                           json={'producto_ids': [p.id]})
        assert resp.status_code == 403


class TestDeshacerImportacion:
    """Una importación aplicada se puede revertir, sin pisar trabajo posterior."""

    def _importar(self, client, filas):
        import io as _io
        resp = client.post(
            '/api/v1/productos/importar',
            data={'archivo': (_io.BytesIO(_build_import_xlsx(filas)), 'carga.xlsx')},
            content_type='multipart/form-data')
        assert resp.status_code == 200, resp.get_json()
        return resp.get_json()

    def test_importar_registra_el_lote(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        body = self._importar(client, [{
            'Código (SKU)': 'UND-1', 'Descripción': 'Tornillo', 'Categoría': 'ZZ Und',
            'Unidad': 'pza', 'Stock Inicial': 5,
        }])
        assert body['importacion_id']
        lista = client.get('/api/v1/productos/importaciones').get_json()
        assert lista[0]['id'] == body['importacion_id']
        assert lista[0]['archivo'] == 'carga.xlsx'
        assert lista[0]['creados'] == 1
        assert lista[0]['puede_deshacerse'] is True

    def test_deshacer_borra_los_productos_creados(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        alm = Almacen(nombre='UND-BOD', qr_code=str(uuid.uuid4()), activo=True)
        db.session.add(alm); db.session.commit()
        body = self._importar(client, [{
            'Código (SKU)': 'UND-2', 'Descripción': 'Tuerca', 'Categoría': 'ZZ Und',
            'Unidad': 'pza', 'Stock Inicial': 8, 'Almacén': 'UND-BOD',
        }])
        p = Producto.query.filter_by(codigo='UND-2').first()
        assert p is not None
        pid = p.id

        resp = client.post(f"/api/v1/productos/importaciones/{body['importacion_id']}/deshacer")
        assert resp.status_code == 200, resp.get_json()
        r = resp.get_json()
        assert r['eliminados'] == 1 and r['desactivados'] == 0
        assert Producto.query.filter_by(codigo='UND-2').first() is None
        # El stock que había depositado también se retira.
        assert StockAlmacenProyecto.query.filter_by(producto_id=pid).first() is None
        assert StockPorAlmacen.query.filter_by(producto_id=pid).first() is None

    def test_deshacer_restaura_los_campos_actualizados(self, client, inv_admin, db):
        from decimal import Decimal
        p = Producto(codigo='UND-3', descripcion='Cinta vieja', categoria='ZZ Und',
                     unidad='pza', stock_actual=Decimal('4'), stock_minimo=Decimal('1'),
                     precio_unitario=Decimal('10'))
        db.session.add(p); db.session.commit()
        _login(client, inv_admin.id, 'admin')
        body = self._importar(client, [{
            'Código (SKU)': 'UND-3', 'Descripción': 'Cinta nueva', 'Categoría': 'ZZ Und',
            'Unidad': 'pza', 'Precio Unitario': 99, 'Stock Mínimo': 7,
        }])
        db.session.refresh(p)
        assert p.descripcion == 'Cinta nueva' and float(p.precio_unitario) == 99.0

        resp = client.post(f"/api/v1/productos/importaciones/{body['importacion_id']}/deshacer")
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['restaurados'] == 1
        db.session.refresh(p)
        assert p.descripcion == 'Cinta vieja'
        assert float(p.precio_unitario) == 10.0
        assert float(p.stock_minimo) == 1.0

    def test_deshacer_respeta_lo_editado_despues(self, client, inv_admin, db):
        """Si alguien editó el campo DESPUÉS de importar, deshacer no lo pisa."""
        from decimal import Decimal
        p = Producto(codigo='UND-4', descripcion='Original', categoria='ZZ Und', unidad='pza',
                     stock_actual=Decimal('1'), stock_minimo=0, precio_unitario=Decimal('5'))
        db.session.add(p); db.session.commit()
        _login(client, inv_admin.id, 'admin')
        body = self._importar(client, [{
            'Código (SKU)': 'UND-4', 'Descripción': 'Importado', 'Categoría': 'ZZ Und',
            'Unidad': 'pza', 'Precio Unitario': 50,
        }])
        # Edición manual posterior: manda sobre el deshacer.
        r = client.put(f'/api/v1/productos/{p.id}', json={'precio_unitario': 77})
        assert r.status_code == 200, r.get_json()

        resp = client.post(f"/api/v1/productos/importaciones/{body['importacion_id']}/deshacer")
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['campos_omitidos'] >= 1
        db.session.refresh(p)
        assert float(p.precio_unitario) == 77.0   # respetado
        assert p.descripcion == 'Original'        # sí revertido

    def test_deshacer_desactiva_si_el_producto_ya_se_movio(self, client, inv_admin, db):
        """Un alta que ya tuvo movimientos no se borra: se da de baja lógica."""
        _login(client, inv_admin.id, 'admin')
        alm = Almacen(nombre='UND-BOD2', qr_code=str(uuid.uuid4()), activo=True)
        db.session.add(alm); db.session.commit()
        body = self._importar(client, [{
            'Código (SKU)': 'UND-5', 'Descripción': 'Clavo', 'Categoría': 'ZZ Und',
            'Unidad': 'pza', 'Stock Inicial': 10, 'Almacén': 'UND-BOD2',
        }])
        p = Producto.query.filter_by(codigo='UND-5').first()
        mov = client.post('/api/v1/movimientos/', json={
            'tipo': 'ENTRADA', 'producto_id': p.id, 'almacen_destino_id': alm.id,
            'cantidad': 5, 'motivo': 'compra',
        })
        assert mov.status_code == 200, mov.get_json()

        resp = client.post(f"/api/v1/productos/importaciones/{body['importacion_id']}/deshacer")
        assert resp.status_code == 200, resp.get_json()
        r = resp.get_json()
        assert r['eliminados'] == 0 and r['desactivados'] == 1
        db.session.refresh(p)
        assert p.activo is False          # sigue existiendo, con su histórico
        assert r['notas']                  # y se explica por qué

    def test_deshacer_nunca_borra_movimientos(self, client, inv_admin, db):
        """Producto.movimientos tiene cascade='all, delete-orphan': borrar el
        producto arrastraría su histórico de movimientos. La guarda que desactiva
        en vez de borrar es lo único que lo impide — si se cae, se pierde
        histórico en silencio."""
        _login(client, inv_admin.id, 'admin')
        alm = Almacen(nombre=f'UND-MOV-{uuid.uuid4().hex[:4]}', qr_code=str(uuid.uuid4()), activo=True)
        db.session.add(alm); db.session.commit()
        body = self._importar(client, [
            {'Código (SKU)': 'UND-MOV-1', 'Descripción': 'Con movimiento',
             'Categoría': 'ZZ Und', 'Unidad': 'pza', 'Stock Inicial': 10,
             'Almacén': alm.nombre},
            {'Código (SKU)': 'UND-MOV-2', 'Descripción': 'Intacto',
             'Categoría': 'ZZ Und', 'Unidad': 'pza'},
        ])
        p1 = Producto.query.filter_by(codigo='UND-MOV-1').first()
        r = client.post('/api/v1/movimientos/', json={
            'tipo': 'SALIDA', 'producto_id': p1.id, 'almacen_origen_id': alm.id,
            'cantidad': 2, 'motivo': 'uso',
        })
        assert r.status_code == 200, r.get_json()
        movs_antes = MovimientoInventario.query.filter_by(producto_id=p1.id).count()
        assert movs_antes >= 1

        resp = client.post(f"/api/v1/productos/importaciones/{body['importacion_id']}/deshacer")
        assert resp.status_code == 200, resp.get_json()
        rr = resp.get_json()
        # El que se movió sobrevive con su histórico; el intacto sí se borra.
        assert rr['desactivados'] == 1 and rr['eliminados'] == 1
        assert MovimientoInventario.query.filter_by(producto_id=p1.id).count() == movs_antes
        assert Producto.query.filter_by(codigo='UND-MOV-2').first() is None

    def test_no_se_puede_deshacer_dos_veces(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        body = self._importar(client, [{
            'Código (SKU)': 'UND-6', 'Descripción': 'Taquete', 'Categoría': 'ZZ Und',
            'Unidad': 'pza',
        }])
        url = f"/api/v1/productos/importaciones/{body['importacion_id']}/deshacer"
        assert client.post(url).status_code == 200
        segunda = client.post(url)
        assert segunda.status_code == 400
        assert 'revertida' in segunda.get_json()['detail'].lower()

    def test_previsualizar_no_registra_lote(self, client, inv_admin, db):
        """El plan no es una importación: no debe aparecer en el historial."""
        import io as _io
        _login(client, inv_admin.id, 'admin')
        antes = len(client.get('/api/v1/productos/importaciones').get_json())
        client.post('/api/v1/productos/importar', data={
            'archivo': (_io.BytesIO(_build_import_xlsx([{
                'Código (SKU)': 'UND-7', 'Descripción': 'X', 'Categoría': 'ZZ Und',
                'Unidad': 'pza',
            }])), 'plan.xlsx'),
            'previsualizar': '1',
        }, content_type='multipart/form-data')
        assert len(client.get('/api/v1/productos/importaciones').get_json()) == antes

    def test_deshacer_sin_bodega_activa_si_borra_el_alta(self, client, inv_admin, db):
        """Sin bodegas activas el stock inicial no cae en ningún bucket. El
        registro debe guardarlo igual: si no, al deshacer el producto parecería
        "ya tocado" (stock 10 vs 0 esperado) y se desactivaría en vez de borrarse.
        """
        _login(client, inv_admin.id, 'admin')
        # Ninguna bodega activa en este test.
        assert Almacen.query.filter_by(activo=True).first() is None
        body = self._importar(client, [{
            'Código (SKU)': 'UND-SB', 'Descripción': 'Sin bodega', 'Categoría': 'ZZ Und',
            'Unidad': 'pza', 'Stock Inicial': 10,
        }])
        assert Producto.query.filter_by(codigo='UND-SB').first() is not None

        resp = client.post(f"/api/v1/productos/importaciones/{body['importacion_id']}/deshacer")
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['eliminados'] == 1
        assert Producto.query.filter_by(codigo='UND-SB').first() is None

    def test_deshacer_masivo_no_consulta_por_producto(self, client, inv_admin, db):
        """Deshacer debe resolver los productos en lotes, no uno por uno: con
        cientos de altas serían cientos de viajes a la base en una sola petición.
        """
        from sqlalchemy import event
        _login(client, inv_admin.id, 'admin')
        N = 120
        body = self._importar(client, [{
            'Código (SKU)': f'UNDM-{i:03d}', 'Descripción': f'Prod {i}',
            'Categoría': 'ZZ Und', 'Unidad': 'pza',
        } for i in range(N)])
        assert body['exitosos'] == N

        consultas = {'n': 0}
        engine = db.engine

        # Se cuentan solo las LECTURAS: los DELETE son uno por producto y son
        # inherentes al borrado. Lo que no debe escalar con N es la resolución
        # de "qué producto es este" y "¿tuvo movimientos?".
        @event.listens_for(engine, 'before_cursor_execute')
        def _contar(conn, cursor, statement, params, context, executemany):
            s = statement.lstrip().upper()
            if s.startswith('SELECT') and (
                    'FROM PRODUCTOS' in s or 'FROM MOVIMIENTOS_INVENTARIO' in s):
                consultas['n'] += 1

        try:
            resp = client.post(f"/api/v1/productos/importaciones/{body['importacion_id']}/deshacer")
        finally:
            event.remove(engine, 'before_cursor_execute', _contar)

        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['eliminados'] == N
        assert consultas['n'] < 20, (
            f'{consultas["n"]} lecturas para deshacer {N} altas: volvió el SELECT por producto')

    def test_detalle_lista_lo_que_hizo_el_lote(self, client, inv_admin, db):
        from decimal import Decimal
        db.session.add(Producto(codigo='UND-8', descripcion='Vieja', categoria='ZZ Und',
                                unidad='pza', stock_actual=Decimal('1'), stock_minimo=0))
        db.session.commit()
        _login(client, inv_admin.id, 'admin')
        body = self._importar(client, [
            {'Código (SKU)': 'UND-8', 'Descripción': 'Nueva', 'Categoría': 'ZZ Und',
             'Unidad': 'pza'},
            {'Código (SKU)': 'UND-9', 'Descripción': 'Alta', 'Categoría': 'ZZ Und',
             'Unidad': 'pza'},
        ])
        det = client.get(f"/api/v1/productos/importaciones/{body['importacion_id']}").get_json()
        assert {c['codigo'] for c in det['detalle']['creados']} == {'UND-9'}
        act = det['detalle']['actualizados'][0]
        assert act['codigo'] == 'UND-8'
        campo = next(c for c in act['campos'] if c['campo'] == 'descripcion')
        assert campo['antes'] == 'Vieja' and campo['despues'] == 'Nueva'


class TestPrevisualizarImportacion:
    """Subir el archivo muestra el plan; hasta confirmar no se escribe nada."""

    def _post(self, client, filas, previsualizar=False):
        import io as _io
        data = {'archivo': (_io.BytesIO(_build_import_xlsx(filas)), 'materiales.xlsx')}
        if previsualizar:
            data['previsualizar'] = '1'
        return client.post('/api/v1/productos/importar', data=data,
                           content_type='multipart/form-data')

    FILA_NUEVA = {
        'Código (SKU)': 'PREV-1', 'Descripción': 'Tornillo nuevo', 'Categoría': 'ZZ Prev',
        'Unidad': 'pza', 'Stock Inicial': 10, 'Stock Mínimo': 2, 'Precio Unitario': 5,
    }

    def test_previsualizar_no_escribe_nada(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        resp = self._post(client, [self.FILA_NUEVA], previsualizar=True)
        assert resp.status_code == 200, resp.get_json()
        body = resp.get_json()
        assert body['previsualizacion'] is True
        assert body['exitosos'] == 1
        assert body['nuevos'][0]['codigo'] == 'PREV-1'
        # Nada tocado: ni el producto ni la categoría nueva.
        assert Producto.query.filter_by(codigo='PREV-1').first() is None
        from app.models.inventario import CategoriaConfig
        assert CategoriaConfig.query.filter_by(nombre='ZZ Prev').first() is None

    def test_previsualizar_no_modifica_al_existente(self, client, inv_admin, db):
        from decimal import Decimal
        p = Producto(codigo='PREV-2', descripcion='Tuerca', categoria='Tuercas', unidad='pza',
                     marca='Vieja', stock_actual=Decimal('5'), stock_minimo=0,
                     precio_unitario=Decimal('2'))
        db.session.add(p); db.session.commit()
        _login(client, inv_admin.id, 'admin')
        resp = self._post(client, [{
            'Código (SKU)': 'PREV-2', 'Descripción': 'Tuerca larga', 'Categoría': 'Tuercas',
            'Unidad': 'pza', 'Precio Unitario': 9,
        }], previsualizar=True)
        assert resp.status_code == 200, resp.get_json()
        body = resp.get_json()
        assert body['actualizados'] == 1
        cambios = body['cambios_detalle'][0]['cambios']
        assert 'descripción' in cambios
        assert any('precio: 2.0 → 9.0' in c for c in cambios)
        db.session.refresh(p)
        assert p.descripcion == 'Tuerca' and float(p.precio_unitario) == 2.0

    def test_el_plan_coincide_con_lo_aplicado(self, client, inv_admin, db):
        """Lo que promete la previsualización es exactamente lo que ocurre."""
        from decimal import Decimal
        db.session.add(Producto(codigo='PREV-3', descripcion='Cinta', categoria='ZZ Prev2',
                                unidad='pza', marca='A', stock_actual=Decimal('1'), stock_minimo=0))
        db.session.commit()
        _login(client, inv_admin.id, 'admin')
        filas = [
            {'Código (SKU)': 'PREV-3', 'Descripción': 'Cinta', 'Categoría': 'ZZ Prev2',
             'Unidad': 'pza', 'Marca': 'B'},
            {'Código (SKU)': 'PREV-4', 'Descripción': 'Clavo', 'Categoría': 'ZZ Prev2',
             'Unidad': 'pza', 'Stock Inicial': 3},
        ]
        plan = self._post(client, filas, previsualizar=True).get_json()
        real = self._post(client, filas).get_json()
        for k in ('exitosos', 'actualizados', 'sin_cambios', 'total_procesadas'):
            assert plan[k] == real[k], f'{k}: plan={plan[k]} real={real[k]}'
        assert plan['cambios_detalle'] == real['cambios_detalle']
        assert Producto.query.filter_by(codigo='PREV-4').first() is not None

    def test_previsualizar_avisa_de_posible_duplicado(self, client, inv_admin, db):
        """Un material nuevo con la MISMA descripción que uno existente (otro
        SKU) se reporta como aviso, no como error."""
        db.session.add(Producto(codigo='DUP-VIEJO', descripcion='Cable THHN 12 negro',
                                categoria='Cable', unidad='M', cable_tipo='THHN',
                                cable_calibre='12', stock_actual=0, stock_minimo=0))
        db.session.commit()
        _login(client, inv_admin.id, 'admin')
        resp = self._post(client, [{
            'Código (SKU)': 'DUP-NUEVO', 'Descripción': 'cable  thhn 12 NEGRO',
            'Categoría': 'Cable', 'Unidad': 'M',
            'Tipo (cable)': 'THHN', 'Tamaño mm²/AWG (cable)': '12',
        }], previsualizar=True)
        assert resp.status_code == 200, resp.get_json()
        body = resp.get_json()
        assert body['errores'] == []  # avisa, no bloquea
        assert len(body['duplicados']) == 1
        assert body['duplicados'][0]['codigo'] == 'DUP-NUEVO'
        assert body['duplicados'][0]['parecido_a'] == 'DUP-VIEJO'

    def test_previsualizar_reporta_errores_sin_aplicar_lo_bueno(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        resp = self._post(client, [
            self.FILA_NUEVA,
            {'Código (SKU)': 'PREV BAD!', 'Descripción': 'Malo', 'Categoría': 'ZZ Prev',
             'Unidad': 'pza'},
        ], previsualizar=True)
        assert resp.status_code == 200, resp.get_json()
        body = resp.get_json()
        assert body['exitosos'] == 1 and len(body['errores']) == 1
        assert Producto.query.filter_by(codigo='PREV-1').first() is None


class TestImportCoherencia:
    """La importación debe seguir las mismas reglas que el alta manual y no
    borrar datos que el archivo simplemente no trae."""

    def _post(self, client, headers, filas):
        """Sube un .xlsx con los encabezados EXACTOS que se le pasen."""
        import io as _io
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for f in filas:
            ws.append([f.get(h, '') for h in headers])
        buf = _io.BytesIO()
        wb.save(buf); buf.seek(0)
        return client.post('/api/v1/productos/importar',
                           data={'archivo': (buf, 'materiales.xlsx')},
                           content_type='multipart/form-data')

    BASE = ['Código (SKU)', 'Descripción', 'Categoría', 'Unidad']

    # ── No borrar lo que el archivo no trae ──────────────────────────────────

    def test_archivo_sin_columna_marca_no_borra_la_marca(self, client, inv_admin, db):
        """Un archivo viejo (sin columna Marca) NO debe vaciar la marca guardada.
        La celda ausente se leía como vacía y borraba el dato de todo el catálogo."""
        p = Producto(codigo='COH-M1', descripcion='Tornillo', categoria='Tornillería',
                     unidad='pza', marca='Truper', stock_actual=0, stock_minimo=0)
        db.session.add(p); db.session.commit()
        _login(client, inv_admin.id, 'admin')
        resp = self._post(client, self.BASE, [{
            'Código (SKU)': 'COH-M1', 'Descripción': 'Tornillo',
            'Categoría': 'Tornillería', 'Unidad': 'pza',
        }])
        assert resp.status_code == 200, resp.get_json()
        db.session.refresh(p)
        assert p.marca == 'Truper'

    def test_archivo_con_columna_marca_vacia_si_la_limpia(self, client, inv_admin, db):
        """Si la columna SÍ viene y la celda está vacía, es una orden de limpiar."""
        p = Producto(codigo='COH-M2', descripcion='Tornillo', categoria='Tornillería',
                     unidad='pza', marca='Truper', stock_actual=0, stock_minimo=0)
        db.session.add(p); db.session.commit()
        _login(client, inv_admin.id, 'admin')
        resp = self._post(client, self.BASE + ['Marca'], [{
            'Código (SKU)': 'COH-M2', 'Descripción': 'Tornillo',
            'Categoría': 'Tornillería', 'Unidad': 'pza', 'Marca': '',
        }])
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['actualizados'] == 1
        db.session.refresh(p)
        assert p.marca is None

    # ── Proveedor habitual (Compras Express) ─────────────────────────────────

    def test_import_guarda_proveedor(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        resp = self._post(client, self.BASE + ['Proveedor', 'Contacto proveedor'], [{
            'Código (SKU)': 'COH-P1', 'Descripción': 'Cinta', 'Categoría': 'Consumibles',
            'Unidad': 'pza', 'Proveedor': 'Ferretería López', 'Contacto proveedor': '5512345678',
        }])
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['exitosos'] == 1
        p = Producto.query.filter_by(codigo='COH-P1').first()
        assert p.proveedor_default_nombre == 'Ferretería López'
        assert p.proveedor_default_contacto == '5512345678'

    def test_export_trae_proveedor_y_reimportarlo_no_cambia_nada(self, client, inv_admin, db):
        import io as _io
        from openpyxl import load_workbook
        p = Producto(codigo='COH-P2', descripcion='Taquete', categoria='Fijación', unidad='pza',
                     stock_actual=0, stock_minimo=0,
                     proveedor_default_nombre='Casa Pérez', proveedor_default_contacto='ventas@x.mx')
        db.session.add(p); db.session.commit()
        _login(client, inv_admin.id, 'admin')

        exp = client.get('/api/v1/productos/exportar')
        assert exp.status_code == 200
        ws = load_workbook(_io.BytesIO(exp.data)).active
        headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
        assert 'Proveedor' in headers and 'Contacto proveedor' in headers
        fila = next(r for r in range(5, ws.max_row + 1)
                    if ws.cell(row=r, column=1).value == 'COH-P2')
        assert ws.cell(row=fila, column=headers.index('Proveedor') + 1).value == 'Casa Pérez'

        imp = client.post('/api/v1/productos/importar',
                          data={'archivo': (_io.BytesIO(exp.data), 'catalogo.xlsx')},
                          content_type='multipart/form-data')
        assert imp.status_code == 200, imp.get_json()
        assert imp.get_json()['actualizados'] == 0, imp.get_json()['cambios_detalle']
        db.session.refresh(p)
        assert p.proveedor_default_nombre == 'Casa Pérez'

    # ── Decimales según la unidad (misma regla que el alta manual) ───────────

    def test_rechaza_stock_decimal_en_unidad_entera(self, client, inv_admin, db):
        """'pza' no admite 2.5 al crear, igual que en el formulario."""
        _login(client, inv_admin.id, 'admin')
        resp = self._post(client, self.BASE + ['Stock Inicial'], [{
            'Código (SKU)': 'COH-D1', 'Descripción': 'Tuerca', 'Categoría': 'Tornillería',
            'Unidad': 'pza', 'Stock Inicial': 2.5,
        }])
        assert resp.status_code == 200, resp.get_json()
        body = resp.get_json()
        assert body['exitosos'] == 0
        assert len(body['errores']) == 1 and 'enteras' in body['errores'][0]
        assert Producto.query.filter_by(codigo='COH-D1').first() is None

    def test_acepta_decimales_en_unidad_continua(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        resp = self._post(client, self.BASE + ['Stock Inicial'], [{
            'Código (SKU)': 'COH-D2', 'Descripción': 'Pintura', 'Categoría': 'Pinturas',
            'Unidad': 'Lts', 'Stock Inicial': 2.5,
        }])
        assert resp.status_code == 200, resp.get_json()
        assert resp.get_json()['exitosos'] == 1

    def test_no_rechaza_decimales_que_la_fila_no_aplica(self, client, inv_admin, db):
        """En un producto EXISTENTE el stock inicial se ignora, así que un decimal
        ahí no debe convertir la fila en error (rompería archivos que hoy suben)."""
        from decimal import Decimal
        p = Producto(codigo='COH-D3', descripcion='Tuerca', categoria='Tornillería',
                     unidad='pza', stock_actual=Decimal('4'), stock_minimo=0)
        db.session.add(p); db.session.commit()
        _login(client, inv_admin.id, 'admin')
        resp = self._post(client, self.BASE + ['Stock Inicial'], [{
            'Código (SKU)': 'COH-D3', 'Descripción': 'Tuerca', 'Categoría': 'Tornillería',
            'Unidad': 'pza', 'Stock Inicial': 2.5,
        }])
        assert resp.status_code == 200, resp.get_json()
        body = resp.get_json()
        assert body['errores'] == []
        assert float(Producto.query.filter_by(codigo='COH-D3').first().stock_actual) == 4.0


class TestExportYPlantillaDestino:
    """El export baja SOLO datos de catálogo; la plantilla baja con el destino
    del stock inicial ya resuelto."""

    def _headers(self, data):
        """Encabezados oficiales de un .xlsx generado (van en la fila 4)."""
        import io as _io
        from openpyxl import load_workbook
        ws = load_workbook(_io.BytesIO(data)).active
        return [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]

    # ── Export ───────────────────────────────────────────────────────────────

    def test_export_sin_columnas_de_stock_ni_destino(self, client, inv_admin, db):
        from decimal import Decimal
        db.session.add(Producto(codigo='EXPC-1', descripcion='Tornillo', categoria='Tornillería',
                                unidad='pza', marca='Truper', stock_actual=Decimal('10'),
                                stock_minimo=Decimal('2'), precio_unitario=Decimal('3.5')))
        db.session.commit()
        _login(client, inv_admin.id, 'admin')
        resp = client.get('/api/v1/productos/exportar')
        assert resp.status_code == 200
        headers = self._headers(resp.data)
        for fuera in ('Stock Inicial', 'Stock Mínimo', 'Almacén', 'Proyecto'):
            assert fuera not in headers, f'{fuera} no debería exportarse'
        for dentro in ('Código (SKU)', 'Descripción', 'Marca', 'Categoría', 'Unidad',
                       'Tipo (cable)', 'Tamaño mm²/AWG (cable)', 'Precio Unitario',
                       'URL Imagen (opcional)'):
            assert dentro in headers, f'falta la columna {dentro}'

    def test_export_alinea_valores_con_sus_columnas(self, client, inv_admin, db):
        """Quitar columnas no debe recorrer los valores de lugar."""
        import io as _io
        from decimal import Decimal
        from openpyxl import load_workbook
        db.session.add(Producto(codigo='EXPC-2', descripcion='Cable THHN', categoria='Cable',
                                unidad='M', marca='Condumex', cable_tipo='THHN',
                                cable_calibre='12', stock_actual=Decimal('250'),
                                stock_minimo=Decimal('5'), precio_unitario=Decimal('18.75')))
        db.session.commit()
        _login(client, inv_admin.id, 'admin')
        resp = client.get('/api/v1/productos/exportar')
        ws = load_workbook(_io.BytesIO(resp.data)).active
        headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
        fila = next(r for r in range(5, ws.max_row + 1)
                    if ws.cell(row=r, column=1).value == 'EXPC-2')
        val = {h: ws.cell(row=fila, column=i + 1).value for i, h in enumerate(headers)}
        assert val['Descripción'] == 'Cable THHN'
        assert val['Marca'] == 'Condumex'
        assert val['Categoría'] == 'Cable'
        assert val['Tipo (cable)'] == 'THHN'
        assert val['Unidad'] == 'M'
        assert float(val['Precio Unitario']) == pytest.approx(18.75, abs=0.001)

    def test_export_respeta_los_filtros_del_catalogo(self, client, inv_admin, db):
        """Con miles de productos, bajar todo para corregir una categoría es
        incómodo: el export acepta los mismos filtros del catálogo."""
        import io as _io
        from decimal import Decimal
        from openpyxl import load_workbook
        db.session.add_all([
            Producto(codigo='FLT-A1', descripcion='Tornillo', categoria='ZZ Alfa', unidad='pza',
                     stock_actual=Decimal('5'), stock_minimo=0),
            Producto(codigo='FLT-A2', descripcion='Tuerca', categoria='ZZ Alfa', unidad='pza',
                     stock_actual=Decimal('5'), stock_minimo=0),
            Producto(codigo='FLT-B1', descripcion='Cinta', categoria='ZZ Beta', unidad='pza',
                     stock_actual=Decimal('5'), stock_minimo=0),
        ])
        db.session.commit()
        _login(client, inv_admin.id, 'admin')

        resp = client.get('/api/v1/productos/exportar?categoria=ZZ Alfa')
        assert resp.status_code == 200
        ws = load_workbook(_io.BytesIO(resp.data)).active
        codigos = {ws.cell(row=r, column=1).value for r in range(5, ws.max_row + 1)}
        assert 'FLT-A1' in codigos and 'FLT-A2' in codigos
        assert 'FLT-B1' not in codigos

        # Búsqueda libre: mismo criterio que el catálogo (código/descripción/…).
        resp = client.get('/api/v1/productos/exportar?q=Cinta')
        ws = load_workbook(_io.BytesIO(resp.data)).active
        codigos = {ws.cell(row=r, column=1).value for r in range(5, ws.max_row + 1)}
        assert 'FLT-B1' in codigos and 'FLT-A1' not in codigos

    def test_reimportar_un_export_filtrado_no_toca_lo_ausente(self, client, inv_admin, db):
        """Subir un archivo con solo una parte del catálogo no debe afectar al
        resto: el importador únicamente toca los SKU que vienen en el archivo."""
        import io as _io
        from decimal import Decimal
        db.session.add_all([
            Producto(codigo='FLT-C1', descripcion='Tornillo', categoria='ZZ Gamma', unidad='pza',
                     marca='Truper', stock_actual=Decimal('5'), stock_minimo=Decimal('3')),
            Producto(codigo='FLT-D1', descripcion='Cinta', categoria='ZZ Delta', unidad='pza',
                     marca='3M', stock_actual=Decimal('7'), stock_minimo=Decimal('2')),
        ])
        db.session.commit()
        _login(client, inv_admin.id, 'admin')

        exp = client.get('/api/v1/productos/exportar?categoria=ZZ Gamma')
        imp = client.post('/api/v1/productos/importar',
                          data={'archivo': (_io.BytesIO(exp.data), 'catalogo_filtrado.xlsx')},
                          content_type='multipart/form-data')
        assert imp.status_code == 200, imp.get_json()
        assert imp.get_json()['total_procesadas'] == 1  # solo el de la categoría filtrada
        otro = Producto.query.filter_by(codigo='FLT-D1').first()
        assert otro.marca == '3M' and float(otro.stock_minimo) == 2.0

    def test_roundtrip_export_conserva_stock_y_minimo(self, client, inv_admin, db):
        """Reimportar el export (ya sin las columnas de stock) NO debe pisar el
        stock actual ni el stock mínimo con 0. Es el caso crítico del cambio."""
        import io as _io
        from decimal import Decimal
        p = Producto(codigo='RTC-1', descripcion='Tuerca', categoria='Tuercas', unidad='pza',
                     marca='Urrea', stock_actual=Decimal('40'), stock_minimo=Decimal('7'),
                     precio_unitario=Decimal('2.5'))
        db.session.add(p); db.session.commit()
        _login(client, inv_admin.id, 'admin')

        exp = client.get('/api/v1/productos/exportar')
        assert exp.status_code == 200
        data = {'archivo': (_io.BytesIO(exp.data), 'catalogo_materiales.xlsx')}
        imp = client.post('/api/v1/productos/importar', data=data,
                          content_type='multipart/form-data')
        assert imp.status_code == 200, imp.get_json()
        body = imp.get_json()
        assert body['actualizados'] == 0, body['cambios_detalle']
        assert body['sin_cambios'] >= 1
        db.session.refresh(p)
        assert float(p.stock_actual) == pytest.approx(40.0, abs=0.01)
        assert float(p.stock_minimo) == pytest.approx(7.0, abs=0.01)

    def test_roundtrip_export_aplica_marca_editada(self, client, inv_admin, db):
        """Editar la marca en el export y reimportar sí debe aplicarse."""
        import io as _io
        from decimal import Decimal
        from openpyxl import load_workbook
        p = Producto(codigo='RTC-2', descripcion='Taquete', categoria='Fijación', unidad='pza',
                     marca='Vieja', stock_actual=Decimal('5'), stock_minimo=Decimal('1'),
                     precio_unitario=Decimal('1'))
        db.session.add(p); db.session.commit()
        _login(client, inv_admin.id, 'admin')

        exp = client.get('/api/v1/productos/exportar')
        wb = load_workbook(_io.BytesIO(exp.data))
        ws = wb.active
        headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
        col_marca = headers.index('Marca') + 1
        fila = next(r for r in range(5, ws.max_row + 1)
                    if ws.cell(row=r, column=1).value == 'RTC-2')
        ws.cell(row=fila, column=col_marca, value='Fischer')
        buf = _io.BytesIO(); wb.save(buf); buf.seek(0)

        imp = client.post('/api/v1/productos/importar',
                          data={'archivo': (buf, 'catalogo_materiales.xlsx')},
                          content_type='multipart/form-data')
        assert imp.status_code == 200, imp.get_json()
        assert imp.get_json()['actualizados'] == 1
        db.session.refresh(p)
        assert p.marca == 'Fischer'
        assert float(p.stock_minimo) == pytest.approx(1.0, abs=0.01)  # intacto

    # ── Plantilla con destino ────────────────────────────────────────────────

    def _destino(self, db):
        alm = Almacen(nombre=f'BOD-{uuid.uuid4().hex[:5]}', qr_code=str(uuid.uuid4()), activo=True)
        proy = Proyecto(numero_proyecto=f'PY-{uuid.uuid4().hex[:5]}', nombre='Nave', activo=True)
        db.session.add_all([alm, proy]); db.session.commit()
        return alm, proy

    def test_plantilla_prellena_destino(self, client, inv_admin, db):
        import io as _io
        from openpyxl import load_workbook
        alm, proy = self._destino(db)
        _login(client, inv_admin.id, 'admin')
        resp = client.get(
            f'/api/v1/productos/plantilla-importar?almacen_id={alm.id}&proyecto_id={proy.id}')
        assert resp.status_code == 200
        wb = load_workbook(_io.BytesIO(resp.data))
        ws = wb.active
        headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
        c_alm = headers.index('Almacén') + 1
        c_proy = headers.index('Proyecto') + 1
        for fila in (5, 6, 50):
            assert ws.cell(row=fila, column=c_alm).value == alm.nombre
            assert ws.cell(row=fila, column=c_proy).value == proy.numero_proyecto
        # Hoja oculta con las listas desplegables.
        assert 'Listas' in wb.sheetnames
        assert wb['Listas'].sheet_state == 'hidden'

    def test_plantilla_sin_destino_sigue_vacia(self, client, inv_admin, db):
        """Sin parámetros la plantilla baja como siempre (nada prellenado)."""
        import io as _io
        from openpyxl import load_workbook
        self._destino(db)
        _login(client, inv_admin.id, 'admin')
        resp = client.get('/api/v1/productos/plantilla-importar')
        assert resp.status_code == 200
        ws = load_workbook(_io.BytesIO(resp.data)).active
        headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
        c_alm = headers.index('Almacén') + 1
        assert ws.cell(row=5, column=c_alm).value is None

    def test_plantilla_destino_inexistente_400(self, client, inv_admin, db):
        _login(client, inv_admin.id, 'admin')
        r1 = client.get('/api/v1/productos/plantilla-importar?almacen_id=999999')
        assert r1.status_code == 400 and 'detail' in r1.get_json()
        r2 = client.get('/api/v1/productos/plantilla-importar?proyecto_id=999999')
        assert r2.status_code == 400 and 'detail' in r2.get_json()

    def test_plantilla_destino_inactivo_400(self, client, inv_admin, db):
        alm, _proy = self._destino(db)
        alm.activo = False
        db.session.commit()
        _login(client, inv_admin.id, 'admin')
        resp = client.get(f'/api/v1/productos/plantilla-importar?almacen_id={alm.id}')
        assert resp.status_code == 400

    def test_plantilla_prellenada_deposita_en_el_bucket(self, client, inv_admin, db):
        """End-to-end: bajar la plantilla con destino, capturar un material y
        subirla debe dejar el stock inicial en ese almacén y proyecto."""
        import io as _io
        from openpyxl import load_workbook
        alm, proy = self._destino(db)
        _login(client, inv_admin.id, 'admin')

        resp = client.get(
            f'/api/v1/productos/plantilla-importar?almacen_id={alm.id}&proyecto_id={proy.id}')
        wb = load_workbook(_io.BytesIO(resp.data))
        ws = wb.active
        headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
        pos = {h: i + 1 for i, h in enumerate(headers)}
        # Solo se captura el material: el destino ya viene lleno.
        ws.cell(row=5, column=pos['Código (SKU)'], value='PLT-DEST-1')
        ws.cell(row=5, column=pos['Descripción'], value='Tornillo hex')
        ws.cell(row=5, column=pos['Categoría'], value='Tornillería')
        ws.cell(row=5, column=pos['Unidad'], value='pza')
        ws.cell(row=5, column=pos['Stock Inicial'], value=25)
        buf = _io.BytesIO(); wb.save(buf); buf.seek(0)

        imp = client.post('/api/v1/productos/importar',
                          data={'archivo': (buf, 'plantilla_materiales.xlsx')},
                          content_type='multipart/form-data')
        assert imp.status_code == 200, imp.get_json()
        body = imp.get_json()
        assert body['exitosos'] == 1, body
        assert body['errores'] == []
        p = Producto.query.filter_by(codigo='PLT-DEST-1').first()
        assert p is not None
        bucket = StockAlmacenProyecto.query.filter_by(
            producto_id=p.id, almacen_id=alm.id, proyecto_id=proy.id).first()
        assert bucket is not None and float(bucket.cantidad) == pytest.approx(25.0, abs=0.01)
        cache = StockPorAlmacen.query.filter_by(producto_id=p.id, almacen_id=alm.id).first()
        assert cache is not None and float(cache.cantidad) == pytest.approx(25.0, abs=0.01)

    def test_reimport_masivo_no_consulta_por_fila(self, client, inv_admin, db):
        """El importador debe resolver los SKU en LOTES, no uno por fila.

        En producción el catálogo pasa de 5 000 productos: un SELECT por renglón
        eran 5 000 viajes a Postgres en una sola petición HTTP — segundos de pura
        latencia y riesgo de que el proxy corte antes de terminar. Este test fija
        esa garantía: el número de consultas no crece con las filas del archivo.
        """
        import io as _io
        from decimal import Decimal
        from sqlalchemy import event

        N = 400
        db.session.bulk_save_objects([
            Producto(codigo=f'MASS-{i:04d}', descripcion=f'Producto {i}', marca='MarcaX',
                     categoria=f'Cat {i % 8}', unidad='pza', stock_actual=Decimal('10'),
                     stock_minimo=Decimal('2'), precio_unitario=Decimal('12.5'))
            for i in range(N)
        ])
        db.session.commit()
        _login(client, inv_admin.id, 'admin')

        exp = client.get('/api/v1/productos/exportar')
        assert exp.status_code == 200

        consultas = {'productos': 0}
        engine = db.engine

        @event.listens_for(engine, 'before_cursor_execute')
        def _contar(conn, cursor, statement, params, context, executemany):
            if 'FROM productos' in statement:
                consultas['productos'] += 1

        try:
            imp = client.post('/api/v1/productos/importar',
                              data={'archivo': (_io.BytesIO(exp.data), 'catalogo.xlsx')},
                              content_type='multipart/form-data')
        finally:
            event.remove(engine, 'before_cursor_execute', _contar)

        assert imp.status_code == 200, imp.get_json()
        assert imp.get_json()['sin_cambios'] == N
        # Con la precarga por lotes son un puñado de consultas; sin ella serían
        # N+1 (>400). El margen es amplio a propósito: lo que se vigila es que no
        # escale con el número de filas.
        assert consultas['productos'] < 25, (
            f'{consultas["productos"]} consultas a productos para {N} filas: '
            'volvió el SELECT por fila'
        )

    def test_filas_prellenadas_vacias_no_generan_errores(self, client, inv_admin, db):
        """Las filas prellenadas que traen SOLO el destino (sin material) no
        deben contar como filas capturadas ni reportar errores."""
        import io as _io
        from openpyxl import load_workbook
        alm, proy = self._destino(db)
        _login(client, inv_admin.id, 'admin')
        resp = client.get(
            f'/api/v1/productos/plantilla-importar?almacen_id={alm.id}&proyecto_id={proy.id}')
        wb = load_workbook(_io.BytesIO(resp.data))
        ws = wb.active
        headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
        pos = {h: i + 1 for i, h in enumerate(headers)}
        ws.cell(row=5, column=pos['Código (SKU)'], value='PLT-DEST-2')
        ws.cell(row=5, column=pos['Descripción'], value='Rondana')
        ws.cell(row=5, column=pos['Categoría'], value='Tornillería')
        ws.cell(row=5, column=pos['Unidad'], value='pza')
        buf = _io.BytesIO(); wb.save(buf); buf.seek(0)

        imp = client.post('/api/v1/productos/importar',
                          data={'archivo': (buf, 'plantilla_materiales.xlsx')},
                          content_type='multipart/form-data')
        assert imp.status_code == 200, imp.get_json()
        body = imp.get_json()
        assert body['exitosos'] == 1
        assert body['errores'] == []
        assert body['total_procesadas'] == 1
