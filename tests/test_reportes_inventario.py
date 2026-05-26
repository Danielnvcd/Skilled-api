"""Tests de Pausa 6 — reportes Excel del módulo Inventario.

Cubre los 5 endpoints `/api/v1/reportes/*.xlsx`:
  - inventario-actual
  - movimientos
  - kardex
  - consumo-proyecto
  - solicitudes

Auth: JWT real (Bearer). Ver memoria `tests-session-vs-jwt`.
"""
import io
import uuid
import datetime
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from werkzeug.security import generate_password_hash

from app.models import (
    User, Almacen, Producto, StockPorAlmacen, MovimientoInventario,
    SolicitudMaterial, SolicitudMaterialDetalle,
)
from app.routes.api_auth import _encode_access_token


XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def _hdr(user):
    return {'Authorization': f'Bearer {_encode_access_token(user)}'}


def _wb(resp):
    return load_workbook(io.BytesIO(resp.data))


# ─── Fixtures ─────────────────────────────────────────────────────────────────

@pytest.fixture
def inv_admin(db):
    u = User(username='rep_admin', password_hash=generate_password_hash('Pass123!'), role='admin')
    db.session.add(u); db.session.commit()
    return u


@pytest.fixture
def solicitante(db):
    u = User(username='rep_sol', password_hash=generate_password_hash('Pass123!'),
             role='solicitante_material')
    db.session.add(u); db.session.commit()
    return u


@pytest.fixture
def outsider(db):
    # Rol no autorizado (coordinador sí entra a inventario lectura desde 05-25).
    u = User(username='rep_out', password_hash=generate_password_hash('Pass123!'),
             role='visitor')
    db.session.add(u); db.session.commit()
    return u


@pytest.fixture
def almacen(db):
    a = Almacen(nombre='Bodega R', qr_code=str(uuid.uuid4()), activo=True)
    db.session.add(a); db.session.commit()
    return a


@pytest.fixture
def setup_inventario(db, inv_admin, almacen):
    """Crea 3 productos con stocks y un par de movimientos."""
    p_alto = Producto(codigo='R-001', descripcion='Producto OK',
                      categoria='General', unidad='pza',
                      stock_actual=Decimal('100'), stock_minimo=Decimal('10'),
                      stock_reservado=Decimal('5'), activo=True)
    p_bajo = Producto(codigo='R-002', descripcion='Producto BAJO',
                      categoria='General', unidad='pza',
                      stock_actual=Decimal('2'), stock_minimo=Decimal('5'),
                      stock_reservado=Decimal('0'), activo=True)
    p_otro = Producto(codigo='R-003', descripcion='Producto otra cat',
                      categoria='Especial', unidad='kg',
                      stock_actual=Decimal('50'), stock_minimo=Decimal('0'),
                      stock_reservado=Decimal('0'), activo=True)
    db.session.add_all([p_alto, p_bajo, p_otro]); db.session.flush()
    db.session.add_all([
        StockPorAlmacen(producto_id=p_alto.id, almacen_id=almacen.id, cantidad=Decimal('100')),
        StockPorAlmacen(producto_id=p_bajo.id, almacen_id=almacen.id, cantidad=Decimal('2')),
        StockPorAlmacen(producto_id=p_otro.id, almacen_id=almacen.id, cantidad=Decimal('50')),
    ])
    # Movimientos: una ENTRADA y una SALIDA para p_alto.
    db.session.add_all([
        MovimientoInventario(tipo='ENTRADA', producto_id=p_alto.id, cantidad=Decimal('50'),
                              almacen_destino_id=almacen.id, usuario_id=inv_admin.id,
                              motivo='Compra inicial'),
        MovimientoInventario(tipo='SALIDA', producto_id=p_alto.id, cantidad=Decimal('10'),
                              almacen_origen_id=almacen.id, usuario_id=inv_admin.id,
                              motivo='Salida prueba'),
    ])
    db.session.commit()
    return {'alto': p_alto, 'bajo': p_bajo, 'otro': p_otro}


# ═══════════════════════════════════════════════════════════════════════════════
# Inventario actual
# ═══════════════════════════════════════════════════════════════════════════════

class TestInventarioActual:

    def test_descarga_ok(self, client, inv_admin, setup_inventario):
        r = client.get('/api/v1/reportes/inventario-actual.xlsx', headers=_hdr(inv_admin))
        assert r.status_code == 200
        assert r.mimetype == XLSX_MIME
        wb = _wb(r)
        assert 'Inventario' in wb.sheetnames
        ws = wb['Inventario']
        headers = [c.value for c in ws[1]]
        assert 'Código' in headers and 'Disponible' in headers and 'Estado' in headers

    def test_filtro_solo_bajo_minimo(self, client, inv_admin, setup_inventario):
        r = client.get(
            '/api/v1/reportes/inventario-actual.xlsx?solo_bajo_minimo=1',
            headers=_hdr(inv_admin),
        )
        assert r.status_code == 200
        ws = _wb(r)['Inventario']
        # Solo p_bajo (stock=2, min=5) debe aparecer
        codigos = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]
        assert codigos == ['R-002']

    def test_filtro_categoria(self, client, inv_admin, setup_inventario):
        r = client.get(
            '/api/v1/reportes/inventario-actual.xlsx?categoria=Especial',
            headers=_hdr(inv_admin),
        )
        assert r.status_code == 200
        ws = _wb(r)['Inventario']
        codigos = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]
        assert codigos == ['R-003']

    def test_403_para_outsider(self, client, outsider, setup_inventario):
        r = client.get('/api/v1/reportes/inventario-actual.xlsx', headers=_hdr(outsider))
        assert r.status_code == 403

    def test_401_sin_token(self, client, setup_inventario):
        r = client.get('/api/v1/reportes/inventario-actual.xlsx')
        assert r.status_code == 401


# ═══════════════════════════════════════════════════════════════════════════════
# Movimientos
# ═══════════════════════════════════════════════════════════════════════════════

class TestMovimientos:

    def test_descarga_ok(self, client, inv_admin, setup_inventario):
        r = client.get('/api/v1/reportes/movimientos.xlsx', headers=_hdr(inv_admin))
        assert r.status_code == 200
        ws = _wb(r)['Movimientos']
        headers = [c.value for c in ws[1]]
        assert 'Tipo' in headers and 'Cantidad' in headers and 'Motivo' in headers
        # 2 filas de datos (1 ENTRADA + 1 SALIDA)
        assert ws.max_row == 3

    def test_filtro_tipo(self, client, inv_admin, setup_inventario):
        r = client.get(
            '/api/v1/reportes/movimientos.xlsx?tipo=ENTRADA',
            headers=_hdr(inv_admin),
        )
        assert r.status_code == 200
        ws = _wb(r)['Movimientos']
        # 1 header + 1 ENTRADA = 2 filas
        assert ws.max_row == 2
        assert ws.cell(row=2, column=2).value == 'ENTRADA'

    def test_filtro_producto_id(self, client, inv_admin, setup_inventario):
        bajo_id = setup_inventario['bajo'].id
        r = client.get(
            f'/api/v1/reportes/movimientos.xlsx?producto_id={bajo_id}',
            headers=_hdr(inv_admin),
        )
        assert r.status_code == 200
        ws = _wb(r)['Movimientos']
        # p_bajo no tiene movimientos → solo header o "Sin datos"
        assert ws.max_row <= 1 or (ws.max_row == 1 and ws.cell(1, 1).value == 'Sin datos para los filtros seleccionados')

    def test_tipo_invalido(self, client, inv_admin, setup_inventario):
        r = client.get(
            '/api/v1/reportes/movimientos.xlsx?tipo=XYZ',
            headers=_hdr(inv_admin),
        )
        assert r.status_code == 422

    def test_rango_invertido(self, client, inv_admin, setup_inventario):
        r = client.get(
            '/api/v1/reportes/movimientos.xlsx?desde=2026-06-01&hasta=2026-01-01',
            headers=_hdr(inv_admin),
        )
        assert r.status_code == 422


# ═══════════════════════════════════════════════════════════════════════════════
# Kardex
# ═══════════════════════════════════════════════════════════════════════════════

class TestKardex:

    def test_descarga_ok(self, client, inv_admin, setup_inventario):
        prod_id = setup_inventario['alto'].id
        r = client.get(
            f'/api/v1/reportes/kardex.xlsx?producto_id={prod_id}',
            headers=_hdr(inv_admin),
        )
        assert r.status_code == 200
        wb = _wb(r)
        assert 'Kardex' in wb.sheetnames and 'Resumen' in wb.sheetnames
        ws = wb['Kardex']
        # 1 header + 1 fila saldo inicial + 2 movimientos
        assert ws.max_row == 4
        # Última fila tiene saldo = stock_actual real (100)
        last_saldo = ws.cell(row=ws.max_row, column=5).value
        assert float(last_saldo) == 100.0

    def test_falta_producto_id(self, client, inv_admin, setup_inventario):
        r = client.get('/api/v1/reportes/kardex.xlsx', headers=_hdr(inv_admin))
        assert r.status_code == 422

    def test_producto_inexistente(self, client, inv_admin, setup_inventario):
        r = client.get(
            '/api/v1/reportes/kardex.xlsx?producto_id=99999',
            headers=_hdr(inv_admin),
        )
        assert r.status_code == 404


# ═══════════════════════════════════════════════════════════════════════════════
# Consumo por proyecto
# ═══════════════════════════════════════════════════════════════════════════════

class TestConsumoProyecto:

    def test_agrupa_por_proyecto(self, client, db, inv_admin, solicitante, setup_inventario, almacen):
        # Crear una solicitud APROBADA con 2 líneas entregadas parcialmente.
        sol = SolicitudMaterial(solicitante_id=solicitante.id, proyecto='Obra A',
                                  estatus='APROBADA')
        db.session.add(sol); db.session.flush()
        db.session.add_all([
            SolicitudMaterialDetalle(
                solicitud_id=sol.id, tipo_item='MATERIAL',
                producto_id=setup_inventario['alto'].id,
                cantidad_solicitada=Decimal('5'),
                cantidad_aprobada=Decimal('5'),
                cantidad_entregada=Decimal('3'),
            ),
            SolicitudMaterialDetalle(
                solicitud_id=sol.id, tipo_item='MATERIAL',
                producto_id=setup_inventario['otro'].id,
                cantidad_solicitada=Decimal('10'),
                cantidad_aprobada=Decimal('10'),
                cantidad_entregada=Decimal('10'),
            ),
        ])
        db.session.commit()

        r = client.get(
            '/api/v1/reportes/consumo-proyecto.xlsx',
            headers=_hdr(inv_admin),
        )
        assert r.status_code == 200
        ws = _wb(r)['Consumo por proyecto']
        # 1 header + 2 filas (una por producto)
        assert ws.max_row == 3
        proyectos = {ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)}
        assert proyectos == {'Obra A'}
        # Verificar cantidades entregadas
        codigos = {ws.cell(row=i, column=2).value: ws.cell(row=i, column=5).value
                    for i in range(2, ws.max_row + 1)}
        assert codigos['R-001'] == 3.0
        assert codigos['R-003'] == 10.0


# ═══════════════════════════════════════════════════════════════════════════════
# Solicitudes
# ═══════════════════════════════════════════════════════════════════════════════

class TestSolicitudes:

    def test_descarga_ok(self, client, db, inv_admin, solicitante, setup_inventario):
        sol = SolicitudMaterial(solicitante_id=solicitante.id, proyecto='Obra X',
                                  estatus='PENDIENTE')
        db.session.add(sol); db.session.flush()
        db.session.add(SolicitudMaterialDetalle(
            solicitud_id=sol.id, tipo_item='MATERIAL',
            producto_id=setup_inventario['alto'].id,
            cantidad_solicitada=Decimal('5'),
        ))
        db.session.commit()

        r = client.get('/api/v1/reportes/solicitudes.xlsx', headers=_hdr(inv_admin))
        assert r.status_code == 200
        ws = _wb(r)['Solicitudes']
        assert ws.max_row == 2  # 1 header + 1 sol
        headers = [c.value for c in ws[1]]
        assert 'Estatus' in headers and 'Total entregado' in headers

    def test_filtro_estatus(self, client, db, inv_admin, solicitante, setup_inventario):
        for estatus in ('PENDIENTE', 'APROBADA', 'ENTREGADA'):
            sol = SolicitudMaterial(solicitante_id=solicitante.id, proyecto='X',
                                      estatus=estatus)
            db.session.add(sol)
        db.session.commit()
        r = client.get(
            '/api/v1/reportes/solicitudes.xlsx?estatus=APROBADA',
            headers=_hdr(inv_admin),
        )
        assert r.status_code == 200
        ws = _wb(r)['Solicitudes']
        assert ws.max_row == 2  # header + 1 APROBADA
        assert ws.cell(2, 4).value == 'APROBADA'

    def test_estatus_invalido(self, client, inv_admin, setup_inventario):
        r = client.get(
            '/api/v1/reportes/solicitudes.xlsx?estatus=XYZ',
            headers=_hdr(inv_admin),
        )
        assert r.status_code == 422
