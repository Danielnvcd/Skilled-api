import pandas as pd
import os
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Todas las columnas posibles del empleado
columnas = [
    # === IDENTIFICADORES ===
    "No. Empleado",
    "Nombre(s)",
    "Apellidos",

    # === LABORALES ===
    "Area",
    "Puesto",
    "Tipo de Movimiento",
    "Tipo de Contrato",
    "Tipo de Jornada",
    "Descripcion de Servicio",
    "Fecha Ingreso (YYYY-MM-DD)",
    "Fecha Inicio (YYYY-MM-DD)",
    "Termino de Prueba (YYYY-MM-DD)",

    # === DATOS PERSONALES ===
    "CURP",
    "RFC",
    "NSS",
    "Fecha Nacimiento (YYYY-MM-DD)",
    "Edad",
    "Sexo (M/F)",
    "Estado Civil",
    "Nacionalidad",
    "Domicilio",
    "Correo",
    "Celular",

    # === DATOS MEDICOS / EMERGENCIA ===
    "Tipo de Sangre",
    "Alergias",
    "Enfermedades Cronicas",
    "Contacto de Emergencia",
    "Parentesco del Contacto",
    "Numero Contacto Emergencia",
    "Usa Lentes (Si/No)",
    "Licencia de Conducir (Tipo)",
    "Estatura",

    # === FINANCIERO ===
    "Tipo de Nomina (Semanal/Por hora/Cuadrado)",
    "Salario Real Pactado por Semana",
    "Sueldo Base (SB)",
    "Salario Diario Integrado (SDI)",
    "Letra",
    "Horas Extra",
    "Infonavit",
    "Caja de Ahorro",
    "Viaticos",
    "Pago Dia Festivo",
    "Pagos Efectivo",
    "Folio Mov IDSE",
    "Tipo Pago",

    # === UBICACION Y OPERACION ===
    "Ubicacion Estado",
    "Observaciones",
]

# Crear un DataFrame vacio solo con las columnas
df = pd.DataFrame(columns=columnas)

# Asegurar que el directorio exista
os.makedirs('static/downloads', exist_ok=True)

# Guardar a excel
writer = pd.ExcelWriter('static/downloads/plantilla_empleados.xlsx', engine='openpyxl')
df.to_excel(writer, index=False, sheet_name='Empleados')

workbook = writer.book
worksheet = writer.sheets['Empleados']

# === ESTILOS PARA ENCABEZADOS ===
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col_idx, col_name in enumerate(columnas, 1):
    cell = worksheet.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Ajustar ancho
for col in worksheet.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = max(20, (max_length + 2) * 1.2)
    worksheet.column_dimensions[column].width = adjusted_width

# Aplicar wrap_text a las celdas para que se ajusten solas
wrap_alignment = Alignment(wrap_text=True, vertical='center')
for row in worksheet.iter_rows(min_row=2, max_row=1000, min_col=1, max_col=len(columnas)):
    for cell in row:
        cell.alignment = wrap_alignment

# === DATA VALIDATIONS (DROPDOWNS) ===

validaciones = {
    "Tipo de Nomina (Semanal/Por hora/Cuadrado)": '"Semanal,Por hora,Cuadrado"',
    "Sexo (M/F)": '"M,F"',
    "Usa Lentes (Si/No)": '"Si,No"',
    "Tipo Pago": '"EFECTIVO,TRANSFERENCIA"',
    "Estado Civil": '"Soltero(a),Casado(a),Union Libre,Divorciado(a),Viudo(a)"',
    "Tipo de Movimiento": '"Alta,Reingreso,Modificacion de salario"',
    "Tipo de Contrato": '"Indefinido,Temporal,Por obra determinada,Periodo de prueba"',
    "Tipo de Jornada": '"Diurna,Nocturna,Mixta"',
    "Tipo de Sangre": '"O+,O-,A+,A-,B+,B-,AB+,AB-"',
    "Licencia de Conducir (Tipo)": '"No tiene,Tipo A,Tipo B,Tipo C,Tipo D,Tipo E"',
    "Nacionalidad": '"Mexicana,Otra"',
}

# Mapear los índices de columnas
def get_col_letter(col_name):
    try:
        idx = columnas.index(col_name) + 1
        return get_column_letter(idx)
    except ValueError:
        return None

for col_name, formula in validaciones.items():
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = f"Por favor selecciona una opción válida para {col_name}"
    dv.errorTitle = "Opción no válida"
    worksheet.add_data_validation(dv)
    
    letra = get_col_letter(col_name)
    if letra:
        dv.add(f'{letra}2:{letra}1000')

# Congelar la primera fila (encabezados)
worksheet.freeze_panes = 'A2'

writer.close()
print("Plantilla (con todos los campos y opciones seleccionables) generada exitosamente.")
