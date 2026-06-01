"""
Endpoints JSON del módulo de Inventario.

Migración de `app/api_fastapi/` a Flask puro para consolidar la stack en una sola
tecnología. Mantiene exactamente el mismo contrato HTTP (`/api/v1/*`) y los mismos
códigos de respuesta que la versión FastAPI original.

Características:
  - Validación con marshmallow (422 en errores de schema).
  - Rate limit con Flask-Limiter (decorador @limiter.limit, comparte storage Redis).
  - Auth con JWT Bearer token (ver app/routes/api_auth.py). Los decoradores
    _require_inventario / _require_inventario_admin validan rol.
  - CSRF exento del blueprint (ver app/__init__.py): la protección viene del JWT
    en el header Authorization. Aplica el mismo patrón que el resto de blueprints
    API (api_auth, api_trabajadores, etc.) consumidos por el SPA React.
"""
import io
import uuid
import logging
import datetime
from decimal import Decimal, InvalidOperation
from functools import wraps

import qrcode
from flask import Blueprint, jsonify, request, session, Response, abort, send_file, render_template, current_app
from marshmallow import Schema, fields, validate, ValidationError, EXCLUDE
from sqlalchemy import distinct as sql_distinct
from sqlalchemy.orm import joinedload, selectinload

from app.extensions import db, limiter, get_real_client_ip_flask
from app.realtime import emit_to_role

# Roles que reciben eventos de inventario (productos, almacenes, movimientos,
# tomas). Coordinador/solicitante ven productos en sus solicitudes pero no
# editan; igual les incluimos para que sus listas (catalogo en SolicitudForm)
# refresquen al cambiar el stock.
_INV_ROLES = ['admin', 'super_admin', 'inventario', 'coordinador', 'solicitante_material']
# Las solicitudes involucran a todos: solicitante las crea, inventario aprueba,
# coordinador puede crearlas también.
_SOL_ROLES = ['admin', 'super_admin', 'inventario', 'coordinador', 'solicitante_material']
from app.models import (
    Almacen, Estante, Producto, MovimientoInventario, StockPorAlmacen,
    ProductoEstante, TomaInventario, TomaInventarioDetalle, ESTADOS_TOMA,
    SolicitudMaterial, SolicitudMaterialDetalle, User, AuditLog, Proyecto,
    CategoriaConfig, Herramienta, HerramientaUnidad, AsignacionHerramienta,
    Trabajador, NotificacionUmbral, crear_notif_inventario,
    crear_evento_herramienta,
)
from app.utils import log_action

logger = logging.getLogger(__name__)

bp = Blueprint('inventario_api', __name__, url_prefix='/api/v1')


# ─── Auth helpers ─────────────────────────────────────────────────────────────

from app.routes.api_auth import jwt_required
from flask import g

def _require_login(view):
    @wraps(view)
    @jwt_required
    def wrapper(*args, **kwargs):
        request.current_user = g._jwt_user
        return view(*args, **kwargs)
    return wrapper


def _require_inventario(view):
    """Lectura: solicitantes, coordinadores, inventario, admin y super_admin.
    Coordinador agregado para que pueda ver catálogo y armar pedidos."""
    @wraps(view)
    @_require_login
    def wrapper(*args, **kwargs):
        if request.current_user.role not in ['inventario', 'solicitante_material', 'coordinador', 'admin', 'super_admin']:
            log_action(f"API 403 lectura '{request.path}' (rol: {request.current_user.role})")
            return jsonify({'detail': 'Forbidden: Required permissions missing'}), 403
        return view(*args, **kwargs)
    return wrapper


def _require_inventario_admin(view):
    """Escritura/borrado: inventario, admin y super_admin."""
    @wraps(view)
    @_require_login
    def wrapper(*args, **kwargs):
        if request.current_user.role not in ['inventario', 'admin', 'super_admin']:
            log_action(f"API 403 escritura '{request.path}' (rol: {request.current_user.role})")
            return jsonify({'detail': 'Se requiere rol de inventario o administrador'}), 403
        return view(*args, **kwargs)
    return wrapper


# ─── Marshmallow schemas ──────────────────────────────────────────────────────

CODIGO_REGEX = r'^[A-Za-z0-9\-_\.\/]+$'

# Anti-SSRF/phishing: cuando guardamos URLs de imagen (categorías, productos) la
# UI las pinta como <img src=...>. Si dejamos cualquier URL, un admin malicioso
# podría meter `javascript:`, `data:text/html`, URLs a otros dominios para
# tracking pixels, o intranet (SSRF si el browser corre detrás de un proxy).
# Forzamos HTTPS + dominios públicos o paths relativos al propio backend.
_IMAGEN_URL_REGEX = r'^(?:https://[A-Za-z0-9.\-_]+(?::\d+)?(?:/[^\s<>"\']*)?|/[A-Za-z0-9.\-_/]+\.(?:png|jpe?g|webp|gif|svg))$'


class _BaseSchema(Schema):
    class Meta:
        # Coincide con el comportamiento por defecto de Pydantic: ignora campos extra.
        unknown = EXCLUDE


class ProductoCreateSchema(_BaseSchema):
    codigo = fields.Str(required=True, validate=[
        validate.Length(min=1, max=50),
        validate.Regexp(CODIGO_REGEX),
    ])
    descripcion = fields.Str(required=True, validate=validate.Length(min=1, max=250))
    categoria = fields.Str(required=True, validate=validate.Length(min=1, max=100))
    unidad = fields.Str(required=True, validate=validate.Length(min=1, max=50))
    stock_actual = fields.Float(load_default=0.0, validate=validate.Range(min=0, max=1_000_000))
    stock_minimo = fields.Float(load_default=0.0, validate=validate.Range(min=0, max=1_000_000))
    imagen_url = fields.Str(load_default=None, allow_none=True, validate=[
        validate.Length(max=500),
        # Anti-XSS/SSRF: solo HTTPS o path absoluto local
        validate.Regexp(_IMAGEN_URL_REGEX, error='imagen_url debe ser HTTPS o un path absoluto a imagen local'),
    ])
    # Pausa 9: proveedor default (opcional al crear)
    proveedor_default_nombre = fields.Str(load_default=None, allow_none=True, validate=validate.Length(max=150))
    proveedor_default_contacto = fields.Str(load_default=None, allow_none=True, validate=validate.Length(max=150))


class ProductoUpdateSchema(_BaseSchema):
    codigo = fields.Str(load_default=None, allow_none=True, validate=[
        validate.Length(min=1, max=50),
        validate.Regexp(CODIGO_REGEX),
    ])
    descripcion = fields.Str(load_default=None, allow_none=True, validate=validate.Length(min=1, max=250))
    categoria = fields.Str(load_default=None, allow_none=True, validate=validate.Length(min=1, max=100))
    unidad = fields.Str(load_default=None, allow_none=True, validate=validate.Length(min=1, max=50))
    stock_actual = fields.Float(load_default=None, allow_none=True, validate=validate.Range(min=0, max=1_000_000))
    stock_minimo = fields.Float(load_default=None, allow_none=True, validate=validate.Range(min=0, max=1_000_000))
    imagen_url = fields.Str(load_default=None, allow_none=True, validate=[
        validate.Length(max=500),
        validate.Regexp(_IMAGEN_URL_REGEX, error='imagen_url debe ser HTTPS o un path absoluto a imagen local'),
    ])
    proveedor_default_nombre = fields.Str(load_default=None, allow_none=True, validate=validate.Length(max=150))
    proveedor_default_contacto = fields.Str(load_default=None, allow_none=True, validate=validate.Length(max=150))


class AlmacenCreateSchema(_BaseSchema):
    nombre = fields.Str(required=True, validate=validate.Length(min=1, max=100))
    ubicacion = fields.Str(load_default=None, allow_none=True, validate=validate.Length(max=250))
    activo = fields.Bool(load_default=True)


class AlmacenUpdateSchema(_BaseSchema):
    nombre = fields.Str(load_default=None, allow_none=True, validate=validate.Length(min=1, max=100))
    ubicacion = fields.Str(load_default=None, allow_none=True, validate=validate.Length(max=250))
    activo = fields.Bool(load_default=None, allow_none=True)


class EstanteCreateSchema(_BaseSchema):
    nombre = fields.Str(required=True, validate=validate.Length(min=1, max=100))
    descripcion = fields.Str(load_default=None, allow_none=True, validate=validate.Length(max=250))
    almacen_id = fields.Int(required=True)


class EstanteUpdateSchema(_BaseSchema):
    nombre = fields.Str(load_default=None, allow_none=True, validate=validate.Length(min=1, max=100))
    descripcion = fields.Str(load_default=None, allow_none=True, validate=validate.Length(max=250))
    almacen_id = fields.Int(load_default=None, allow_none=True)


class MovimientoCreateSchema(_BaseSchema):
    tipo = fields.Str(required=True, validate=validate.OneOf(['ENTRADA', 'SALIDA', 'AJUSTE', 'TRASPASO']))
    producto_id = fields.Int(required=True)
    cantidad = fields.Float(required=True, validate=validate.Range(min=-100_000, max=100_000))
    almacen_origen_id = fields.Int(load_default=None, allow_none=True)
    almacen_destino_id = fields.Int(load_default=None, allow_none=True)
    estante_id = fields.Int(load_default=None, allow_none=True)
    motivo = fields.Str(load_default=None, allow_none=True, validate=validate.Length(max=250))


class SolicitudDetalleCreateSchema(_BaseSchema):
    # XOR a nivel app: si tipo_item=MATERIAL → producto_id; si HERRAMIENTA → herramienta_id
    tipo_item = fields.Str(load_default='MATERIAL', validate=validate.OneOf(['MATERIAL', 'HERRAMIENTA']))
    producto_id = fields.Int(load_default=None, allow_none=True)
    herramienta_id = fields.Int(load_default=None, allow_none=True)
    cantidad_solicitada = fields.Float(required=True, validate=validate.Range(min=0.0001, max=10_000))
    fecha_uso_inicio = fields.Date(load_default=None, allow_none=True)
    fecha_uso_fin = fields.Date(load_default=None, allow_none=True)
    justificacion = fields.Str(load_default=None, allow_none=True, validate=validate.Length(max=2000))
    complementos = fields.Str(load_default=None, allow_none=True, validate=validate.Length(max=500))


class SolicitudCreateSchema(_BaseSchema):
    proyecto = fields.Str(load_default=None, allow_none=True, validate=validate.Length(max=200))
    detalles = fields.List(
        fields.Nested(SolicitudDetalleCreateSchema),
        required=True,
        validate=validate.Length(min=1, max=100),
    )


class SolicitudUpdateEstadoSchema(_BaseSchema):
    estatus = fields.Str(required=True, validate=validate.OneOf(['APROBADA', 'RECHAZADA', 'ENTREGADA', 'PENDIENTE']))


# Pausa 8b — Editar cantidad_aprobada de una línea.
class SolicitudDetallePatchSchema(_BaseSchema):
    cantidad_aprobada = fields.Float(required=True, validate=validate.Range(min=0, max=100_000))


# Pausa 8b — Entrega total o parcial de una solicitud APROBADA.
class EntregaItemSchema(_BaseSchema):
    detalle_id = fields.Int(required=True)
    cantidad_entregada = fields.Float(required=True, validate=validate.Range(min=0, max=100_000))


class EntregarSolicitudSchema(_BaseSchema):
    almacen_origen_id = fields.Int(load_default=None, allow_none=True)
    motivo = fields.Str(load_default=None, allow_none=True, validate=validate.Length(max=250))
    # Para líneas HERRAMIENTA: fecha de devolución prevista que se copia a cada
    # AsignacionHerramienta generada. Opcional.
    fecha_devolucion_prevista = fields.DateTime(load_default=None, allow_none=True)
    entregas = fields.List(
        fields.Nested(EntregaItemSchema),
        required=True,
        validate=validate.Length(min=1, max=100),
    )


class CategoriaConfigUpsertSchema(_BaseSchema):
    imagen_url = fields.Str(load_default=None, allow_none=True, validate=[
        validate.Length(max=500),
        validate.Regexp(_IMAGEN_URL_REGEX, error='imagen_url debe ser HTTPS o un path absoluto a imagen local'),
    ])


# ─── Serializers ──────────────────────────────────────────────────────────────

def _producto_to_dict(p: Producto) -> dict:
    actual = float(p.stock_actual or 0)
    reservado = float(p.stock_reservado or 0)
    return {
        'id': p.id,
        'codigo': p.codigo,
        'descripcion': p.descripcion,
        'categoria': p.categoria,
        'unidad': p.unidad,
        'stock_actual': actual,
        'stock_reservado': reservado,         # Pausa 2-bis: apartado por solicitudes APROBADAS
        'stock_disponible': actual - reservado,  # lo que sí se puede mover
        'stock_minimo': float(p.stock_minimo or 0),
        'imagen_url': p.imagen_url,
        # Pausa 9: proveedor default para Compras express.
        'proveedor_default_nombre': p.proveedor_default_nombre,
        'proveedor_default_contacto': p.proveedor_default_contacto,
        'activo': bool(p.activo),
        'created_at': p.created_at.isoformat() if p.created_at else None,
        'updated_at': p.updated_at.isoformat() if p.updated_at else None,
        'created_by_id': p.created_by_id,
    }


def _almacen_to_dict(a: Almacen) -> dict:
    return {
        'id': a.id,
        'nombre': a.nombre,
        'ubicacion': a.ubicacion,
        'activo': bool(a.activo),
        'qr_code': a.qr_code,
    }


def _estante_to_dict(e: Estante) -> dict:
    return {
        'id': e.id,
        'nombre': e.nombre,
        'descripcion': e.descripcion,
        'almacen_id': e.almacen_id,
        'qr_code': e.qr_code,
        'activo': bool(e.activo),
        'created_at': e.created_at.isoformat() if e.created_at else None,
    }


def _movimiento_to_dict(m: MovimientoInventario) -> dict:
    return {
        'id': m.id,
        'tipo': m.tipo,
        'producto_id': m.producto_id,
        'cantidad': float(m.cantidad or 0),
        'fecha': m.fecha.isoformat() if m.fecha else None,
        'almacen_origen_id': m.almacen_origen_id,
        'almacen_destino_id': m.almacen_destino_id,
        'usuario_id': m.usuario_id,
        'motivo': m.motivo,
    }


def _solicitud_detalle_to_dict(d: SolicitudMaterialDetalle) -> dict:
    tipo = (d.tipo_item or 'MATERIAL').upper()
    base = {
        'id': d.id,
        'tipo_item': tipo,
        'cantidad_solicitada': float(d.cantidad_solicitada or 0),
        'cantidad_aprobada': float(d.cantidad_aprobada or 0),
        'cantidad_entregada': float(d.cantidad_entregada or 0),
        'fecha_uso_inicio': d.fecha_uso_inicio.isoformat() if d.fecha_uso_inicio else None,
        'fecha_uso_fin': d.fecha_uso_fin.isoformat() if d.fecha_uso_fin else None,
        'justificacion': d.justificacion,
        'complementos': d.complementos,
    }
    if tipo == 'HERRAMIENTA':
        base['herramienta_id'] = d.herramienta_id
        base['producto_id'] = None
        base['item_descripcion'] = d.herramienta.descripcion if d.herramienta else 'Herramienta eliminada'
        base['item_codigo'] = d.herramienta.sku if d.herramienta else '---'
        base['item_unidad'] = d.herramienta.unidad if d.herramienta else 'pza'
        # Compat: claves antiguas para que el SPA siga renderizando
        base['producto_descripcion'] = base['item_descripcion']
        base['producto_codigo'] = base['item_codigo']
        base['producto_unidad'] = base['item_unidad']
    else:
        base['producto_id'] = d.producto_id
        base['herramienta_id'] = None
        base['item_descripcion'] = d.producto.descripcion if d.producto else 'Producto eliminado'
        base['item_codigo'] = d.producto.codigo if d.producto else '---'
        base['item_unidad'] = d.producto.unidad if d.producto else 'pza'
        base['producto_descripcion'] = base['item_descripcion']
        base['producto_codigo'] = base['item_codigo']
        base['producto_unidad'] = base['item_unidad']
    return base


def _solicitud_to_dict(s: SolicitudMaterial) -> dict:
    return {
        'id': s.id,
        'solicitante_id': s.solicitante_id,
        'proyecto': s.proyecto,
        'estatus': s.estatus,
        'fecha_creacion': s.fecha_creacion.isoformat() if s.fecha_creacion else None,
        'fecha_cierre': s.fecha_cierre.isoformat() if s.fecha_cierre else None,
        'solicitante_nombre': s.solicitante.username if s.solicitante else 'Desconocido',
        'detalles': [_solicitud_detalle_to_dict(d) for d in s.detalles],
    }


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _parse_or_422(schema: Schema, data):
    """Valida `data` con `schema`. Si falla, devuelve tuple (None, response_422).
    Si pasa, devuelve (dict, None).
    """
    if not isinstance(data, dict):
        return None, (jsonify({'detail': 'Payload debe ser un objeto JSON'}), 422)
    try:
        return schema.load(data), None
    except ValidationError as err:
        return None, (jsonify({'detail': err.messages}), 422)


def _int_arg(name: str, default: int, minimum: int, maximum: int):
    """Lee un query param int, lo recorta al rango y devuelve (valor, error_response).
    Devuelve 422 si el valor no es numérico o queda fuera del rango.
    """
    raw = request.args.get(name)
    if raw is None:
        return default, None
    try:
        val = int(raw)
    except (TypeError, ValueError):
        return None, (jsonify({'detail': f"Parámetro '{name}' debe ser entero"}), 422)
    if val < minimum or val > maximum:
        return None, (jsonify({'detail': f"Parámetro '{name}' fuera de rango"}), 422)
    return val, None


# ─── Reservas de stock (Pausa 2-bis del plan) ─────────────────────────────────

def _reservas_de_solicitud(sol: 'SolicitudMaterial') -> dict[int, Decimal]:
    """Suma por producto la cantidad que esta solicitud debe tener APARTADA.

    Solo cuenta detalles MATERIAL (no HERRAMIENTAS).

    Pausa 8b: la reserva por línea es `cantidad_aprobada - cantidad_entregada`.
    Si `cantidad_aprobada` es 0 (caso recién-aprobada sin tocar líneas), cae
    al fallback `cantidad_solicitada` para preservar el comportamiento previo
    a 8b (todas las solicitudes pre-8b se aprobaban "tal cual se solicitó").

    Una solicitud puede repetir el mismo producto en varias líneas — los
    agrupamos para hacer un solo update por producto.
    """
    out: dict[int, Decimal] = {}
    for d in (sol.detalles or []):
        if d.producto_id is None:
            continue
        cant_aprob = Decimal(str(d.cantidad_aprobada or 0))
        cant_sol = Decimal(str(d.cantidad_solicitada or 0))
        cant_ent = Decimal(str(d.cantidad_entregada or 0))
        base = cant_aprob if cant_aprob > 0 else cant_sol
        reserva = base - cant_ent
        if reserva > 0:
            out[d.producto_id] = out.get(d.producto_id, Decimal('0')) + reserva
    return out


def _intentar_reservar(reservas: dict[int, Decimal]) -> list[str]:
    """Locks + valida disponibilidad. Aplica las reservas si TODOS los productos
    alcanzan. Si alguno no, NO aplica nada (caller debe hacer rollback) y
    devuelve lista de errores legibles para el SPA.
    Disponible = stock_actual − stock_reservado."""
    if not reservas:
        return []
    errores = []
    a_aplicar = []
    for prod_id, cant in reservas.items():
        prod = (
            Producto.query
            .with_for_update(nowait=True)
            .filter(Producto.id == prod_id)
            .first()
        )
        if not prod:
            errores.append(f"Producto #{prod_id} no encontrado")
            continue
        actual = Decimal(str(prod.stock_actual or 0))
        reservado = Decimal(str(prod.stock_reservado or 0))
        disponible = actual - reservado
        if disponible < cant:
            errores.append(
                f"{prod.codigo} — {prod.descripcion}: requiere {cant} {prod.unidad} "
                f"pero solo hay {disponible} disponibles (stock {actual}, ya apartado {reservado})"
            )
            continue
        a_aplicar.append((prod, cant))
    if errores:
        return errores
    for prod, cant in a_aplicar:
        prod.stock_reservado = (prod.stock_reservado or Decimal('0')) + cant
    return []


def _liberar_reservas(reservas: dict[int, Decimal]):
    """Resta reservas (con clamp a 0 por seguridad). No falla si el producto
    no existe — la reserva ya quedó liberada conceptualmente."""
    for prod_id, cant in reservas.items():
        prod = (
            Producto.query
            .with_for_update(nowait=True)
            .filter(Producto.id == prod_id)
            .first()
        )
        if not prod:
            continue
        actual = Decimal(str(prod.stock_reservado or 0))
        nuevo = actual - cant
        prod.stock_reservado = nuevo if nuevo > 0 else Decimal('0')


# ─── Stock por almacén (Pausa 2 del plan) ─────────────────────────────────────

def _almacen_default_id() -> int | None:
    """Devuelve el id del almacén activo de menor id. Sirve como fallback cuando
    un movimiento llega sin almacén ni estante (clientes viejos del SPA antes
    del refactor a stock por almacén). Devuelve None si no hay ninguno."""
    row = (
        db.session.query(Almacen.id)
        .filter(Almacen.activo == True)  # noqa: E712
        .order_by(Almacen.id.asc())
        .first()
    )
    return row[0] if row else None


def _lock_stock(producto_id: int, almacen_id: int) -> StockPorAlmacen:
    """SELECT ... FOR UPDATE sobre la fila (producto, almacen). Crea la fila
    en 0 si no existe — útil para ENTRADAs hacia bodegas nuevas que aún no
    tienen registro de este producto."""
    fila = (
        db.session.query(StockPorAlmacen)
        .with_for_update(nowait=True)
        .filter(
            StockPorAlmacen.producto_id == producto_id,
            StockPorAlmacen.almacen_id == almacen_id,
        )
        .first()
    )
    if fila is None:
        fila = StockPorAlmacen(producto_id=producto_id, almacen_id=almacen_id, cantidad=Decimal('0'))
        db.session.add(fila)
        db.session.flush()  # asegura que existe antes de bloquear
    return fila


def _recalcular_cache_stock(producto: Producto):
    """Actualiza el cache denormalizado `Producto.stock_actual` con la suma
    de todas las filas de stock_por_almacen del producto. Se llama dentro de
    la misma transacción que modificó stock_por_almacen, así nunca queda
    desfasado en commits exitosos."""
    total = (
        db.session.query(db.func.coalesce(db.func.sum(StockPorAlmacen.cantidad), 0))
        .filter(StockPorAlmacen.producto_id == producto.id)
        .scalar()
    )
    producto.stock_actual = Decimal(str(total or 0))


def _audit(user: User, action: str):
    """Escribe AuditLog usando la IP real (anti-spoofing en get_real_client_ip_flask).
    MED-03: pasa por _safe_log_value para evitar log forging vía CRLF."""
    from app.extensions import get_real_client_ip_flask
    from app.utils import _safe_log_value
    ip = get_real_client_ip_flask()
    entry = AuditLog(
        user=_safe_log_value(user.username, 80),
        action=_safe_log_value(action, 200),
        ip=ip,
    )
    db.session.add(entry)


# ─── Health ───────────────────────────────────────────────────────────────────

@bp.route('/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'ok'})


# ─── Productos ────────────────────────────────────────────────────────────────

@bp.route('/productos/by-codigo/<string:codigo>', methods=['GET'])
@_require_inventario
def get_producto_por_codigo(codigo: str):
    """Lookup de producto por su código (usado por el scanner móvil cuando
    el QR escaneado no es estante ni herramienta)."""
    codigo = (codigo or '').strip()
    if not codigo:
        return jsonify({'detail': 'codigo requerido'}), 422
    prod = Producto.query.filter(Producto.codigo == codigo, Producto.activo == True).first()
    if not prod:
        return jsonify({'detail': f'Producto {codigo} no encontrado'}), 404
    return jsonify(_producto_to_dict(prod))


@bp.route('/productos/', methods=['GET'])
@_require_inventario
def get_productos():
    skip, err = _int_arg('skip', 0, 0, 1_000_000)
    if err: return err
    limit, err = _int_arg('limit', 200, 0, 1000)
    if err: return err

    productos = (
        Producto.query
        .filter(Producto.activo == True)
        .offset(skip)
        .limit(limit)
        .all()
    )
    return jsonify([_producto_to_dict(p) for p in productos])


@bp.route('/productos/bajo-minimo/', methods=['GET'])
@_require_inventario
def get_productos_bajo_minimo():
    """Productos en o bajo el mínimo, con consumo promedio y días restantes (Pausa 5).

    consumo_promedio_30d = SUM(SALIDAs últimos 30 días) / 30 (unidades/día).
    dias_de_stock_restante = stock_actual / consumo_promedio_30d (None si consumo=0).
    Orden: mayor urgencia primero (menos días restantes; consumo=0 al final).
    """
    productos = (
        Producto.query
        .filter(Producto.activo == True, Producto.stock_actual <= Producto.stock_minimo)  # noqa: E712
        .all()
    )
    if not productos:
        return jsonify([])

    # Una sola query para el consumo de todos los productos bajo mínimo
    # (en lugar de N queries dentro del loop).
    hace_30 = datetime.datetime.now() - datetime.timedelta(days=30)
    ids = [p.id for p in productos]
    consumos = dict(
        db.session.query(
            MovimientoInventario.producto_id,
            db.func.coalesce(db.func.sum(MovimientoInventario.cantidad), 0),
        )
        .filter(
            MovimientoInventario.producto_id.in_(ids),
            MovimientoInventario.tipo == 'SALIDA',
            MovimientoInventario.fecha >= hace_30,
        )
        .group_by(MovimientoInventario.producto_id)
        .all()
    )

    out = []
    for p in productos:
        consumo_total = float(consumos.get(p.id, 0) or 0)
        consumo_diario = round(consumo_total / 30.0, 2)
        stock = float(p.stock_actual or 0)
        minimo = float(p.stock_minimo or 0)
        # División por cero: si no hay consumo, no hay forma de estimar días.
        if consumo_diario > 0:
            dias_restantes = round(stock / consumo_diario, 1)
        else:
            dias_restantes = None
        # Urgencia para que el SPA coloree sin recalcular.
        if dias_restantes is None:
            urgencia = 'estatico'  # bajo mínimo pero sin consumo: producto parado
        elif dias_restantes < 7:
            urgencia = 'critico'
        elif dias_restantes < 14:
            urgencia = 'alto'
        else:
            urgencia = 'medio'
        out.append({
            'id': p.id,
            'codigo': p.codigo,
            'descripcion': p.descripcion,
            'categoria': p.categoria,
            'unidad': p.unidad,
            'stock_actual': stock,
            'stock_minimo': minimo,
            'faltante': max(0.0, minimo - stock),
            'consumo_promedio_30d': consumo_diario,
            'dias_de_stock_restante': dias_restantes,
            'urgencia': urgencia,
        })

    # Orden: críticos → altos → medios → estáticos. Dentro de cada grupo,
    # menos días primero.
    URGENCIA_ORDEN = {'critico': 0, 'alto': 1, 'medio': 2, 'estatico': 3}
    out.sort(key=lambda x: (
        URGENCIA_ORDEN[x['urgencia']],
        x['dias_de_stock_restante'] if x['dias_de_stock_restante'] is not None else 99999,
    ))
    return jsonify(out)


@bp.route('/productos/', methods=['POST'])
@_require_inventario_admin
def create_producto():
    data, err = _parse_or_422(ProductoCreateSchema(), request.get_json(silent=True))
    if err: return err

    if Producto.query.filter(Producto.codigo == data['codigo']).first():
        return jsonify({'detail': 'El código de producto ya existe'}), 400

    user = request.current_user
    stock_inicial = Decimal(str(data['stock_actual']))
    nuevo = Producto(
        codigo=data['codigo'],
        descripcion=data['descripcion'],
        categoria=data['categoria'],
        unidad=data['unidad'],
        stock_actual=stock_inicial,
        stock_minimo=Decimal(str(data['stock_minimo'])),
        imagen_url=data.get('imagen_url') or None,
        proveedor_default_nombre=(data.get('proveedor_default_nombre') or None),
        proveedor_default_contacto=(data.get('proveedor_default_contacto') or None),
        created_by_id=user.id,
    )
    db.session.add(nuevo)
    db.session.flush()  # obtener nuevo.id

    # Pausa 2: depositar el stock inicial en la bodega default. Sin esto,
    # Producto.stock_actual (cache) y stock_por_almacen (verdad) divergen
    # desde el primer movimiento.
    if stock_inicial > 0:
        default_id = _almacen_default_id()
        if default_id:
            db.session.add(StockPorAlmacen(
                producto_id=nuevo.id,
                almacen_id=default_id,
                cantidad=stock_inicial,
            ))

    _audit(user, f"Producto creado: {data['codigo']} — {data['descripcion']}")
    db.session.commit()
    db.session.refresh(nuevo)
    emit_to_role(_INV_ROLES, 'producto:changed', {
        'id': nuevo.id, 'action': 'created',
    })
    return jsonify(_producto_to_dict(nuevo))


@bp.route('/productos/<int:producto_id>', methods=['PUT'])
@_require_inventario_admin
def update_producto(producto_id: int):
    data, err = _parse_or_422(ProductoUpdateSchema(), request.get_json(silent=True))
    if err: return err

    prod = Producto.query.filter(Producto.id == producto_id, Producto.activo == True).first()
    if not prod:
        return jsonify({'detail': 'Producto no encontrado'}), 404

    cambios = []
    if data.get('codigo') is not None and data['codigo'] != prod.codigo:
        if Producto.query.filter(Producto.codigo == data['codigo']).first():
            return jsonify({'detail': 'El código ya existe en otro producto'}), 400
        cambios.append(f"codigo: {prod.codigo}→{data['codigo']}")
        prod.codigo = data['codigo']
    if data.get('descripcion') is not None:
        cambios.append("descripcion actualizada")
        prod.descripcion = data['descripcion']
    if data.get('categoria') is not None: prod.categoria = data['categoria']
    if data.get('unidad') is not None: prod.unidad = data['unidad']
    if data.get('imagen_url') is not None: prod.imagen_url = data['imagen_url'] or None
    if data.get('stock_actual') is not None:
        cambios.append(f"stock_actual: {prod.stock_actual}→{data['stock_actual']}")
        prod.stock_actual = Decimal(str(data['stock_actual']))
    if data.get('stock_minimo') is not None:
        prod.stock_minimo = Decimal(str(data['stock_minimo']))
    if data.get('proveedor_default_nombre') is not None:
        prod.proveedor_default_nombre = data['proveedor_default_nombre'] or None
    if data.get('proveedor_default_contacto') is not None:
        prod.proveedor_default_contacto = data['proveedor_default_contacto'] or None

    if cambios:
        _audit(request.current_user, f"Producto #{producto_id} editado: {'; '.join(cambios)}")

    db.session.commit()
    db.session.refresh(prod)
    emit_to_role(_INV_ROLES, 'producto:changed', {
        'id': prod.id, 'action': 'updated',
    })
    return jsonify(_producto_to_dict(prod))


@bp.route('/productos/<int:producto_id>/stocks', methods=['GET'])
@_require_inventario_admin
def get_producto_stocks(producto_id: int):
    """Desglose de stock por almacén para un producto (Pausa 2).

    Devuelve solo filas con cantidad > 0 por defecto. Pasar
    `?incluir_vacios=1` para ver también las bodegas con cantidad 0 (útil al
    decidir destino de un TRASPASO).
    """
    if not Producto.query.filter(Producto.id == producto_id, Producto.activo == True).first():  # noqa: E712
        return jsonify({'detail': 'Producto no encontrado'}), 404

    incluir_vacios = request.args.get('incluir_vacios') in ('1', 'true', 'yes')

    q = (
        db.session.query(StockPorAlmacen, Almacen)
        .join(Almacen, Almacen.id == StockPorAlmacen.almacen_id)
        .filter(StockPorAlmacen.producto_id == producto_id, Almacen.activo == True)  # noqa: E712
        .order_by(Almacen.nombre)
    )
    if not incluir_vacios:
        q = q.filter(StockPorAlmacen.cantidad > 0)

    rows = q.all()
    total = sum(float(s.cantidad or 0) for s, _ in rows)
    return jsonify({
        'producto_id': producto_id,
        'total': total,
        'stocks': [
            {
                'almacen_id': a.id,
                'almacen_nombre': a.nombre,
                'almacen_ubicacion': a.ubicacion or '',
                'cantidad': float(s.cantidad or 0),
                'updated_at': s.updated_at.isoformat() if s.updated_at else None,
            }
            for s, a in rows
        ],
    })


@bp.route('/productos/<int:producto_id>/disponibilidad', methods=['GET'])
@_require_inventario_admin
def get_producto_disponibilidad(producto_id: int):
    """Stock real / reservado / disponible de un producto (Pausa 2-bis).
    Incluye lista de solicitudes APROBADAS no entregadas que están apartando
    stock, para que el SPA muestre por qué hay reservas."""
    p = Producto.query.filter(Producto.id == producto_id).first()
    if not p:
        return jsonify({'detail': 'Producto no encontrado'}), 404

    actual = float(p.stock_actual or 0)
    reservado = float(p.stock_reservado or 0)
    disponible = actual - reservado

    # Solicitudes que generan la reserva: APROBADAS con detalles de este producto.
    rows = (
        db.session.query(
            SolicitudMaterial.id,
            SolicitudMaterial.proyecto,
            SolicitudMaterial.fecha_creacion,
            User.username,
            db.func.sum(
                db.func.greatest(
                    SolicitudMaterialDetalle.cantidad_solicitada
                    - db.func.coalesce(SolicitudMaterialDetalle.cantidad_entregada, 0),
                    0,
                )
            ).label('pendiente'),
        )
        .join(SolicitudMaterialDetalle, SolicitudMaterialDetalle.solicitud_id == SolicitudMaterial.id)
        .outerjoin(User, User.id == SolicitudMaterial.solicitante_id)
        .filter(
            SolicitudMaterial.estatus == 'APROBADA',
            SolicitudMaterialDetalle.producto_id == producto_id,
        )
        .group_by(SolicitudMaterial.id, SolicitudMaterial.proyecto,
                  SolicitudMaterial.fecha_creacion, User.username)
        .order_by(SolicitudMaterial.fecha_creacion.desc())
        .all()
    )

    return jsonify({
        'producto_id': producto_id,
        'codigo': p.codigo,
        'unidad': p.unidad,
        'stock_actual': actual,
        'stock_reservado': reservado,
        'stock_disponible': disponible,
        'reservas': [
            {
                'solicitud_id': r.id,
                'folio': f'SOL-{r.id:06d}',
                'proyecto': r.proyecto or '',
                'solicitante': r.username or '',
                'fecha': r.fecha_creacion.isoformat() if r.fecha_creacion else None,
                'cantidad': float(r.pendiente or 0),
            }
            for r in rows
        ],
    })


@bp.route('/productos/<int:producto_id>/kardex', methods=['GET'])
@_require_inventario_admin
def get_producto_kardex(producto_id: int):
    """Kardex (historial cronológico con saldo corrido) de un producto — Pausa 3.

    Query params:
      - desde (YYYY-MM-DD): default 30 días atrás.
      - hasta (YYYY-MM-DD): default hoy.
      - tipo: filtra ENTRADA/SALIDA/AJUSTE/TRASPASO (opcional).
      - limit: tope de filas (1..2000, default 500).

    Cálculo del saldo:
      saldo_inicial = stock_actual − Σ(deltas posteriores a `desde`)
    Luego se aplica el delta de cada movimiento en orden cronológico ascendente
    para obtener `saldo` por fila. TRASPASO no cambia el saldo total (mueve
    entre bodegas) pero se muestra para trazabilidad.
    """
    producto = Producto.query.filter(Producto.id == producto_id).first()
    if not producto:
        return jsonify({'detail': 'Producto no encontrado'}), 404

    # Rango por defecto: últimos 30 días.
    hoy = datetime.date.today()
    try:
        desde_str = request.args.get('desde')
        desde = datetime.date.fromisoformat(desde_str) if desde_str else (hoy - datetime.timedelta(days=30))
    except (TypeError, ValueError):
        return jsonify({'detail': "Parámetro 'desde' debe ser YYYY-MM-DD"}), 422
    try:
        hasta_str = request.args.get('hasta')
        hasta = datetime.date.fromisoformat(hasta_str) if hasta_str else hoy
    except (TypeError, ValueError):
        return jsonify({'detail': "Parámetro 'hasta' debe ser YYYY-MM-DD"}), 422
    if desde > hasta:
        return jsonify({'detail': "'desde' no puede ser mayor que 'hasta'"}), 422

    limit, err = _int_arg('limit', 500, 1, 2000)
    if err: return err

    tipo_filtro = request.args.get('tipo')
    if tipo_filtro and tipo_filtro not in ('ENTRADA', 'SALIDA', 'AJUSTE', 'TRASPASO'):
        return jsonify({'detail': "Parámetro 'tipo' inválido"}), 422

    # Helper: convierte (tipo, cantidad) en delta firmado para el saldo total.
    def _delta(mov: MovimientoInventario) -> Decimal:
        cant = mov.cantidad or Decimal('0')
        if mov.tipo == 'ENTRADA':
            return cant
        if mov.tipo == 'SALIDA':
            return -cant
        if mov.tipo == 'AJUSTE':
            return cant  # ya viene firmada
        return Decimal('0')  # TRASPASO no altera total

    # Datetime para filtros (incluyendo todo el día 'hasta').
    desde_dt = datetime.datetime.combine(desde, datetime.time.min)
    hasta_dt = datetime.datetime.combine(hasta, datetime.time.max)

    # 1) Calcular saldo inicial: stock_actual − Σ deltas posteriores a 'desde'.
    movs_post = (
        MovimientoInventario.query
        .filter(
            MovimientoInventario.producto_id == producto_id,
            MovimientoInventario.fecha >= desde_dt,
        )
        .all()
    )
    delta_post = sum((_delta(m) for m in movs_post), Decimal('0'))
    saldo_inicial = (producto.stock_actual or Decimal('0')) - delta_post

    # 2) Cargar movimientos del rango (con join a usuarios y almacenes para
    # evitar lazy queries en la serialización).
    q = (
        MovimientoInventario.query
        .options(
            joinedload(MovimientoInventario.usuario),
            joinedload(MovimientoInventario.almacen_origen),
            joinedload(MovimientoInventario.almacen_destino),
        )
        .filter(
            MovimientoInventario.producto_id == producto_id,
            MovimientoInventario.fecha >= desde_dt,
            MovimientoInventario.fecha <= hasta_dt,
        )
    )
    if tipo_filtro:
        q = q.filter(MovimientoInventario.tipo == tipo_filtro)

    # Orden ASC para calcular saldo corrido; al frontend le viene útil ASC
    # para timeline cronológico de arriba a abajo, pero también lo invertimos
    # opcionalmente en la UI.
    movs = q.order_by(MovimientoInventario.fecha.asc(), MovimientoInventario.id.asc()).limit(limit).all()

    saldo = saldo_inicial
    filas = []
    for m in movs:
        d = _delta(m)
        saldo = saldo + d
        filas.append({
            'id': m.id,
            'fecha': m.fecha.isoformat() if m.fecha else None,
            'tipo': m.tipo,
            'cantidad': float(m.cantidad or 0),
            'delta': float(d),
            'saldo': float(saldo),
            'almacen_origen': m.almacen_origen.nombre if m.almacen_origen else None,
            'almacen_destino': m.almacen_destino.nombre if m.almacen_destino else None,
            'usuario': m.usuario.username if m.usuario else None,
            'motivo': m.motivo or '',
        })

    return jsonify({
        'producto': {
            'id': producto.id,
            'codigo': producto.codigo,
            'descripcion': producto.descripcion,
            'unidad': producto.unidad,
            'categoria': producto.categoria,
            'stock_actual': float(producto.stock_actual or 0),
            'stock_minimo': float(producto.stock_minimo or 0),
        },
        'desde': desde.isoformat(),
        'hasta': hasta.isoformat(),
        'saldo_inicial': float(saldo_inicial),
        'saldo_final': float(saldo),
        'total_movimientos': len(filas),
        'movimientos': filas,
    })


@bp.route('/productos/<int:producto_id>', methods=['DELETE'])
@_require_inventario_admin
def delete_producto(producto_id: int):
    prod = Producto.query.filter(Producto.id == producto_id).first()
    if not prod:
        return jsonify({'detail': 'Producto no encontrado'}), 404
    prod.activo = False  # Soft delete: mantener histórico de movimientos/solicitudes
    _audit(request.current_user, f"Producto #{producto_id} ({prod.codigo}) desactivado (soft delete)")
    db.session.commit()
    emit_to_role(_INV_ROLES, 'producto:changed', {
        'id': producto_id, 'action': 'deleted',
    })
    return Response(status=204)


# ─── Almacenes ────────────────────────────────────────────────────────────────

@bp.route('/almacenes/', methods=['GET'])
@_require_inventario
def get_almacenes():
    almacenes = Almacen.query.filter(Almacen.activo == True).all()
    return jsonify([_almacen_to_dict(a) for a in almacenes])


@bp.route('/almacenes/', methods=['POST'])
@_require_inventario_admin
def create_almacen():
    data, err = _parse_or_422(AlmacenCreateSchema(), request.get_json(silent=True))
    if err: return err

    nuevo = Almacen(
        nombre=data['nombre'],
        ubicacion=data.get('ubicacion'),
        activo=data.get('activo', True),
        qr_code=str(uuid.uuid4()),
    )
    db.session.add(nuevo)
    _audit(request.current_user, f"Almacén creado: {data['nombre']}")
    db.session.commit()
    db.session.refresh(nuevo)
    return jsonify(_almacen_to_dict(nuevo))


@bp.route('/almacenes/<int:almacen_id>', methods=['PUT'])
@_require_inventario_admin
def update_almacen(almacen_id: int):
    data, err = _parse_or_422(AlmacenUpdateSchema(), request.get_json(silent=True))
    if err: return err

    alm = Almacen.query.filter(Almacen.id == almacen_id).first()
    if not alm:
        return jsonify({'detail': 'Bodega no encontrada'}), 404

    if data.get('nombre') is not None: alm.nombre = data['nombre']
    if data.get('ubicacion') is not None: alm.ubicacion = data['ubicacion']
    if data.get('activo') is not None: alm.activo = data['activo']
    _audit(request.current_user, f"Almacén #{almacen_id} editado")
    db.session.commit()
    db.session.refresh(alm)
    return jsonify(_almacen_to_dict(alm))


@bp.route('/almacenes/<int:almacen_id>', methods=['DELETE'])
@_require_inventario_admin
def delete_almacen(almacen_id: int):
    alm = Almacen.query.filter(Almacen.id == almacen_id).first()
    if not alm:
        return jsonify({'detail': 'Bodega no encontrada'}), 404
    alm.activo = False
    _audit(request.current_user, f"Almacén #{almacen_id} ({alm.nombre}) desactivado (soft delete)")
    db.session.commit()
    return Response(status=204)


@bp.route('/almacenes/<qr_code>/validar', methods=['GET'])
@_require_inventario
def validar_almacen(qr_code: str):
    alm = Almacen.query.filter(Almacen.qr_code == qr_code).first()
    if not alm:
        return jsonify({'detail': 'Almacén no encontrado o QR inválido'}), 404
    return jsonify(_almacen_to_dict(alm))


@bp.route('/almacenes/<int:almacen_id>/estantes', methods=['GET'])
@_require_inventario
def get_estantes_por_almacen(almacen_id: int):
    estantes = (
        Estante.query
        .filter(Estante.almacen_id == almacen_id, Estante.activo == True)
        .all()
    )
    return jsonify([_estante_to_dict(e) for e in estantes])


# ─── Estantes ─────────────────────────────────────────────────────────────────

@bp.route('/estantes/', methods=['GET'])
@_require_inventario
def get_estantes():
    estantes = Estante.query.filter(Estante.activo == True).all()
    return jsonify([_estante_to_dict(e) for e in estantes])


@bp.route('/estantes/', methods=['POST'])
@_require_inventario_admin
def create_estante():
    data, err = _parse_or_422(EstanteCreateSchema(), request.get_json(silent=True))
    if err: return err

    almacen = Almacen.query.filter(Almacen.id == data['almacen_id']).first()
    if not almacen:
        return jsonify({'detail': 'Almacén no encontrado'}), 404

    nuevo = Estante(
        nombre=data['nombre'],
        descripcion=data.get('descripcion'),
        almacen_id=data['almacen_id'],
        qr_code=str(uuid.uuid4()),
    )
    db.session.add(nuevo)
    _audit(request.current_user, f"Estante creado: {data['nombre']} en almacén #{data['almacen_id']}")
    db.session.commit()
    db.session.refresh(nuevo)
    return jsonify(_estante_to_dict(nuevo))


@bp.route('/estantes/<int:estante_id>', methods=['PUT'])
@_require_inventario_admin
def update_estante(estante_id: int):
    data, err = _parse_or_422(EstanteUpdateSchema(), request.get_json(silent=True))
    if err: return err

    est = Estante.query.filter(Estante.id == estante_id, Estante.activo == True).first()
    if not est:
        return jsonify({'detail': 'Estante no encontrado'}), 404

    if data.get('nombre') is not None: est.nombre = data['nombre']
    if data.get('descripcion') is not None: est.descripcion = data['descripcion']
    if data.get('almacen_id') is not None:
        almacen = Almacen.query.filter(Almacen.id == data['almacen_id']).first()
        if not almacen:
            return jsonify({'detail': 'Bodega destino no encontrada'}), 404
        est.almacen_id = data['almacen_id']
    _audit(request.current_user, f"Estante #{estante_id} editado")
    db.session.commit()
    db.session.refresh(est)
    return jsonify(_estante_to_dict(est))


@bp.route('/estantes/<int:estante_id>', methods=['DELETE'])
@_require_inventario_admin
def delete_estante(estante_id: int):
    est = Estante.query.filter(Estante.id == estante_id).first()
    if not est:
        return jsonify({'detail': 'Estante no encontrado'}), 404
    est.activo = False
    _audit(request.current_user, f"Estante #{estante_id} ({est.nombre}) desactivado (soft delete)")
    db.session.commit()
    return Response(status=204)


@bp.route('/estantes/<qr_code>/validar', methods=['GET'])
@_require_inventario
def validar_estante(qr_code: str):
    est = Estante.query.filter(Estante.qr_code == qr_code, Estante.activo == True).first()
    if not est:
        return jsonify({'detail': 'Estante no encontrado o QR inválido'}), 404
    return jsonify(_estante_to_dict(est))


@bp.route('/estantes/<qr_code>/inventario', methods=['GET'])
@_require_inventario
def inventario_estante(qr_code: str):
    """Devuelve los productos asignados al estante via ProductoEstante (Pausa 4).
    Si el estante no tiene productos asignados, lista vacía."""
    est = Estante.query.filter(Estante.qr_code == qr_code, Estante.activo == True).first()
    if not est:
        return jsonify({'detail': 'Estante no encontrado o QR inválido'}), 404
    productos = (
        Producto.query
        .join(ProductoEstante, ProductoEstante.producto_id == Producto.id)
        .filter(ProductoEstante.estante_id == est.id, Producto.activo == True)
        .order_by(Producto.codigo)
        .all()
    )
    return jsonify({
        'estante': _estante_to_dict(est),
        'productos': [_producto_to_dict(p) for p in productos],
    })


@bp.route('/estantes/<int:estante_id>/productos', methods=['GET'])
@_require_inventario
def estante_productos(estante_id: int):
    """Lista de productos asignados a un estante (por id, para la UI de admin)."""
    est = Estante.query.get_or_404(estante_id)
    productos = (
        Producto.query
        .join(ProductoEstante, ProductoEstante.producto_id == Producto.id)
        .filter(ProductoEstante.estante_id == est.id)
        .order_by(Producto.codigo)
        .all()
    )
    return jsonify([_producto_to_dict(p) for p in productos])


@bp.route('/estantes/<int:estante_id>/productos', methods=['PUT'])
@_require_inventario_admin
def set_estante_productos(estante_id: int):
    """Reemplaza la lista de productos asignados al estante.
    Body: `{producto_ids: [int]}`. Idempotente."""
    est = Estante.query.get_or_404(estante_id)
    body = request.get_json(silent=True) or {}
    ids = body.get('producto_ids')
    if not isinstance(ids, list):
        return jsonify({'detail': 'producto_ids debe ser una lista'}), 422
    ids = [int(x) for x in ids if isinstance(x, int) or (isinstance(x, str) and x.isdigit())]
    if ids:
        existentes = {p.id for p in Producto.query.filter(Producto.id.in_(ids)).all()}
        faltantes = set(ids) - existentes
        if faltantes:
            return jsonify({'detail': f'Productos inexistentes: {sorted(faltantes)}'}), 404

    ProductoEstante.query.filter_by(estante_id=est.id).delete()
    for pid in set(ids):
        db.session.add(ProductoEstante(producto_id=pid, estante_id=est.id))
    _audit(request.current_user, f"Estante #{est.id} ({est.nombre}): {len(set(ids))} productos asignados")
    db.session.commit()
    return jsonify({'success': True, 'asignados': len(set(ids))})


@bp.route('/estantes/<int:estante_id>/qr-image', methods=['GET'])
@_require_inventario
def get_estante_qr_image(estante_id: int):
    est = Estante.query.filter(Estante.id == estante_id).first()
    if not est:
        return jsonify({'detail': 'Estante no encontrado'}), 404

    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(est.qr_code)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')

    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return Response(buf.getvalue(), mimetype='image/png')


# ─── Movimientos ──────────────────────────────────────────────────────────────

@bp.route('/movimientos/', methods=['GET'])
@_require_inventario_admin
def get_movimientos():
    producto_id = request.args.get('producto_id', type=int)
    tipo = request.args.get('tipo', type=str)
    limit, err = _int_arg('limit', 200, 0, 1000)
    if err: return err

    q = MovimientoInventario.query
    if producto_id:
        q = q.filter(MovimientoInventario.producto_id == producto_id)
    if tipo:
        if len(tipo) > 20:
            return jsonify({'detail': "Parámetro 'tipo' demasiado largo"}), 422
        q = q.filter(MovimientoInventario.tipo == tipo.upper())
    movs = q.order_by(MovimientoInventario.fecha.desc()).limit(limit).all()
    return jsonify([_movimiento_to_dict(m) for m in movs])


@bp.route('/movimientos/', methods=['POST'])
@limiter.limit(
    "20/minute",
    # key por IP real para que requests sin sesión también cuenten al contador.
    # Si pusiéramos @_require_inventario antes, los 401 no incrementarían el contador
    # y un atacante anónimo podría martillear el endpoint sin freno hasta el límite global.
    key_func=lambda: f"ip:{get_real_client_ip_flask()}",
)
@_require_inventario_admin
def create_movimiento():
    """Crea un movimiento de inventario alterando StockPorAlmacen (Pausa 2).

    Reglas de almacén por tipo:
      - ENTRADA: requiere almacen_destino (donde llega el stock).
      - SALIDA:  requiere almacen_origen (de donde sale).
      - AJUSTE:  usa almacen_destino si cantidad>0, almacen_origen si <0.
      - TRASPASO: requiere ambos y deben ser distintos.

    Compatibilidad con clientes viejos: si no se manda ni almacén ni estante,
    se infiere de la bodega activa de menor id (no rompemos integraciones que
    aún tratan al stock como global). Loguea warning para detectar quién sigue
    sin mandar almacén.
    """
    data, err = _parse_or_422(MovimientoCreateSchema(), request.get_json(silent=True))
    if err: return err
    return _perform_movimiento(data, request.current_user)


def _perform_movimiento(data: dict, user):
    """Lógica central de creación de movimiento — usada por POST /movimientos/
    y POST /movimientos/rapido. `data` ya viene validado por
    MovimientoCreateSchema."""
    tipo = data['tipo']
    cantidad_raw = data['cantidad']

    # ENTRADA/SALIDA/TRASPASO requieren cantidad estrictamente positiva.
    # AJUSTE permite negativo (mermas) — eso lo controla la lógica de stock más abajo.
    if tipo in ['ENTRADA', 'SALIDA', 'TRASPASO'] and cantidad_raw <= 0:
        return jsonify({'detail': 'La cantidad debe ser positiva para este tipo de movimiento'}), 422
    if tipo not in ('ENTRADA', 'SALIDA', 'AJUSTE', 'TRASPASO'):
        return jsonify({'detail': 'Tipo de movimiento inválido'}), 400

    # 1) Resolver almacenes (con inferencia desde estante si solo vino estante_id).
    almacen_destino_id = data.get('almacen_destino_id')
    almacen_origen_id = data.get('almacen_origen_id')
    estante_id = data.get('estante_id')
    if estante_id and (not almacen_destino_id and not almacen_origen_id):
        estante = Estante.query.filter(Estante.id == estante_id).first()
        if estante:
            if tipo in ('ENTRADA',) or (tipo == 'AJUSTE' and cantidad_raw > 0):
                almacen_destino_id = estante.almacen_id
            else:
                almacen_origen_id = estante.almacen_id

    # Fallback compat: si todavía no hay almacén, usar la bodega default.
    if not (almacen_origen_id or almacen_destino_id):
        default_id = _almacen_default_id()
        if not default_id:
            return jsonify({'detail': 'No hay bodegas registradas. Crea un almacén antes de mover stock.'}), 400
        current_app.logger.warning(
            "Movimiento %s sin almacén explícito; usando bodega default #%s (producto=%s)",
            tipo, default_id, data['producto_id'],
        )
        if tipo == 'ENTRADA' or (tipo == 'AJUSTE' and cantidad_raw > 0):
            almacen_destino_id = default_id
        else:
            almacen_origen_id = default_id

    # 2) Validar combinación tipo/almacén.
    if tipo == 'ENTRADA' and not almacen_destino_id:
        return jsonify({'detail': 'ENTRADA requiere almacen_destino_id'}), 422
    if tipo == 'SALIDA' and not almacen_origen_id:
        return jsonify({'detail': 'SALIDA requiere almacen_origen_id'}), 422
    if tipo == 'TRASPASO':
        if not (almacen_origen_id and almacen_destino_id):
            return jsonify({'detail': 'TRASPASO requiere almacen_origen_id y almacen_destino_id'}), 422
        if almacen_origen_id == almacen_destino_id:
            return jsonify({'detail': 'TRASPASO requiere bodegas distintas'}), 422

    # 3) Verificar que el producto existe (sin lock todavía; el lock va sobre
    # StockPorAlmacen, que es la fuente de verdad).
    producto = Producto.query.filter(Producto.id == data['producto_id']).first()
    if not producto:
        return jsonify({'detail': 'Producto no encontrado'}), 404

    cantidad_decimal = Decimal(str(cantidad_raw))

    # Capturar stock previo para detectar cruce del umbral mínimo (Pausa 5).
    # TRASPASO no altera el total → no puede cruzar.
    stock_antes = Decimal(str(producto.stock_actual or 0))
    stock_minimo = Decimal(str(producto.stock_minimo or 0))

    try:
        # 4) Lock + alteración de filas de stock_por_almacen.
        # with_for_update previene over-selling cuando dos requests intentan
        # reducir el mismo stock al mismo tiempo.
        if tipo == 'ENTRADA':
            stock_dest = _lock_stock(producto.id, almacen_destino_id)
            stock_dest.cantidad = (stock_dest.cantidad or Decimal('0')) + cantidad_decimal

        elif tipo == 'SALIDA':
            stock_orig = _lock_stock(producto.id, almacen_origen_id)
            if (stock_orig.cantidad or Decimal('0')) < cantidad_decimal:
                db.session.rollback()
                return jsonify({
                    'detail': f'Stock insuficiente en bodega #{almacen_origen_id}. '
                              f'Disponible: {stock_orig.cantidad}'
                }), 400
            # Pausa 2-bis: respetar reservas globales (no sacar lo apartado).
            reservado = Decimal(str(producto.stock_reservado or 0))
            disponible_global = stock_antes - reservado
            if disponible_global < cantidad_decimal:
                db.session.rollback()
                return jsonify({
                    'detail': (
                        f'Hay {reservado} {producto.unidad} apartados por solicitudes aprobadas. '
                        f'Disponible para sacar: {disponible_global}. '
                        f'Libera/rechaza solicitudes antes o entrega vía el flujo de solicitudes.'
                    ),
                }), 409
            stock_orig.cantidad = stock_orig.cantidad - cantidad_decimal

        elif tipo == 'AJUSTE':
            # Positivo → sube destino; negativo → baja origen (sin pasar de 0).
            if cantidad_decimal >= 0:
                stock_dest = _lock_stock(producto.id, almacen_destino_id)
                stock_dest.cantidad = (stock_dest.cantidad or Decimal('0')) + cantidad_decimal
            else:
                stock_orig = _lock_stock(producto.id, almacen_origen_id)
                disponible = stock_orig.cantidad or Decimal('0')
                if disponible + cantidad_decimal < 0:
                    db.session.rollback()
                    return jsonify({
                        'detail': f'Ajuste provocaría stock negativo en bodega #{almacen_origen_id}'
                    }), 400
                # Pausa 2-bis: respetar reservas globales también.
                reservado = Decimal(str(producto.stock_reservado or 0))
                disponible_global_post = stock_antes + cantidad_decimal - reservado
                if disponible_global_post < 0:
                    db.session.rollback()
                    return jsonify({
                        'detail': (
                            f'Ajuste invadiría stock apartado. Reservado: {reservado}, '
                            f'stock tras ajuste: {stock_antes + cantidad_decimal}.'
                        ),
                    }), 409
                stock_orig.cantidad = disponible + cantidad_decimal

        elif tipo == 'TRASPASO':
            # Lock en orden determinístico (menor id primero) para evitar deadlocks
            # entre dos TRASPASOs cruzados (A→B y B→A simultáneos).
            ids = sorted([almacen_origen_id, almacen_destino_id])
            stock_a = _lock_stock(producto.id, ids[0])
            stock_b = _lock_stock(producto.id, ids[1])
            stock_orig = stock_a if ids[0] == almacen_origen_id else stock_b
            stock_dest = stock_b if ids[0] == almacen_origen_id else stock_a
            if (stock_orig.cantidad or Decimal('0')) < cantidad_decimal:
                db.session.rollback()
                return jsonify({
                    'detail': f'Stock insuficiente para traspaso en bodega #{almacen_origen_id}. '
                              f'Disponible: {stock_orig.cantidad}'
                }), 400
            stock_orig.cantidad = stock_orig.cantidad - cantidad_decimal
            stock_dest.cantidad = (stock_dest.cantidad or Decimal('0')) + cantidad_decimal

    except Exception as exc:
        db.session.rollback()
        # nowait=True levanta si la fila ya estaba bloqueada por otra transacción.
        # Devolvemos 409 para que el cliente reintente.
        if 'could not obtain lock' in str(exc).lower():
            return jsonify({'detail': 'Stock bloqueado por otra operación, reintenta'}), 409
        raise

    # 5) Actualizar cache desnormalizado Producto.stock_actual. Lo hacemos
    # DENTRO de la misma transacción para que nunca quede desfasado en
    # commits exitosos. TRASPASO no cambia el total, pero igual recalculamos
    # por seguridad (el costo es despreciable).
    _recalcular_cache_stock(producto)

    # 5b) Pausa 5: si este movimiento CRUZÓ el umbral mínimo, notificar a
    # inventario. Solo notifica al cruzar (de OK a bajo mínimo); movimientos
    # adicionales bajo mínimo no spamean. Idempotencia diaria con tabla
    # NotificacionUmbral. Try/except: la notif no debe romper el movimiento.
    stock_despues = Decimal(str(producto.stock_actual or 0))
    cruzo_umbral = (
        stock_minimo > 0
        and stock_antes > stock_minimo
        and stock_despues <= stock_minimo
    )
    if cruzo_umbral:
        try:
            hoy = datetime.date.today()
            ya_notificado = db.session.get(NotificacionUmbral, (producto.id, hoy))
            if not ya_notificado:
                db.session.add(NotificacionUmbral(producto_id=producto.id, fecha=hoy))
                crear_notif_inventario(
                    tipo='STOCK_BAJO',
                    titulo=f'Stock bajo: {producto.codigo}',
                    mensaje=(
                        f'{producto.descripcion} quedó en {stock_despues} {producto.unidad} '
                        f'(mínimo: {stock_minimo}).'
                    ),
                    url='/inventario/bajo-minimo',
                )
        except Exception:
            current_app.logger.warning("No se pudo crear notificación STOCK_BAJO", exc_info=True)

    # 6) Registrar el movimiento histórico.
    nuevo_mov = MovimientoInventario(
        tipo=tipo,
        producto_id=data['producto_id'],
        cantidad=cantidad_decimal,
        almacen_origen_id=almacen_origen_id,
        almacen_destino_id=almacen_destino_id,
        motivo=data.get('motivo') or (f"Estante #{estante_id}" if estante_id else None),
        usuario_id=user.id,
    )
    db.session.add(nuevo_mov)
    _audit(user, f"Movimiento {tipo} — producto #{data['producto_id']} — cantidad: {cantidad_raw}")
    db.session.commit()
    db.session.refresh(nuevo_mov)
    emit_to_role(_INV_ROLES, 'movimiento:changed', {
        'id': nuevo_mov.id, 'producto_id': nuevo_mov.producto_id, 'tipo': tipo,
    })
    return jsonify(_movimiento_to_dict(nuevo_mov))


@bp.route('/movimientos/rapido', methods=['POST'])
@limiter.limit(
    "30/minute",
    key_func=lambda: f"ip:{get_real_client_ip_flask()}",
)
@_require_inventario_admin
def create_movimiento_rapido():
    """Atajo de PWA: registra movimiento resolviendo producto por su código/QR
    y almacén por estante_qr (Pausa 4).

    Body: `{producto_qr: str, estante_qr?: str, tipo: ENTRADA|SALIDA|AJUSTE,
            cantidad: number, motivo?: str}`.

    - `producto_qr` se resuelve contra `Producto.codigo`.
    - `estante_qr` se resuelve contra `Estante.qr_code` para inferir el almacén.
      Si no se manda, se usa el almacén default.
    - Para AJUSTE acepta cantidades negativas.

    Devuelve el mismo shape que POST /movimientos/.
    """
    body = request.get_json(silent=True) or {}
    producto_qr = (body.get('producto_qr') or '').strip()
    estante_qr = (body.get('estante_qr') or '').strip()
    tipo = (body.get('tipo') or '').strip().upper()
    cantidad_raw = body.get('cantidad')
    motivo = (body.get('motivo') or 'Movimiento rápido desde PWA').strip()[:250]

    if not producto_qr:
        return jsonify({'detail': 'producto_qr es requerido'}), 422
    if tipo not in ('ENTRADA', 'SALIDA', 'AJUSTE'):
        return jsonify({'detail': "tipo debe ser ENTRADA, SALIDA o AJUSTE"}), 422

    producto = Producto.query.filter(Producto.codigo == producto_qr, Producto.activo == True).first()
    if not producto:
        return jsonify({'detail': f'Producto con código {producto_qr} no encontrado'}), 404

    almacen_id = None
    if estante_qr:
        est = Estante.query.filter(Estante.qr_code == estante_qr, Estante.activo == True).first()
        if not est:
            return jsonify({'detail': f'Estante {estante_qr} no encontrado'}), 404
        almacen_id = est.almacen_id
    else:
        almacen_id = _almacen_default_id()
        if not almacen_id:
            return jsonify({'detail': 'No hay almacén default configurado'}), 422

    try:
        cant_decimal = Decimal(str(cantidad_raw))
    except Exception:
        return jsonify({'detail': 'cantidad inválida'}), 422

    data = {
        'tipo': tipo,
        'producto_id': producto.id,
        'cantidad': cant_decimal,
        'motivo': motivo,
    }
    if tipo == 'ENTRADA' or (tipo == 'AJUSTE' and cant_decimal >= 0):
        data['almacen_destino_id'] = almacen_id
    else:
        data['almacen_origen_id'] = almacen_id

    return _perform_movimiento(data, request.current_user)


# ─── Solicitudes ──────────────────────────────────────────────────────────────

@bp.route('/solicitudes/', methods=['POST'])
@limiter.limit(
    "10/minute",
    # key por IP real: ver comentario en create_movimiento sobre por qué el limiter
    # debe correr ANTES del check de auth (replica el patrón del FastAPI original).
    key_func=lambda: f"ip:{get_real_client_ip_flask()}",
)
@_require_login
def create_solicitud():
    user = request.current_user
    if user.role not in ['solicitante_material', 'coordinador', 'admin', 'inventario']:
        return jsonify({'detail': 'No tienes permiso para crear solicitudes'}), 403

    data, err = _parse_or_422(SolicitudCreateSchema(), request.get_json(silent=True))
    if err: return err

    nueva = SolicitudMaterial(
        solicitante_id=user.id,
        proyecto=data.get('proyecto'),
        estatus='PENDIENTE',
    )
    db.session.add(nueva)
    db.session.flush()  # Necesario para obtener nueva.id antes de crear detalles

    errores_detalle = []
    for idx, det in enumerate(data['detalles']):
        tipo = (det.get('tipo_item') or 'MATERIAL').upper()

        # XOR: cada línea es MATERIAL o HERRAMIENTA, no ambas.
        if tipo == 'MATERIAL':
            if not det.get('producto_id') or det.get('herramienta_id'):
                errores_detalle.append(f"Línea {idx+1}: MATERIAL requiere producto_id y no herramienta_id")
                continue
            producto = Producto.query.filter(Producto.id == det['producto_id'],
                                              Producto.activo == True).first()
            if not producto:
                errores_detalle.append(f"Línea {idx+1}: producto_id {det['producto_id']} no existe o inactivo")
                continue
            db.session.add(SolicitudMaterialDetalle(
                solicitud_id=nueva.id,
                tipo_item='MATERIAL',
                producto_id=det['producto_id'],
                cantidad_solicitada=Decimal(str(det['cantidad_solicitada'])),
                justificacion=det.get('justificacion'),
            ))
        elif tipo == 'HERRAMIENTA':
            if not det.get('herramienta_id') or det.get('producto_id'):
                errores_detalle.append(f"Línea {idx+1}: HERRAMIENTA requiere herramienta_id y no producto_id")
                continue
            herr = Herramienta.query.filter(Herramienta.id == det['herramienta_id'],
                                              Herramienta.activo == True).first()
            if not herr:
                errores_detalle.append(f"Línea {idx+1}: herramienta_id {det['herramienta_id']} no existe o inactiva")
                continue
            fi = det.get('fecha_uso_inicio')
            ff = det.get('fecha_uso_fin')
            if fi and ff and fi > ff:
                errores_detalle.append(f"Línea {idx+1}: fecha_uso_inicio > fecha_uso_fin")
                continue
            if fi and ff and (ff - fi).days > 365:
                errores_detalle.append(f"Línea {idx+1}: rango de uso mayor a 365 días")
                continue
            db.session.add(SolicitudMaterialDetalle(
                solicitud_id=nueva.id,
                tipo_item='HERRAMIENTA',
                herramienta_id=det['herramienta_id'],
                cantidad_solicitada=Decimal(str(det['cantidad_solicitada'])),
                fecha_uso_inicio=fi,
                fecha_uso_fin=ff,
                justificacion=det.get('justificacion'),
                complementos=det.get('complementos'),
            ))

    if errores_detalle:
        db.session.rollback()
        return jsonify({'detail': errores_detalle}), 400

    _audit(user, f"Nueva solicitud — proyecto: {data.get('proyecto') or 'Sin proyecto'}")
    db.session.commit()
    db.session.refresh(nueva)
    emit_to_role(_SOL_ROLES, 'solicitud:changed', {
        'id': nueva.id, 'action': 'created',
    })

    return jsonify(_solicitud_to_dict(nueva))


@bp.route('/solicitudes/', methods=['GET'])
@_require_login
def get_solicitudes():
    user = request.current_user
    skip, err = _int_arg('skip', 0, 0, 1_000_000)
    if err: return err
    limit, err = _int_arg('limit', 200, 0, 500)
    if err: return err

    query = SolicitudMaterial.query
    # solicitante_material y coordinador solo ven sus propias solicitudes.
    # inventario/admin/super_admin ven todas. Otros roles no entran.
    if user.role in ('solicitante_material', 'coordinador'):
        query = query.filter(SolicitudMaterial.solicitante_id == user.id)
    elif user.role not in ('inventario', 'admin', 'super_admin'):
        return jsonify({'detail': 'No tienes permiso'}), 403

    solicitudes = (
        query
        .options(
            joinedload(SolicitudMaterial.solicitante),
            selectinload(SolicitudMaterial.detalles).joinedload(SolicitudMaterialDetalle.producto),
            selectinload(SolicitudMaterial.detalles).joinedload(SolicitudMaterialDetalle.herramienta),
        )
        .order_by(SolicitudMaterial.fecha_creacion.desc())
        .offset(skip).limit(limit)
        .all()
    )
    return jsonify([_solicitud_to_dict(s) for s in solicitudes])


@bp.route('/solicitudes/<int:sol_id>/estado', methods=['PATCH'])
@_require_inventario_admin
def update_solicitud_estado(sol_id: int):
    """Cambia el estatus de una solicitud aplicando reservas de stock (Pausa 2-bis).

    Transiciones y efecto en stock_reservado:
      - PENDIENTE → APROBADA:  RESERVA (puede fallar 409 si no hay disponible).
      - APROBADA → RECHAZADA:  LIBERA reservas.
      - APROBADA → PENDIENTE:  LIBERA reservas (se re-aprobará después).
      - APROBADA → ENTREGADA:  LIBERA reservas. NO descuenta stock — la SALIDA
        real la registra el almacenista por separado en /movimientos. Cuando
        llegue Pausa 8b (entrega parcial), ese endpoint sí descuenta stock.
      - ENTREGADA → PENDIENTE: RE-RESERVA (puede fallar 409).
      - RECHAZADA → PENDIENTE: sin efecto (no había reserva).
      - PENDIENTE → RECHAZADA: sin efecto.
    """
    data, err = _parse_or_422(SolicitudUpdateEstadoSchema(), request.get_json(silent=True))
    if err: return err

    sol = (
        SolicitudMaterial.query
        .options(selectinload(SolicitudMaterial.detalles))
        .filter(SolicitudMaterial.id == sol_id)
        .first()
    )
    if not sol:
        return jsonify({'detail': 'Solicitud no encontrada'}), 404

    estado_previo = sol.estatus
    nuevo_estado = data['estatus']

    TRANSICIONES_VALIDAS = {
        'PENDIENTE':  {'APROBADA', 'RECHAZADA'},
        'APROBADA':   {'ENTREGADA', 'RECHAZADA', 'PENDIENTE'},
        'RECHAZADA':  {'PENDIENTE'},
        'ENTREGADA':  {'PENDIENTE'},
    }
    permitidas = TRANSICIONES_VALIDAS.get(estado_previo, set())
    if nuevo_estado != estado_previo and nuevo_estado not in permitidas:
        return jsonify({
            'detail': f"Transición inválida: {estado_previo} → {nuevo_estado}",
            'permitidas': sorted(permitidas),
        }), 409

    # Pausa 8b: al APROBAR, sembramos cantidad_aprobada = cantidad_solicitada en
    # cada línea MATERIAL que aún esté en 0 (default del modelo). Así la
    # reserva, la entrega parcial y el PATCH de detalle trabajan sobre un
    # campo explícito y dejamos de depender del fallback a cantidad_solicitada.
    if estado_previo == 'PENDIENTE' and nuevo_estado == 'APROBADA':
        for d in (sol.detalles or []):
            if d.tipo_item != 'MATERIAL' or not d.producto_id:
                continue
            if Decimal(str(d.cantidad_aprobada or 0)) == 0:
                d.cantidad_aprobada = Decimal(str(d.cantidad_solicitada or 0))

    # Reservas que esta solicitud "debería" tener apartadas.
    reservas = _reservas_de_solicitud(sol)

    try:
        # ── Aplicar efecto en stock_reservado según transición ──
        if estado_previo == 'PENDIENTE' and nuevo_estado == 'APROBADA':
            errs = _intentar_reservar(reservas)
            if errs:
                db.session.rollback()
                return jsonify({
                    'detail': 'No se puede aprobar: stock insuficiente',
                    'errores': errs,
                }), 409

        elif estado_previo == 'ENTREGADA' and nuevo_estado == 'PENDIENTE':
            # Reabrir entregada: re-reservar. Si el stock ya se movió a otra
            # solicitud entre tanto, falla.
            errs = _intentar_reservar(reservas)
            if errs:
                db.session.rollback()
                return jsonify({
                    'detail': 'No se puede reabrir (entregada): stock ya no disponible',
                    'errores': errs,
                }), 409

        elif estado_previo == 'APROBADA' and nuevo_estado in ('RECHAZADA', 'PENDIENTE', 'ENTREGADA'):
            # Liberar lo que se había reservado al aprobar.
            _liberar_reservas(reservas)

        # Resto de transiciones (PENDIENTE↔RECHAZADA, RECHAZADA→PENDIENTE): sin efecto.

        sol.estatus = nuevo_estado
        if nuevo_estado == 'PENDIENTE':
            sol.fecha_cierre = None
        else:
            sol.fecha_cierre = datetime.datetime.now()

        if estado_previo != nuevo_estado:
            _audit(request.current_user, f"Solicitud #{sol_id} estatus: {estado_previo} → {nuevo_estado}")

        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        if 'could not obtain lock' in str(exc).lower():
            return jsonify({'detail': 'Stock bloqueado por otra operación, reintenta'}), 409
        raise

    db.session.refresh(sol)
    _ = list(sol.detalles)
    if estado_previo != nuevo_estado:
        emit_to_role(_SOL_ROLES, 'solicitud:changed', {
            'id': sol.id, 'action': f'estado:{nuevo_estado}',
        })
    return jsonify(_solicitud_to_dict(sol))


# ─── Pausa 8b: edición de cantidad aprobada y entrega parcial ────────────────

@bp.route('/solicitudes/<int:sol_id>/detalles/<int:det_id>', methods=['PATCH'])
@_require_inventario_admin
def patch_solicitud_detalle(sol_id: int, det_id: int):
    """Edita `cantidad_aprobada` de una línea de solicitud APROBADA (Pausa 8b).

    Reglas:
      - Solicitud debe estar en APROBADA.
      - Línea debe ser MATERIAL con producto_id.
      - 0 ≤ cantidad_aprobada ≤ cantidad_solicitada.
      - cantidad_aprobada ≥ cantidad_entregada (no se aprueba menos de lo ya salido).
      - Ajusta `Producto.stock_reservado` por el delta:
          delta > 0 → intenta reservar (puede fallar 409 si no hay disponible).
          delta < 0 → libera.
    """
    data, err = _parse_or_422(SolicitudDetallePatchSchema(), request.get_json(silent=True))
    if err: return err

    det = (
        SolicitudMaterialDetalle.query
        .filter(
            SolicitudMaterialDetalle.id == det_id,
            SolicitudMaterialDetalle.solicitud_id == sol_id,
        )
        .first()
    )
    if not det:
        return jsonify({'detail': 'Detalle no encontrado'}), 404
    if (det.tipo_item or 'MATERIAL').upper() != 'MATERIAL' or not det.producto_id:
        return jsonify({'detail': 'Solo líneas de MATERIAL pueden editarse aquí'}), 422

    sol = SolicitudMaterial.query.filter(SolicitudMaterial.id == sol_id).first()
    if not sol:
        return jsonify({'detail': 'Solicitud no encontrada'}), 404
    if sol.estatus != 'APROBADA':
        return jsonify({
            'detail': f'Solo solicitudes APROBADAS permiten editar cantidad_aprobada (actual: {sol.estatus})'
        }), 409

    nueva_aprob = Decimal(str(data['cantidad_aprobada']))
    cant_sol = Decimal(str(det.cantidad_solicitada or 0))
    cant_ent = Decimal(str(det.cantidad_entregada or 0))
    cant_aprob_actual = Decimal(str(det.cantidad_aprobada or 0))

    if nueva_aprob > cant_sol:
        return jsonify({
            'detail': f'cantidad_aprobada ({nueva_aprob}) no puede exceder cantidad_solicitada ({cant_sol})'
        }), 422
    if nueva_aprob < cant_ent:
        return jsonify({
            'detail': f'cantidad_aprobada ({nueva_aprob}) no puede ser menor a cantidad_entregada ({cant_ent})'
        }), 422

    # Baseline para reserva previa: lo que el código antiguo asumía si cant_aprob=0.
    baseline_anterior = cant_aprob_actual if cant_aprob_actual > 0 else cant_sol
    reserva_anterior = baseline_anterior - cant_ent
    reserva_nueva = nueva_aprob - cant_ent
    delta = reserva_nueva - reserva_anterior

    try:
        if delta > 0:
            errs = _intentar_reservar({det.producto_id: delta})
            if errs:
                db.session.rollback()
                return jsonify({
                    'detail': 'No se puede aumentar la cantidad aprobada: stock insuficiente',
                    'errores': errs,
                }), 409
        elif delta < 0:
            _liberar_reservas({det.producto_id: -delta})

        det.cantidad_aprobada = nueva_aprob
        _audit(
            request.current_user,
            f"Solicitud #{sol_id} det #{det_id} cantidad_aprobada: {cant_aprob_actual} → {nueva_aprob}",
        )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        if 'could not obtain lock' in str(exc).lower():
            return jsonify({'detail': 'Stock bloqueado por otra operación, reintenta'}), 409
        raise

    db.session.refresh(det)
    emit_to_role(_SOL_ROLES, 'solicitud:changed', {
        'id': sol_id, 'detalle_id': det_id, 'action': 'detalle_updated',
    })
    return jsonify(_solicitud_detalle_to_dict(det))


@bp.route('/solicitudes/<int:sol_id>/entregar', methods=['POST'])
@limiter.limit(
    "20/minute",
    key_func=lambda: f"ip:{get_real_client_ip_flask()}",
)
@_require_inventario_admin
def entregar_solicitud(sol_id: int):
    """Entrega total o parcial de una solicitud APROBADA (Pausa 8b).

    Body: `{ almacen_origen_id?, motivo?, entregas: [{detalle_id, cantidad_entregada}, ...] }`.

    Crea SALIDA por cada línea con cantidad > 0, libera la porción de
    `stock_reservado` correspondiente y descuenta el stock físico del almacén.

    Si tras la entrega todas las líneas MATERIAL tienen
    cantidad_entregada == cantidad_aprobada, la solicitud queda ENTREGADA.
    En caso contrario sigue APROBADA (entrega parcial).

    Reglas de validación:
      - sol.estatus debe ser 'APROBADA'.
      - Cada detalle debe pertenecer a la solicitud y ser MATERIAL.
      - cantidad_entregada por línea ≤ (cantidad_aprobada − cantidad_entregada actual).
      - Stock en el almacén origen ≥ suma de entregas por producto.
      - Lock con `with_for_update(nowait=True)` en orden determinístico (id asc).
    """
    data, err = _parse_or_422(EntregarSolicitudSchema(), request.get_json(silent=True))
    if err: return err

    sol = (
        SolicitudMaterial.query
        .options(selectinload(SolicitudMaterial.detalles))
        .filter(SolicitudMaterial.id == sol_id)
        .first()
    )
    if not sol:
        return jsonify({'detail': 'Solicitud no encontrada'}), 404
    if sol.estatus != 'APROBADA':
        return jsonify({
            'detail': f'Solo solicitudes APROBADAS pueden entregarse (actual: {sol.estatus})'
        }), 409

    detalles_por_id = {d.id: d for d in (sol.detalles or [])}
    vistos: set[int] = set()
    # MATERIAL: (det, delta, baseline). HERRAMIENTA: (det, delta_int, baseline_int).
    entregas_material: list[tuple[SolicitudMaterialDetalle, Decimal, Decimal]] = []
    entregas_herramienta: list[tuple[SolicitudMaterialDetalle, int, int]] = []

    for item in data['entregas']:
        det_id = item['detalle_id']
        if det_id in vistos:
            return jsonify({'detail': f'Detalle #{det_id} duplicado en el payload'}), 422
        vistos.add(det_id)

        det = detalles_por_id.get(det_id)
        if not det:
            return jsonify({'detail': f'Detalle #{det_id} no pertenece a la solicitud #{sol_id}'}), 422

        delta = Decimal(str(item['cantidad_entregada']))
        if delta < 0:
            return jsonify({'detail': f'Detalle #{det_id}: cantidad_entregada no puede ser negativa'}), 422
        if delta == 0:
            continue  # el front puede mandar 0 para "no entregar esta línea ahora"

        tipo = (det.tipo_item or 'MATERIAL').upper()

        if tipo == 'HERRAMIENTA':
            if not det.herramienta_id:
                return jsonify({
                    'detail': f'Detalle #{det_id}: línea HERRAMIENTA sin herramienta_id',
                }), 422
            # Herramientas son enteras (1 unidad física = 1 asignación).
            if delta != delta.to_integral_value():
                return jsonify({
                    'detail': f'Detalle #{det_id}: cantidad de herramientas debe ser entera',
                }), 422
            cant_aprob_h = int(det.cantidad_aprobada or 0)
            cant_ent_h = int(det.cantidad_entregada or 0)
            baseline_h = cant_aprob_h if cant_aprob_h > 0 else int(det.cantidad_solicitada or 0)
            pendiente_h = baseline_h - cant_ent_h
            delta_int = int(delta)
            if delta_int > pendiente_h:
                return jsonify({
                    'detail': (
                        f'Detalle #{det_id}: cantidad_entregada ({delta_int}) excede el '
                        f'pendiente ({pendiente_h}). Aprobada: {baseline_h}, '
                        f'ya entregada: {cant_ent_h}.'
                    ),
                }), 422
            entregas_herramienta.append((det, delta_int, baseline_h))
            continue

        # MATERIAL
        if not det.producto_id:
            return jsonify({
                'detail': f'Detalle #{det_id}: línea MATERIAL sin producto_id',
            }), 422

        cant_aprob = Decimal(str(det.cantidad_aprobada or 0))
        cant_ent_actual = Decimal(str(det.cantidad_entregada or 0))
        # Compat con solicitudes pre-8b que aprobaron sin sembrar cantidad_aprobada.
        baseline = cant_aprob if cant_aprob > 0 else Decimal(str(det.cantidad_solicitada or 0))
        pendiente = baseline - cant_ent_actual
        if delta > pendiente:
            return jsonify({
                'detail': (
                    f'Detalle #{det_id}: cantidad_entregada ({delta}) excede el pendiente '
                    f'({pendiente}). Aprobada: {baseline}, ya entregada: {cant_ent_actual}.'
                ),
            }), 422

        entregas_material.append((det, delta, baseline))

    if not entregas_material and not entregas_herramienta:
        return jsonify({'detail': 'Ninguna línea con cantidad mayor a 0 para entregar'}), 422

    # Si hay líneas HERRAMIENTA, el solicitante DEBE tener un trabajador
    # asociado: la asignación se hace al Trabajador, no al User.
    trab_solicitante = None
    if entregas_herramienta:
        if not sol.solicitante or not sol.solicitante.trabajador_id:
            return jsonify({
                'detail': (
                    'El solicitante no tiene un trabajador asociado. '
                    'Liga la cuenta a un trabajador desde Usuarios para poder '
                    'entregar las herramientas, o asígnalas manualmente desde '
                    'Asignaciones de Herramienta.'
                ),
            }), 400
        trab_solicitante = Trabajador.query.filter(
            Trabajador.id == sol.solicitante.trabajador_id,
            Trabajador.activo == True,  # noqa: E712
        ).first()
        if not trab_solicitante:
            return jsonify({
                'detail': (
                    f'El trabajador #{sol.solicitante.trabajador_id} asociado al '
                    f'solicitante no existe o está inactivo.'
                ),
            }), 400

    # Resolver almacén (solo necesario si hay líneas MATERIAL). Fallback al
    # default. Si solo hay HERRAMIENTAS no exigimos bodega.
    almacen_id = None
    if entregas_material:
        almacen_id = data.get('almacen_origen_id') or _almacen_default_id()
        if not almacen_id:
            return jsonify({'detail': 'No hay bodegas registradas para descontar stock'}), 400
        almacen = Almacen.query.filter(Almacen.id == almacen_id, Almacen.activo == True).first()  # noqa: E712
        if not almacen:
            return jsonify({'detail': f'Almacén #{almacen_id} no existe o está inactivo'}), 404

    # Sumar deltas por producto para validar stock una sola vez por producto.
    delta_por_producto: dict[int, Decimal] = {}
    for det, delta, _ in entregas_material:
        delta_por_producto[det.producto_id] = (
            delta_por_producto.get(det.producto_id, Decimal('0')) + delta
        )

    # Sumar unidades por herramienta para validar disponibilidad una sola vez.
    delta_por_herramienta: dict[int, int] = {}
    for det, delta_int, _ in entregas_herramienta:
        delta_por_herramienta[det.herramienta_id] = (
            delta_por_herramienta.get(det.herramienta_id, 0) + delta_int
        )

    user = request.current_user
    motivo_base = (data.get('motivo') or '').strip() or f'Entrega solicitud #{sol_id}'
    fecha_dev_prevista = data.get('fecha_devolucion_prevista')

    try:
        # Lock determinístico (id asc) sobre Producto + StockPorAlmacen.
        productos_locked: dict[int, Producto] = {}
        for prod_id in sorted(delta_por_producto.keys()):
            producto = (
                Producto.query
                .with_for_update(nowait=True)
                .filter(Producto.id == prod_id)
                .first()
            )
            if not producto:
                db.session.rollback()
                return jsonify({'detail': f'Producto #{prod_id} no encontrado'}), 404
            productos_locked[prod_id] = producto

            cant_total = delta_por_producto[prod_id]
            stock_almacen = _lock_stock(prod_id, almacen_id)
            disponible_almacen = stock_almacen.cantidad or Decimal('0')
            if disponible_almacen < cant_total:
                db.session.rollback()
                return jsonify({
                    'detail': (
                        f'Stock insuficiente en bodega #{almacen_id} para {producto.codigo}: '
                        f'requiere {cant_total} {producto.unidad}, disponible {disponible_almacen}.'
                    ),
                }), 409

            # Descontar stock físico del almacén y liberar reserva equivalente.
            # Liberamos lo apartado por esta solicitud (clamp a 0 por seguridad).
            stock_almacen.cantidad = disponible_almacen - cant_total
            reservado = Decimal(str(producto.stock_reservado or 0))
            nuevo_res = reservado - cant_total
            producto.stock_reservado = nuevo_res if nuevo_res > 0 else Decimal('0')

        # Reservar unidades DISPONIBLES por herramienta (id asc anti-deadlock).
        # Tomamos exactamente `cant_total` unidades por herramienta — si no hay
        # suficientes, rollback con 409.
        unidades_por_herramienta: dict[int, list[HerramientaUnidad]] = {}
        for h_id in sorted(delta_por_herramienta.keys()):
            cant_unidades = delta_por_herramienta[h_id]
            unidades = (
                HerramientaUnidad.query
                .with_for_update(nowait=True)
                .filter(
                    HerramientaUnidad.herramienta_id == h_id,
                    HerramientaUnidad.estado == 'DISPONIBLE',
                )
                .order_by(HerramientaUnidad.id.asc())
                .limit(cant_unidades)
                .all()
            )
            if len(unidades) < cant_unidades:
                db.session.rollback()
                herr = Herramienta.query.get(h_id)
                nombre = herr.descripcion if herr else f'#{h_id}'
                return jsonify({
                    'detail': (
                        f'No hay unidades suficientes DISPONIBLES de "{nombre}": '
                        f'requiere {cant_unidades}, disponibles {len(unidades)}.'
                    ),
                }), 409
            unidades_por_herramienta[h_id] = unidades

        # Aplicar deltas por línea MATERIAL + registrar movimientos SALIDA.
        for det, delta, baseline in entregas_material:
            # Si la línea era pre-8b (cant_aprob=0), formalizar la aprobación con
            # el baseline para que la lógica de completa/reservas sea consistente.
            if Decimal(str(det.cantidad_aprobada or 0)) == 0 and baseline > 0:
                det.cantidad_aprobada = baseline

            det.cantidad_entregada = Decimal(str(det.cantidad_entregada or 0)) + delta

            mov = MovimientoInventario(
                tipo='SALIDA',
                producto_id=det.producto_id,
                cantidad=delta,
                almacen_origen_id=almacen_id,
                motivo=motivo_base,
                usuario_id=user.id,
            )
            db.session.add(mov)

        # Refrescar cache stock_actual por cada producto tocado.
        for producto in productos_locked.values():
            _recalcular_cache_stock(producto)

        # Aplicar deltas por línea HERRAMIENTA: crear N asignaciones al
        # trabajador del solicitante, marcar unidades como ASIGNADA y registrar
        # evento de cambio de estado.
        asignaciones_creadas = 0
        for det, delta_int, baseline_h in entregas_herramienta:
            if int(det.cantidad_aprobada or 0) == 0 and baseline_h > 0:
                det.cantidad_aprobada = baseline_h

            # Consumimos del pool de unidades reservado arriba para esta herramienta.
            pool = unidades_por_herramienta[det.herramienta_id]
            for _ in range(delta_int):
                unidad = pool.pop(0)
                estado_anterior = unidad.estado
                asig = AsignacionHerramienta(
                    unidad_id=unidad.id,
                    trabajador_id=trab_solicitante.id,
                    solicitud_id=sol.id,
                    proyecto=(sol.proyecto or None),
                    fecha_entrega=datetime.datetime.utcnow(),
                    fecha_devolucion_prevista=fecha_dev_prevista,
                    estado='ACTIVA',
                    condicion_entrega='BUENA',
                    observaciones_entrega=motivo_base,
                    entregado_por_id=user.id,
                )
                db.session.add(asig)
                db.session.flush()  # para tener asig.id en el evento

                unidad.estado = 'ASIGNADA'
                unidad.asignado_trabajador_id = trab_solicitante.id
                crear_evento_herramienta(
                    unidad, 'ASIGNACION', user,
                    observaciones=(
                        f'Entrega solicitud #{sol_id} → '
                        f'{trab_solicitante.nombre_completo}'
                    ),
                    estado_anterior=estado_anterior, estado_nuevo='ASIGNADA',
                    referencia_id=asig.id, referencia_tipo='asignacion',
                )
                asignaciones_creadas += 1

            det.cantidad_entregada = int(det.cantidad_entregada or 0) + delta_int

        # ¿Quedó completa? Ahora considera AMBOS tipos: cada línea debe tener
        # entregada ≥ aprobada (cuando aprobada > 0).
        completa = True
        for d in (sol.detalles or []):
            aprob = Decimal(str(d.cantidad_aprobada or 0))
            ent = Decimal(str(d.cantidad_entregada or 0))
            if aprob > 0 and ent < aprob:
                completa = False
                break

        if completa:
            sol.estatus = 'ENTREGADA'
            sol.fecha_cierre = datetime.datetime.now()
            # Por seguridad liberamos cualquier reserva sobrante (debería ser 0).
            restos = _reservas_de_solicitud(sol)
            if restos:
                _liberar_reservas(restos)

        total_lineas = len(entregas_material) + len(entregas_herramienta)
        _audit(
            user,
            (
                f"Solicitud #{sol_id} {'ENTREGADA' if completa else 'entrega parcial'} "
                f"({total_lineas} líneas: {len(entregas_material)} mat, "
                f"{asignaciones_creadas} herr asignadas"
                f"{f', almacén #{almacen_id}' if almacen_id else ''})"
            ),
        )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        if 'could not obtain lock' in str(exc).lower():
            return jsonify({'detail': 'Stock bloqueado por otra operación, reintenta'}), 409
        raise

    db.session.refresh(sol)
    _ = list(sol.detalles)
    emit_to_role(_SOL_ROLES, 'solicitud:changed', {
        'id': sol.id, 'action': 'entregada' if sol.estatus == 'ENTREGADA' else 'entrega_parcial',
    })
    # Una entrega real genera SALIDAs de material → notificar también que
    # cambió stock para que el catalogo/bajo-minimo refresquen.
    if entregas_material:
        emit_to_role(_INV_ROLES, 'movimiento:changed', {
            'origen': 'solicitud_entrega', 'solicitud_id': sol.id,
        })
    return jsonify(_solicitud_to_dict(sol))


# ─── Proyectos ────────────────────────────────────────────────────────────────

@bp.route('/proyectos/', methods=['GET'])
@_require_login
def get_proyectos():
    proyectos = (
        Proyecto.query
        .filter(Proyecto.activo == True)
        .order_by(Proyecto.numero_proyecto)
        .all()
    )
    return jsonify([
        {'id': p.id, 'numero_proyecto': p.numero_proyecto, 'nombre': p.nombre or ''}
        for p in proyectos
    ])


# ─── Categorías ───────────────────────────────────────────────────────────────

@bp.route('/categorias/', methods=['GET'])
@_require_inventario
def get_categorias():
    """Devuelve la unión de categorías presentes en el catálogo de productos
    y las registradas en `categorias_config` (admin pudo crear categorías sin
    haber capturado aún ningún producto)."""
    prod_rows = (
        db.session.query(sql_distinct(Producto.categoria))
        .filter(Producto.activo == True, Producto.categoria != None, Producto.categoria != '')
        .all()
    )
    cfg_rows = db.session.query(CategoriaConfig.nombre).all()
    nombres = {r[0] for r in prod_rows} | {r[0] for r in cfg_rows}
    return jsonify(sorted(nombres))


# ─── CategoriaConfig (metadatos visuales por categoría) ──────────────────────

def _categoria_config_to_dict(c: CategoriaConfig) -> dict:
    return {
        'nombre': c.nombre,
        'imagen_url': c.imagen_url,
        'updated_at': c.updated_at.isoformat() if c.updated_at else None,
    }


@bp.route('/categorias-config/', methods=['GET'])
@_require_login
def get_categorias_config():
    """Lista todas las configuraciones (imagen, etc.) por nombre de categoría.
    Lectura abierta a cualquier usuario autenticado: el dashboard de inventario
    también lo consume desde el rol solicitante_material."""
    rows = CategoriaConfig.query.order_by(CategoriaConfig.nombre).all()
    return jsonify([_categoria_config_to_dict(c) for c in rows])


@bp.route('/categorias-config/<string:nombre>', methods=['PUT'])
@_require_inventario_admin
def upsert_categoria_config(nombre: str):
    """Crea o actualiza la config de la categoría con `nombre`. Si imagen_url
    viene null o vacío, persiste null (UI lo trata como "quitar imagen")."""
    nombre = (nombre or '').strip()
    if not nombre or len(nombre) > 100:
        return jsonify({'detail': "Nombre de categoría inválido (1..100 caracteres)"}), 422

    data, err = _parse_or_422(CategoriaConfigUpsertSchema(), request.get_json(silent=True))
    if err: return err

    imagen = (data.get('imagen_url') or '').strip() or None

    cfg = CategoriaConfig.query.filter(CategoriaConfig.nombre == nombre).first()
    if cfg is None:
        cfg = CategoriaConfig(
            nombre=nombre,
            imagen_url=imagen,
            created_by_id=request.current_user.id,
        )
        db.session.add(cfg)
        _audit(request.current_user, f"Categoría '{nombre}' creada/actualizada")
    else:
        cfg.imagen_url = imagen
        _audit(request.current_user, f"Categoría '{nombre}' actualizada")

    db.session.commit()
    db.session.refresh(cfg)
    return jsonify(_categoria_config_to_dict(cfg))


@bp.route('/categorias-config/<string:nombre>', methods=['DELETE'])
@_require_inventario_admin
def delete_categoria_config(nombre: str):
    """Elimina la fila de config (no afecta productos, solo limpia los metadatos visuales)."""
    cfg = CategoriaConfig.query.filter(CategoriaConfig.nombre == nombre).first()
    if not cfg:
        return jsonify({'detail': 'Categoría no encontrada en config'}), 404
    db.session.delete(cfg)
    _audit(request.current_user, f"Categoría '{nombre}' config eliminada")
    db.session.commit()
    return Response(status=204)


# ─── Importar materiales desde Excel ─────────────────────────────────────────

# Encabezados oficiales de la plantilla. El importador acepta también una
# variante en lowercase/sin tildes para no romper si el usuario edita los headers.
PLANTILLA_HEADERS = [
    'Código (SKU)', 'Descripción', 'Categoría', 'Unidad',
    'Stock Inicial', 'Stock Mínimo', 'URL Imagen (opcional)',
]
PLANTILLA_MAX_FILAS = 5000  # Tope para evitar DoS por archivo gigante


def _norm_header(s: str) -> str:
    """Normaliza un header para comparación: lower, sin acentos, sin espacios extra."""
    import unicodedata
    s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode('ascii')
    return s.strip().lower().replace('  ', ' ')


def _cell_str(value, maxlen=None) -> str:
    """Convierte valor de pandas a str limpio (NaN→'', strip). Trunca si maxlen."""
    import math
    if value is None:
        return ''
    try:
        if isinstance(value, float) and math.isnan(value):
            return ''
    except Exception:
        pass
    s = str(value).strip()
    if s.lower() == 'nan':
        return ''
    if maxlen and len(s) > maxlen:
        s = s[:maxlen]
    return s


def _norm_categoria(s: str) -> str:
    """Clave de comparación case/acento-insensitiva para nombres de categoría.
    Sirve solo como índice: el nombre original (con tildes y mayúsculas) se
    conserva como 'nombre canónico' en la primera ocurrencia."""
    import unicodedata
    s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode('ascii')
    return ' '.join(s.strip().lower().split())


def _cell_number(value, default=0.0):
    """Convierte celda a float. Acepta '1,234.56' (formato MX), devuelve (valor, error_str_o_None)."""
    import math
    if value is None or value == '':
        return default, None
    try:
        if isinstance(value, float) and math.isnan(value):
            return default, None
    except Exception:
        pass
    if isinstance(value, (int, float)):
        return float(value), None
    s = str(value).strip().replace(',', '').replace('$', '').replace(' ', '')
    if not s:
        return default, None
    try:
        return float(s), None
    except ValueError:
        return default, f'no es un número válido ({value!r})'


@bp.route('/productos/plantilla-importar', methods=['GET'])
@_require_inventario
def get_plantilla_materiales():
    """Genera y sirve un Excel de plantilla para carga masiva de productos.
    Incluye instrucciones, validaciones de celda (dropdown unidad, números) y
    columna opcional de URL de imagen."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.comments import Comment
    except ImportError:
        return jsonify({'detail': 'openpyxl no instalado en el servidor'}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Materiales'

    # Estilos
    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    title_font = Font(bold=True, color='111827', size=14)
    instr_font = Font(italic=True, color='6B7280', size=10)
    thin = Side(border_style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Fila 1: título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(PLANTILLA_HEADERS))
    c = ws.cell(row=1, column=1, value='Plantilla de Importación de Materiales — SKILLED')
    c.font = title_font
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Fila 2: instrucciones
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(PLANTILLA_HEADERS))
    c = ws.cell(row=2, column=1,
        value='⚠ No alteres los encabezados de la fila 4. URL Imagen es opcional y solo acepta HTTPS o ruta /static/...png. Los duplicados de SKU se ignoran.')
    c.font = instr_font
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Fila 3 vacía como separador
    ws.row_dimensions[3].height = 6

    # Fila 4: encabezados oficiales
    for col, header in enumerate(PLANTILLA_HEADERS, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        # Anchos diferenciados
        widths = {1: 18, 2: 36, 3: 18, 4: 10, 5: 14, 6: 14, 7: 42}
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 18)

    # Comentarios (tooltips) en cada header
    tooltips = {
        1: 'Código único del producto (SKU). Acepta letras, números, -_./',
        2: 'Descripción corta del producto (máx 250 caracteres)',
        3: 'Categoría (ej: Tornillería, Eléctrico, Pinturas)',
        4: 'Unidad de medida: Pza, Kg, Mts, Lts, Caja, Bote, etc.',
        5: 'Cantidad inicial en almacén (número >= 0)',
        6: 'Cantidad mínima antes de alertar de bajo stock (número >= 0)',
        7: 'OPCIONAL — URL HTTPS de la imagen del producto. Ej: https://cdn.miempresa.com/tornillo.jpg',
    }
    for col, txt in tooltips.items():
        ws.cell(row=4, column=col).comment = Comment(txt, 'Plantilla SKILLED')

    ws.row_dimensions[4].height = 36
    ws.freeze_panes = 'A5'

    # Sin filas de ejemplo: la plantilla viene vacía para que el usuario
    # solo capture sus productos reales (no se importen los ejemplos por error).
    # Las instrucciones en filas 1-2 + los tooltips en los headers ya muestran
    # el formato esperado.

    # Data validations (solo aplican a filas 5..1004 para mantener archivo ligero)
    rango = '5:1004'

    # Stock inicial y mínimo: número >= 0
    num_dv = DataValidation(type='decimal', operator='greaterThanOrEqual', formula1=0,
                             showErrorMessage=True,
                             errorTitle='Stock inválido',
                             error='El stock debe ser un número mayor o igual a 0.')
    num_dv.add(f'E5:E1004')
    num_dv.add(f'F5:F1004')
    ws.add_data_validation(num_dv)

    # Longitudes
    desc_dv = DataValidation(type='textLength', operator='between', formula1=1, formula2=250,
                              showErrorMessage=True,
                              errorTitle='Descripción inválida',
                              error='Entre 1 y 250 caracteres.')
    desc_dv.add(f'B5:B1004')
    ws.add_data_validation(desc_dv)

    sku_dv = DataValidation(type='textLength', operator='between', formula1=1, formula2=50,
                              showErrorMessage=True,
                              errorTitle='SKU inválido',
                              error='Entre 1 y 50 caracteres.')
    sku_dv.add(f'A5:A1004')
    ws.add_data_validation(sku_dv)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='plantilla_materiales.xlsx',
    )


@bp.route('/productos/importar', methods=['POST'])
@_require_inventario_admin
@limiter.limit('5 per minute')
def importar_materiales():
    """Importa productos en masa desde un archivo Excel.

    Validaciones a prueba de tontos:
      - Tamaño max 5 MB.
      - Extensión .xlsx/.xls.
      - Headers tolerantes (acepta variantes sin tildes/lowercase, encuentra la
        fila de headers automáticamente).
      - Trim, normalización y truncado por seguridad.
      - SKU debe pasar CODIGO_REGEX (mismo que crear manual).
      - Descripción, categoría, unidad: validación de longitud.
      - Stocks: parseo tolerante ('1,234.56', '$50', etc.) con error claro.
      - URL imagen pasa por _IMAGEN_URL_REGEX (anti-XSS/SSRF).
      - Duplicados dentro del mismo archivo y contra DB.
      - Filas vacías se ignoran silenciosamente.
      - Tope de PLANTILLA_MAX_FILAS para evitar DoS.
    """
    try:
        import pandas as pd
    except ImportError:
        return jsonify({'detail': 'pandas no instalado en el servidor'}), 500

    file = request.files.get('archivo') or request.files.get('archivo_excel')
    if not file or not file.filename:
        return jsonify({'detail': 'No se envió archivo'}), 400
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'detail': 'Formato no válido. Debe ser .xlsx o .xls'}), 400

    # Validar tamaño antes de cargar a pandas (evita DoS).
    file.stream.seek(0, io.SEEK_END)
    size = file.stream.tell()
    file.stream.seek(0)
    MAX_BYTES = 5 * 1024 * 1024
    if size > MAX_BYTES:
        return jsonify({'detail': f'Archivo demasiado grande. Máximo {MAX_BYTES // (1024*1024)} MB.'}), 413
    if size < 100:
        return jsonify({'detail': 'Archivo vacío o corrupto.'}), 400

    # Buscar la fila de encabezados (acepta hasta fila 10 — la plantilla los pone en fila 4).
    # Si los encabezados no se encuentran, fallamos con mensaje claro.
    try:
        raw = pd.read_excel(file, header=None, nrows=PLANTILLA_MAX_FILAS + 20)
    except Exception as e:
        return jsonify({
            'detail': 'No se pudo leer el Excel. Asegúrate de usar la plantilla y guardar como .xlsx.',
            'tecnico': str(e)[:200],
        }), 400

    expected_norm = {_norm_header(h): h for h in PLANTILLA_HEADERS}
    header_row_idx = None
    column_map = {}  # idx_columna_excel -> nombre_oficial
    for ridx in range(min(10, len(raw))):
        row_vals = [_norm_header(v) for v in raw.iloc[ridx].tolist()]
        matches = {expected_norm[h]: i for i, h in enumerate(row_vals) if h in expected_norm}
        # Necesitamos al menos los 4 obligatorios para considerar esa fila como header
        obligatorios = {'Código (SKU)', 'Descripción', 'Categoría', 'Unidad'}
        if obligatorios.issubset(matches.keys()):
            header_row_idx = ridx
            column_map = {matches[h]: h for h in matches}
            break

    if header_row_idx is None:
        return jsonify({
            'detail': 'No se encontraron los encabezados esperados. '
                      'Descarga la plantilla nueva (debe incluir Código (SKU), Descripción, Categoría, Unidad).',
        }), 400

    # Datos = todo lo que está debajo del header
    data_df = raw.iloc[header_row_idx + 1:].reset_index(drop=True)
    if len(data_df) > PLANTILLA_MAX_FILAS:
        return jsonify({
            'detail': f'Demasiadas filas ({len(data_df)}). Máximo {PLANTILLA_MAX_FILAS} por importación.',
        }), 400

    user = request.current_user
    exitosos = 0
    errores = []
    skus_en_archivo = set()  # detectar duplicados intra-archivo

    import re
    codigo_re = re.compile(CODIGO_REGEX)
    imagen_re = re.compile(_IMAGEN_URL_REGEX)

    # Pausa 2: bodega default donde caerá el stock inicial de productos nuevos.
    # Si no hay ninguna bodega activa, el stock se importa pero queda
    # huérfano (cache lleno, stock_por_almacen vacío) hasta que se cree una.
    bodega_default_id = _almacen_default_id()

    # Cache de categorías existentes para resolver case/acento-insensitivamente.
    # Unimos Producto.categoria (catálogo real) + CategoriaConfig (metadatos
    # visuales). El usuario puede capturar "tornilleria" en el Excel y, si ya
    # existe "Tornillería", reutilizamos el nombre canónico en vez de crear un
    # duplicado por dedazo. Cargamos una sola vez (evita N+1 consultas).
    cat_canon: dict[str, str] = {}  # clave normalizada -> nombre canónico
    for (nombre,) in db.session.query(sql_distinct(Producto.categoria)).filter(
        Producto.categoria != None, Producto.categoria != ''
    ).all():
        if nombre:
            cat_canon.setdefault(_norm_categoria(nombre), nombre)
    for (nombre,) in db.session.query(CategoriaConfig.nombre).all():
        if nombre:
            cat_canon.setdefault(_norm_categoria(nombre), nombre)

    categorias_creadas: list[str] = []  # nombres de las categorías nuevas (para reportar al SPA)

    def _g(row, header_oficial):
        """Lee la celda por header oficial, no por posición."""
        for col_idx, oficial in column_map.items():
            if oficial == header_oficial:
                return row.iloc[col_idx] if col_idx < len(row) else None
        return None

    for offset, row in data_df.iterrows():
        fila_excel = header_row_idx + offset + 2  # +1 por header + 1 porque Excel es 1-indexed
        try:
            codigo      = _cell_str(_g(row, 'Código (SKU)'), maxlen=50)
            descripcion = _cell_str(_g(row, 'Descripción'), maxlen=250)
            categoria   = _cell_str(_g(row, 'Categoría'), maxlen=100)
            unidad      = _cell_str(_g(row, 'Unidad'), maxlen=50)
            imagen_url  = _cell_str(_g(row, 'URL Imagen (opcional)'), maxlen=500)

            # Fila completamente vacía: ignorar sin reportar
            if not (codigo or descripcion or categoria or unidad):
                continue

            problemas = []
            if not codigo:
                problemas.append('falta SKU')
            elif not codigo_re.match(codigo):
                problemas.append('SKU contiene caracteres no permitidos (usa A-Z 0-9 - _ . /)')
            if not descripcion:
                problemas.append('falta descripción')
            if not categoria:
                problemas.append('falta categoría')
            if not unidad:
                unidad = 'Pza'  # default suave

            stock_actual, err_si = _cell_number(_g(row, 'Stock Inicial'), default=0.0)
            if err_si:
                problemas.append(f'stock inicial {err_si}')
            elif stock_actual < 0:
                problemas.append('stock inicial debe ser >= 0')
            elif stock_actual > 1_000_000:
                problemas.append('stock inicial fuera de rango (máx 1M)')

            stock_minimo, err_sm = _cell_number(_g(row, 'Stock Mínimo'), default=0.0)
            if err_sm:
                problemas.append(f'stock mínimo {err_sm}')
            elif stock_minimo < 0:
                problemas.append('stock mínimo debe ser >= 0')
            elif stock_minimo > 1_000_000:
                problemas.append('stock mínimo fuera de rango (máx 1M)')

            # Validar URL imagen (opcional). Si viene, debe ser HTTPS o /path.png.
            imagen_final = None
            if imagen_url:
                if not imagen_re.match(imagen_url):
                    problemas.append('URL imagen inválida (solo HTTPS o /static/...png)')
                else:
                    imagen_final = imagen_url

            if problemas:
                errores.append(f'Fila {fila_excel}: ' + '; '.join(problemas))
                continue

            # Duplicado intra-archivo
            sku_lower = codigo.lower()
            if sku_lower in skus_en_archivo:
                errores.append(f'Fila {fila_excel}: SKU "{codigo}" duplicado en este archivo')
                continue
            skus_en_archivo.add(sku_lower)

            # Duplicado contra DB
            if Producto.query.filter(Producto.codigo == codigo).first():
                errores.append(f'Fila {fila_excel}: SKU "{codigo}" ya existe en el catálogo')
                continue

            # Resolver categoría a prueba de tontos:
            #  - Si ya existe una equivalente (case/acento-insensitiva), usar el
            #    nombre canónico para no romper el agrupado del dashboard.
            #  - Si es nueva, registrarla en CategoriaConfig (sin imagen) y
            #    agregarla al cache para que el resto del archivo la reutilice.
            cat_key = _norm_categoria(categoria)
            categoria_canonica = cat_canon.get(cat_key)
            if categoria_canonica is None:
                categoria_canonica = categoria  # primera ocurrencia → este es el canónico
                cat_canon[cat_key] = categoria_canonica
                db.session.add(CategoriaConfig(
                    nombre=categoria_canonica,
                    imagen_url=None,
                    created_by_id=user.id,
                ))
                categorias_creadas.append(categoria_canonica)

            stock_inicial_dec = Decimal(str(stock_actual))
            nuevo = Producto(
                codigo=codigo,
                descripcion=descripcion,
                categoria=categoria_canonica,
                unidad=unidad,
                stock_actual=stock_inicial_dec,
                stock_minimo=Decimal(str(stock_minimo)),
                imagen_url=imagen_final,
                created_by_id=user.id,
            )
            db.session.add(nuevo)
            db.session.flush()  # asegura nuevo.id antes de crear StockPorAlmacen

            # Pausa 2: depositar stock inicial en bodega default si hay y >0.
            if stock_inicial_dec > 0 and bodega_default_id:
                db.session.add(StockPorAlmacen(
                    producto_id=nuevo.id,
                    almacen_id=bodega_default_id,
                    cantidad=stock_inicial_dec,
                ))

            exitosos += 1

        except Exception as e:
            errores.append(f'Fila {fila_excel}: error inesperado — {str(e)[:80]}')

    try:
        msg = f'Importación masiva: {exitosos} productos creados, {len(errores)} errores'
        if categorias_creadas:
            msg += f', {len(categorias_creadas)} categorías nuevas'
        _audit(user, msg)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error commit importar_materiales')
        return jsonify({'detail': f'Error al guardar en base de datos: {str(e)[:100]}'}), 500

    return jsonify({
        'exitosos': exitosos,
        'errores': errores,
        'total_procesadas': exitosos + len(errores),
        'categorias_creadas': categorias_creadas,
    })


# ─── PDF de solicitudes ──────────────────────────────────────────────────────

def _render_solicitud_pdf(*, folio, fecha_str, solicitante, proyecto, notas, materiales, herramientas):
    """Helper común: ya recibe los dicts normalizados y devuelve BytesIO con el PDF."""
    try:
        from xhtml2pdf import pisa
    except ImportError:
        return None
    import os as _os
    logo_path = _os.path.join(current_app.static_folder, 'imagenes', 'skilled (1).png')
    html_salida = render_template(
        'solicitud_pedido_pdf.html',
        folio=folio,
        fecha=fecha_str,
        solicitante=solicitante,
        proyecto=proyecto,
        notas=notas,
        materiales=materiales,
        herramientas=herramientas,
        logo_path=logo_path if _os.path.exists(logo_path) else None,
    )
    buf = io.BytesIO()
    status = pisa.CreatePDF(io.BytesIO(html_salida.encode('utf-8')), dest=buf)
    if status.err:
        return None
    buf.seek(0)
    return buf


@bp.route('/solicitudes/<int:sol_id>/pdf', methods=['GET'])
@_require_login
def imprimir_solicitud(sol_id: int):
    """Genera el PDF de una solicitud ya guardada. Solicitante solo puede
    imprimir las suyas; inventario/admin/super_admin pueden imprimir todas."""
    user = request.current_user
    sol = (
        SolicitudMaterial.query
        .options(
            joinedload(SolicitudMaterial.solicitante),
            selectinload(SolicitudMaterial.detalles).joinedload(SolicitudMaterialDetalle.producto),
            selectinload(SolicitudMaterial.detalles).joinedload(SolicitudMaterialDetalle.herramienta),
        )
        .filter(SolicitudMaterial.id == sol_id).first()
    )
    if not sol:
        return jsonify({'detail': 'Solicitud no encontrada'}), 404

    # AuthZ: solicitante y coordinador solo pueden imprimir las suyas.
    if user.role in ('solicitante_material', 'coordinador') and sol.solicitante_id != user.id:
        return jsonify({'detail': 'Forbidden'}), 403
    if user.role not in ('solicitante_material', 'coordinador', 'inventario', 'admin', 'super_admin'):
        return jsonify({'detail': 'Forbidden'}), 403

    materiales, herramientas = [], []
    for d in sol.detalles:
        tipo = (d.tipo_item or 'MATERIAL').upper()
        cantidad = float(d.cantidad_solicitada or 0)
        cantidad = int(cantidad) if cantidad % 1 == 0 else cantidad
        if tipo == 'HERRAMIENTA':
            herramientas.append({
                'descripcion': (d.herramienta.descripcion if d.herramienta else 'Herramienta eliminada')[:250],
                'sku': (d.herramienta.sku if d.herramienta else '---')[:50],
                'cantidad': cantidad,
                'fecha_uso_inicio': d.fecha_uso_inicio.isoformat() if d.fecha_uso_inicio else '',
                'fecha_uso_fin': d.fecha_uso_fin.isoformat() if d.fecha_uso_fin else '',
                'justificacion': (d.justificacion or '')[:2000],
                'complementos': (d.complementos or '')[:500],
            })
        else:
            materiales.append({
                'descripcion': (d.producto.descripcion if d.producto else 'Producto eliminado')[:250],
                'codigo': (d.producto.codigo if d.producto else '---')[:50],
                'categoria': (d.producto.categoria if d.producto else '')[:100],
                'unidad': (d.producto.unidad if d.producto else '')[:50],
                'cantidad': cantidad,
            })

    folio = f'SOL-{sol.id:06d}'
    fecha_str = sol.fecha_creacion.strftime('%d/%m/%Y %H:%M') if sol.fecha_creacion else ''
    solicitante = sol.solicitante.full_name or sol.solicitante.username if sol.solicitante else '—'

    pdf = _render_solicitud_pdf(
        folio=folio, fecha_str=fecha_str, solicitante=solicitante,
        proyecto=sol.proyecto or '', notas='',
        materiales=materiales, herramientas=herramientas,
    )
    if pdf is None:
        return jsonify({'detail': 'Error al generar el PDF'}), 500

    return send_file(pdf, mimetype='application/pdf', as_attachment=False, download_name=f'{folio}.pdf')


# ─── PDF preview de solicitud (sin persistir) ────────────────────────────────

@bp.route('/solicitudes/preview-pdf', methods=['POST'])
@limiter.limit('10/minute', key_func=lambda: f"ip:{get_real_client_ip_flask()}")
@_require_login
def preview_solicitud_pdf():
    """Genera un PDF a partir del carrito actual del usuario, SIN guardar la
    solicitud. Sirve para que el solicitante pueda imprimir/firmar antes de
    enviar al almacén. Mismo mecanismo que prenómina: xhtml2pdf → send_file.
    """
    user = request.current_user
    if user.role not in ('solicitante_material', 'coordinador', 'inventario', 'admin', 'super_admin'):
        return jsonify({'detail': 'Forbidden'}), 403

    payload = request.get_json(silent=True) or {}
    materiales_raw = payload.get('materiales') or []
    herramientas_raw = payload.get('herramientas') or []

    if not materiales_raw and not herramientas_raw:
        return jsonify({'detail': 'Agrega al menos un material o herramienta'}), 422
    if len(materiales_raw) > 200 or len(herramientas_raw) > 200:
        return jsonify({'detail': 'Demasiados ítems en una sola solicitud'}), 422

    # Normalizar / sanitizar (xhtml2pdf escapa automáticamente vía Jinja autoescape)
    def _clean(s, maxlen=500):
        return (str(s or '')[:maxlen]).strip()

    materiales = []
    for m in materiales_raw:
        try:
            cantidad = float(m.get('cantidad') or 0)
        except (TypeError, ValueError):
            cantidad = 0
        materiales.append({
            'descripcion': _clean(m.get('descripcion'), 250),
            'codigo': _clean(m.get('codigo'), 50),
            'categoria': _clean(m.get('categoria'), 100),
            'unidad': _clean(m.get('unidad'), 50),
            'cantidad': cantidad if cantidad % 1 else int(cantidad),
        })

    herramientas = []
    for h in herramientas_raw:
        try:
            cantidad = int(float(h.get('cantidad') or 0))
        except (TypeError, ValueError):
            cantidad = 0
        herramientas.append({
            'descripcion': _clean(h.get('descripcion'), 250),
            'sku': _clean(h.get('sku'), 50),
            'cantidad': cantidad,
            'fecha_uso_inicio': _clean(h.get('fecha_uso_inicio'), 20),
            'fecha_uso_fin': _clean(h.get('fecha_uso_fin'), 20),
            'justificacion': _clean(h.get('justificacion'), 2000),
            'complementos': _clean(h.get('complementos'), 500),
        })

    proyecto = _clean(payload.get('proyecto'), 200)
    notas = _clean(payload.get('notas'), 2000)
    folio = 'SOL-' + datetime.datetime.now().strftime('%y%m%d%H%M%S')
    fecha_str = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
    solicitante = user.full_name or user.username

    pdf = _render_solicitud_pdf(
        folio=folio, fecha_str=fecha_str, solicitante=solicitante,
        proyecto=proyecto, notas=notas,
        materiales=materiales, herramientas=herramientas,
    )
    if pdf is None:
        return jsonify({'detail': 'Error al generar el PDF (xhtml2pdf no disponible)'}), 500

    _audit(user, f"Vista previa PDF solicitud ({len(materiales)} mat, {len(herramientas)} herr)")
    db.session.commit()

    return send_file(pdf, mimetype='application/pdf', as_attachment=False, download_name=f'{folio}.pdf')


# ─── Reportes Excel (Pausa 6) ────────────────────────────────────────────────

REPORTES_MAX_FILAS = 10_000


def _aplicar_estilos_ws(ws, money_cols: set[str] | None = None):
    """Aplica encabezado azul + zebra striping + freeze panes + auto-width.

    `money_cols`: nombres de columna (exactos) que reciben formato '$#,##0.00'.
    El resto de los numéricos quedan sin formato (importante: stocks, días,
    cantidades enteras o decimales sin moneda).
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    money_cols = money_cols or set()

    headers = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
    money_idx = {i + 1 for i, h in enumerate(headers) if h in money_cols}

    if ws.max_row > 0:
        ws.freeze_panes = 'A2'

    for row_idx in range(1, ws.max_row + 1):
        is_header = (row_idx == 1)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if is_header:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
            else:
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill
                if col_idx in money_idx and isinstance(cell.value, (int, float)):
                    cell.number_format = '"$"#,##0.00'

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    n = len(str(cell.value))
                    if n > max_length:
                        max_length = n
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)


def _stream_excel(sheets: dict, filename: str, money_cols: set[str] | None = None):
    """Genera un .xlsx multi-hoja en memoria y lo devuelve con send_file.

    `sheets`: dict[nombre_hoja, list[dict]]. Los keys de la primera fila se
    usan como encabezados (orden estable de Python 3.7+).
    Aplica saneo anti CSV-injection con `safe_excel_value`.
    """
    from openpyxl import Workbook
    from app.utils import safe_excel_value

    wb = Workbook()
    # Workbook trae una hoja "Sheet" por default; la quitamos para que solo
    # queden las hojas que pasamos en `sheets`.
    default = wb.active
    wb.remove(default)

    for raw_name, rows in sheets.items():
        # Excel limita nombre de hoja a 31 chars y prohíbe / \ ? * [ ]
        safe = (raw_name or 'Hoja')[:31]
        for ch in r'/\?*[]:':
            safe = safe.replace(ch, '_')
        ws = wb.create_sheet(title=safe)
        if not rows:
            ws['A1'] = 'Sin datos para los filtros seleccionados'
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([safe_excel_value(row.get(h)) for h in headers])
        _aplicar_estilos_ws(ws, money_cols=money_cols)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


def _parse_fecha_arg(name: str, default: datetime.date | None):
    """Lee un query param YYYY-MM-DD o devuelve `default` si está vacío.
    Devuelve (date, error_response). Si error_response es None, todo bien.
    """
    raw = request.args.get(name)
    if not raw:
        return default, None
    try:
        return datetime.date.fromisoformat(raw), None
    except ValueError:
        return None, (jsonify({'detail': f"Parámetro '{name}' debe ser YYYY-MM-DD"}), 422)


@bp.route('/reportes/inventario-actual.xlsx', methods=['GET'])
@limiter.limit(
    "10/minute",
    key_func=lambda: f"ip:{get_real_client_ip_flask()}",
)
@_require_inventario_admin
def reporte_inventario_actual():
    """Reporte de stock actual de todos los productos activos.

    Query params:
      - categoria: filtra por categoría exacta (opcional).
      - solo_bajo_minimo: 1 → solo productos con stock_actual ≤ stock_minimo.
    """
    q = Producto.query.filter(Producto.activo == True)  # noqa: E712
    categoria = (request.args.get('categoria') or '').strip()
    if categoria:
        q = q.filter(Producto.categoria == categoria)
    if request.args.get('solo_bajo_minimo') in ('1', 'true', 'True'):
        q = q.filter(Producto.stock_actual <= Producto.stock_minimo)
    productos = q.order_by(Producto.categoria, Producto.codigo).limit(REPORTES_MAX_FILAS).all()

    rows = []
    for p in productos:
        actual = float(p.stock_actual or 0)
        reservado = float(p.stock_reservado or 0)
        minimo = float(p.stock_minimo or 0)
        rows.append({
            'Código': p.codigo,
            'Descripción': p.descripcion,
            'Categoría': p.categoria,
            'Unidad': p.unidad,
            'Stock actual': actual,
            'Reservado': reservado,
            'Disponible': actual - reservado,
            'Mínimo': minimo,
            'Diferencia vs mínimo': actual - minimo,
            'Estado': 'BAJO' if actual <= minimo else 'OK',
        })

    _audit(request.current_user, f"Reporte inventario actual ({len(rows)} filas)")
    db.session.commit()

    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    return _stream_excel(
        {'Inventario': rows},
        f'inventario_actual_{ts}.xlsx',
    )


@bp.route('/reportes/movimientos.xlsx', methods=['GET'])
@limiter.limit(
    "10/minute",
    key_func=lambda: f"ip:{get_real_client_ip_flask()}",
)
@_require_inventario_admin
def reporte_movimientos():
    """Reporte de movimientos de inventario con filtros.

    Query params:
      - desde, hasta (YYYY-MM-DD): default últimos 30 días.
      - tipo: ENTRADA / SALIDA / AJUSTE / TRASPASO (opcional).
      - producto_id (opcional).
      - usuario_id (opcional).
    """
    hoy = datetime.date.today()
    desde, err = _parse_fecha_arg('desde', hoy - datetime.timedelta(days=30))
    if err: return err
    hasta, err = _parse_fecha_arg('hasta', hoy)
    if err: return err
    if desde > hasta:
        return jsonify({'detail': "'desde' no puede ser mayor que 'hasta'"}), 422

    tipo = request.args.get('tipo')
    if tipo and tipo not in ('ENTRADA', 'SALIDA', 'AJUSTE', 'TRASPASO'):
        return jsonify({'detail': "Parámetro 'tipo' inválido"}), 422

    prod_id, err = _int_arg('producto_id', 0, 0, 1_000_000_000)
    if err: return err
    usr_id, err = _int_arg('usuario_id', 0, 0, 1_000_000_000)
    if err: return err

    desde_dt = datetime.datetime.combine(desde, datetime.time.min)
    hasta_dt = datetime.datetime.combine(hasta, datetime.time.max)

    q = (
        MovimientoInventario.query
        .options(
            joinedload(MovimientoInventario.producto),
            joinedload(MovimientoInventario.almacen_origen),
            joinedload(MovimientoInventario.almacen_destino),
            joinedload(MovimientoInventario.usuario),
        )
        .filter(MovimientoInventario.fecha >= desde_dt, MovimientoInventario.fecha <= hasta_dt)
    )
    if tipo:
        q = q.filter(MovimientoInventario.tipo == tipo)
    if prod_id:
        q = q.filter(MovimientoInventario.producto_id == prod_id)
    if usr_id:
        q = q.filter(MovimientoInventario.usuario_id == usr_id)

    movs = q.order_by(MovimientoInventario.fecha.desc()).limit(REPORTES_MAX_FILAS).all()

    rows = []
    for m in movs:
        rows.append({
            'Fecha': m.fecha.strftime('%Y-%m-%d %H:%M') if m.fecha else '',
            'Tipo': m.tipo,
            'Código': m.producto.codigo if m.producto else '',
            'Descripción': m.producto.descripcion if m.producto else '',
            'Cantidad': float(m.cantidad or 0),
            'Unidad': m.producto.unidad if m.producto else '',
            'Almacén origen': m.almacen_origen.nombre if m.almacen_origen else '',
            'Almacén destino': m.almacen_destino.nombre if m.almacen_destino else '',
            'Usuario': m.usuario.username if m.usuario else '',
            'Motivo': m.motivo or '',
        })

    _audit(
        request.current_user,
        f"Reporte movimientos {desde} a {hasta} ({len(rows)} filas)",
    )
    db.session.commit()

    return _stream_excel(
        {'Movimientos': rows},
        f'movimientos_{desde}_{hasta}.xlsx',
    )


@bp.route('/reportes/kardex.xlsx', methods=['GET'])
@limiter.limit(
    "10/minute",
    key_func=lambda: f"ip:{get_real_client_ip_flask()}",
)
@_require_inventario_admin
def reporte_kardex_xlsx():
    """Kardex de un producto exportado a Excel.

    Query params:
      - producto_id (requerido).
      - desde, hasta (YYYY-MM-DD): default últimos 30 días.
    """
    prod_id, err = _int_arg('producto_id', 0, 0, 1_000_000_000)
    if err: return err
    if not prod_id:
        return jsonify({'detail': "Parámetro 'producto_id' es requerido"}), 422

    producto = Producto.query.filter(Producto.id == prod_id).first()
    if not producto:
        return jsonify({'detail': 'Producto no encontrado'}), 404

    hoy = datetime.date.today()
    desde, err = _parse_fecha_arg('desde', hoy - datetime.timedelta(days=30))
    if err: return err
    hasta, err = _parse_fecha_arg('hasta', hoy)
    if err: return err
    if desde > hasta:
        return jsonify({'detail': "'desde' no puede ser mayor que 'hasta'"}), 422

    # Reutiliza la misma fórmula de saldo corrido del endpoint JSON.
    def _delta(m: MovimientoInventario) -> Decimal:
        cant = m.cantidad or Decimal('0')
        if m.tipo == 'ENTRADA':
            return cant
        if m.tipo == 'SALIDA':
            return -cant
        if m.tipo == 'AJUSTE':
            return cant
        return Decimal('0')  # TRASPASO no altera total

    desde_dt = datetime.datetime.combine(desde, datetime.time.min)
    hasta_dt = datetime.datetime.combine(hasta, datetime.time.max)

    # Saldo inicial = stock_actual − Σ deltas posteriores a 'desde'.
    movs_post = (
        MovimientoInventario.query
        .filter(
            MovimientoInventario.producto_id == prod_id,
            MovimientoInventario.fecha >= desde_dt,
        )
        .all()
    )
    delta_post = sum((_delta(m) for m in movs_post), Decimal('0'))
    saldo_inicial = (producto.stock_actual or Decimal('0')) - delta_post

    movs = (
        MovimientoInventario.query
        .options(
            joinedload(MovimientoInventario.usuario),
            joinedload(MovimientoInventario.almacen_origen),
            joinedload(MovimientoInventario.almacen_destino),
        )
        .filter(
            MovimientoInventario.producto_id == prod_id,
            MovimientoInventario.fecha >= desde_dt,
            MovimientoInventario.fecha <= hasta_dt,
        )
        .order_by(MovimientoInventario.fecha.asc(), MovimientoInventario.id.asc())
        .limit(REPORTES_MAX_FILAS)
        .all()
    )

    rows = [{
        'Fecha': '',
        'Tipo': '— Saldo inicial —',
        'Cantidad': '',
        'Delta': '',
        'Saldo': float(saldo_inicial),
        'Almacén origen': '',
        'Almacén destino': '',
        'Usuario': '',
        'Motivo': '',
    }]
    saldo = saldo_inicial
    for m in movs:
        d = _delta(m)
        saldo = saldo + d
        rows.append({
            'Fecha': m.fecha.strftime('%Y-%m-%d %H:%M') if m.fecha else '',
            'Tipo': m.tipo,
            'Cantidad': float(m.cantidad or 0),
            'Delta': float(d),
            'Saldo': float(saldo),
            'Almacén origen': m.almacen_origen.nombre if m.almacen_origen else '',
            'Almacén destino': m.almacen_destino.nombre if m.almacen_destino else '',
            'Usuario': m.usuario.username if m.usuario else '',
            'Motivo': m.motivo or '',
        })

    info = [
        {'Campo': 'Producto', 'Valor': f"{producto.codigo} — {producto.descripcion}"},
        {'Campo': 'Unidad', 'Valor': producto.unidad},
        {'Campo': 'Stock actual', 'Valor': float(producto.stock_actual or 0)},
        {'Campo': 'Periodo', 'Valor': f"{desde} a {hasta}"},
        {'Campo': 'Movimientos', 'Valor': len(movs)},
    ]

    _audit(request.current_user, f"Reporte kardex prod #{prod_id} {desde} a {hasta}")
    db.session.commit()

    return _stream_excel(
        {'Resumen': info, 'Kardex': rows},
        f'kardex_{producto.codigo}_{desde}_{hasta}.xlsx',
    )


@bp.route('/reportes/consumo-proyecto.xlsx', methods=['GET'])
@limiter.limit(
    "10/minute",
    key_func=lambda: f"ip:{get_real_client_ip_flask()}",
)
@_require_inventario_admin
def reporte_consumo_proyecto():
    """Consumo por proyecto agrupado, basado en solicitudes entregadas o con
    entrega parcial dentro del rango.

    Query params:
      - desde, hasta (YYYY-MM-DD): default últimos 90 días.
      - estatus: por defecto ['ENTREGADA','APROBADA'] (incluye parciales).
    """
    hoy = datetime.date.today()
    desde, err = _parse_fecha_arg('desde', hoy - datetime.timedelta(days=90))
    if err: return err
    hasta, err = _parse_fecha_arg('hasta', hoy)
    if err: return err
    if desde > hasta:
        return jsonify({'detail': "'desde' no puede ser mayor que 'hasta'"}), 422

    desde_dt = datetime.datetime.combine(desde, datetime.time.min)
    hasta_dt = datetime.datetime.combine(hasta, datetime.time.max)

    estatus_arg = (request.args.get('estatus') or '').strip()
    if estatus_arg:
        estatus_list = [e.strip().upper() for e in estatus_arg.split(',') if e.strip()]
        if any(e not in ('PENDIENTE', 'APROBADA', 'RECHAZADA', 'ENTREGADA') for e in estatus_list):
            return jsonify({'detail': "Parámetro 'estatus' inválido"}), 422
    else:
        estatus_list = ['ENTREGADA', 'APROBADA']

    sols = (
        SolicitudMaterial.query
        .options(
            selectinload(SolicitudMaterial.detalles).joinedload(SolicitudMaterialDetalle.producto),
        )
        .filter(
            SolicitudMaterial.fecha_creacion >= desde_dt,
            SolicitudMaterial.fecha_creacion <= hasta_dt,
            SolicitudMaterial.estatus.in_(estatus_list),
        )
        .all()
    )

    # Agregamos por (proyecto, código). Usamos cantidad_entregada (real); para
    # ENTREGADAs pre-8b sin cantidad_entregada, caemos a cantidad_solicitada.
    agg: dict[tuple, dict] = {}
    for s in sols:
        proyecto = (s.proyecto or 'Sin proyecto').strip() or 'Sin proyecto'
        for d in (s.detalles or []):
            if (d.tipo_item or 'MATERIAL').upper() != 'MATERIAL' or not d.producto_id:
                continue
            cant_ent = Decimal(str(d.cantidad_entregada or 0))
            if cant_ent <= 0 and s.estatus == 'ENTREGADA':
                # Legacy: solicitudes pre-8b marcadas como ENTREGADA sin descontar.
                cant_ent = Decimal(str(d.cantidad_solicitada or 0))
            if cant_ent <= 0:
                continue
            codigo = d.producto.codigo if d.producto else f'(#{d.producto_id})'
            descripcion = d.producto.descripcion if d.producto else 'Producto eliminado'
            unidad = d.producto.unidad if d.producto else ''
            key = (proyecto, codigo)
            if key not in agg:
                agg[key] = {
                    'Proyecto': proyecto,
                    'Código': codigo,
                    'Descripción': descripcion,
                    'Unidad': unidad,
                    'Cantidad entregada': 0.0,
                    'Solicitudes': set(),
                }
            agg[key]['Cantidad entregada'] += float(cant_ent)
            agg[key]['Solicitudes'].add(s.id)

    rows = []
    for v in sorted(agg.values(), key=lambda r: (r['Proyecto'], r['Código'])):
        rows.append({
            'Proyecto': v['Proyecto'],
            'Código': v['Código'],
            'Descripción': v['Descripción'],
            'Unidad': v['Unidad'],
            'Cantidad entregada': round(v['Cantidad entregada'], 4),
            '# Solicitudes': len(v['Solicitudes']),
        })

    _audit(
        request.current_user,
        f"Reporte consumo por proyecto {desde} a {hasta} ({len(rows)} filas)",
    )
    db.session.commit()

    return _stream_excel(
        {'Consumo por proyecto': rows},
        f'consumo_proyecto_{desde}_{hasta}.xlsx',
    )


@bp.route('/reportes/solicitudes.xlsx', methods=['GET'])
@limiter.limit(
    "10/minute",
    key_func=lambda: f"ip:{get_real_client_ip_flask()}",
)
@_require_inventario_admin
def reporte_solicitudes():
    """Reporte de solicitudes del periodo con totales por solicitud.

    Query params:
      - desde, hasta (YYYY-MM-DD): default últimos 30 días.
      - estatus: PENDIENTE / APROBADA / RECHAZADA / ENTREGADA (opcional).
    """
    hoy = datetime.date.today()
    desde, err = _parse_fecha_arg('desde', hoy - datetime.timedelta(days=30))
    if err: return err
    hasta, err = _parse_fecha_arg('hasta', hoy)
    if err: return err
    if desde > hasta:
        return jsonify({'detail': "'desde' no puede ser mayor que 'hasta'"}), 422

    estatus = (request.args.get('estatus') or '').strip().upper()
    if estatus and estatus not in ('PENDIENTE', 'APROBADA', 'RECHAZADA', 'ENTREGADA'):
        return jsonify({'detail': "Parámetro 'estatus' inválido"}), 422

    desde_dt = datetime.datetime.combine(desde, datetime.time.min)
    hasta_dt = datetime.datetime.combine(hasta, datetime.time.max)

    q = (
        SolicitudMaterial.query
        .options(
            joinedload(SolicitudMaterial.solicitante),
            selectinload(SolicitudMaterial.detalles),
        )
        .filter(
            SolicitudMaterial.fecha_creacion >= desde_dt,
            SolicitudMaterial.fecha_creacion <= hasta_dt,
        )
    )
    if estatus:
        q = q.filter(SolicitudMaterial.estatus == estatus)
    sols = q.order_by(SolicitudMaterial.fecha_creacion.desc()).limit(REPORTES_MAX_FILAS).all()

    rows = []
    for s in sols:
        total_sol = sum(float(d.cantidad_solicitada or 0) for d in (s.detalles or []))
        total_aprob = sum(float(d.cantidad_aprobada or 0) for d in (s.detalles or []))
        total_ent = sum(float(d.cantidad_entregada or 0) for d in (s.detalles or []))
        rows.append({
            'ID': s.id,
            'Fecha': s.fecha_creacion.strftime('%Y-%m-%d %H:%M') if s.fecha_creacion else '',
            'Cierre': s.fecha_cierre.strftime('%Y-%m-%d %H:%M') if s.fecha_cierre else '',
            'Estatus': s.estatus,
            'Solicitante': s.solicitante.username if s.solicitante else '',
            'Proyecto': s.proyecto or '',
            'Líneas': len(s.detalles or []),
            'Total solicitado': round(total_sol, 4),
            'Total aprobado': round(total_aprob, 4),
            'Total entregado': round(total_ent, 4),
        })

    _audit(
        request.current_user,
        f"Reporte solicitudes {desde} a {hasta} ({len(rows)} filas)",
    )
    db.session.commit()

    return _stream_excel(
        {'Solicitudes': rows},
        f'solicitudes_{desde}_{hasta}.xlsx',
    )


# ─── Etiquetas imprimibles (Pausa 8a) ────────────────────────────────────────

class _EtiquetaItemSchema(_BaseSchema):
    producto_id = fields.Int(required=True)
    cantidad = fields.Int(required=True, validate=validate.Range(min=1, max=500))


class EtiquetasPdfSchema(_BaseSchema):
    formato = fields.Str(load_default='avery_5160',
                          validate=validate.OneOf(['avery_5160', 'avery_5163']))
    tipo = fields.Str(load_default='barcode',
                       validate=validate.OneOf(['barcode', 'qr']))
    items = fields.List(
        fields.Nested(_EtiquetaItemSchema),
        required=True,
        validate=validate.Length(min=1, max=200),
    )


# Especificaciones físicas de las hojas Avery más comunes (página Letter 8.5"×11").
# Medidas en pulgadas; convertimos a puntos con `inch` al usar reportlab.
ETIQUETA_FORMATOS = {
    # 30 etiquetas/hoja (3 columnas × 10 filas), 2.625" × 1".
    'avery_5160': {
        'cols': 3, 'rows': 10,
        'label_w': 2.625, 'label_h': 1.0,
        'col_gap': 0.125, 'row_gap': 0.0,
        'top_margin': 0.5, 'left_margin': 0.1875,
        'descripcion': 'Avery 5160 — 30 etiquetas/hoja',
    },
    # 10 etiquetas/hoja (2 columnas × 5 filas), 4" × 2".
    'avery_5163': {
        'cols': 2, 'rows': 5,
        'label_w': 4.0, 'label_h': 2.0,
        'col_gap': 0.125, 'row_gap': 0.0,
        'top_margin': 0.5, 'left_margin': 0.15625,
        'descripcion': 'Avery 5163 — 10 etiquetas/hoja',
    },
}

# Tope de seguridad: PDFs con > 500 etiquetas se rechazan (evita DoS por
# generación masiva accidental — un usuario puede pedir 200 productos × 500
# cantidad sin este check). 500 cabe en ~17 hojas Avery 5160.
ETIQUETAS_MAX_TOTAL = 500


def _truncate_text(s: str, max_chars: int) -> str:
    s = (s or '').strip()
    if len(s) <= max_chars:
        return s
    return s[:max_chars - 1].rstrip() + '…'


def _draw_etiqueta(c, x, y, w, h, prod, tipo):
    """Dibuja una etiqueta en (x, y) (esquina inferior-izquierda en coord. reportlab).

    Layout:
      - barcode: descripción arriba (2 líneas), código de barras al centro,
        texto del código abajo.
      - qr: QR cuadrado a la izquierda, descripción y código a la derecha.
    """
    from reportlab.lib.units import inch
    from reportlab.graphics.barcode.code128 import Code128
    from reportlab.lib.utils import ImageReader

    pad = 0.06 * inch
    codigo = prod.codigo or ''
    descripcion = prod.descripcion or ''
    categoria = prod.categoria or ''
    unidad = prod.unidad or ''

    if tipo == 'qr':
        # QR ocupa ~ alto de la etiqueta menos padding. Cuadrado.
        qr_size = h - 2 * pad
        # Generamos PIL Image y la metemos al canvas como ImageReader.
        img = qrcode.make(codigo)
        img_buf = io.BytesIO()
        img.save(img_buf, format='PNG')
        img_buf.seek(0)
        c.drawImage(ImageReader(img_buf), x + pad, y + pad,
                    width=qr_size, height=qr_size, preserveAspectRatio=True)

        # Texto a la derecha del QR.
        text_x = x + pad + qr_size + 0.05 * inch
        avail_w = w - (text_x - x) - pad
        max_chars = max(8, int(avail_w / (0.06 * inch)))  # heurística
        c.setFont('Helvetica-Bold', 9)
        c.drawString(text_x, y + h - 0.2 * inch, _truncate_text(descripcion, max_chars))
        c.setFont('Helvetica', 7)
        c.drawString(text_x, y + h - 0.32 * inch, _truncate_text(categoria, max_chars))
        c.setFont('Helvetica-Bold', 10)
        c.drawString(text_x, y + pad + 0.08 * inch, _truncate_text(codigo, max_chars))
        c.setFont('Helvetica', 6)
        c.drawString(text_x, y + pad, f'Unidad: {_truncate_text(unidad, 12)}')
    else:
        # Barcode Code128 + texto.
        # Descripción arriba.
        max_desc = 32 if w < 3.5 * 72 else 50
        c.setFont('Helvetica-Bold', 8)
        c.drawString(x + pad, y + h - 0.16 * inch, _truncate_text(descripcion, max_desc))
        c.setFont('Helvetica', 6)
        c.drawString(x + pad, y + h - 0.27 * inch, _truncate_text(categoria, max_desc))

        # Code128 centrado horizontalmente.
        # barWidth ajustado para que quepa "razonablemente" en 2.4" (5160) o 3.8" (5163).
        bar_h = 0.30 * inch if h < 1.5 * inch else 0.50 * inch
        bar_w = 0.011 * inch if w < 3.5 * inch else 0.014 * inch
        bc = Code128(codigo, barHeight=bar_h, barWidth=bar_w, humanReadable=False)
        # Centrar
        bc_w = bc.width
        bc_x = x + max(pad, (w - bc_w) / 2)
        bc_y = y + (0.20 * inch if h < 1.5 * inch else 0.40 * inch)
        bc.drawOn(c, bc_x, bc_y)

        # Texto del código abajo (centrado).
        c.setFont('Helvetica-Bold', 8)
        c.drawCentredString(x + w / 2, y + 0.08 * inch, _truncate_text(codigo, 24))


def _generar_etiquetas_pdf(productos_expandidos, formato: str, tipo: str) -> io.BytesIO:
    """productos_expandidos: lista plana de Producto, uno por etiqueta a imprimir.
    El orden determina la posición (izq→der, arriba→abajo, hoja por hoja).
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas

    fmt = ETIQUETA_FORMATOS[formato]
    page_w, page_h = letter

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    c.setTitle(f'Etiquetas {fmt["descripcion"]}')

    label_w = fmt['label_w'] * inch
    label_h = fmt['label_h'] * inch
    col_gap = fmt['col_gap'] * inch
    row_gap = fmt['row_gap'] * inch
    top_margin = fmt['top_margin'] * inch
    left_margin = fmt['left_margin'] * inch
    cols = fmt['cols']
    rows = fmt['rows']
    per_page = cols * rows

    for idx, prod in enumerate(productos_expandidos):
        pos = idx % per_page
        if pos == 0 and idx > 0:
            c.showPage()
        row = pos // cols
        col = pos % cols
        x = left_margin + col * (label_w + col_gap)
        # En reportlab origen = bottom-left, así que para fila r desde arriba:
        y = page_h - top_margin - (row + 1) * label_h - row * row_gap
        _draw_etiqueta(c, x, y, label_w, label_h, prod, tipo)

    c.save()
    buf.seek(0)
    return buf


@bp.route('/etiquetas/pdf', methods=['POST'])
@limiter.limit(
    "10/minute",
    key_func=lambda: f"ip:{get_real_client_ip_flask()}",
)
@_require_inventario
def generar_etiquetas_pdf():
    """Genera un PDF de etiquetas Avery (5160 o 5163) con código de barras o QR.

    Body:
      - formato: 'avery_5160' (default, 30/hoja) | 'avery_5163' (10/hoja).
      - tipo: 'barcode' (Code128, default) | 'qr'.
      - items: [{producto_id, cantidad}], ≥1 línea, ≤200 líneas.

    Reglas:
      - Tope global: 500 etiquetas por PDF.
      - Producto debe existir y estar activo.
    """
    data, err = _parse_or_422(EtiquetasPdfSchema(), request.get_json(silent=True))
    if err: return err

    items = data['items']
    total = sum(int(it['cantidad']) for it in items)
    if total > ETIQUETAS_MAX_TOTAL:
        return jsonify({
            'detail': f'Total de etiquetas ({total}) excede el tope de {ETIQUETAS_MAX_TOTAL}'
        }), 422

    # Cargamos todos los productos en una sola query.
    ids = [it['producto_id'] for it in items]
    productos = {
        p.id: p for p in Producto.query.filter(Producto.id.in_(ids), Producto.activo == True).all()  # noqa: E712
    }
    faltantes = [i for i in ids if i not in productos]
    if faltantes:
        return jsonify({
            'detail': f'Productos no encontrados o inactivos: {faltantes}',
        }), 404

    # Expandimos a una lista plana (una entrada por etiqueta) respetando el
    # orden del payload — el front controla en qué orden salen las hojas.
    expandidos = []
    for it in items:
        prod = productos[it['producto_id']]
        expandidos.extend([prod] * int(it['cantidad']))

    pdf = _generar_etiquetas_pdf(expandidos, data['formato'], data['tipo'])

    _audit(
        request.current_user,
        f"Etiquetas PDF ({data['formato']}, {data['tipo']}, {total} etiquetas)",
    )
    db.session.commit()

    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'etiquetas_{data["formato"]}_{ts}.pdf'
    return send_file(
        pdf,
        mimetype='application/pdf',
        as_attachment=False,
        download_name=filename,
    )


# ─── Compras express (Pausa 9) ───────────────────────────────────────────────
#
# Cierra el ciclo Bajo mínimo → Orden de Compra sin construir aún el módulo
# completo de proveedores. La sugerencia de cantidad usa la misma fórmula de
# consumo que Bajo mínimo (Pausa 5) — la diferencia es que aquí se agrupa por
# `Producto.proveedor_default_nombre` para que un PDF cubra a un solo
# proveedor por vez (más práctico para enviar por WhatsApp).
#
# Endpoints:
#   POST /ordenes-compra/express/sugerencia  → JSON con grupos por proveedor
#   POST /ordenes-compra/express/pdf         → PDF (binario) + WhatsApp link
#                                              en header X-Whatsapp-Link
#
# Sin migración nueva más allá de las columnas `proveedor_default_*` ya
# agregadas. El número de OC es un folio efímero (`OCE-YYYYMMDDHHMMSS`); no se
# persiste la orden porque la consideramos un PDF "throw-away" — el inventario
# real se actualiza cuando llega la entrada al almacén.

# Topes generosos para soportar inventarios grandes (2000+ productos bajo
# mínimo en un solo proveedor o en "Sin proveedor"). El PDF con miles de
# ítems funciona, pero xhtml2pdf no es rápido con tablas enormes — la
# generación puede tardar varios segundos. Si se vuelve un problema,
# considerar migrar a reportlab nativo o paginar en cliente.
OC_EXPRESS_MAX_PDF_ITEMS = 10_000
OC_EXPRESS_MAX_SUGERENCIA_ITEMS = 10_000


class _OCExpressItemSchema(_BaseSchema):
    """Línea de la orden, ya editada por el usuario en el modal de preview."""
    producto_id = fields.Int(required=True)
    cantidad = fields.Float(required=True, validate=validate.Range(min=0.01, max=1_000_000))


class OCExpressPdfSchema(_BaseSchema):
    proveedor = fields.Str(required=True, validate=validate.Length(min=1, max=150))
    contacto = fields.Str(load_default='', allow_none=True, validate=validate.Length(max=150))
    notas = fields.Str(load_default='', allow_none=True, validate=validate.Length(max=2000))
    items = fields.List(
        fields.Nested(_OCExpressItemSchema),
        required=True,
        validate=validate.Length(min=1, max=OC_EXPRESS_MAX_PDF_ITEMS),
    )


class OCExpressSugerenciaSchema(_BaseSchema):
    producto_ids = fields.List(
        fields.Int(),
        required=True,
        validate=validate.Length(min=1, max=OC_EXPRESS_MAX_SUGERENCIA_ITEMS),
    )


def _cantidad_sugerida_para_30d(stock_actual: float, stock_minimo: float,
                                  consumo_diario: float) -> float:
    """Cantidad a comprar para cubrir 30 días + reponer el mínimo.

    Formula:   (consumo_diario * 30) - stock_actual + stock_minimo
    Si el resultado da negativo (stock ya cubre el mes), devolvemos
    `max(0, stock_minimo - stock_actual)` para al menos volver al mínimo.
    Redondeo: hacia arriba a 2 decimales.
    """
    necesidad_30d = (consumo_diario * 30.0) - stock_actual + stock_minimo
    if necesidad_30d <= 0:
        # Aún hay buffer para el mes, pero podemos estar bajo el mínimo.
        necesidad_30d = max(0.0, stock_minimo - stock_actual)
    # Redondeo a 2 decimales hacia arriba.
    import math
    return math.ceil(necesidad_30d * 100) / 100.0


@bp.route('/ordenes-compra/express/sugerencia', methods=['POST'])
@limiter.limit('20/minute', key_func=lambda: f"ip:{get_real_client_ip_flask()}")
@_require_inventario_admin
def oc_express_sugerencia():
    """Calcula sugerencia de compra para un set de productos y los agrupa
    por proveedor default. El frontend usa esta respuesta para armar el modal
    de preview antes de generar el PDF.

    Body: `{ producto_ids: [int, ...] }` (1..100).

    Response 200:
    ```
    {
      "grupos": [
        {
          "proveedor": "Cementos del Norte" | "Sin proveedor",
          "contacto": "55 1234 5678" | "",
          "items": [
            {
              "producto_id": 1, "codigo": "CEM-50", "descripcion": "...",
              "unidad": "saco", "stock_actual": 5.0, "stock_minimo": 20.0,
              "consumo_promedio_30d": 1.2, "cantidad_sugerida": 27.0
            }
          ]
        }
      ]
    }
    ```
    """
    data, err = _parse_or_422(OCExpressSugerenciaSchema(), request.get_json(silent=True))
    if err: return err

    ids = list(dict.fromkeys(data['producto_ids']))  # dedupe preservando orden
    productos = {
        p.id: p for p in Producto.query.filter(
            Producto.id.in_(ids),
            Producto.activo == True,  # noqa: E712
        ).all()
    }
    faltantes = [i for i in ids if i not in productos]
    if faltantes:
        return jsonify({
            'detail': f'Productos no encontrados o inactivos: {faltantes}',
        }), 404

    # Consumo en una sola query (anti N+1), igual que el endpoint bajo-mínimo.
    hace_30 = datetime.datetime.now() - datetime.timedelta(days=30)
    consumos = dict(
        db.session.query(
            MovimientoInventario.producto_id,
            db.func.coalesce(db.func.sum(MovimientoInventario.cantidad), 0),
        )
        .filter(
            MovimientoInventario.producto_id.in_(ids),
            MovimientoInventario.tipo == 'SALIDA',
            MovimientoInventario.fecha >= hace_30,
        )
        .group_by(MovimientoInventario.producto_id)
        .all()
    )

    # Construimos los items y los agrupamos por proveedor default.
    grupos: dict[str, dict] = {}
    for pid in ids:
        p = productos[pid]
        consumo_total = float(consumos.get(p.id, 0) or 0)
        consumo_diario = round(consumo_total / 30.0, 2)
        stock = float(p.stock_actual or 0)
        minimo = float(p.stock_minimo or 0)
        sugerida = _cantidad_sugerida_para_30d(stock, minimo, consumo_diario)

        # Normalizamos el nombre del proveedor: vacío/None → "Sin proveedor"
        # para que igual aparezca en el preview (el usuario completa después).
        prov = (p.proveedor_default_nombre or '').strip() or 'Sin proveedor'
        contacto = (p.proveedor_default_contacto or '').strip()

        if prov not in grupos:
            grupos[prov] = {'proveedor': prov, 'contacto': contacto, 'items': []}
        else:
            # Si dos productos del grupo tienen contactos distintos, conservamos
            # el primero — el usuario puede sobreescribir en el modal.
            if not grupos[prov]['contacto'] and contacto:
                grupos[prov]['contacto'] = contacto

        grupos[prov]['items'].append({
            'producto_id': p.id,
            'codigo': p.codigo,
            'descripcion': p.descripcion,
            'unidad': p.unidad,
            'stock_actual': stock,
            'stock_minimo': minimo,
            'consumo_promedio_30d': consumo_diario,
            'cantidad_sugerida': sugerida,
        })

    return jsonify({'grupos': list(grupos.values())})


def _whatsapp_link(proveedor: str, contacto: str, folio: str,
                    items: list[dict]) -> str:
    """Construye un enlace `https://wa.me/<num>?text=...` con un resumen
    listo para enviar al proveedor.

    - Si el contacto incluye dígitos, los extrae y los usa como número.
      Caso contrario, `wa.me/?text=` (el usuario elige el chat).
    - Solo soporta números MX +52 cuando el contacto trae 10 dígitos sin
      prefijo (`5512345678` → `525512345678`).
    """
    from urllib.parse import quote
    import re as _re

    digitos = _re.sub(r'\D', '', contacto or '')
    if digitos and len(digitos) == 10:
        digitos = '52' + digitos  # asumimos MX
    elif len(digitos) > 15:
        digitos = digitos[:15]  # E.164 máximo

    lineas = [f'Orden de compra {folio}', f'Proveedor: {proveedor}', '']
    for it in items[:40]:  # WhatsApp truncará si pasa de ~2k chars
        lineas.append(f"• {it['codigo']} — {it['descripcion']} · {it['cantidad']} {it['unidad']}")
    if len(items) > 40:
        lineas.append(f'... y {len(items) - 40} ítems más (ver PDF adjunto)')
    texto = '\n'.join(lineas)

    base = f'https://wa.me/{digitos}' if digitos else 'https://wa.me/'
    return f'{base}?text={quote(texto)}'


def _render_oc_express_pdf(*, folio: str, fecha_str: str, proveedor: str,
                            contacto: str, notas: str, solicitante: str,
                            items: list[dict]) -> io.BytesIO | None:
    """Genera el PDF con xhtml2pdf reutilizando el estilo de
    `solicitud_pedido_pdf.html` (mismo header azul, mismas tablas).
    Devuelve None si xhtml2pdf no está disponible.
    """
    try:
        from xhtml2pdf import pisa
    except ImportError:
        return None
    import os as _os
    logo_path = _os.path.join(current_app.static_folder, 'imagenes', 'skilled (1).png')
    html_salida = render_template(
        'orden_compra_express_pdf.html',
        folio=folio,
        fecha=fecha_str,
        proveedor=proveedor,
        contacto=contacto,
        notas=notas,
        solicitante=solicitante,
        items=items,
        logo_path=logo_path if _os.path.exists(logo_path) else None,
    )
    buf = io.BytesIO()
    status = pisa.CreatePDF(io.BytesIO(html_salida.encode('utf-8')), dest=buf)
    if status.err:
        return None
    buf.seek(0)
    return buf


@bp.route('/ordenes-compra/express/pdf', methods=['POST'])
@limiter.limit('10/minute', key_func=lambda: f"ip:{get_real_client_ip_flask()}")
@_require_inventario_admin
def oc_express_pdf():
    """Genera el PDF de la orden de compra express y devuelve binario PDF +
    el link de WhatsApp en el header `X-Whatsapp-Link` (URL-encoded).

    Body:
    ```
    {
      "proveedor": "Cementos del Norte",
      "contacto": "55 1234 5678",
      "notas": "Entregar en planta 2",
      "items": [{"producto_id": 1, "cantidad": 27.0}, ...]
    }
    ```

    Reglas:
      - Productos deben existir y estar activos.
      - Sin items → 422.
      - Tope `OC_EXPRESS_MAX_PDF_ITEMS` (anti-DoS).
      - NO persiste la orden — es un PDF "throw-away" para enviar al proveedor.
      - El descuento/entrada al stock se hace por separado vía /movimientos.
    """
    data, err = _parse_or_422(OCExpressPdfSchema(), request.get_json(silent=True))
    if err: return err

    # Dedupe + carga
    ids = list({it['producto_id'] for it in data['items']})
    if len(ids) != len(data['items']):
        return jsonify({'detail': 'Productos duplicados en items'}), 422

    productos = {
        p.id: p for p in Producto.query.filter(
            Producto.id.in_(ids),
            Producto.activo == True,  # noqa: E712
        ).all()
    }
    faltantes = [i for i in ids if i not in productos]
    if faltantes:
        return jsonify({
            'detail': f'Productos no encontrados o inactivos: {faltantes}',
        }), 404

    user = request.current_user
    folio = 'OCE-' + datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    fecha_str = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
    proveedor = data['proveedor'].strip()
    contacto = (data.get('contacto') or '').strip()
    notas = (data.get('notas') or '').strip()
    solicitante = user.full_name or user.username

    # Items decorados (con código/descripción/unidad para el PDF).
    items_view = []
    for it in data['items']:
        p = productos[it['producto_id']]
        cant = float(it['cantidad'])
        cant_int_or_float = int(cant) if cant % 1 == 0 else round(cant, 2)
        items_view.append({
            'codigo': p.codigo,
            'descripcion': p.descripcion,
            'unidad': p.unidad,
            'cantidad': cant_int_or_float,
        })

    pdf = _render_oc_express_pdf(
        folio=folio, fecha_str=fecha_str,
        proveedor=proveedor, contacto=contacto, notas=notas,
        solicitante=solicitante, items=items_view,
    )
    if pdf is None:
        return jsonify({'detail': 'Error al generar el PDF (xhtml2pdf no disponible)'}), 500

    wa_link = _whatsapp_link(proveedor, contacto, folio, items_view)

    _audit(user, f"OC express PDF {folio} → {proveedor} ({len(items_view)} ítems)")
    db.session.commit()

    response = send_file(
        pdf,
        mimetype='application/pdf',
        as_attachment=False,
        download_name=f'{folio}.pdf',
    )
    response.headers['X-Whatsapp-Link'] = wa_link
    response.headers['X-Folio'] = folio
    # Expose para que el SPA pueda leer estos headers desde JS (CORS).
    response.headers['Access-Control-Expose-Headers'] = 'X-Whatsapp-Link, X-Folio'
    return response


# ═══════════════════════════════════════════════════════════════════════════
# Pausa 10 — Conteo físico / Toma de inventario
# ═══════════════════════════════════════════════════════════════════════════

def _toma_to_dict(t: TomaInventario, *, include_detalles: bool = False) -> dict:
    total = len(t.detalles or [])
    capturados = sum(1 for d in (t.detalles or []) if d.cantidad_fisica is not None)
    out = {
        'id': t.id,
        'almacen_id': t.almacen_id,
        'almacen_nombre': t.almacen.nombre if t.almacen else None,
        'fecha_inicio': t.fecha_inicio.isoformat() if t.fecha_inicio else None,
        'fecha_cierre': t.fecha_cierre.isoformat() if t.fecha_cierre else None,
        'usuario_id': t.usuario_id,
        'usuario_nombre': (t.usuario.full_name or t.usuario.username) if t.usuario else None,
        'cerrada_por_id': t.cerrada_por_id,
        'cerrada_por_nombre': (t.cerrada_por.full_name or t.cerrada_por.username) if t.cerrada_por else None,
        'estatus': t.estatus,
        'notas': t.notas,
        'total_lineas': total,
        'lineas_capturadas': capturados,
        'progreso': (capturados / total) if total else 0,
    }
    if include_detalles:
        out['detalles'] = [_toma_detalle_to_dict(d) for d in (t.detalles or [])]
    return out


def _toma_detalle_to_dict(d: TomaInventarioDetalle) -> dict:
    p = d.producto
    cant_fis = None if d.cantidad_fisica is None else float(d.cantidad_fisica)
    cant_sis = float(d.cantidad_sistema or 0)
    diff = None if cant_fis is None else (cant_fis - cant_sis)
    return {
        'id': d.id,
        'toma_id': d.toma_id,
        'producto_id': d.producto_id,
        'producto_codigo': p.codigo if p else None,
        'producto_descripcion': p.descripcion if p else None,
        'producto_unidad': p.unidad if p else None,
        'cantidad_sistema': cant_sis,
        'cantidad_fisica': cant_fis,
        'diferencia': diff,
        'capturado_por_id': d.capturado_por_id,
        'capturado_en': d.capturado_en.isoformat() if d.capturado_en else None,
    }


@bp.route('/tomas/', methods=['POST'])
@_require_inventario_admin
def create_toma():
    """Inicia una toma física para un almacén. Snapshotea StockPorAlmacen
    para todos los productos activos con stock cacheado en ese almacén.
    Si un producto activo no tiene fila en StockPorAlmacen para esa bodega,
    arranca con cantidad_sistema=0 (puede que aparezca al contar)."""
    body = request.get_json(silent=True) or {}
    almacen_id = body.get('almacen_id')
    notas = (body.get('notas') or '').strip()[:1000] or None

    if not almacen_id:
        return jsonify({'detail': 'almacen_id es requerido'}), 422
    alm = Almacen.query.get(almacen_id)
    if not alm or not alm.activo:
        return jsonify({'detail': 'Almacén no encontrado o inactivo'}), 404

    abierta = TomaInventario.query.filter_by(almacen_id=alm.id, estatus='ABIERTA').first()
    if abierta:
        return jsonify({
            'detail': f'Ya hay una toma ABIERTA (#{abierta.id}) para este almacén',
            'toma_id': abierta.id,
        }), 409

    user = request.current_user
    nueva = TomaInventario(
        almacen_id=alm.id,
        usuario_id=user.id,
        estatus='ABIERTA',
        notas=notas,
    )
    db.session.add(nueva)
    db.session.flush()

    # Snapshot: por cada producto activo que tenga fila StockPorAlmacen en
    # este almacén, copiamos la cantidad. También incluimos productos activos
    # sin fila para que el almacenista pueda registrarlos si los encuentra.
    stocks = {
        s.producto_id: s.cantidad
        for s in StockPorAlmacen.query.filter_by(almacen_id=alm.id).all()
    }
    productos = Producto.query.filter(Producto.activo == True).all()
    for p in productos:
        cant = stocks.get(p.id, Decimal('0'))
        db.session.add(TomaInventarioDetalle(
            toma_id=nueva.id,
            producto_id=p.id,
            cantidad_sistema=cant,
        ))

    _audit(user, f"Toma #{nueva.id} iniciada en almacén {alm.nombre} ({len(productos)} líneas)")
    db.session.commit()
    db.session.refresh(nueva)
    return jsonify(_toma_to_dict(nueva))


@bp.route('/tomas/', methods=['GET'])
@_require_inventario_admin
def list_tomas():
    estatus = request.args.get('estatus', '').strip().upper()
    almacen_id = request.args.get('almacen_id', type=int)
    q = TomaInventario.query
    if estatus and estatus in ESTADOS_TOMA:
        q = q.filter(TomaInventario.estatus == estatus)
    if almacen_id:
        q = q.filter(TomaInventario.almacen_id == almacen_id)
    tomas = q.order_by(TomaInventario.fecha_inicio.desc()).limit(200).all()
    return jsonify([_toma_to_dict(t) for t in tomas])


@bp.route('/tomas/<int:toma_id>', methods=['GET'])
@_require_inventario_admin
def get_toma(toma_id: int):
    t = TomaInventario.query.get_or_404(toma_id)
    return jsonify(_toma_to_dict(t, include_detalles=True))


@bp.route('/tomas/<int:toma_id>/detalles/<int:det_id>', methods=['PATCH'])
@_require_inventario_admin
def patch_toma_detalle(toma_id: int, det_id: int):
    """Captura `cantidad_fisica` en una línea. Body: `{cantidad_fisica}`.
    Acepta null para limpiar la captura."""
    t = TomaInventario.query.get_or_404(toma_id)
    if t.estatus != 'ABIERTA':
        return jsonify({'detail': f'Toma {t.estatus.lower()} — no se puede modificar'}), 409
    det = TomaInventarioDetalle.query.filter_by(id=det_id, toma_id=t.id).first_or_404()

    body = request.get_json(silent=True) or {}
    raw = body.get('cantidad_fisica')
    if raw is None or raw == '':
        det.cantidad_fisica = None
        det.capturado_por_id = None
        det.capturado_en = None
    else:
        try:
            cant = Decimal(str(raw))
        except Exception:
            return jsonify({'detail': 'cantidad_fisica inválida'}), 422
        if cant < 0:
            return jsonify({'detail': 'cantidad_fisica no puede ser negativa'}), 422
        det.cantidad_fisica = cant
        det.capturado_por_id = request.current_user.id
        det.capturado_en = datetime.datetime.utcnow()

    db.session.commit()
    db.session.refresh(det)
    return jsonify(_toma_detalle_to_dict(det))


@bp.route('/tomas/<int:toma_id>/detalles/por-codigo', methods=['PATCH'])
@_require_inventario_admin
def patch_toma_detalle_por_codigo(toma_id: int):
    """Atajo para PWA scanner: captura por código de producto.
    Body: `{codigo, cantidad_fisica}`. Devuelve el detalle actualizado."""
    t = TomaInventario.query.get_or_404(toma_id)
    if t.estatus != 'ABIERTA':
        return jsonify({'detail': f'Toma {t.estatus.lower()} — no se puede modificar'}), 409

    body = request.get_json(silent=True) or {}
    codigo = (body.get('codigo') or '').strip()
    if not codigo:
        return jsonify({'detail': 'codigo es requerido'}), 422

    prod = Producto.query.filter_by(codigo=codigo, activo=True).first()
    if not prod:
        return jsonify({'detail': f'Producto {codigo} no encontrado'}), 404

    det = TomaInventarioDetalle.query.filter_by(toma_id=t.id, producto_id=prod.id).first()
    if not det:
        # producto que no estaba en el snapshot (activo nuevo) — lo agregamos
        det = TomaInventarioDetalle(toma_id=t.id, producto_id=prod.id, cantidad_sistema=Decimal('0'))
        db.session.add(det)
        db.session.flush()

    raw = body.get('cantidad_fisica')
    try:
        cant = Decimal(str(raw))
    except Exception:
        return jsonify({'detail': 'cantidad_fisica inválida'}), 422
    if cant < 0:
        return jsonify({'detail': 'cantidad_fisica no puede ser negativa'}), 422

    det.cantidad_fisica = cant
    det.capturado_por_id = request.current_user.id
    det.capturado_en = datetime.datetime.utcnow()
    db.session.commit()
    db.session.refresh(det)
    return jsonify(_toma_detalle_to_dict(det))


@bp.route('/tomas/<int:toma_id>/cerrar', methods=['POST'])
@_require_inventario_admin
def cerrar_toma(toma_id: int):
    """Cierra la toma generando AJUSTES por cada línea con diferencia.
    Líneas sin captura (cantidad_fisica null) se asumen iguales al sistema
    — NO se ajustan, pero quedan registradas como "no contadas".

    El llamador puede pasar `{omitir_no_capturados: true}` (default) o false
    para forzar que las no capturadas se traten como cantidad_fisica=0 (riesgoso).
    """
    t = TomaInventario.query.get_or_404(toma_id)
    if t.estatus != 'ABIERTA':
        return jsonify({'detail': f'Toma ya está {t.estatus.lower()}'}), 409

    body = request.get_json(silent=True) or {}
    asumir_cero = bool(body.get('asumir_cero_no_capturados', False))

    user = request.current_user
    ajustes_creados = 0
    errores = []

    for det in t.detalles:
        cant_fis = det.cantidad_fisica
        if cant_fis is None:
            if not asumir_cero:
                continue
            cant_fis = Decimal('0')
        cant_sis = Decimal(str(det.cantidad_sistema or 0))
        diff = Decimal(str(cant_fis)) - cant_sis
        if diff == 0:
            continue
        # AJUSTE: positivo sube destino, negativo baja origen.
        data = {
            'tipo': 'AJUSTE',
            'producto_id': det.producto_id,
            'cantidad': diff,
            'motivo': f'Toma física #{t.id}',
        }
        if diff > 0:
            data['almacen_destino_id'] = t.almacen_id
        else:
            data['almacen_origen_id'] = t.almacen_id
        resp = _perform_movimiento(data, user)
        # _perform_movimiento puede devolver (response, status) o response.
        # Si es tuple con código de error, abortamos.
        if isinstance(resp, tuple):
            body_resp, status = resp
            if status >= 400:
                errores.append({
                    'producto_id': det.producto_id,
                    'detail': body_resp.get_json().get('detail') if hasattr(body_resp, 'get_json') else str(body_resp),
                    'status': status,
                })
                continue
        ajustes_creados += 1

    if errores:
        db.session.rollback()
        return jsonify({
            'detail': 'No se pudieron generar todos los ajustes — toma sigue ABIERTA',
            'errores': errores,
            'ajustes_intentados': ajustes_creados,
        }), 409

    t.estatus = 'CERRADA'
    t.fecha_cierre = datetime.datetime.utcnow()
    t.cerrada_por_id = user.id
    _audit(user, f"Toma #{t.id} cerrada — {ajustes_creados} ajustes generados")
    db.session.commit()
    db.session.refresh(t)
    return jsonify({**_toma_to_dict(t), 'ajustes_creados': ajustes_creados})


@bp.route('/tomas/<int:toma_id>/cancelar', methods=['POST'])
@_require_inventario_admin
def cancelar_toma(toma_id: int):
    """Cancela una toma sin aplicar ajustes."""
    t = TomaInventario.query.get_or_404(toma_id)
    if t.estatus != 'ABIERTA':
        return jsonify({'detail': f'Toma ya está {t.estatus.lower()}'}), 409
    t.estatus = 'CANCELADA'
    t.fecha_cierre = datetime.datetime.utcnow()
    t.cerrada_por_id = request.current_user.id
    _audit(request.current_user, f"Toma #{t.id} cancelada")
    db.session.commit()
    return jsonify(_toma_to_dict(t))


@bp.route('/tomas/<int:toma_id>/pdf', methods=['GET'])
@_require_inventario_admin
def get_toma_pdf(toma_id: int):
    """PDF de acta de toma con diferencias y firmas."""
    try:
        from xhtml2pdf import pisa
    except ImportError:
        return jsonify({'detail': 'xhtml2pdf no instalado'}), 500

    t = TomaInventario.query.get_or_404(toma_id)
    detalles_out = []
    capturadas = 0
    con_diff = 0
    no_cap = 0
    for d in sorted(t.detalles, key=lambda x: (x.producto.codigo if x.producto else '')):
        cant_fis = None if d.cantidad_fisica is None else float(d.cantidad_fisica)
        cant_sis = float(d.cantidad_sistema or 0)
        if cant_fis is None:
            no_cap += 1
            diff = None
        else:
            capturadas += 1
            diff = cant_fis - cant_sis
            if abs(diff) > 1e-9:
                con_diff += 1
        detalles_out.append({
            'codigo': d.producto.codigo if d.producto else '—',
            'descripcion': d.producto.descripcion if d.producto else '—',
            'unidad': d.producto.unidad if d.producto else None,
            'sistema': cant_sis,
            'fisico': cant_fis,
            'diff': diff if diff is not None else 0,
        })

    import os as _os
    logo_path = _os.path.join(current_app.static_folder, 'imagenes', 'skilled (1).png')
    html_salida = render_template(
        'toma_inventario_pdf.html',
        toma=t,
        almacen_nombre=t.almacen.nombre if t.almacen else None,
        iniciada_por=(t.usuario.full_name or t.usuario.username) if t.usuario else None,
        cerrada_por=(t.cerrada_por.full_name or t.cerrada_por.username) if t.cerrada_por else None,
        estatus=t.estatus,
        fecha_inicio=t.fecha_inicio.strftime('%Y-%m-%d %H:%M') if t.fecha_inicio else '',
        detalles=detalles_out,
        resumen={
            'total': len(detalles_out),
            'capturadas': capturadas,
            'con_diferencia': con_diff,
            'no_capturadas': no_cap,
        },
        logo_path=logo_path if _os.path.exists(logo_path) else None,
    )
    buf = io.BytesIO()
    status = pisa.CreatePDF(io.BytesIO(html_salida.encode('utf-8')), dest=buf)
    if status.err:
        return jsonify({'detail': 'Error generando PDF'}), 500
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/pdf',
        as_attachment=False,
        download_name=f'toma-{t.id}.pdf',
    )
