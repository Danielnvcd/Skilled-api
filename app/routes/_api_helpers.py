"""Helpers compartidos por los blueprints `api_*` (consumidos por el SPA React).

Centraliza patrones que estaban duplicados en 10+ archivos:

  current_user()             usuario autenticado actual (alias del legacy `_u()`)
  is_admin()                 True si rol ∈ {admin, super_admin} → operación RRHH
  is_super_admin()           True si rol == super_admin
  is_sistemas()              True si rol == sistemas
  puede_gestionar_sistema()  True si rol ∈ {sistemas, super_admin}
  require_admin()            devuelve (resp, code) si NO es admin; None si pasa
  require_roles(*roles)      idem pero para sets arbitrarios de roles
  require_gestion_usuarios() gate de alta/baja de cuentas (rol `sistemas`)
  require_panel_sistemas()   idem + exige 2FA activo (panel de sistemas)
  api_transactional          decorador: rollback + log + 500 ante excepciones no HTTP

Los dos ejes de permiso son deliberadamente independientes:
  - `admin`/`super_admin` → datos del negocio (nómina, empleados, préstamos).
  - `sistemas`/`super_admin` → el sistema (cuentas, sesiones, servidor).
Solo `super_admin` cruza ambos, y existe como cuenta de recuperación.

Todas asumen que el endpoint ya pasó por `@jwt_required` (de api_auth), que
deja el User cargado en `flask.g._jwt_user`. Si se llaman sin JWT activo,
disparan `AttributeError` — comportamiento equivalente al de los `_u()`
originales (fallaban igual).

No se cambia ningún contrato externo: las funciones devuelven los MISMOS
mensajes y códigos HTTP que los helpers locales que reemplazan.
"""
from __future__ import annotations

import traceback
from functools import wraps
from typing import Iterable

from flask import current_app, g, jsonify, send_file
from werkzeug.exceptions import HTTPException

from app.extensions import db
from app.models import User
from app.utils import log_action, safe_excel_value


def current_user() -> User:
    """Usuario autenticado actual (puesto por `@jwt_required`)."""
    return g._jwt_user


def is_admin() -> bool:
    """True si el rol da acceso a la operación de RRHH/nómina.

    OJO — `sistemas` NO entra aquí a propósito. Ese rol administra el SISTEMA
    (usuarios, sesiones, estado del servidor, eventos de seguridad), no los
    datos de nómina: sueldos, préstamos, expedientes de empleados. Es la
    separación clásica entre quien opera la infraestructura y quien puede leer
    la información sensible del negocio.

    Que un rol pueda crear cuentas no debe implicar que pueda leer los sueldos.
    Si se quisiera lo contrario, el cambio es agregar 'sistemas' a esta tupla —
    pero entonces una sola cuenta comprometida expone también toda la nómina.
    """
    return current_user().role in ('admin', 'super_admin')


def is_super_admin() -> bool:
    return current_user().role == 'super_admin'


def is_sistemas() -> bool:
    return current_user().role == 'sistemas'


# Roles que administran el sistema: el rol dedicado `sistemas` y `super_admin`
# como salida de emergencia (si la cuenta de sistemas queda bloqueada, sin esto
# nadie podría crear usuarios ni destrabar nada).
_ROLES_GESTION = ('sistemas', 'super_admin')


def puede_gestionar_sistema() -> bool:
    return current_user().role in _ROLES_GESTION


def require_gestion_usuarios(message: str = 'Solo el rol sistemas puede administrar usuarios'):
    """Gate para el alta/baja/edición de usuarios y el panel de sistemas.

    Sustituye a `require_admin()` en `api_users`: la gestión de cuentas salió
    de admin (que quedó para RRHH) y vive en el rol `sistemas`.
    """
    if not puede_gestionar_sistema():
        return jsonify({'error': message}), 403
    return None


def require_panel_sistemas():
    """Gate del panel de sistemas: rol correcto **y** 2FA activo.

    Por qué el 2FA es obligatorio justo aquí y no en el resto de la app: este
    rol puede crear cuentas y revocar sesiones, o sea que puede fabricarse un
    acceso administrativo. Es el objetivo más valioso del sistema, así que una
    contraseña filtrada no puede ser suficiente para entrar.

    Se verifica en el ENDPOINT, no en el login: el usuario puede iniciar sesión
    normalmente y el SPA lo manda a inscribir su TOTP con el flag
    `requiere_2fa` que devolvemos aquí. Así nadie queda fuera de la aplicación
    por no tener el segundo factor todavía — solo fuera del panel.
    """
    err = require_gestion_usuarios('Solo el rol sistemas puede entrar al panel')
    if err:
        return err
    u = current_user()
    if not u.totp_secret:
        return jsonify({
            'error': (
                'Este panel requiere autenticación de dos factores. '
                'Actívala en Mi Perfil para continuar.'
            ),
            'requiere_2fa': True,
        }), 403
    return None


def require_admin(message: str = 'Acceso denegado'):
    """Gate de admin para usar dentro del cuerpo del endpoint.

    Patrón de uso (mantiene el shape de los `_admin_only()` originales):

        @bp.route(...)
        @jwt_required
        def vista():
            err = require_admin()
            if err:
                return err
            ...

    Devuelve la tupla `(response, 403)` si el usuario no es admin, o `None`
    si pasa la verificación.
    """
    if not is_admin():
        return jsonify({'error': message}), 403
    return None


def require_roles(roles: Iterable[str], message: str = 'Acceso denegado'):
    """Gate genérico: pasa si el rol del usuario está en `roles`."""
    if current_user().role not in set(roles):
        return jsonify({'error': message}), 403
    return None


def api_transactional(error_message: str = 'Ocurrió un error en el servidor'):
    """Decorador para endpoints que escriben a BD.

    Equivale a este patrón que aparece duplicado ~50 veces en los blueprints
    `api_*`:

        try:
            ...lógica...
            return jsonify(...), 2xx
        except Exception as e:
            db.session.rollback()
            current_app.logger.error('Error ...: %s\\n%s', e, traceback.format_exc())
            return jsonify({'error': '...'}), 500

    Importante:
    - `HTTPException` (404, 422, abort(...)) **se deja pasar** — Flask la
      convierte en su respuesta normal sin romper el contrato HTTP.
    - Cualquier otra excepción dispara rollback + log con traceback completo
      y devuelve 500 con `error_message`.

    NO reemplaza catches específicos que devuelven mensajes / códigos
    distintos: esos siguen viviendo dentro del endpoint.
    """
    def wrapper(view):
        @wraps(view)
        def inner(*args, **kwargs):
            try:
                return view(*args, **kwargs)
            except HTTPException:
                raise
            except Exception as exc:
                db.session.rollback()
                current_app.logger.error(
                    '[%s] %s\n%s', view.__name__, exc, traceback.format_exc()
                )
                return jsonify({'error': error_message}), 500
        return inner
    return wrapper


# ── Excel: saneo + estilos comunes ──────────────────────────────────────────
# Lo usan 5 paquetes distintos (api_ajustes, api_historico, api_prenomina,
# api_prestamos, api_proyecto_total). Vive aquí para que no hayan dos copias
# y para no atar el saneo a un blueprint en particular.

def _sanitize_rows(rows):
    """Sanea cada celda de cada fila contra inyección de fórmulas Excel
    (campos como 'Nombre', 'Motivo' o 'Notas' pueden venir con un '=' al inicio)."""
    for row in rows:
        for k, v in row.items():
            row[k] = safe_excel_value(v)
    return rows


def _aplicar_estilos_y_retornar(writer, output, filename):
    """Aplica un diseño premium y retorna la respuesta de Flask para descargar el Excel."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Dark Blue
    header_font = Font(name="Calibri", bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")

    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")  # Very light slate

    total_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")  # Light Blue
    total_font = Font(name="Calibri", bold=True, color="1E3A8A")
    top_border = Border(top=Side(style='medium', color="1E3A8A"))

    workbook = writer.book
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        worksheet.freeze_panes = 'A2'
        max_row = worksheet.max_row

        for row in range(1, max_row + 1):
            is_header = (row == 1)
            is_total = False

            first_cell = worksheet.cell(row=row, column=1)
            if first_cell.value and str(first_cell.value).upper() == 'TOTAL':
                is_total = True

            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)

                if is_header:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                else:
                    if not is_total and row % 2 == 0:
                        cell.fill = zebra_fill

                    if is_total:
                        cell.fill = total_fill
                        cell.font = total_font
                        cell.border = top_border

                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '"$"#,##0.00'

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = min(adjusted_width, 40)

    writer.close()
    output.seek(0)

    log_action(f"Exportó Reporte Excel: {filename}")

    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
