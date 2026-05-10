from flask import Blueprint, render_template, request, flash, redirect, url_for, jsonify, current_app, send_from_directory, send_file, session
import io
from sqlalchemy import or_, func
from sqlalchemy.orm import selectinload
from app.extensions import db, limiter
from app.models import Trabajador, CredencialPlanta, DocumentoTrabajador, Proyecto
from app.utils import login_required, log_action, admin_required, allowed_file, allowed_image_file, validate_lengths
from werkzeug.utils import secure_filename
import traceback
import json
import os
import re
import time
from datetime import datetime as dt
import pandas as pd

_CURP_RE = re.compile(r'^[A-Z]{4}[0-9]{6}[HM][A-Z]{5}[A-Z0-9]{2}$')
_RFC_RE  = re.compile(r'^[A-Z&Ñ]{3,4}[0-9]{6}[A-Z0-9]{3}$')



bp = Blueprint('trabajadores', __name__, url_prefix='/trabajadores')

def is_authorized_for_worker(t):
    from flask import session
    role = session.get('role')
    if role in ['admin', 'super_admin']:
        return True
    
    user_id = session.get('user_id')
    for p in t.proyectos.all():
        if p.activo and p.coordinador_id == user_id:
            return True
    return False

def _parse_date(value):
    """Convierte string YYYY-MM-DD a date. Retorna None si está vacío o es inválido."""
    if not value:
        return None
    try:
        return dt.strptime(value, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None

@bp.route('/', methods=['GET'])
@login_required
def index():
    page = request.args.get('page', 1, type=int)
    q = request.args.get('q', '').strip()
    
    query = Trabajador.query.options(
        selectinload(Trabajador.credenciales),
        selectinload(Trabajador.documentos)
    ).filter(Trabajador.activo == True, Trabajador.fecha_baja == None)

    if q:
        query = query.filter(or_(
            Trabajador.nombre.ilike(f'%{q}%'),
            Trabajador.nombre_apellidos.ilike(f'%{q}%'),
            Trabajador.no_empleado.ilike(f'%{q}%'),
            Trabajador.rfc.ilike(f'%{q}%')
        ))

    pagination = query.order_by(func.lower(Trabajador.nombre)).paginate(page=page, per_page=20, error_out=False)
    trabajadores = pagination.items

    return render_template('trabajadores.html', trabajadores=trabajadores, pagination=pagination, q=q)

@bp.route('/bajas', methods=['GET'])
@login_required
@admin_required
def bajas():
    page = request.args.get('page', 1, type=int)
    q = request.args.get('q', '').strip()

    query = Trabajador.query.options(
        selectinload(Trabajador.credenciales),
        selectinload(Trabajador.documentos)
    ).filter(or_(Trabajador.activo == False, Trabajador.fecha_baja != None))
    
    if q:
        query = query.filter(or_(
            Trabajador.nombre.ilike(f'%{q}%'),
            Trabajador.nombre_apellidos.ilike(f'%{q}%'),
            Trabajador.no_empleado.ilike(f'%{q}%'),
            Trabajador.rfc.ilike(f'%{q}%')
        ))
        
    pagination = query.order_by(func.lower(Trabajador.nombre)).paginate(page=page, per_page=20, error_out=False)
    trabajadores = pagination.items

    return render_template('trabajadores_bajas.html', trabajadores=trabajadores, pagination=pagination, q=q)

@bp.route('/agregar', methods=['POST'])
@login_required
@limiter.limit("10 per minute")
def agregar():
    try:
        data = request.form
        
        no_empleado = data.get('no_empleado', '').strip()
        if not no_empleado:
            flash('Error: El Número de Empleado es obligatorio.', 'danger')
            return redirect(url_for('trabajadores.index'))
            
        errores_longitud = validate_lengths(data)
        if errores_longitud:
            flash("Error de longitud en los campos:<br>" + "<br>".join(errores_longitud), 'danger')
            return redirect(url_for('trabajadores.index'))

        # Validación de formato CURP y RFC (advertencia, no bloqueo, por registros históricos)
        _curp_val = (data.get('curp') or '').strip().upper()
        _rfc_val  = (data.get('rfc') or '').strip().upper()
        if _curp_val and not _CURP_RE.match(_curp_val):
            flash('Advertencia: El formato de CURP no es válido. Verifica antes de continuar.', 'warning')
        if _rfc_val and not _RFC_RE.match(_rfc_val):
            flash('Advertencia: El formato de RFC no es válido. Verifica antes de continuar.', 'warning')

        # Basic validation
        if Trabajador.query.filter_by(no_empleado=no_empleado).first():
            flash('Error: El Número de Empleado ya existe.', 'danger')
            return redirect(url_for('trabajadores.index'))

        try:
            salario = float(data.get('salario_real_pactado_x_sem') or 0)
            if salario < 0:
                flash('Error: El salario no puede ser negativo.', 'danger')
                return redirect(url_for('trabajadores.index'))
        except ValueError:
            flash('Error: El salario ingresado es inválido.', 'danger')
            return redirect(url_for('trabajadores.index'))

        nuevo_trabajador = Trabajador(
            # Identificadores
            no_empleado=no_empleado,
            nombre_apellidos=data.get('nombre_apellidos'),
            nombre=data.get('nombre'),
            
            # Laborales
            tipo_mov=data.get('tipo_mov'),
            tipo_cont=data.get('tipo_cont'),
            area=data.get('area'),
            puesto=data.get('puesto'),
            tipo_jornada=data.get('tipo_jornada'),
            fecha_ingreso=_parse_date(data.get('fecha_ingreso')),
            descripcion_servicio=data.get('descripcion_servicio'),
            inicio=_parse_date(data.get('inicio')),
            termino_prueba=_parse_date(data.get('termino_prueba')),
            fecha_baja=_parse_date(data.get('fecha_baja')),
            
            # Generales
            curp=data.get('curp'),
            rfc=data.get('rfc'),
            nss=data.get('nss'),
            fecha_nacimiento=_parse_date(data.get('fecha_nacimiento')),
            sexo=data.get('sexo'),
            estado_civil=data.get('estado_civil'),
            nacionalidad=data.get('nacionalidad'),
            edad=data.get('edad') or None,
            domicilio=data.get('domicilio'),
            
            # Datos de contacto
            correo=data.get('correo'),
            celular=data.get('celular'),
            
            # Médicos
            tipo_sangre=data.get('tipo_sangre'),
            alergias=data.get('alergias'),
            enfermedades_cronicas=data.get('enfermedades_cronicas'),
            contacto_emergencia=data.get('contacto_emergencia'),
            parentesco_contacto=data.get('parentesco_contacto'),
            numero_contacto_emerg=data.get('numero_contacto_emerg'),
            lentes=data.get('lentes'),
            licencia_conducir=data.get('licencia_conducir'),
            estatura=data.get('estatura'),
            
            # Sueldos (Finanzas)
            salario_real_pactado_x_sem=float(data.get('salario_real_pactado_x_sem') or 0),
            tipo_pago=data.get('tipo_pago'),
            tipo_nomina=data.get('tipo_nomina'),
            sb=data.get('sb') or 0,
            sdi=data.get('sdi') or 0,
            letra=data.get('letra'),
            hr_extra=data.get('hr_extra') or 0,
            infonavit=data.get('infonavit') or 0,
            ajuste_inbursa=data.get('ajuste_inbursa') or 0,
            caja_ahorro=data.get('caja_ahorro') or 0,
            viaticos=data.get('viaticos') or 0,
            pago_dia_festivo=data.get('pago_dia_festivo') or 0,
            pagos_efectivo=data.get('pagos_efectivo') or 0,
            folio_mov_idse=data.get('folio_mov_idse'),
            
            # Ubicación y Operación (no_proyecto, ubicacion_actual, coord_a_cargo se manejan en Proyectos)
            ubicacion_estado=data.get('ubicacion_estado'),
            observaciones=data.get('observaciones')
        )
        
        # Handle Profile Picture (solo JPG/PNG reales)
        if 'foto_perfil' in request.files:
            file = request.files['foto_perfil']
            if file and file.filename != '':
                if not allowed_image_file(file):
                    flash('Foto de perfil rechazada: solo se permiten imágenes JPG o PNG reales.', 'danger')
                    return redirect(url_for('trabajadores.index'))
                unique_filename = f"pp_{int(time.time())}_{secure_filename(file.filename)}"
                pp_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'perfiles')
                nuevo_trabajador.foto_perfil = _save_profile_picture(file, pp_folder, unique_filename)
        
        # Parse Credentials JSON
        try:
            credenciales_str = data.get('credenciales_json', '[]')
            credenciales_data = json.loads(credenciales_str)
            for c_data in credenciales_data:
                nueva_credencial = CredencialPlanta(
                    planta=c_data.get('planta', '').upper(),
                    credencial_id=c_data.get('credencial_id', '')[:40],
                    fecha_caducidad=_parse_date(c_data.get('fecha_caducidad'))
                )
                nuevo_trabajador.credenciales.append(nueva_credencial)
        except Exception as json_err:
            current_app.logger.error(f"Error parsing credentials: {json_err}")
            # we continue even if credentials fail parsing, though we can flash a warning
            flash('Hubo un problema guardando algunas credenciales de planta.', 'warning')
        
        db.session.add(nuevo_trabajador)
        db.session.commit()
        log_action(f"Agregó al trabajador {nuevo_trabajador.nombre} ({nuevo_trabajador.no_empleado})")
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': 'Trabajador agregado exitosamente.', 'id': nuevo_trabajador.id})
            
        flash('Trabajador agregado exitosamente.', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error saving worker: {e}\n{traceback.format_exc()}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'Ocurrió un error al guardar el trabajador. Verifica los datos.'}), 500
        flash('Ocurrió un error al guardar el trabajador. Verifica los datos.', 'danger')
        
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': False, 'message': 'Operación fallida.'}), 400
    return redirect(url_for('trabajadores.index'))

@bp.route('/importar', methods=['GET'])
@login_required
def importar():
    return render_template('importar_empleados.html')

@bp.route('/procesar_importacion', methods=['POST'])
@login_required
@limiter.limit("5 per minute")
def procesar_importacion():
    if 'archivo_excel' not in request.files:
        flash('No se seleccionó ningún archivo.', 'danger')
        return redirect(url_for('trabajadores.importar'))
        
    file = request.files['archivo_excel']
    if file.filename == '':
        flash('El archivo está vacío.', 'danger')
        return redirect(url_for('trabajadores.importar'))
        
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Formato no válido. Debe ser .xlsx o .xls', 'danger')
        return redirect(url_for('trabajadores.importar'))
        
    # Validar magic bytes para prevenir subida de archivos maliciosos ocultos como Excel
    import filetype
    header = file.read(2048)
    file.seek(0)
    kind = filetype.guess(header)
    if not kind or kind.extension not in ['xlsx', 'xls', 'zip', 'cfb']:
        flash('El archivo fue rechazado por razones de seguridad: no parece ser un documento Excel válido.', 'danger')
        return redirect(url_for('trabajadores.importar'))
        
    # DoS protection: read file size without putting all in memory
    file.seek(0, os.SEEK_END)
    size = file.tell()
    file.seek(0)
    if size > 5 * 1024 * 1024:
        flash('El archivo excede el tamaño máximo permitido (5MB).', 'danger')
        return redirect(url_for('trabajadores.importar'))
        
    try:
        df = pd.read_excel(file)
        
        errores = []
        exitosos = 0
        
        for index, row in df.iterrows():
            no_emp = str(row.get('No. Empleado', '')).strip()
            if no_emp == 'nan' or not no_emp:
                errores.append(f"Fila {index+2}: Falta No. Empleado")
                continue
                
            if Trabajador.query.filter_by(no_empleado=no_emp).first():
                errores.append(f"Fila {index+2}: El empleado {no_emp} ya existe.")
                continue
                
            nombre = str(row.get('Nombre(s)', '')).strip()
            apellidos = str(row.get('Apellidos', '')).strip()
            
            if nombre == 'nan': nombre = ''
            if apellidos == 'nan': apellidos = ''
            
            # Flags for current row errors
            row_errors = []

            _FORMULA_CHARS = ('=', '+', '-', '@', '\t', '\r')

            def get_str(col_name, max_len=None):
                val = row.get(col_name, '')
                s = str(val).strip() if pd.notna(val) else ''
                # Sanitizar formula injection: remover caracteres de inicio de fórmula CSV/Excel
                if s and s[0] in _FORMULA_CHARS:
                    s = "'" + s
                if max_len and len(s) > max_len:
                    row_errors.append(f"El campo '{col_name}' excede el límite de {max_len} caracteres (recibidos: {len(s)}).")
                return s
                
            def get_float(col_name):
                val = row.get(col_name, 0)
                if pd.isna(val) or val == '':
                    return 0.0
                try:
                    return float(val)
                except ValueError:
                    row_errors.append(f"El valor '{val}' en '{col_name}' no es un número válido.")
                    return 0.0
                    
            def get_date(col_name):
                raw = row.get(col_name)
                if pd.isna(raw) or raw == '':
                    return None
                try:
                    return pd.to_datetime(raw).date()
                except Exception:
                    row_errors.append(f"La fecha '{raw}' en '{col_name}' no tiene el formato correcto (YYYY-MM-DD).")
                    return None

            t_nuevo = Trabajador(
                # Identificadores
                no_empleado=get_str('No. Empleado', 50),
                nombre=get_str('Nombre(s)', 250),
                nombre_apellidos=get_str('Apellidos', 250),
                
                # Laborales
                tipo_mov=get_str('Tipo de Movimiento', 100),
                tipo_cont=get_str('Tipo de Contrato', 100),
                area=get_str('Area', 150),
                puesto=get_str('Puesto', 150),
                tipo_jornada=get_str('Tipo de Jornada', 100),
                fecha_ingreso=get_date('Fecha Ingreso (YYYY-MM-DD)'),
                descripcion_servicio=get_str('Descripcion de Servicio'),
                inicio=get_date('Fecha Inicio (YYYY-MM-DD)'),
                termino_prueba=get_date('Termino de Prueba (YYYY-MM-DD)'),
                
                # Generales
                rfc=get_str('RFC', 13),
                curp=get_str('CURP', 18),
                nss=get_str('NSS', 20),
                fecha_nacimiento=get_date('Fecha Nacimiento (YYYY-MM-DD)'),
                sexo=get_str('Sexo (M/F)', 20),
                estado_civil=get_str('Estado Civil', 50),
                nacionalidad=get_str('Nacionalidad', 100),
                edad=int(get_float('Edad')) if get_float('Edad') > 0 else None,
                domicilio=get_str('Domicilio'),
                
                # Contacto
                correo=get_str('Correo', 150),
                celular=get_str('Celular', 20),
                
                # Medicos
                tipo_sangre=get_str('Tipo de Sangre', 10),
                alergias=get_str('Alergias'),
                enfermedades_cronicas=get_str('Enfermedades Cronicas'),
                contacto_emergencia=get_str('Contacto de Emergencia', 200),
                parentesco_contacto=get_str('Parentesco del Contacto', 100),
                numero_contacto_emerg=get_str('Numero Contacto Emergencia', 20),
                lentes=get_str('Usa Lentes (Si/No)', 20),
                licencia_conducir=get_str('Licencia de Conducir (Tipo)', 50),
                estatura=get_str('Estatura', 20),
                
                # Finanzas
                salario_real_pactado_x_sem=get_float('Salario Real Pactado por Semana'),
                tipo_pago=get_str('Tipo Pago', 50),
                tipo_nomina=get_str('Tipo de Nomina (Semanal/Por hora/Cuadrado)', 100),
                sb=get_float('Sueldo Base (SB)'),
                sdi=get_float('Salario Diario Integrado (SDI)'),
                letra=get_str('Letra', 100),
                hr_extra=get_float('Horas Extra'),
                infonavit=get_float('Infonavit'),
                caja_ahorro=get_float('Caja de Ahorro'),
                viaticos=get_float('Viaticos'),
                pago_dia_festivo=get_float('Pago Dia Festivo'),
                pagos_efectivo=get_float('Pagos Efectivo'),
                folio_mov_idse=get_str('Folio Mov IDSE', 100),
                
                # Operacion
                ubicacion_estado=get_str('Ubicacion Estado', 100),
                observaciones=get_str('Observaciones')
            )
            
            if row_errors:
                # Si hay errores de formato en esta fila, abortamos su inserción y anotamos
                errores.append(f"Fila {index+2} ({no_emp}): " + ", ".join(row_errors))
            else:
                db.session.add(t_nuevo)
                exitosos += 1
            
        if exitosos > 0:
            db.session.commit()
            log_action(f"Importó masivamente {exitosos} trabajadores desde Excel")
            flash(f'¡Éxito! Se importaron {exitosos} empleados correctamente.', 'success')
            
        if errores:
            mensaje_error = "Errores encontrados:<br>" + "<br>".join(errores[:10])
            if len(errores) > 10:
                mensaje_error += f"<br>...y {len(errores)-10} errores más."
            flash(mensaje_error, 'warning')
            
        if exitosos > 0 and not errores:
            return redirect(url_for('trabajadores.index'))
            
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error procesando Excel: {e}\n{traceback.format_exc()}")
        flash('Error al leer el archivo Excel. Asegúrate de usar la plantilla correcta.', 'danger')

    return redirect(url_for('trabajadores.importar'))

def _generate_thumbnail(original_path: str, thumb_path: str, size: tuple = (100, 100)) -> None:
    """Genera un thumbnail JPEG de `size` px a partir de `original_path`."""
    from PIL import Image
    with Image.open(original_path) as img:
        img = img.convert('RGB')
        img.thumbnail(size, Image.LANCZOS)
        img.save(thumb_path, 'JPEG', quality=85, optimize=True)


def _save_profile_picture(file, pp_folder: str, unique_filename: str) -> str:
    """
    Guarda la foto original y genera su thumbnail.
    Devuelve la ruta relativa al UPLOAD_FOLDER de la imagen original
    (p.ej. 'perfiles/pp_123_foto.jpg').
    El thumbnail se guarda como 'perfiles/thumb_pp_123_foto.jpg'.
    """
    os.makedirs(pp_folder, exist_ok=True)
    original_path = os.path.join(pp_folder, unique_filename)
    file.save(original_path)

    thumb_filename = f"thumb_{unique_filename}"
    thumb_path = os.path.join(pp_folder, thumb_filename)
    try:
        _generate_thumbnail(original_path, thumb_path)
    except Exception as e:
        current_app.logger.warning(f"No se pudo generar thumbnail para {unique_filename}: {e}")

    return f"perfiles/{unique_filename}"


def _delete_profile_picture(foto_perfil_rel: str) -> None:
    """Elimina la foto original y su thumbnail dado el path relativo almacenado en BD."""
    from flask import current_app
    upload_folder = current_app.config['UPLOAD_FOLDER']
    original_path = os.path.join(upload_folder, foto_perfil_rel)
    try:
        if os.path.exists(original_path):
            os.remove(original_path)
    except Exception as e:
        current_app.logger.error(f"Error eliminando foto original: {e}")

    # thumb vive en la misma carpeta con prefijo 'thumb_'
    folder, filename = os.path.split(original_path)
    thumb_path = os.path.join(folder, f"thumb_{filename}")
    try:
        if os.path.exists(thumb_path):
            os.remove(thumb_path)
    except Exception as e:
        current_app.logger.error(f"Error eliminando thumbnail: {e}")


def _mask_pii(value: str, visible: int = 4) -> str:
    """Oculta los caracteres centrales de un campo PII, dejando `visible` al inicio y al final."""
    if not value or len(value) <= visible * 2:
        return value
    return value[:visible] + '*' * (len(value) - visible * 2) + value[-visible:]


@bp.route('/get/<int:id>')
@login_required
def get_trabajador(id):
    from flask import session
    t = (Trabajador.query
         .options(
             selectinload(Trabajador.credenciales),
             selectinload(Trabajador.documentos),
             # proyectos usa lazy='dynamic' (query object), no es compatible con selectinload
         )
         .filter_by(id=id)
         .first_or_404())

    credenciales = [{'planta': c.planta, 'credencial_id': c.credencial_id, 'fecha_caducidad': c.fecha_caducidad.isoformat() if c.fecha_caducidad else None} for c in t.credenciales]
    documentos = [d.to_dict() for d in t.documentos]

    # Extraer coordinadores de los proyectos activos asignados
    # joinedload evita N+1: carga coordinador en el mismo query en lugar de 1 query por proyecto
    from sqlalchemy.orm import joinedload as _jl
    proyectos_activos = (t.proyectos
        .options(_jl(Proyecto.coordinador))
        .filter_by(activo=True)
        .all())
    coordinadores_set = set()
    for p in proyectos_activos:
        if p.coordinador:
            coordinadores_set.add(p.coordinador.full_name or p.coordinador.username)

    coordinadores_asignados = ", ".join(coordinadores_set) if coordinadores_set else "Ninguno asignado"

    # Solo admin y super_admin ven PII completa; otros roles reciben datos enmascarados
    is_admin = session.get('role') in ('admin', 'super_admin')
    curp = t.curp or ''
    rfc  = t.rfc  or ''
    nss  = t.nss  or ''
    if not is_admin:
        curp = _mask_pii(curp)
        rfc  = _mask_pii(rfc)
        nss  = _mask_pii(nss)

    data = {
        'id': t.id,
        'no_empleado': t.no_empleado,
        'nombre_apellidos': t.nombre_apellidos,
        'nombre': t.nombre,

        # Laborales
        'tipo_mov': t.tipo_mov or '',
        'tipo_cont': t.tipo_cont or '',
        'area': t.area or '',
        'puesto': t.puesto or '',
        'tipo_jornada': t.tipo_jornada or '',
        'fecha_ingreso': t.fecha_ingreso.isoformat() if t.fecha_ingreso else '',
        'descripcion_servicio': t.descripcion_servicio or '',
        'inicio': t.inicio.isoformat() if t.inicio else '',
        'termino_prueba': t.termino_prueba.isoformat() if t.termino_prueba else '',
        'fecha_baja': t.fecha_baja.isoformat() if t.fecha_baja else '',

        # Generales
        'curp': curp,
        'rfc': rfc,
        'nss': nss,
        'fecha_nacimiento': t.fecha_nacimiento.isoformat() if t.fecha_nacimiento else '',
        'sexo': t.sexo or '',
        'estado_civil': t.estado_civil or '',
        'nacionalidad': t.nacionalidad or '',
        'edad': t.edad or '',
        'domicilio': t.domicilio or '',
        
        # Contacto
        'correo': t.correo or '',
        'celular': t.celular or '',
        
        # Medicos
        'tipo_sangre': t.tipo_sangre or '',
        'alergias': t.alergias or '',
        'enfermedades_cronicas': t.enfermedades_cronicas or '',
        'contacto_emergencia': t.contacto_emergencia or '',
        'parentesco_contacto': t.parentesco_contacto or '',
        'numero_contacto_emerg': t.numero_contacto_emerg or '',
        'lentes': t.lentes or '',
        'licencia_conducir': t.licencia_conducir or '',
        'estatura': t.estatura or '',
        
        # Finanzas
        'salario_real_pactado_x_sem': str(t.salario_real_pactado_x_sem) if t.salario_real_pactado_x_sem else '',
        'tipo_pago': t.tipo_pago or '',
        'tipo_nomina': t.tipo_nomina or '',
        'sb': str(t.sb) if t.sb else '',
        'sdi': str(t.sdi) if t.sdi else '',
        'letra': t.letra or '',
        'hr_extra': str(t.hr_extra) if t.hr_extra else '',
        'infonavit': str(t.infonavit) if t.infonavit else '',
        'ajuste_inbursa': str(t.ajuste_inbursa) if t.ajuste_inbursa else '',
        'caja_ahorro': str(t.caja_ahorro) if t.caja_ahorro else '',
        'viaticos': str(t.viaticos) if t.viaticos else '',
        'pago_dia_festivo': str(t.pago_dia_festivo) if t.pago_dia_festivo else '',
        'pagos_efectivo': str(t.pagos_efectivo) if t.pagos_efectivo else '',
        'folio_mov_idse': t.folio_mov_idse or '',
        
        # Operacion
        'ubicacion_actual': t.ubicacion_actual or '',
        'ubicacion_estado': t.ubicacion_estado or '',
        'coordinadores_actuales': coordinadores_asignados,
        'no_proyecto': t.no_proyecto or '',
        'observaciones': t.observaciones or '',
        
        'foto_perfil': t.foto_perfil or '',
        'credenciales': credenciales,
        'documentos': documentos
    }
    return jsonify(data)

@bp.route('/editar/<int:id>', methods=['POST'])
@login_required
def editar(id):
    try:
        t = Trabajador.query.get_or_404(id)

        if not is_authorized_for_worker(t):
            flash('Acceso denegado.', 'danger')
            return redirect(url_for('trabajadores.index'))

        data = request.form

        # Check uniqueness of no_empleado if it changed
        new_no = data.get('no_empleado')
        if new_no != t.no_empleado and Trabajador.query.filter_by(no_empleado=new_no).first():
            flash('Error: El Número de Empleado ya existe.', 'danger')
            return redirect(url_for('trabajadores.index'))
            
        errores_longitud = validate_lengths(data)
        if errores_longitud:
            flash("Error de longitud en los campos:<br>" + "<br>".join(errores_longitud), 'danger')
            return redirect(url_for('trabajadores.index'))

        # Validación de formato CURP y RFC (advertencia, no bloqueo, por registros históricos)
        _curp_val = (data.get('curp') or '').strip().upper()
        _rfc_val  = (data.get('rfc') or '').strip().upper()
        if _curp_val and not _CURP_RE.match(_curp_val):
            flash('Advertencia: El formato de CURP no es válido. Verifica antes de continuar.', 'warning')
        if _rfc_val and not _RFC_RE.match(_rfc_val):
            flash('Advertencia: El formato de RFC no es válido. Verifica antes de continuar.', 'warning')

        t.no_empleado = new_no
        t.nombre_apellidos = data.get('nombre_apellidos')
        t.nombre = data.get('nombre')
        
        # Laborales
        t.tipo_mov = data.get('tipo_mov')
        t.tipo_cont = data.get('tipo_cont')
        t.area = data.get('area')
        t.puesto = data.get('puesto')
        t.tipo_jornada = data.get('tipo_jornada')
        t.fecha_ingreso = _parse_date(data.get('fecha_ingreso'))
        t.descripcion_servicio = data.get('descripcion_servicio')
        t.inicio = _parse_date(data.get('inicio'))
        t.termino_prueba = _parse_date(data.get('termino_prueba'))
        t.fecha_baja = _parse_date(data.get('fecha_baja'))
        
        # Generales
        t.curp = data.get('curp')
        t.rfc = data.get('rfc')
        t.nss = data.get('nss')
        t.fecha_nacimiento = _parse_date(data.get('fecha_nacimiento'))
        t.sexo = data.get('sexo')
        t.estado_civil = data.get('estado_civil')
        t.nacionalidad = data.get('nacionalidad')
        t.edad = data.get('edad') or None
        t.domicilio = data.get('domicilio')
        
        # Contacto
        t.correo = data.get('correo')
        t.celular = data.get('celular')
        
        # Medicos
        t.tipo_sangre = data.get('tipo_sangre')
        t.alergias = data.get('alergias')
        t.enfermedades_cronicas = data.get('enfermedades_cronicas')
        t.contacto_emergencia = data.get('contacto_emergencia')
        t.parentesco_contacto = data.get('parentesco_contacto')
        t.numero_contacto_emerg = data.get('numero_contacto_emerg')
        t.lentes = data.get('lentes')
        t.licencia_conducir = data.get('licencia_conducir')
        t.estatura = data.get('estatura')
        
        # Finanzas
        t.salario_real_pactado_x_sem = data.get('salario_real_pactado_x_sem') or 0
        t.tipo_pago = data.get('tipo_pago')
        t.tipo_nomina = data.get('tipo_nomina')
        t.sb = data.get('sb') or 0
        t.sdi = data.get('sdi') or 0
        t.letra = data.get('letra')
        t.hr_extra = data.get('hr_extra') or 0
        t.infonavit = data.get('infonavit') or 0
        t.ajuste_inbursa = data.get('ajuste_inbursa') or 0
        t.caja_ahorro = data.get('caja_ahorro') or 0
        t.viaticos = data.get('viaticos') or 0
        t.pago_dia_festivo = data.get('pago_dia_festivo') or 0
        t.pagos_efectivo = data.get('pagos_efectivo') or 0
        t.folio_mov_idse = data.get('folio_mov_idse')
        
        # Operacion (no_proyecto, ubicacion_actual, coord_a_cargo se manejan en Proyectos)
        t.ubicacion_estado = data.get('ubicacion_estado')
        t.observaciones = data.get('observaciones')
        
        # Handle Profile Picture Edit (solo JPG/PNG reales)
        if 'foto_perfil' in request.files:
            file = request.files['foto_perfil']
            if file and file.filename != '':
                if not allowed_image_file(file):
                    flash('Foto de perfil rechazada: solo se permiten imágenes JPG o PNG reales.', 'danger')
                    return redirect(url_for('trabajadores.index'))
                if t.foto_perfil:
                    _delete_profile_picture(t.foto_perfil)
                unique_filename = f"pp_{int(time.time())}_{secure_filename(file.filename)}"
                pp_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'perfiles')
                t.foto_perfil = _save_profile_picture(file, pp_folder, unique_filename)
        
        # Update credentials
        try:
            credenciales_str = data.get('credenciales_json', '[]')
            credenciales_data = json.loads(credenciales_str)
            # Remove old credentials
            t.credenciales = []
            
            for c_data in credenciales_data:
                planta = str(c_data.get('planta', '')).strip().upper()
                credencial_id = str(c_data.get('credencial_id', '')).strip()
                
                if len(planta) > 100 or len(credencial_id) > 40:
                    raise ValueError(f"Longitud inválida en credencial ({len(planta)}/100, {len(credencial_id)}/40)")
                
                nueva_credencial = CredencialPlanta(
                    planta=planta,
                    credencial_id=credencial_id,
                    fecha_caducidad=_parse_date(c_data.get('fecha_caducidad'))
                )
                t.credenciales.append(nueva_credencial)
        except ValueError as val_err:
            current_app.logger.error(f"Validation error in credentials: {val_err}")
            flash(f'Error validando credenciales: {val_err}', 'danger')
            return redirect(url_for('trabajadores.index'))
        except Exception as json_err:
            current_app.logger.error(f"Error parsing credentials on edit: {json_err}")
            flash('Hubo un problema actualizando algunas credenciales.', 'warning')
            
        db.session.commit()
        log_action(f"Actualizó al trabajador {t.nombre} ({t.no_empleado})")
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': 'Trabajador actualizado exitosamente.', 'id': t.id})
            
        flash('Trabajador actualizado exitosamente.', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating worker: {e}\n{traceback.format_exc()}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'Ocurrió un error al actualizar el trabajador.'}), 500
        flash('Ocurrió un error al actualizar el trabajador.', 'danger')
        
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': False, 'message': 'Operación fallida.'}), 400
    return redirect(url_for('trabajadores.index'))

@bp.route('/credenciales/<int:id>', methods=['POST'])
@login_required
@admin_required
def guardar_credenciales(id):
    try:
        t = Trabajador.query.get_or_404(id)
        data = request.form
        
        # Update observaciones
        if 'observaciones' in data:
            t.observaciones = data.get('observaciones')

        # Update credentials
        credenciales_str = data.get('credenciales_json', '[]')
        credenciales_data = json.loads(credenciales_str)
        
        # Remove old credentials
        t.credenciales = []
        
        for c_data in credenciales_data:
            planta = str(c_data.get('planta', '')).strip().upper()
            credencial_id = str(c_data.get('credencial_id', '')).strip()
            
            if len(planta) > 100 or len(credencial_id) > 40:
                flash(f"Error: Longitud inválida en credencial. Planta max 100 (recibidos: {len(planta)}), ID max 40 (recibidos: {len(credencial_id)}).", 'danger')
                return redirect(url_for('trabajadores.index'))
            
            nueva_credencial = CredencialPlanta(
                planta=planta,
                credencial_id=credencial_id,
                fecha_caducidad=_parse_date(c_data.get('fecha_caducidad'))
            )
            t.credenciales.append(nueva_credencial)
            
        db.session.commit()
        log_action(f"Actualizó las credenciales del trabajador {t.nombre} ({t.no_empleado})")
        return jsonify({'success': True, 'message': 'Credenciales actualizadas correctamente.'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating credentials: {e}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'error': 'Ocurrió un error al actualizar las credenciales.'}), 500

@bp.route('/eliminar/<int:id>', methods=['POST'])
@login_required
@admin_required
def eliminar(id):
    try:
        from datetime import date
        t = Trabajador.query.get_or_404(id)
        nombre_trabajador = t.nombre
        no_emp = t.no_empleado
        
        # Soft delete
        t.activo = False
        t.fecha_baja = date.today()
        
        db.session.commit()
        log_action(f"Dio de baja al trabajador {nombre_trabajador} ({no_emp})")
        flash('Trabajador dado de baja exitosamente.', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error inactivating worker: {e}\n{traceback.format_exc()}")
        flash('Ocurrió un error al dar de baja el trabajador.', 'danger')
        
    return redirect(url_for('trabajadores.index'))

@bp.route('/reactivar/<int:id>', methods=['POST'])
@login_required
@admin_required
def reactivar(id):
    try:
        t = Trabajador.query.get_or_404(id)
        nombre_trabajador = t.nombre
        no_emp = t.no_empleado
        
        t.activo = True
        t.fecha_baja = None
        
        db.session.commit()
        log_action(f"Reactivó al trabajador {nombre_trabajador} ({no_emp})")
        flash('Trabajador reactivado exitosamente.', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error reactivating worker: {e}\n{traceback.format_exc()}")
        flash('Ocurrió un error al reactivar el trabajador.', 'danger')
        
    return redirect(url_for('trabajadores.bajas'))


@bp.route('/exportar/<int:id>', methods=['GET'])
@login_required
def exportar_excel(id):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import date as date_type

    t = (Trabajador.query
         .options(selectinload(Trabajador.credenciales))
         .filter_by(id=id).first_or_404())

    is_admin = session.get('role') in ('admin', 'super_admin')

    def fmt_date(d):
        return d.strftime('%d/%m/%Y') if d else ''

    def fmt_money(v):
        return float(v) if v else ''

    def mask(val):
        if not val or is_admin:
            return val or ''
        return val[:3] + '***' + val[-2:] if len(val) > 5 else '***'

    wb = Workbook()
    ws = wb.active
    ws.title = "Empleado"

    AZUL_OSC = PatternFill("solid", fgColor="1E3A5F")
    AZUL_HDR = PatternFill("solid", fgColor="2563EB")
    BLANCO   = PatternFill("solid", fgColor="FFFFFF")

    thin  = Side(style="thin",   color="CBD5E1")
    thick = Side(style="medium", color="1E3A5F")
    BORDER     = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
    BORDER_HDR = Border(left=thick, right=thick, top=thick, bottom=thick)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    HEADERS = [
        "No. Empleado", "Nombre(s)", "Apellidos", "RFC", "CURP", "NSS",
        "Fecha Nacimiento", "Sexo", "Estado Civil", "Nacionalidad", "Edad",
        "Correo", "Celular", "Domicilio",
        "Área", "Puesto", "Tipo Movimiento", "Tipo Contrato",
        "Fecha Ingreso", "Tipo Jornada", "Tipo Nómina", "Tipo Pago",
        "Folio IDSE", "Desc. Servicio",
        "Tipo Sangre", "Lentes", "Alergias", "Estatura", "Enf. Crónicas", "Licencia Conducir",
        "Contacto Emergencia", "Parentesco", "Tel. Emergencia",
        "Salario Pactado/Sem", "Salario Base", "SDI", "Letra/Categoría",
        "Hrs Extra", "Infonavit", "Caja Ahorro", "Viáticos", "Día Festivo", "Pagos Efectivo",
        "Ubicación Actual", "Estado Ubic.", "No. Proyecto", "Coord. a Cargo", "Observaciones",
    ]

    # ── Título ────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 36
    ws.merge_cells(f'A1:{get_column_letter(len(HEADERS))}1')
    c = ws['A1']
    c.value = f"{t.nombre} {t.nombre_apellidos}".upper() + f"  —  No. {t.no_empleado}  —  Generado: {date_type.today().strftime('%d/%m/%Y')}"
    c.fill = AZUL_OSC
    c.font = Font(name="Calibri", color="FFFFFF", bold=True, size=13)
    c.alignment = CENTER
    c.border = BORDER_HDR

    # ── Cabecera ─────────────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 22
    for col_idx, header in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=col_idx, value=header)
        c.fill = AZUL_HDR
        c.font = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
        c.alignment = CENTER
        c.border = BORDER_HDR

    # ── Fila de datos ─────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 18
    values = [
        t.no_empleado, t.nombre, t.nombre_apellidos,
        mask(t.rfc), mask(t.curp), mask(t.nss),
        fmt_date(t.fecha_nacimiento), t.sexo, t.estado_civil, t.nacionalidad,
        t.edad, t.correo, t.celular, t.domicilio,
        t.area, t.puesto, t.tipo_mov, t.tipo_cont,
        fmt_date(t.fecha_ingreso), t.tipo_jornada, t.tipo_nomina, t.tipo_pago,
        t.folio_mov_idse, t.descripcion_servicio,
        t.tipo_sangre, t.lentes, t.alergias, t.estatura,
        t.enfermedades_cronicas, t.licencia_conducir,
        t.contacto_emergencia, t.parentesco_contacto, t.numero_contacto_emerg,
        fmt_money(t.salario_real_pactado_x_sem), fmt_money(t.sb), fmt_money(t.sdi),
        t.letra, fmt_money(t.hr_extra), fmt_money(t.infonavit),
        fmt_money(t.caja_ahorro), fmt_money(t.viaticos),
        fmt_money(t.pago_dia_festivo), fmt_money(t.pagos_efectivo),
        t.ubicacion_actual, t.ubicacion_estado, t.no_proyecto, t.coord_a_cargo,
        t.observaciones,
    ]
    for col_idx, val in enumerate(values, start=1):
        c = ws.cell(row=3, column=col_idx, value=val if val is not None else '')
        c.fill = BLANCO
        c.font = Font(name="Calibri", color="374151", size=10)
        c.alignment = LEFT
        c.border = BORDER

    # ── Credenciales (hoja aparte si las hay) ─────────────────────────────────
    if t.credenciales:
        ws2 = wb.create_sheet("Credenciales")
        ws2.row_dimensions[1].height = 22
        for col_idx, label in enumerate(['Planta', 'ID Credencial', 'Caducidad', 'Estado'], start=1):
            c = ws2.cell(row=1, column=col_idx, value=label)
            c.fill = AZUL_HDR
            c.font = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
            c.alignment = CENTER
            c.border = BORDER_HDR
        today = date_type.today()
        VERDE_CLR = PatternFill("solid", fgColor="D1FAE5")
        ROJO_CLR  = PatternFill("solid", fgColor="FEE2E2")
        for i, cred in enumerate(t.credenciales, start=2):
            caducidad = cred.fecha_caducidad
            cad_str = caducidad.strftime('%d/%m/%Y') if caducidad else ''
            vencida = caducidad and caducidad < today
            estado = "CADUCADA" if vencida else "VIGENTE"
            row_fill = ROJO_CLR if vencida else VERDE_CLR
            for col_idx, val in enumerate([cred.planta, cred.credencial_id, cad_str, estado], start=1):
                c = ws2.cell(row=i, column=col_idx, value=val or '')
                c.fill = row_fill
                c.font = Font(name="Calibri", color="7F1D1D" if vencida else "065F46", size=10)
                c.alignment = LEFT
                c.border = BORDER
        for col_idx, w in enumerate([20, 20, 14, 12], start=1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Anchos ───────────────────────────────────────────────────────────────
    col_widths = [14,18,20,14,20,14,14,8,14,14,6,28,14,32,
                  16,20,16,16,14,14,14,12,12,20,
                  12,8,16,10,16,16,
                  22,14,14,
                  16,14,12,14,12,12,12,12,12,14,
                  20,16,14,20,24]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = f"{t.no_empleado}_{(t.nombre_apellidos or '').replace(' ', '_')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=safe_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@bp.route('/exportar-todos', methods=['GET'])
@login_required
def exportar_todos():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import date as date_type

    is_admin = session.get('role') in ('admin', 'super_admin')
    trabajadores = (Trabajador.query
                    .filter(Trabajador.activo == True, Trabajador.fecha_baja == None)
                    .order_by(func.lower(Trabajador.nombre))
                    .all())

    def fmt_date(d):
        return d.strftime('%d/%m/%Y') if d else ''

    def fmt_money(v):
        return float(v) if v else ''

    def mask(val):
        if not val or is_admin:
            return val or ''
        return val[:3] + '***' + val[-2:] if len(val) > 5 else '***'

    wb = Workbook()
    ws = wb.active
    ws.title = "Empleados"

    AZUL_OSC = PatternFill("solid", fgColor="1E3A5F")
    AZUL_HDR = PatternFill("solid", fgColor="2563EB")
    GRIS_ALT = PatternFill("solid", fgColor="F1F5F9")
    BLANCO   = PatternFill("solid", fgColor="FFFFFF")

    thin  = Side(style="thin",   color="CBD5E1")
    thick = Side(style="medium", color="1E3A5F")
    BORDER     = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
    BORDER_HDR = Border(left=thick, right=thick, top=thick, bottom=thick)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    HEADERS = [
        "No. Empleado", "Nombre(s)", "Apellidos", "RFC", "CURP", "NSS",
        "Fecha Nacimiento", "Sexo", "Estado Civil", "Nacionalidad", "Edad",
        "Correo", "Celular", "Domicilio",
        "Área", "Puesto", "Tipo Movimiento", "Tipo Contrato",
        "Fecha Ingreso", "Tipo Jornada", "Tipo Nómina", "Tipo Pago",
        "Folio IDSE", "Desc. Servicio",
        "Tipo Sangre", "Lentes", "Alergias", "Estatura", "Enf. Crónicas", "Licencia Conducir",
        "Contacto Emergencia", "Parentesco", "Tel. Emergencia",
        "Salario Pactado/Sem", "Salario Base", "SDI", "Letra/Categoría",
        "Hrs Extra", "Infonavit", "Caja Ahorro", "Viáticos", "Día Festivo", "Pagos Efectivo",
        "Ubicación Actual", "Estado Ubic.", "No. Proyecto", "Coord. a Cargo",
    ]

    # ── Fila de título ────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 38
    ws.merge_cells(f'A1:{get_column_letter(len(HEADERS))}1')
    c = ws['A1']
    c.value = f"LISTADO DE EMPLEADOS — Generado: {date_type.today().strftime('%d/%m/%Y')}"
    c.fill = AZUL_OSC
    c.font = Font(name="Calibri", color="FFFFFF", bold=True, size=14)
    c.alignment = CENTER
    c.border = BORDER_HDR

    # ── Cabecera de columnas ──────────────────────────────────────────────────
    ws.row_dimensions[2].height = 22
    for col_idx, header in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=col_idx, value=header)
        c.fill = AZUL_HDR
        c.font = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
        c.alignment = CENTER
        c.border = BORDER_HDR

    # ── Datos ─────────────────────────────────────────────────────────────────
    for row_idx, t in enumerate(trabajadores, start=3):
        alt = (row_idx % 2 == 0)
        fill = GRIS_ALT if alt else BLANCO
        ws.row_dimensions[row_idx].height = 16

        values = [
            t.no_empleado, t.nombre, t.nombre_apellidos,
            mask(t.rfc), mask(t.curp), mask(t.nss),
            fmt_date(t.fecha_nacimiento), t.sexo, t.estado_civil, t.nacionalidad,
            t.edad, t.correo, t.celular, t.domicilio,
            t.area, t.puesto, t.tipo_mov, t.tipo_cont,
            fmt_date(t.fecha_ingreso), t.tipo_jornada, t.tipo_nomina, t.tipo_pago,
            t.folio_mov_idse, t.descripcion_servicio,
            t.tipo_sangre, t.lentes, t.alergias, t.estatura,
            t.enfermedades_cronicas, t.licencia_conducir,
            t.contacto_emergencia, t.parentesco_contacto, t.numero_contacto_emerg,
            fmt_money(t.salario_real_pactado_x_sem), fmt_money(t.sb), fmt_money(t.sdi),
            t.letra, fmt_money(t.hr_extra), fmt_money(t.infonavit),
            fmt_money(t.caja_ahorro), fmt_money(t.viaticos),
            fmt_money(t.pago_dia_festivo), fmt_money(t.pagos_efectivo),
            t.ubicacion_actual, t.ubicacion_estado, t.no_proyecto, t.coord_a_cargo,
        ]

        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val if val is not None else '')
            c.fill = fill
            c.font = Font(name="Calibri", color="374151", size=10)
            c.alignment = LEFT
            c.border = BORDER

    # ── Anchos automáticos (estimados) ───────────────────────────────────────
    col_widths = [14,18,20,14,20,14,14,8,14,14,6,28,14,32,
                  16,20,16,16,14,14,14,12,12,20,
                  12,8,16,10,16,16,
                  22,14,14,
                  16,14,12,14,12,12,12,12,12,14,
                  20,16,14,20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"empleados_{date_type.today().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@bp.route('/<int:id>/documentos', methods=['POST'])
@login_required
@limiter.limit("10 per minute")
def upload_documento(id):
    t = Trabajador.query.get_or_404(id)
    
    if 'documento' not in request.files:
        return jsonify({'error': 'No file part'}), 400
        
    file = request.files['documento']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
        
    if file:
        if not allowed_file(file):
            return jsonify({'error': 'File format not allowed or file is corrupt'}), 400
        filename = secure_filename(file.filename)
        # Create a unique filename to prevent overwrites
        unique_filename = f"{int(time.time())}_{filename}"
        
        # Save path: uploads/trabajadores/<id>/
        upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'trabajadores', str(t.id))
        os.makedirs(upload_folder, exist_ok=True)
        
        file_path = os.path.join(upload_folder, unique_filename)
        file.save(file_path)
        
        # Parse Dates
        fecha_inicio_str = request.form.get('fecha_inicio')
        fecha_fin_str = request.form.get('fecha_fin')
        f_inicio = _parse_date(fecha_inicio_str) if fecha_inicio_str else None
        f_fin = _parse_date(fecha_fin_str) if fecha_fin_str else None
        
        # Parse optional tipo_documento (allow custom types up to 100 chars)
        tipo_doc = request.form.get('tipo_documento', '').strip() or None
        if tipo_doc and len(tipo_doc) > 100:
            return jsonify({'error': 'El tipo de documento es demasiado largo (máximo 100 caracteres)'}), 400
        
        # Relate to DB path
        db_path = f"trabajadores/{t.id}/{unique_filename}"
        
        nuevo_doc = DocumentoTrabajador(
            trabajador_id=t.id,
            nombre_archivo=filename,
            ruta_archivo=db_path,
            tipo_documento=tipo_doc,
            fecha_inicio=f_inicio,
            fecha_fin=f_fin
        )
        db.session.add(nuevo_doc)
        db.session.commit()
        log_action(f"Subió documento {filename} para {t.nombre} ({t.no_empleado})")
        
        return jsonify(nuevo_doc.to_dict()), 201

@bp.route('/documento/<int:doc_id>', methods=['GET'])
@login_required
def get_documento(doc_id):
    doc = DocumentoTrabajador.query.get_or_404(doc_id)
    t = doc.trabajador
    if not is_authorized_for_worker(t):
        return jsonify({'error': 'Acceso denegado'}), 403
    return send_from_directory(
        current_app.config['UPLOAD_FOLDER'],
        doc.ruta_archivo,
        as_attachment=True,
    )

@bp.route('/foto/<int:id>', methods=['GET'])
@login_required
def get_foto_perfil(id):
    t = Trabajador.query.get_or_404(id)
    if not is_authorized_for_worker(t):
        return jsonify({'error': 'Acceso denegado'}), 403
    if not t.foto_perfil:
        return jsonify({'error': 'No encontrado'}), 404
    return send_from_directory(current_app.config['UPLOAD_FOLDER'], t.foto_perfil)


@bp.route('/foto/<int:id>/thumb', methods=['GET'])
@login_required
def get_foto_perfil_thumb(id):
    t = Trabajador.query.get_or_404(id)
    if not is_authorized_for_worker(t):
        return jsonify({'error': 'Acceso denegado'}), 403
    if not t.foto_perfil:
        return jsonify({'error': 'No encontrado'}), 404
    folder, filename = os.path.split(t.foto_perfil)
    thumb_rel = os.path.join(folder, f"thumb_{filename}").replace('\\', '/')
    upload_folder = current_app.config['UPLOAD_FOLDER']
    if os.path.exists(os.path.join(upload_folder, thumb_rel)):
        return send_from_directory(upload_folder, thumb_rel)
    # Fallback a la imagen original si el thumbnail no existe (fotos pre-migración)
    return send_from_directory(upload_folder, t.foto_perfil)
    
@bp.route('/documento/<int:doc_id>', methods=['DELETE'])
@login_required
def delete_documento(doc_id):
    doc = DocumentoTrabajador.query.get_or_404(doc_id)
    t = doc.trabajador
    if not is_authorized_for_worker(t):
        return jsonify({'error': 'Acceso denegado'}), 403
        
    file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], doc.ruta_archivo)
    
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except Exception as e:
        current_app.logger.error(f"Error physically deleting file {file_path}: {e}")
        
    log_action(f"Eliminó documento {doc.nombre_archivo} del trabajador ID {doc.trabajador_id}")
    db.session.delete(doc)
    db.session.commit()
    
    return jsonify({'success': True}), 200
