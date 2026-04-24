from flask import Blueprint, render_template, session, redirect, url_for, flash, jsonify, make_response, current_app, request, send_file
from app.utils import login_required, log_action
from app.models import User, Estante, Proyecto, SolicitudMaterial, Producto
from app.extensions import db, limiter
import io
import os
import traceback

bp = Blueprint('inventario_ui', __name__, url_prefix='/inventario')

@bp.route('/movil')
@login_required
def movil():
    user = User.query.get(session.get('user_id'))
    if user.role not in ['inventario', 'admin']:
        flash('No tienes permiso para ver esta página.', 'danger')
        return redirect(url_for('main.home'))
    return render_template('inventario_movil.html', user=user)

@bp.route('/web')
@login_required
def web():
    user = User.query.get(session.get('user_id'))
    if user.role not in ['inventario', 'admin']:
        flash('No tienes permiso para ver esta página.', 'danger')
        return redirect(url_for('main.home'))
    # Redirigir a catálogo por defecto si alguien entra a /web
    return redirect(url_for('inventario_ui.catalogo'))

@bp.route('/catalogo')
@login_required
def catalogo():
    user = User.query.get(session.get('user_id'))
    if user.role not in ['inventario', 'admin']:
        flash('No tienes permiso.', 'danger')
        return redirect(url_for('main.home'))
    return render_template('inventario_catalogo.html', user=user)

@bp.route('/catalogo/<categoria>')
@login_required
def catalogo_categoria(categoria):
    user = User.query.get(session.get('user_id'))
    if user.role not in ['inventario', 'admin']:
        flash('No tienes permiso.', 'danger')
        return redirect(url_for('main.home'))
    return render_template('inventario_categoria.html', user=user, categoria=categoria)

@bp.route('/estantes')
@login_required
def estantes():
    user = User.query.get(session.get('user_id'))
    if user.role not in ['inventario', 'admin']:
        flash('No tienes permiso.', 'danger')
        return redirect(url_for('main.home'))
    return render_template('inventario_estantes.html', user=user)

@bp.route('/qr/estante/<int:estante_id>')
@login_required
def qr_estante(estante_id):
    """Página de impresión del QR de un estante."""
    user = User.query.get(session.get('user_id'))
    if user.role not in ['inventario', 'admin']:
        flash('No tienes permiso.', 'danger')
        return redirect(url_for('main.home'))
    estante = Estante.query.get_or_404(estante_id)
    return render_template('inventario_qr_print.html', estante=estante)

@bp.route('/solicitar')
@login_required
def solicitar():
    """Formulario para que los solicitantes pidan material."""
    user = User.query.get(session.get('user_id'))
    if user.role not in ['solicitante_material', 'admin', 'inventario']:
        flash('No tienes permiso para solicitar material.', 'danger')
        return redirect(url_for('main.home'))
    return render_template('inventario_solicitar.html', user=user)

@bp.route('/api/proyectos')
@login_required
def api_proyectos():
    proyectos = Proyecto.query.filter_by(activo=True).order_by(Proyecto.numero_proyecto).all()
    return jsonify([{
        'id': p.id,
        'numero_proyecto': p.numero_proyecto,
        'nombre': p.nombre or ''
    } for p in proyectos])

@bp.route('/solicitudes')
@login_required
def solicitudes():
    """Pantalla para gestionar solicitudes (aprobar/denegar)."""
    user = User.query.get(session.get('user_id'))
    if user.role not in ['inventario', 'admin']:
        flash('Acceso restringido a personal de inventario.', 'danger')
        return redirect(url_for('main.home'))
    return render_template('inventario_solicitudes.html', user=user)

@bp.route('/solicitudes/<int:solicitud_id>/pdf')
@login_required
def solicitud_pdf(solicitud_id):
    from xhtml2pdf import pisa
    import os

    s = SolicitudMaterial.query.get_or_404(solicitud_id)
    fecha = s.fecha_creacion.strftime('%d/%m/%Y %H:%M') if s.fecha_creacion else '—'
    estatus = s.estatus or 'PENDIENTE'

    logo_path = os.path.join(current_app.root_path, '..', 'static', 'imagenes', 'skilled (1).png')
    logo_path = os.path.normpath(logo_path)

    STATUS_COLORS = {
        'PENDIENTE': ('#92400E', '#D97706', '#FFFBEB'),
        'APROBADA':  ('#065F46', '#10B981', '#ECFDF5'),
        'RECHAZADA': ('#991B1B', '#EF4444', '#FEF2F2'),
        'ENTREGADA': ('#1E40AF', '#3B82F6', '#EFF6FF'),
    }
    sc, sb, sbg = STATUS_COLORS.get(estatus, ('#374151', '#6B7280', '#F9FAFB'))

    filas_html = ''.join(f"""
        <tr>
            <td class="td {'td-alt' if i%2==0 else ''}" align="center">{i+1}</td>
            <td class="td {'td-alt' if i%2==0 else ''}" style="font-weight:bold;">{d.producto.descripcion if d.producto else '—'}</td>
            <td class="td {'td-alt' if i%2==0 else ''}" style="font-size:10px;color:#6B7280;font-weight:bold;text-transform:uppercase;">{d.producto.codigo if d.producto else '—'}</td>
            <td class="td {'td-alt' if i%2==0 else ''}" align="center" style="font-weight:bold;font-size:13px;color:#4F46E5;">{int(d.cantidad_solicitada)}</td>
            <td class="td {'td-alt' if i%2==0 else ''}" style="color:#9CA3AF;">{d.producto.unidad if d.producto else 'pza'}</td>
        </tr>""" for i, d in enumerate(s.detalles))

    solicitante_nombre = s.solicitante.username if s.solicitante else '—'
    proyecto = s.proyecto or 'General'

    logo_uri = logo_path.replace('\\', '/')

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<style>
  body {{ font-family: Arial, sans-serif; color: #111; font-size: 12px; margin: 0; padding: 0; }}
  @page {{ margin: 2cm 2cm 2cm 2cm; size: letter; }}
  .lbl {{ font-size: 8px; font-weight: bold; color: #888; text-transform: uppercase; }}
  .val {{ font-size: 12px; font-weight: bold; color: #111; border-bottom: 1px solid #ddd; padding-bottom: 3px; }}
  .sec {{ font-size: 9px; font-weight: bold; color: #6B7280; text-transform: uppercase; letter-spacing: 1px; margin: 14px 0 6px 0; }}
  .th {{ background-color: #1f2937; color: #fff; font-size: 9px; font-weight: bold; text-transform: uppercase; padding: 8px 10px; text-align: left; }}
  .td {{ font-size: 11px; padding: 8px 10px; border-bottom: 1px solid #f1f5f9; }}
  .td-alt {{ background-color: #f9fafb; }}
  .sig-line {{ border-top: 1px solid #374151; padding-top: 6px; font-size: 9px; color: #6B7280; font-weight: bold; text-transform: uppercase; text-align: center; }}
  .status {{ font-size: 9px; font-weight: bold; text-transform: uppercase; padding: 3px 10px; border: 1px solid {sb}; color: {sc}; background-color: {sbg}; }}
</style>
</head>
<body>

<!-- HEADER -->
<table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:4px;">
  <tr>
    <td width="50%" valign="top">
      <img src="file:///{logo_uri}" height="46" alt="Skilled">
      <br>
      <span style="font-size:9px; color:#888; font-weight:bold; text-transform:uppercase;">Solicitud de Materiales</span>
    </td>
    <td width="50%" valign="top" align="right">
      <div style="font-size:16px; font-weight:bold; color:#111;">Solicitud #{s.id}</div>
      <div style="font-size:10px; color:#888; margin-top:3px;">{fecha}</div>
      <div style="margin-top:5px;"><span class="status">{estatus}</span></div>
    </td>
  </tr>
</table>

<hr style="border:none; border-top:2px solid #111; margin:10px 0 14px 0;">

<!-- INFO -->
<table width="100%" cellpadding="4" cellspacing="0" style="margin-bottom:14px;">
  <tr>
    <td width="50%" valign="top">
      <div class="lbl">Solicitante</div>
      <div class="val">{solicitante_nombre}</div>
    </td>
    <td width="50%" valign="top">
      <div class="lbl">Proyecto</div>
      <div class="val">{proyecto}</div>
    </td>
  </tr>
  <tr>
    <td width="50%" valign="top" style="padding-top:8px;">
      <div class="lbl">Total de materiales</div>
      <div class="val">{len(s.detalles)} material{'es' if len(s.detalles) != 1 else ''}</div>
    </td>
    <td width="50%" valign="top" style="padding-top:8px;">
      <div class="lbl">Estado</div>
      <div class="val">{estatus}</div>
    </td>
  </tr>
</table>

<div class="sec">Materiales solicitados</div>
<table width="100%" cellpadding="0" cellspacing="0" style="border: 1px solid #e5e7eb;">
  <thead>
    <tr>
      <th class="th" width="5%" align="center">#</th>
      <th class="th" width="45%">Descripción</th>
      <th class="th" width="25%">Código</th>
      <th class="th" width="15%" align="center">Cantidad</th>
      <th class="th" width="10%">Unidad</th>
    </tr>
  </thead>
  <tbody>
    {filas_html}
  </tbody>
</table>

<br><br>
<table width="100%" cellpadding="0" cellspacing="0" style="margin-top:32px;">
  <tr>
    <td width="33%" style="padding: 0 12px;"><div class="sig-line">Solicitante</div></td>
    <td width="33%" style="padding: 0 12px;"><div class="sig-line">Autorizado por</div></td>
    <td width="33%" style="padding: 0 12px;"><div class="sig-line">Entregado por</div></td>
  </tr>
</table>

</body>
</html>"""

    static_dir = os.path.normpath(os.path.join(current_app.root_path, '..', 'static'))

    def link_callback(uri, rel):
        if uri.startswith('file:///'):
            return uri[8:]
        if uri.startswith('/static/'):
            return os.path.join(static_dir, uri[8:])
        return uri

    buf = io.BytesIO()
    pisa.CreatePDF(html, dest=buf, link_callback=link_callback)
    buf.seek(0)
    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="Solicitud_{s.id}.pdf"'
    return response


@bp.route('/movimientos')
@login_required
def movimientos():
    user = User.query.get(session.get('user_id'))
    if user.role not in ['inventario', 'admin']:
        flash('No tienes permiso para ver esta página.', 'danger')
        return redirect(url_for('main.home'))
    from app.models import Producto, Almacen
    productos = Producto.query.filter_by(activo=True).order_by(Producto.descripcion).all()
    almacenes = Almacen.query.filter_by(activo=True).order_by(Almacen.nombre).all()
    return render_template('inventario_movimientos.html', user=user, productos=productos, almacenes=almacenes)

@bp.route('/mis-pedidos')
@login_required
def mis_pedidos():
    """Historial de pedidos del solicitante."""
    user = User.query.get(session.get('user_id'))
    if user.role not in ['solicitante_material', 'admin', 'inventario']:
        flash('No tienes permiso para ver esta página.', 'danger')
        return redirect(url_for('main.home'))
    return render_template('inventario_mis_pedidos.html', user=user)


# ─── IMPORTACIÓN MASIVA DE MATERIALES ────────────────────────────────────────

_CATS_BASE = [
    'Tornillería', 'Tuercas', 'Rondanas', 'Pijas',
    'Abrazaderas', 'Soportería', 'Tubería/Accesorios',
]


def _normalizar_categoria(val, categorias_disponibles):
    """Devuelve la categoría oficial más cercana usando coincidencia difusa."""
    from difflib import get_close_matches
    if not val:
        return val
    if val in categorias_disponibles:
        return val
    for c in categorias_disponibles:
        if c.lower() == val.lower():
            return c
    matches = get_close_matches(val, categorias_disponibles, n=1, cutoff=0.75)
    return matches[0] if matches else val


@bp.route('/importar')
@login_required
def importar_materiales():
    user = User.query.get(session.get('user_id'))
    if user.role not in ['inventario', 'admin']:
        flash('No tienes permiso.', 'danger')
        return redirect(url_for('main.home'))
    return render_template('importar_materiales.html', user=user)


@bp.route('/plantilla_materiales')
@login_required
def plantilla_materiales():
    """Genera y descarga la plantilla Excel para importar materiales."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = 'Materiales'

    columnas = ['Código (SKU)', 'Categoría', 'Descripción', 'Unidad', 'Stock Inicial', 'Stock Mínimo']

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    for col_idx, col_name in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[1].height = 30
    col_widths = [18, 24, 44, 14, 16, 16]
    for i, w in enumerate(col_widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w

    wrap_align = Alignment(wrap_text=True, vertical='center')
    for row in ws.iter_rows(min_row=2, max_row=500, min_col=1, max_col=len(columnas)):
        for cell in row:
            cell.alignment = wrap_align

    # Dropdown: Categoría
    cats_formula = '"' + ','.join(_CATS_BASE) + '"'
    dv_cat = DataValidation(type="list", formula1=cats_formula, allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_cat)
    dv_cat.add('B2:B500')

    # Dropdown: Unidad
    dv_unidad = DataValidation(
        type="list",
        formula1='"Pza,Kg,Lt,Mt,Caja,Rollo,Juego,Par,Bolsa,Tramo"',
        allow_blank=True, showErrorMessage=False,
    )
    ws.add_data_validation(dv_unidad)
    dv_unidad.add('D2:D500')

    ws.freeze_panes = 'A2'

    # Hoja de instrucciones
    ws2 = wb.create_sheet(title='Instrucciones')
    instrucciones = [
        ('A1', 'INSTRUCCIONES DE USO', True, 14),
        ('A3', '• Código (SKU): identificador único — obligatorio, sin duplicados.', False, 11),
        ('A4', '• Categoría: usa el desplegable o escribe el nombre; el sistema corregirá errores leves.', False, 11),
        ('A5', '• Descripción: nombre completo del material — obligatorio.', False, 11),
        ('A6', '• Unidad: Pza, Kg, Lt, Mt, Caja, etc. (usa el desplegable).', False, 11),
        ('A7', '• Stock Inicial: cantidad disponible al importar (número, puede ser 0).', False, 11),
        ('A8', '• Stock Mínimo: alerta de reorden cuando el stock baje de este valor.', False, 11),
        ('A10', '• No alteres ni renombres los encabezados de la fila 1.', False, 11),
        ('A11', '• Categorías base disponibles: ' + ', '.join(_CATS_BASE), False, 11),
    ]
    for cell_ref, text, bold, size in instrucciones:
        cell = ws2[cell_ref]
        cell.value = text
        cell.font = Font(bold=bold, size=size)
    ws2.column_dimensions['A'].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name='plantilla_materiales.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@bp.route('/procesar_importacion', methods=['POST'])
@login_required
@limiter.limit("5 per minute")
def procesar_importacion_materiales():
    import filetype as ft
    import pandas as pd

    user = User.query.get(session.get('user_id'))
    if user.role not in ['inventario', 'admin']:
        flash('No tienes permiso.', 'danger')
        return redirect(url_for('main.home'))

    if 'archivo_excel' not in request.files:
        flash('No se seleccionó ningún archivo.', 'danger')
        return redirect(url_for('inventario_ui.importar_materiales'))

    file = request.files['archivo_excel']
    if not file.filename:
        flash('El archivo está vacío.', 'danger')
        return redirect(url_for('inventario_ui.importar_materiales'))

    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        flash('Formato no válido. Debe ser .xlsx o .xls', 'danger')
        return redirect(url_for('inventario_ui.importar_materiales'))

    header = file.read(2048)
    file.seek(0)
    kind = ft.guess(header)
    if not kind or kind.extension not in ['xlsx', 'xls', 'zip', 'cfb']:
        flash('El archivo fue rechazado: no parece ser un documento Excel válido.', 'danger')
        return redirect(url_for('inventario_ui.importar_materiales'))

    file.seek(0, os.SEEK_END)
    size = file.tell()
    file.seek(0)
    if size > 5 * 1024 * 1024:
        flash('El archivo excede el tamaño máximo permitido (5 MB).', 'danger')
        return redirect(url_for('inventario_ui.importar_materiales'))

    # Categorías disponibles: base + las ya registradas en BD
    cats_db = [r[0] for r in db.session.query(Producto.categoria).distinct().all() if r[0]]
    categorias_disponibles = list(dict.fromkeys(_CATS_BASE + cats_db))

    try:
        df = pd.read_excel(file)
        errores = []
        correcciones = []
        exitosos = 0
        _FORMULA_CHARS = ('=', '+', '-', '@', '\t', '\r')

        for index, row in df.iterrows():
            row_errors = []

            def get_str(col_name, max_len=None, _row=row, _re=row_errors):
                val = _row.get(col_name, '')
                s = str(val).strip() if pd.notna(val) else ''
                if s and s[0] in _FORMULA_CHARS:
                    s = s.lstrip('=+-@\t\r')
                if max_len and len(s) > max_len:
                    _re.append(f"'{col_name}' excede {max_len} caracteres.")
                return s

            def get_num(col_name, _row=row, _re=row_errors):
                val = _row.get(col_name, 0)
                if pd.isna(val) or val == '':
                    return 0.0
                try:
                    return float(val)
                except (ValueError, TypeError):
                    _re.append(f"'{col_name}' no es un número válido.")
                    return 0.0

            codigo = get_str('Código (SKU)', 100)
            if not codigo or codigo == 'nan':
                errores.append(f"Fila {index+2}: Falta el Código (SKU).")
                continue

            if Producto.query.filter_by(codigo=codigo).first():
                errores.append(f"Fila {index+2}: El código '{codigo}' ya existe en el catálogo.")
                continue

            descripcion = get_str('Descripción', 250)
            if not descripcion or descripcion == 'nan':
                errores.append(f"Fila {index+2} ({codigo}): Falta la Descripción.")
                continue

            cat_raw = get_str('Categoría', 100)
            if cat_raw and cat_raw != 'nan':
                cat_norm = _normalizar_categoria(cat_raw, categorias_disponibles)
                if cat_norm != cat_raw:
                    correcciones.append(f"Fila {index+2}: Categoría «{cat_raw}» → «{cat_norm}»")
            else:
                cat_norm = 'Sin categoría'

            unidad = get_str('Unidad', 50)
            if not unidad or unidad == 'nan':
                unidad = 'Pza'

            if row_errors:
                errores.append(f"Fila {index+2} ({codigo}): " + "; ".join(row_errors))
                continue

            db.session.add(Producto(
                codigo=codigo,
                descripcion=descripcion,
                categoria=cat_norm,
                unidad=unidad,
                stock_actual=get_num('Stock Inicial'),
                stock_minimo=get_num('Stock Mínimo'),
                created_by_id=user.id,
            ))
            exitosos += 1

        if exitosos > 0:
            db.session.commit()
            log_action(f"Importó masivamente {exitosos} materiales desde Excel")
            flash(f'¡Éxito! Se importaron {exitosos} materiales correctamente.', 'success')

        if correcciones:
            msg = f"Se corrigieron automáticamente {len(correcciones)} categoría(s):<br>" + "<br>".join(correcciones[:10])
            flash(msg, 'info')

        if errores:
            msg = "Errores encontrados:<br>" + "<br>".join(errores[:10])
            if len(errores) > 10:
                msg += f"<br>...y {len(errores)-10} más."
            flash(msg, 'warning')

        if exitosos > 0 and not errores:
            return redirect(url_for('inventario_ui.catalogo'))

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error procesando Excel materiales: {e}\n{traceback.format_exc()}")
        flash('Error al leer el archivo Excel. Asegúrate de usar la plantilla correcta.', 'danger')

    return redirect(url_for('inventario_ui.importar_materiales'))
