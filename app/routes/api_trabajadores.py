"""API JSON para el módulo de Empleados (consumida por el SPA React).

Reusa el modelo `Trabajador` y los helpers internos del blueprint clásico
`trabajadores.py` (parser de fechas, guardado de foto, validaciones). La
autorización viene del JWT, no de la sesión: la función `_authorized` replica
`is_authorized_for_worker` leyendo el rol/usuario desde el JWT.
"""
import io
import json
import os
import time
import traceback
from datetime import date, datetime as dt

from flask import Blueprint, current_app, g, jsonify, request, send_file, send_from_directory
from sqlalchemy import func, or_
from sqlalchemy.orm import joinedload, selectinload
from werkzeug.utils import secure_filename

from app.extensions import db, limiter
from app.models import CredencialPlanta, DocumentoTrabajador, Proyecto, Trabajador
from app.realtime import emit_to_role
from app.routes.api_auth import jwt_required
from app.routes.trabajadores import (
    _CURP_RE,
    _RFC_RE,
    _delete_profile_picture,
    _mask_pii,
    _parse_date,
    _save_profile_picture,
)
from app.utils import (
    allowed_file,
    allowed_image_file,
    log_action,
    safe_excel_value,
    validate_lengths,
)

bp = Blueprint('api_trabajadores', __name__, url_prefix='/api/trabajadores')


# ── Helpers ────────────────────────────────────────────────────────────────────

def _u():
    """Atajo: usuario autenticado actual (puesto por jwt_required)."""
    return g._jwt_user


def _is_admin() -> bool:
    return _u().role in ('admin', 'super_admin')


def _authorized(t: Trabajador) -> bool:
    """Versión JWT de `trabajadores.is_authorized_for_worker`."""
    if _is_admin():
        return True
    uid = _u().id
    for p in t.proyectos.all():
        if p.activo and p.coordinador_id == uid:
            return True
    return False


def _row_summary(t: Trabajador) -> dict:
    """Subset usado por la lista (no manda PII completa, ni objetos relacionados)."""
    return {
        'id': t.id,
        'no_empleado': t.no_empleado,
        'nombre': t.nombre,
        'nombre_apellidos': t.nombre_apellidos,
        'area': t.area,
        'puesto': t.puesto,
        'tipo_nomina': t.tipo_nomina,
        'salario_real_pactado_x_sem': float(t.salario_real_pactado_x_sem) if t.salario_real_pactado_x_sem else None,
        'fecha_ingreso': t.fecha_ingreso.isoformat() if t.fecha_ingreso else None,
        'fecha_baja': t.fecha_baja.isoformat() if t.fecha_baja else None,
        'activo': t.activo,
        'foto_perfil': t.foto_perfil,
    }


def _full_detail(t: Trabajador) -> dict:
    """Detalle completo (idéntico al endpoint clásico /trabajadores/get/<id>)."""
    credenciales = [
        {
            'planta': c.planta,
            'credencial_id': c.credencial_id,
            'fecha_caducidad': c.fecha_caducidad.isoformat() if c.fecha_caducidad else None,
        }
        for c in t.credenciales
    ]
    documentos = [d.to_dict() for d in t.documentos]

    proyectos_activos = (
        t.proyectos.options(joinedload(Proyecto.coordinador)).filter_by(activo=True).all()
    )
    coordinadores_set = set()
    for p in proyectos_activos:
        if p.coordinador:
            coordinadores_set.add(p.coordinador.full_name or p.coordinador.username)

    curp, rfc, nss = (t.curp or ''), (t.rfc or ''), (t.nss or '')
    if not _is_admin():
        curp = _mask_pii(curp)
        rfc = _mask_pii(rfc)
        nss = _mask_pii(nss)

    return {
        'id': t.id,
        'no_empleado': t.no_empleado,
        'nombre': t.nombre,
        'nombre_apellidos': t.nombre_apellidos,
        # Laborales
        'tipo_mov': t.tipo_mov or '',
        'tipo_cont': t.tipo_cont or '',
        'area': t.area or '',
        'puesto': t.puesto or '',
        'tipo_jornada': t.tipo_jornada or '',
        'fecha_ingreso': t.fecha_ingreso.isoformat() if t.fecha_ingreso else '',
        'descripcion_servicio': t.descripcion_servicio or '',
        'inicio': t.inicio.isoformat() if t.inicio else '',
        'termino_prueba': t.termino_prueba.isoformat() if t.termino_prueba else '',
        'fecha_baja': t.fecha_baja.isoformat() if t.fecha_baja else '',
        # Generales
        'curp': curp,
        'rfc': rfc,
        'nss': nss,
        'fecha_nacimiento': t.fecha_nacimiento.isoformat() if t.fecha_nacimiento else '',
        'sexo': t.sexo or '',
        'estado_civil': t.estado_civil or '',
        'nacionalidad': t.nacionalidad or '',
        'edad': t.edad or '',
        'domicilio': t.domicilio or '',
        # Contacto
        'correo': t.correo or '',
        'celular': t.celular or '',
        # Médicos
        'tipo_sangre': t.tipo_sangre or '',
        'alergias': t.alergias or '',
        'enfermedades_cronicas': t.enfermedades_cronicas or '',
        'contacto_emergencia': t.contacto_emergencia or '',
        'parentesco_contacto': t.parentesco_contacto or '',
        'numero_contacto_emerg': t.numero_contacto_emerg or '',
        'lentes': t.lentes or '',
        'licencia_conducir': t.licencia_conducir or '',
        'estatura': t.estatura or '',
        # Finanzas
        'salario_real_pactado_x_sem': str(t.salario_real_pactado_x_sem) if t.salario_real_pactado_x_sem else '',
        'tipo_pago': t.tipo_pago or '',
        'tipo_nomina': t.tipo_nomina or '',
        'sb': str(t.sb) if t.sb else '',
        'sdi': str(t.sdi) if t.sdi else '',
        'letra': t.letra or '',
        'hr_extra': str(t.hr_extra) if t.hr_extra else '',
        'infonavit': str(t.infonavit) if t.infonavit else '',
        'ajuste_inbursa': str(t.ajuste_inbursa) if t.ajuste_inbursa else '',
        'caja_ahorro': str(t.caja_ahorro) if t.caja_ahorro else '',
        'viaticos': str(t.viaticos) if t.viaticos else '',
        'pago_dia_festivo': str(t.pago_dia_festivo) if t.pago_dia_festivo else '',
        'pagos_efectivo': str(t.pagos_efectivo) if t.pagos_efectivo else '',
        'folio_mov_idse': t.folio_mov_idse or '',
        # Operación
        'ubicacion_actual': t.ubicacion_actual or '',
        'ubicacion_estado': t.ubicacion_estado or '',
        'coordinadores_actuales': ', '.join(sorted(coordinadores_set)) if coordinadores_set else 'Ninguno asignado',
        'no_proyecto': t.no_proyecto or '',
        'observaciones': t.observaciones or '',
        # Multimedia / Relaciones
        'foto_perfil': t.foto_perfil or '',
        'activo': t.activo,
        'credenciales': credenciales,
        'documentos': documentos,
    }


# ── Whitelist de campos editables por rol ────────────────────────────────────
# CRIT-03 fix: `_apply_payload` antes asignaba CUALQUIER campo del form al modelo
# sin discriminar por rol. Un coordinador "autorizado" (vía proyecto) podía
# modificar salario, RFC, CURP, NSS, etc. Ahora separamos los campos en buckets
# y solo admin/super_admin puede tocar financieros y PII fiscal.

_ADMIN_ONLY_FIELDS = {
    # Identificación principal
    'no_empleado',
    # PII fiscal / regulada
    'curp', 'rfc', 'nss',
    # Datos personales formales (los coordinadores no deben tocarlos)
    'nombre', 'nombre_apellidos', 'fecha_nacimiento', 'sexo', 'estado_civil',
    'nacionalidad', 'edad', 'domicilio',
    # Laborales / contractuales
    'tipo_mov', 'tipo_cont', 'tipo_jornada', 'fecha_ingreso',
    'descripcion_servicio', 'inicio', 'termino_prueba', 'fecha_baja',
    'area', 'puesto',
    # Finanzas — la TOTALIDAD bloqueada para no-admin
    'salario_real_pactado_x_sem', 'tipo_pago', 'tipo_nomina',
    'sb', 'sdi', 'letra', 'hr_extra', 'infonavit', 'ajuste_inbursa',
    'caja_ahorro', 'viaticos', 'pago_dia_festivo', 'pagos_efectivo',
    'folio_mov_idse',
    # Contacto formal (admin)
    'correo',
}

# Campos editables por coordinador (operativos / contacto en campo):
_COORD_ALLOWED_FIELDS = {
    'celular',
    'tipo_sangre', 'alergias', 'enfermedades_cronicas',
    'contacto_emergencia', 'parentesco_contacto', 'numero_contacto_emerg',
    'lentes', 'licencia_conducir', 'estatura',
    'ubicacion_estado', 'observaciones',
}


def _apply_payload(t: Trabajador, data, *, actor_is_admin: bool) -> list[str]:
    """Aplica los campos del formulario al modelo respetando la whitelist por rol.

    - admin/super_admin: puede tocar todos los campos.
    - coordinador: solo `_COORD_ALLOWED_FIELDS` (operativos + datos médicos +
      contacto de emergencia). Cualquier intento de modificar campos admin-only
      se ignora silenciosamente y se loggea como warning.
    """
    warnings = []

    if actor_is_admin:
        editable = _ADMIN_ONLY_FIELDS | _COORD_ALLOWED_FIELDS
    else:
        # Coord: detectar y advertir si está intentando tocar campos prohibidos
        bloqueados = [k for k in data.keys() if k in _ADMIN_ONLY_FIELDS]
        if bloqueados:
            warnings.append(
                'Algunos campos solo pueden ser editados por un administrador y '
                f'fueron ignorados: {", ".join(sorted(bloqueados))}.'
            )
        editable = _COORD_ALLOWED_FIELDS

    # Validar formato CURP/RFC solo si se está actualizando (admin)
    if 'curp' in editable:
        curp_v = (data.get('curp') or '').strip().upper()
        if curp_v and not _CURP_RE.match(curp_v):
            warnings.append('El formato de CURP no es válido.')
    if 'rfc' in editable:
        rfc_v = (data.get('rfc') or '').strip().upper()
        if rfc_v and not _RFC_RE.match(rfc_v):
            warnings.append('El formato de RFC no es válido.')

    # ── Asignaciones por bucket (solo si el campo es editable para este rol) ──

    # Identificación / strings simples
    _STR_FIELDS = {
        'no_empleado', 'nombre', 'nombre_apellidos',
        'tipo_mov', 'tipo_cont', 'area', 'puesto', 'tipo_jornada',
        'descripcion_servicio', 'curp', 'rfc', 'nss',
        'sexo', 'estado_civil', 'nacionalidad', 'domicilio',
        'correo', 'celular',
        'tipo_sangre', 'alergias', 'enfermedades_cronicas',
        'contacto_emergencia', 'parentesco_contacto', 'numero_contacto_emerg',
        'lentes', 'licencia_conducir', 'estatura',
        'tipo_pago', 'tipo_nomina', 'letra', 'folio_mov_idse',
        'ubicacion_estado', 'observaciones',
    }
    for f in _STR_FIELDS:
        if f in editable and f in data:
            val = data.get(f)
            # no_empleado se strip-ea
            if f == 'no_empleado':
                val = (val or '').strip()
            setattr(t, f, val)

    # Fechas
    _DATE_FIELDS = {'fecha_ingreso', 'inicio', 'termino_prueba', 'fecha_baja', 'fecha_nacimiento'}
    for f in _DATE_FIELDS:
        if f in editable and f in data:
            setattr(t, f, _parse_date(data.get(f)))

    # Numéricos (con default 0)
    _NUM_FIELDS_ZERO = {
        'sb', 'sdi', 'hr_extra', 'infonavit', 'ajuste_inbursa',
        'caja_ahorro', 'viaticos', 'pago_dia_festivo', 'pagos_efectivo',
    }
    for f in _NUM_FIELDS_ZERO:
        if f in editable and f in data:
            setattr(t, f, data.get(f) or 0)

    # Edad: int o None
    if 'edad' in editable and 'edad' in data:
        t.edad = data.get('edad') or None

    # Salario: float estricto
    if 'salario_real_pactado_x_sem' in editable and 'salario_real_pactado_x_sem' in data:
        try:
            t.salario_real_pactado_x_sem = float(data.get('salario_real_pactado_x_sem') or 0)
        except (TypeError, ValueError):
            warnings.append('Salario inválido — se mantuvo el valor anterior.')

    return warnings


def _replace_credenciales(t: Trabajador, payload_str: str) -> None:
    """Reemplaza completamente las credenciales del trabajador con el array recibido."""
    if not payload_str:
        return
    credenciales_data = json.loads(payload_str)
    t.credenciales = []
    for c in credenciales_data:
        planta = str(c.get('planta', '')).strip().upper()
        credencial_id = str(c.get('credencial_id', '')).strip()
        if len(planta) > 100 or len(credencial_id) > 40:
            raise ValueError(
                f'Longitud inválida en credencial (planta {len(planta)}/100, id {len(credencial_id)}/40)'
            )
        t.credenciales.append(CredencialPlanta(
            planta=planta,
            credencial_id=credencial_id,
            fecha_caducidad=_parse_date(c.get('fecha_caducidad')),
        ))


def _save_foto(t: Trabajador, file) -> None:
    if not allowed_image_file(file):
        raise ValueError('Foto rechazada: solo se permiten imágenes JPG o PNG reales.')
    if t.foto_perfil:
        _delete_profile_picture(t.foto_perfil)
    unique_filename = f"pp_{int(time.time())}_{secure_filename(file.filename)}"
    pp_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'perfiles')
    t.foto_perfil = _save_profile_picture(file, pp_folder, unique_filename)


# ── Endpoints CRUD ─────────────────────────────────────────────────────────────

@bp.route('', methods=['GET'])
@jwt_required
def listar():
    page = request.args.get('page', 1, type=int)
    per_page = min(request.args.get('per_page', 20, type=int), 10000)
    q = (request.args.get('q') or '').strip()
    estado = (request.args.get('estado') or 'activos').lower()  # activos|bajas|todos

    if estado == 'bajas':
        if not _is_admin():
            return jsonify({'error': 'Solo admin puede ver bajas'}), 403
        query = Trabajador.query.filter(or_(Trabajador.activo == False, Trabajador.fecha_baja != None))  # noqa: E712
    elif estado == 'todos':
        if not _is_admin():
            return jsonify({'error': 'Solo admin puede ver todos'}), 403
        query = Trabajador.query
    else:
        query = Trabajador.query.filter(Trabajador.activo == True, Trabajador.fecha_baja == None)  # noqa: E712

    if _u().role == 'coordinador':
        mis_proyectos = Proyecto.query.options(selectinload(Proyecto.participantes)).filter_by(
            activo=True, coordinador_id=_u().id,
        ).all()
        ids_trabajadores = {t.id for p in mis_proyectos for t in p.participantes}
        query = query.filter(Trabajador.id.in_(ids_trabajadores))

    if q:
        like = f'%{q}%'
        query = query.filter(or_(
            Trabajador.nombre.ilike(like),
            Trabajador.nombre_apellidos.ilike(like),
            Trabajador.no_empleado.ilike(like),
            Trabajador.rfc.ilike(like),
        ))

    pagination = query.order_by(func.lower(Trabajador.nombre)).paginate(
        page=page, per_page=per_page, error_out=False,
    )
    return jsonify({
        'items': [_row_summary(t) for t in pagination.items],
        'page': pagination.page,
        'per_page': pagination.per_page,
        'total': pagination.total,
        'pages': pagination.pages,
        'has_next': pagination.has_next,
        'has_prev': pagination.has_prev,
    })


@bp.route('/ficha-tecnica', methods=['GET'])
@jwt_required
def ficha_tecnica():
    """Listado de trabajadores con info médica y contacto de emergencia.

    Coordinadores: solo los trabajadores de sus proyectos activos.
    Admin/super_admin: todos los activos.
    Otros roles: 403.
    """
    role = _u().role
    if role not in ('coordinador', 'admin', 'super_admin'):
        return jsonify({'error': 'Acceso denegado'}), 403

    if role == 'coordinador':
        mis_proyectos = Proyecto.query.options(selectinload(Proyecto.participantes)).filter_by(
            activo=True, coordinador_id=_u().id,
        ).all()
        ids = {t.id for p in mis_proyectos for t in p.participantes if t.activo}
        if not ids:
            return jsonify({'items': []})
        trabajadores = (
            Trabajador.query.options(selectinload(Trabajador.documentos))
            .filter(Trabajador.id.in_(ids), Trabajador.activo == True)  # noqa: E712
            .order_by(func.lower(Trabajador.nombre_apellidos))
            .all()
        )
    else:
        trabajadores = (
            Trabajador.query.options(selectinload(Trabajador.documentos))
            .filter(Trabajador.activo == True, Trabajador.fecha_baja == None)  # noqa: E712
            .order_by(func.lower(Trabajador.nombre_apellidos))
            .all()
        )

    return jsonify({
        'items': [
            {
                'id': t.id,
                'no_empleado': t.no_empleado,
                'nombre': t.nombre,
                'nombre_completo': t.nombre_completo,
                'foto_perfil': t.foto_perfil or '',
                'tipo_sangre': t.tipo_sangre or '',
                'alergias': t.alergias or '',
                'enfermedades_cronicas': t.enfermedades_cronicas or '',
                'contacto_emergencia': t.contacto_emergencia or '',
                'parentesco_contacto': t.parentesco_contacto or '',
                'numero_contacto_emerg': t.numero_contacto_emerg or '',
                'documentos': [
                    {'id': d.id, 'nombre_archivo': d.nombre_archivo}
                    for d in t.documentos
                ],
            }
            for t in trabajadores
        ],
    })


@bp.route('/<int:id>', methods=['GET'])
@jwt_required
def obtener(id):
    t = (
        Trabajador.query.options(
            selectinload(Trabajador.credenciales),
            selectinload(Trabajador.documentos),
        )
        .filter_by(id=id)
        .first()
    )
    if not t:
        return jsonify({'error': 'No encontrado'}), 404
    if not _authorized(t):
        return jsonify({'error': 'Acceso denegado'}), 403
    return jsonify(_full_detail(t))


@bp.route('', methods=['POST'])
@jwt_required
@limiter.limit('10 per minute')
def crear():
    # SEGURIDAD: la creación de trabajadores cambia salarios y datos fiscales;
    # solo admin/super_admin puede invocarla. Antes era abierto a cualquier
    # autenticado y la autorización dependía solo de la asignación a proyecto.
    if not _is_admin():
        return jsonify({'error': 'Solo admin puede crear trabajadores'}), 403
    try:
        data = request.form
        if not (data.get('no_empleado') or '').strip():
            return jsonify({'error': 'El Número de Empleado es obligatorio'}), 400
        errores_longitud = validate_lengths(data)
        if errores_longitud:
            return jsonify({'error': 'Error de longitud en campos', 'details': errores_longitud}), 400
        if Trabajador.query.filter_by(no_empleado=data.get('no_empleado').strip()).first():
            return jsonify({'error': 'El Número de Empleado ya existe'}), 409
        try:
            salario = float(data.get('salario_real_pactado_x_sem') or 0)
            if salario < 0:
                return jsonify({'error': 'El salario no puede ser negativo'}), 400
        except ValueError:
            return jsonify({'error': 'Salario inválido'}), 400

        t = Trabajador()
        warnings = _apply_payload(t, data, actor_is_admin=True)

        try:
            _replace_credenciales(t, data.get('credenciales_json', '[]'))
        except ValueError as e:
            return jsonify({'error': str(e)}), 400
        except Exception as e:
            current_app.logger.error('Error parsing credentials on create: %s', e)
            warnings.append('Hubo un problema guardando algunas credenciales de planta.')

        foto = request.files.get('foto_perfil')
        if foto and foto.filename:
            try:
                _save_foto(t, foto)
            except ValueError as e:
                return jsonify({'error': str(e)}), 400

        db.session.add(t)
        db.session.commit()
        log_action(f'Agregó al trabajador {t.nombre} ({t.no_empleado})')
        emit_to_role(['admin', 'super_admin'], 'empleado:changed', {
            'id': t.id, 'action': 'created',
        })
        return jsonify({'id': t.id, 'warnings': warnings}), 201
    except Exception as e:
        db.session.rollback()
        current_app.logger.error('Error creando trabajador: %s\n%s', e, traceback.format_exc())
        return jsonify({'error': 'Error al guardar el trabajador'}), 500


@bp.route('/<int:id>', methods=['PUT', 'POST'])
@jwt_required
def actualizar(id):
    t = Trabajador.query.get(id)
    if not t:
        return jsonify({'error': 'No encontrado'}), 404
    if not _authorized(t):
        return jsonify({'error': 'Acceso denegado'}), 403

    actor_is_admin = _is_admin()
    try:
        data = request.form
        new_no = (data.get('no_empleado') or '').strip()
        # SEGURIDAD: el coordinador no puede cambiar no_empleado (admin-only).
        # Si lo intenta y el valor difiere del actual, lo bloqueamos en vez de
        # ignorar silenciosamente — el form del coord no debería ni mandarlo.
        if not actor_is_admin and new_no and new_no != t.no_empleado:
            return jsonify({'error': 'No puedes modificar el número de empleado'}), 403
        if actor_is_admin and new_no and new_no != t.no_empleado and Trabajador.query.filter_by(no_empleado=new_no).first():
            return jsonify({'error': 'El Número de Empleado ya existe'}), 409
        errores_longitud = validate_lengths(data)
        if errores_longitud:
            return jsonify({'error': 'Error de longitud en campos', 'details': errores_longitud}), 400

        warnings = _apply_payload(t, data, actor_is_admin=actor_is_admin)

        try:
            _replace_credenciales(t, data.get('credenciales_json', '[]'))
        except ValueError as e:
            return jsonify({'error': str(e)}), 400
        except Exception as e:
            current_app.logger.error('Error parsing credentials on update: %s', e)
            warnings.append('Hubo un problema actualizando algunas credenciales.')

        foto = request.files.get('foto_perfil')
        if foto and foto.filename:
            try:
                _save_foto(t, foto)
            except ValueError as e:
                return jsonify({'error': str(e)}), 400

        db.session.commit()
        log_action(f'Actualizó al trabajador {t.nombre} ({t.no_empleado})')
        # Coordinador también dispara la invalidación: edita campos médicos /
        # contacto que aparecen en la lista de empleados de admin.
        emit_to_role(['admin', 'super_admin'], 'empleado:changed', {
            'id': t.id, 'action': 'updated',
        })
        return jsonify({'id': t.id, 'warnings': warnings})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error('Error actualizando trabajador: %s\n%s', e, traceback.format_exc())
        return jsonify({'error': 'Error al actualizar el trabajador'}), 500


@bp.route('/<int:id>', methods=['DELETE'])
@jwt_required
def dar_baja(id):
    if not _is_admin():
        return jsonify({'error': 'Solo admin puede dar de baja'}), 403
    t = Trabajador.query.get(id)
    if not t:
        return jsonify({'error': 'No encontrado'}), 404
    try:
        t.activo = False
        t.fecha_baja = date.today()
        db.session.commit()
        log_action(f'Dio de baja al trabajador {t.nombre} ({t.no_empleado})')
        emit_to_role(['admin', 'super_admin'], 'empleado:changed', {
            'id': t.id, 'action': 'baja',
        })
        return jsonify({'ok': True})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error('Error baja: %s', e)
        return jsonify({'error': 'Error al dar de baja'}), 500


@bp.route('/<int:id>/reactivar', methods=['POST'])
@jwt_required
def reactivar(id):
    if not _is_admin():
        return jsonify({'error': 'Solo admin puede reactivar'}), 403
    t = Trabajador.query.get(id)
    if not t:
        return jsonify({'error': 'No encontrado'}), 404
    try:
        t.activo = True
        t.fecha_baja = None
        db.session.commit()
        log_action(f'Reactivó al trabajador {t.nombre} ({t.no_empleado})')
        emit_to_role(['admin', 'super_admin'], 'empleado:changed', {
            'id': t.id, 'action': 'reactivado',
        })
        return jsonify({'ok': True})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error('Error reactivar: %s', e)
        return jsonify({'error': 'Error al reactivar'}), 500


# ── Credenciales (bulk) ────────────────────────────────────────────────────────

def _credencial_row(t: Trabajador) -> dict:
    """Fila para la pantalla de Credenciales: incluye credenciales + datos de la ficha."""
    proyectos_activos = (
        t.proyectos.options(joinedload(Proyecto.coordinador)).filter_by(activo=True).all()
    )
    proyectos_nombres = [p.nombre for p in proyectos_activos if p.nombre]
    coordinadores = []
    for p in proyectos_activos:
        if p.coordinador:
            nombre = p.coordinador.full_name or p.coordinador.username
            if nombre and nombre not in coordinadores:
                coordinadores.append(nombre)

    return {
        'id': t.id,
        'no_empleado': t.no_empleado,
        'nombre': t.nombre,
        'nombre_apellidos': t.nombre_apellidos,
        'area': t.area,
        'puesto': t.puesto,
        'tipo_nomina': t.tipo_nomina,
        'celular': t.celular,
        'ubicacion_actual': t.ubicacion_actual,
        'ubicacion_estado': t.ubicacion_estado,
        'observaciones': t.observaciones,
        'foto_perfil': t.foto_perfil,
        'coord_a_cargo': ', '.join(coordinadores) if coordinadores else (t.coord_a_cargo or ''),
        'proyectos_activos': ', '.join(proyectos_nombres) if proyectos_nombres else '',
        'credenciales': [
            {
                'planta': c.planta,
                'credencial_id': c.credencial_id,
                'fecha_caducidad': c.fecha_caducidad.isoformat() if c.fecha_caducidad else None,
            }
            for c in t.credenciales
        ],
    }


@bp.route('/credenciales-lista', methods=['GET'])
@jwt_required
def listar_credenciales():
    page = request.args.get('page', 1, type=int)
    per_page = min(request.args.get('per_page', 20, type=int), 100)
    q = (request.args.get('q') or '').strip()

    query = Trabajador.query.options(
        selectinload(Trabajador.credenciales),
    ).filter(Trabajador.activo == True, Trabajador.fecha_baja == None)  # noqa: E712

    if _u().role == 'coordinador':
        mis_proyectos = Proyecto.query.options(selectinload(Proyecto.participantes)).filter_by(
            activo=True, coordinador_id=_u().id,
        ).all()
        ids_trabajadores = {t.id for p in mis_proyectos for t in p.participantes}
        query = query.filter(Trabajador.id.in_(ids_trabajadores))

    if q:
        like = f'%{q}%'
        query = query.filter(or_(
            Trabajador.nombre.ilike(like),
            Trabajador.nombre_apellidos.ilike(like),
            Trabajador.no_empleado.ilike(like),
            Trabajador.rfc.ilike(like),
        ))

    pagination = query.order_by(func.lower(Trabajador.nombre)).paginate(
        page=page, per_page=per_page, error_out=False,
    )
    return jsonify({
        'items': [_credencial_row(t) for t in pagination.items],
        'page': pagination.page,
        'per_page': pagination.per_page,
        'total': pagination.total,
        'pages': pagination.pages,
        'has_next': pagination.has_next,
        'has_prev': pagination.has_prev,
    })


@bp.route('/<int:id>/credenciales', methods=['POST'])
@jwt_required
def guardar_credenciales(id):
    if not _is_admin():
        return jsonify({'error': 'Solo admin puede editar credenciales'}), 403
    t = Trabajador.query.get(id)
    if not t:
        return jsonify({'error': 'No encontrado'}), 404
    try:
        data = request.get_json(silent=True) or {}
        if 'observaciones' in data:
            t.observaciones = data.get('observaciones')
        _replace_credenciales(t, json.dumps(data.get('credenciales', [])))
        db.session.commit()
        log_action(f'Actualizó credenciales del trabajador {t.nombre} ({t.no_empleado})')
        emit_to_role(['admin', 'super_admin', 'coordinador'], 'credencial:changed', {
            'trabajador_id': t.id,
        })
        return jsonify({'ok': True})
    except ValueError as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        current_app.logger.error('Error credenciales: %s', e)
        return jsonify({'error': 'Error al guardar credenciales'}), 500


# ── Foto de perfil ─────────────────────────────────────────────────────────────

@bp.route('/<int:id>/foto', methods=['POST'])
@jwt_required
def subir_foto(id):
    t = Trabajador.query.get(id)
    if not t:
        return jsonify({'error': 'No encontrado'}), 404
    if not _authorized(t):
        return jsonify({'error': 'Acceso denegado'}), 403
    foto = request.files.get('foto_perfil') or request.files.get('foto')
    if not foto or not foto.filename:
        return jsonify({'error': 'No se envió archivo'}), 400
    try:
        _save_foto(t, foto)
        db.session.commit()
        log_action(f'Actualizó foto de perfil del trabajador {t.nombre} ({t.no_empleado})')
        return jsonify({'foto_perfil': t.foto_perfil})
    except ValueError as e:
        return jsonify({'error': str(e)}), 400


@bp.route('/<int:id>/foto', methods=['GET'])
@jwt_required
def get_foto(id):
    t = Trabajador.query.get(id)
    if not t:
        return jsonify({'error': 'No encontrado'}), 404
    if not _authorized(t):
        return jsonify({'error': 'Acceso denegado'}), 403
    if not t.foto_perfil:
        return jsonify({'error': 'Sin foto'}), 404
    return send_from_directory(current_app.config['UPLOAD_FOLDER'], t.foto_perfil)


@bp.route('/<int:id>/foto/thumb', methods=['GET'])
@jwt_required
def get_foto_thumb(id):
    t = Trabajador.query.get(id)
    if not t:
        return jsonify({'error': 'No encontrado'}), 404
    if not _authorized(t):
        return jsonify({'error': 'Acceso denegado'}), 403
    if not t.foto_perfil:
        return jsonify({'error': 'Sin foto'}), 404
    folder, filename = os.path.split(t.foto_perfil)
    thumb_rel = os.path.join(folder, f'thumb_{filename}').replace('\\', '/')
    upload_folder = current_app.config['UPLOAD_FOLDER']
    if os.path.exists(os.path.join(upload_folder, thumb_rel)):
        return send_from_directory(upload_folder, thumb_rel)
    return send_from_directory(upload_folder, t.foto_perfil)


# ── Documentos ────────────────────────────────────────────────────────────────

# HIGH-03 fix: para documentos de trabajadores restringimos drásticamente los
# tipos permitidos. El set global `ALLOWED_EXTENSIONS` incluye mp4/mp3/wav y
# formatos de Office con macros (xlsm/docx/pptx) que abren vías de RCE cuando
# el admin descarga y abre el archivo. Los documentos legales de RRHH son PDF
# o imagen — limitarlos baja la superficie sin perder funcionalidad real.
_DOCUMENTO_TRABAJADOR_EXTS = {'pdf', 'jpg', 'jpeg', 'png', 'heic'}
_DOCUMENTO_MAX_BYTES = 20 * 1024 * 1024  # 20 MB por archivo (vs los 50 MB globales)


@bp.route('/<int:id>/documentos', methods=['POST'])
@jwt_required
@limiter.limit('10 per minute')
def subir_documento(id):
    t = Trabajador.query.get(id)
    if not t:
        return jsonify({'error': 'No encontrado'}), 404
    if not _authorized(t):
        return jsonify({'error': 'Acceso denegado'}), 403

    file = request.files.get('documento') or request.files.get('file')
    if not file or not file.filename:
        return jsonify({'error': 'No se envió archivo'}), 400

    # HIGH-03: chequeo de tamaño antes que el de magic-bytes (más barato y
    # corta DoS por upload grande).
    file.seek(0, os.SEEK_END)
    size = file.tell()
    file.seek(0)
    if size > _DOCUMENTO_MAX_BYTES:
        return jsonify({
            'error': f'El archivo excede {_DOCUMENTO_MAX_BYTES // (1024 * 1024)} MB.'
        }), 400

    if not allowed_file(file, allowed_exts=_DOCUMENTO_TRABAJADOR_EXTS):
        return jsonify({
            'error': 'Tipo de archivo no permitido. Solo PDF, JPG, PNG o HEIC.'
        }), 400

    tipo_doc = (request.form.get('tipo_documento') or '').strip() or None
    if tipo_doc and len(tipo_doc) > 100:
        return jsonify({'error': 'tipo_documento demasiado largo (máx. 100)'}), 400

    filename = secure_filename(file.filename)
    unique_filename = f'{int(time.time())}_{filename}'
    upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'trabajadores', str(t.id))
    os.makedirs(upload_folder, exist_ok=True)
    file.save(os.path.join(upload_folder, unique_filename))

    doc = DocumentoTrabajador(
        trabajador_id=t.id,
        nombre_archivo=filename,
        ruta_archivo=f'trabajadores/{t.id}/{unique_filename}',
        tipo_documento=tipo_doc,
        fecha_inicio=_parse_date(request.form.get('fecha_inicio')),
        fecha_fin=_parse_date(request.form.get('fecha_fin')),
    )
    db.session.add(doc)
    db.session.commit()
    log_action(f'Subió documento {filename} para {t.nombre} ({t.no_empleado})')
    return jsonify(doc.to_dict()), 201


@bp.route('/documentos/<int:doc_id>', methods=['GET'])
@jwt_required
def descargar_documento(doc_id):
    doc = DocumentoTrabajador.query.get(doc_id)
    if not doc:
        return jsonify({'error': 'No encontrado'}), 404
    if not _authorized(doc.trabajador):
        return jsonify({'error': 'Acceso denegado'}), 403
    # secure_filename para defender contra header-injection vía nombre con
    # CRLF o caracteres de control que algún antiguo browser podría
    # interpretar mal en el Content-Disposition.
    safe_name = secure_filename(doc.nombre_archivo) or f'documento_{doc_id}'
    return send_from_directory(
        current_app.config['UPLOAD_FOLDER'],
        doc.ruta_archivo,
        as_attachment=True,
        download_name=safe_name,
    )


@bp.route('/documentos/<int:doc_id>', methods=['DELETE'])
@jwt_required
def eliminar_documento(doc_id):
    doc = DocumentoTrabajador.query.get(doc_id)
    if not doc:
        return jsonify({'error': 'No encontrado'}), 404
    if not _authorized(doc.trabajador):
        return jsonify({'error': 'Acceso denegado'}), 403
    file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], doc.ruta_archivo)
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except Exception as e:
        current_app.logger.error('Error eliminando archivo físico: %s', e)
    log_action(f'Eliminó documento {doc.nombre_archivo} del trabajador {doc.trabajador_id}')
    db.session.delete(doc)
    db.session.commit()
    return jsonify({'ok': True})


# ── Import / Export Excel ──────────────────────────────────────────────────────

@bp.route('/plantilla-importar', methods=['GET'])
@jwt_required
def descargar_plantilla_importar():
    """Sirve el archivo de plantilla de importación. Reusa el .xlsx que vive en
    static/downloads/ (mismo que usa la UI clásica)."""
    base_dir = current_app.config['BASE_DIR']
    folder = os.path.join(base_dir, 'static', 'downloads')
    filename = 'plantilla_empleados.xlsx'
    if not os.path.exists(os.path.join(folder, filename)):
        return jsonify({'error': 'Plantilla no encontrada en el servidor'}), 404
    return send_from_directory(
        folder, filename,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@bp.route('/importar', methods=['POST'])
@jwt_required
@limiter.limit('5 per minute')
def importar_excel():
    if not _is_admin():
        return jsonify({'error': 'Solo admin puede importar'}), 403
    file = request.files.get('archivo') or request.files.get('archivo_excel')
    if not file or not file.filename:
        return jsonify({'error': 'No se envió archivo'}), 400
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Formato no válido. Debe ser .xlsx o .xls'}), 400

    import filetype
    header = file.read(2048)
    file.seek(0)
    kind = filetype.guess(header)
    if not kind or kind.extension not in ['xlsx', 'xls', 'zip', 'cfb']:
        return jsonify({'error': 'El archivo no parece ser Excel válido'}), 400

    file.seek(0, os.SEEK_END)
    size = file.tell()
    file.seek(0)
    if size > 5 * 1024 * 1024:
        return jsonify({'error': 'El archivo excede 5MB'}), 400

    import pandas as pd
    try:
        df = pd.read_excel(file)
    except Exception as e:
        current_app.logger.error('Error leyendo Excel: %s', e)
        return jsonify({'error': 'Error al leer el Excel. Usa la plantilla correcta.'}), 400

    _FORMULA_CHARS = ('=', '+', '-', '@', '\t', '\r')
    errores: list[str] = []
    exitosos = 0

    for index, row in df.iterrows():
        no_emp = str(row.get('No. Empleado', '')).strip()
        if no_emp in ('', 'nan'):
            errores.append(f'Fila {index + 2}: Falta No. Empleado')
            continue
        if Trabajador.query.filter_by(no_empleado=no_emp).first():
            errores.append(f'Fila {index + 2}: El empleado {no_emp} ya existe')
            continue

        row_errors: list[str] = []

        def get_str(col, max_len=None):
            val = row.get(col, '')
            import pandas as _pd
            s = str(val).strip() if _pd.notna(val) else ''
            if s and s[0] in _FORMULA_CHARS:
                s = "'" + s
            if max_len and len(s) > max_len:
                row_errors.append(f"'{col}' excede {max_len} (recibidos {len(s)})")
            return s

        def get_float(col):
            val = row.get(col, 0)
            import pandas as _pd
            if _pd.isna(val) or val == '':
                return 0.0
            try:
                return float(val)
            except ValueError:
                row_errors.append(f"'{col}' no es número válido: {val}")
                return 0.0

        def get_date(col):
            import pandas as _pd
            raw = row.get(col)
            if _pd.isna(raw) or raw == '':
                return None
            try:
                return _pd.to_datetime(raw).date()
            except Exception:
                row_errors.append(f"'{col}' fecha inválida: {raw}")
                return None

        t = Trabajador(
            no_empleado=get_str('No. Empleado', 50),
            nombre=get_str('Nombre(s)', 250),
            nombre_apellidos=get_str('Apellidos', 250),
            tipo_mov=get_str('Tipo de Movimiento', 100),
            tipo_cont=get_str('Tipo de Contrato', 100),
            area=get_str('Area', 150),
            puesto=get_str('Puesto', 150),
            tipo_jornada=get_str('Tipo de Jornada', 100),
            fecha_ingreso=get_date('Fecha Ingreso (YYYY-MM-DD)'),
            descripcion_servicio=get_str('Descripcion de Servicio'),
            inicio=get_date('Fecha Inicio (YYYY-MM-DD)'),
            termino_prueba=get_date('Termino de Prueba (YYYY-MM-DD)'),
            rfc=get_str('RFC', 13),
            curp=get_str('CURP', 18),
            nss=get_str('NSS', 20),
            fecha_nacimiento=get_date('Fecha Nacimiento (YYYY-MM-DD)'),
            sexo=get_str('Sexo (M/F)', 20),
            estado_civil=get_str('Estado Civil', 50),
            nacionalidad=get_str('Nacionalidad', 100),
            edad=int(get_float('Edad')) if get_float('Edad') > 0 else None,
            domicilio=get_str('Domicilio'),
            correo=get_str('Correo', 150),
            celular=get_str('Celular', 20),
            tipo_sangre=get_str('Tipo de Sangre', 10),
            alergias=get_str('Alergias'),
            enfermedades_cronicas=get_str('Enfermedades Cronicas'),
            contacto_emergencia=get_str('Contacto de Emergencia', 200),
            parentesco_contacto=get_str('Parentesco del Contacto', 100),
            numero_contacto_emerg=get_str('Numero Contacto Emergencia', 20),
            lentes=get_str('Usa Lentes (Si/No)', 20),
            licencia_conducir=get_str('Licencia de Conducir (Tipo)', 50),
            estatura=get_str('Estatura', 20),
            salario_real_pactado_x_sem=get_float('Salario Real Pactado por Semana'),
            tipo_pago=get_str('Tipo Pago', 50),
            tipo_nomina=get_str('Tipo de Nomina (Semanal/Por hora/Cuadrado)', 100),
            sb=get_float('Sueldo Base (SB)'),
            sdi=get_float('Salario Diario Integrado (SDI)'),
            letra=get_str('Letra', 100),
            hr_extra=get_float('Horas Extra'),
            infonavit=get_float('Infonavit'),
            caja_ahorro=get_float('Caja de Ahorro'),
            viaticos=get_float('Viaticos'),
            pago_dia_festivo=get_float('Pago Dia Festivo'),
            pagos_efectivo=get_float('Pagos Efectivo'),
            folio_mov_idse=get_str('Folio Mov IDSE', 100),
            ubicacion_estado=get_str('Ubicacion Estado', 100),
            observaciones=get_str('Observaciones'),
        )
        if row_errors:
            errores.append(f"Fila {index + 2} ({no_emp}): " + ', '.join(row_errors))
        else:
            db.session.add(t)
            exitosos += 1

    if exitosos:
        try:
            db.session.commit()
            log_action(f'Importó masivamente {exitosos} trabajadores desde Excel')
            emit_to_role(['admin', 'super_admin'], 'empleado:changed', {
                'action': 'imported', 'count': exitosos,
            })
        except Exception as e:
            db.session.rollback()
            current_app.logger.error('Error commit import: %s', e)
            return jsonify({'error': 'Error guardando los registros'}), 500

    return jsonify({'exitosos': exitosos, 'errores': errores})


def _build_export_styles():
    """Estilos compartidos para los dos exports XLSX."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    return {
        'AZUL_OSC': PatternFill('solid', fgColor='1E3A5F'),
        'AZUL_HDR': PatternFill('solid', fgColor='2563EB'),
        'BLANCO': PatternFill('solid', fgColor='FFFFFF'),
        'GRIS_ALT': PatternFill('solid', fgColor='F1F5F9'),
        'BORDER': Border(*(Side(style='thin', color='CBD5E1') for _ in range(4))),
        'BORDER_HDR': Border(*(Side(style='medium', color='1E3A5F') for _ in range(4))),
        'CENTER': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'LEFT': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'FONT_WHITE_BOLD': Font(name='Calibri', color='FFFFFF', bold=True, size=10),
        'FONT_BODY': Font(name='Calibri', color='374151', size=10),
        'FONT_TITLE': Font(name='Calibri', color='FFFFFF', bold=True, size=13),
    }


_HEADERS_EXPORT = [
    'No. Empleado', 'Nombre(s)', 'Apellidos', 'RFC', 'CURP', 'NSS',
    'Fecha Nacimiento', 'Sexo', 'Estado Civil', 'Nacionalidad', 'Edad',
    'Correo', 'Celular', 'Domicilio',
    'Área', 'Puesto', 'Tipo Movimiento', 'Tipo Contrato',
    'Fecha Ingreso', 'Tipo Jornada', 'Tipo Nómina', 'Tipo Pago',
    'Folio IDSE', 'Desc. Servicio',
    'Tipo Sangre', 'Lentes', 'Alergias', 'Estatura', 'Enf. Crónicas', 'Licencia Conducir',
    'Contacto Emergencia', 'Parentesco', 'Tel. Emergencia',
    'Salario Pactado/Sem', 'Salario Base', 'SDI', 'Letra/Categoría',
    'Hrs Extra', 'Infonavit', 'Caja Ahorro', 'Viáticos', 'Día Festivo', 'Pagos Efectivo',
    'Ubicación Actual', 'Estado Ubic.', 'No. Proyecto', 'Coord. a Cargo',
]


def _row_values(t: Trabajador, mask_pii: bool):
    def fmt_d(d):
        return d.strftime('%d/%m/%Y') if d else ''

    def fmt_m(v):
        return float(v) if v else ''

    def mask(val):
        if not val or not mask_pii:
            return val or ''
        return val[:3] + '***' + val[-2:] if len(val) > 5 else '***'

    return [
        t.no_empleado, t.nombre, t.nombre_apellidos,
        mask(t.rfc), mask(t.curp), mask(t.nss),
        fmt_d(t.fecha_nacimiento), t.sexo, t.estado_civil, t.nacionalidad,
        t.edad, t.correo, t.celular, t.domicilio,
        t.area, t.puesto, t.tipo_mov, t.tipo_cont,
        fmt_d(t.fecha_ingreso), t.tipo_jornada, t.tipo_nomina, t.tipo_pago,
        t.folio_mov_idse, t.descripcion_servicio,
        t.tipo_sangre, t.lentes, t.alergias, t.estatura,
        t.enfermedades_cronicas, t.licencia_conducir,
        t.contacto_emergencia, t.parentesco_contacto, t.numero_contacto_emerg,
        fmt_m(t.salario_real_pactado_x_sem), fmt_m(t.sb), fmt_m(t.sdi),
        t.letra, fmt_m(t.hr_extra), fmt_m(t.infonavit),
        fmt_m(t.caja_ahorro), fmt_m(t.viaticos),
        fmt_m(t.pago_dia_festivo), fmt_m(t.pagos_efectivo),
        t.ubicacion_actual, t.ubicacion_estado, t.no_proyecto, t.coord_a_cargo,
    ]


@bp.route('/<int:id>/exportar', methods=['GET'])
@jwt_required
def exportar_uno(id):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    t = (
        Trabajador.query.options(selectinload(Trabajador.credenciales))
        .filter_by(id=id)
        .first()
    )
    if not t:
        return jsonify({'error': 'No encontrado'}), 404
    if not _authorized(t):
        return jsonify({'error': 'Acceso denegado'}), 403

    S = _build_export_styles()
    headers = _HEADERS_EXPORT + ['Observaciones']
    mask_pii = not _is_admin()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Empleado'
    ws.row_dimensions[1].height = 36
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    c = ws['A1']
    c.value = (
        f"{t.nombre} {t.nombre_apellidos}".upper()
        + f"  —  No. {t.no_empleado}  —  Generado: {date.today().strftime('%d/%m/%Y')}"
    )
    c.fill = S['AZUL_OSC']; c.font = S['FONT_TITLE']; c.alignment = S['CENTER']; c.border = S['BORDER_HDR']

    ws.row_dimensions[2].height = 22
    for i, header in enumerate(headers, 1):
        cc = ws.cell(row=2, column=i, value=header)
        cc.fill = S['AZUL_HDR']; cc.font = S['FONT_WHITE_BOLD']; cc.alignment = S['CENTER']; cc.border = S['BORDER_HDR']

    ws.row_dimensions[3].height = 18
    values = _row_values(t, mask_pii) + [t.observaciones]
    for i, val in enumerate(values, 1):
        cc = ws.cell(row=3, column=i, value=safe_excel_value(val) if val is not None else '')
        cc.fill = S['BLANCO']; cc.font = S['FONT_BODY']; cc.alignment = S['LEFT']; cc.border = S['BORDER']

    if t.credenciales:
        from openpyxl.styles import Font, PatternFill
        ws2 = wb.create_sheet('Credenciales')
        ws2.row_dimensions[1].height = 22
        for i, label in enumerate(['Planta', 'ID Credencial', 'Caducidad', 'Estado'], 1):
            cc = ws2.cell(row=1, column=i, value=label)
            cc.fill = S['AZUL_HDR']; cc.font = S['FONT_WHITE_BOLD']; cc.alignment = S['CENTER']; cc.border = S['BORDER_HDR']
        today = date.today()
        verde = PatternFill('solid', fgColor='D1FAE5')
        rojo = PatternFill('solid', fgColor='FEE2E2')
        for i, cred in enumerate(t.credenciales, 2):
            vencida = cred.fecha_caducidad and cred.fecha_caducidad < today
            cad_str = cred.fecha_caducidad.strftime('%d/%m/%Y') if cred.fecha_caducidad else ''
            estado = 'CADUCADA' if vencida else 'VIGENTE'
            row_fill = rojo if vencida else verde
            for j, val in enumerate([cred.planta, cred.credencial_id, cad_str, estado], 1):
                cc = ws2.cell(row=i, column=j, value=val or '')
                cc.fill = row_fill
                cc.font = Font(name='Calibri', color='7F1D1D' if vencida else '065F46', size=10)
                cc.alignment = S['LEFT']; cc.border = S['BORDER']
        for i, w in enumerate([20, 20, 14, 12], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

    widths = [14, 18, 20, 14, 20, 14, 14, 8, 14, 14, 6, 28, 14, 32,
              16, 20, 16, 16, 14, 14, 14, 12, 12, 20,
              12, 8, 16, 10, 16, 16,
              22, 14, 14,
              16, 14, 12, 14, 12, 12, 12, 12, 12, 14,
              20, 16, 14, 20, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    safe_name = f"{t.no_empleado}_{(t.nombre_apellidos or '').replace(' ', '_')}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=safe_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@bp.route('/exportar-todos', methods=['GET'])
@jwt_required
def exportar_todos():
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    base = Trabajador.query.filter(Trabajador.activo == True, Trabajador.fecha_baja == None)  # noqa: E712
    if _u().role == 'coordinador':
        mis = Proyecto.query.filter_by(activo=True, coordinador_id=_u().id).all()
        ids = {t.id for p in mis for t in p.participantes}
        base = base.filter(Trabajador.id.in_(ids))

    trabajadores = base.order_by(func.lower(Trabajador.nombre)).all()
    S = _build_export_styles()
    mask_pii = not _is_admin()

    wb = Workbook(); ws = wb.active; ws.title = 'Empleados'
    ws.row_dimensions[1].height = 38
    ws.merge_cells(f'A1:{get_column_letter(len(_HEADERS_EXPORT))}1')
    c = ws['A1']
    c.value = f"LISTADO DE EMPLEADOS — Generado: {date.today().strftime('%d/%m/%Y')}"
    c.fill = S['AZUL_OSC']; c.font = S['FONT_TITLE']; c.alignment = S['CENTER']; c.border = S['BORDER_HDR']

    ws.row_dimensions[2].height = 22
    for i, header in enumerate(_HEADERS_EXPORT, 1):
        cc = ws.cell(row=2, column=i, value=header)
        cc.fill = S['AZUL_HDR']; cc.font = S['FONT_WHITE_BOLD']; cc.alignment = S['CENTER']; cc.border = S['BORDER_HDR']

    for row_idx, t in enumerate(trabajadores, start=3):
        fill = S['GRIS_ALT'] if (row_idx % 2 == 0) else S['BLANCO']
        ws.row_dimensions[row_idx].height = 16
        for i, val in enumerate(_row_values(t, mask_pii), 1):
            cc = ws.cell(row=row_idx, column=i, value=safe_excel_value(val) if val is not None else '')
            cc.fill = fill; cc.font = S['FONT_BODY']; cc.alignment = S['LEFT']; cc.border = S['BORDER']

    widths = [14, 18, 20, 14, 20, 14, 14, 8, 14, 14, 6, 28, 14, 32,
              16, 20, 16, 16, 14, 14, 14, 12, 12, 20,
              12, 8, 16, 10, 16, 16,
              22, 14, 14,
              16, 14, 12, 14, 12, 12, 12, 12, 12, 14,
              20, 16, 14, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"empleados_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
