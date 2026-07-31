"""Armado de los .xlsx de reportes: estilos, streaming y parseo de fechas.

Todos los reportes salen con el mismo formato (encabezado azul, columnas
autoajustadas, moneda con dos decimales) porque comparten `_stream_excel`.
"""
import datetime
import io

from flask import jsonify, request, send_file

# Tope de filas por reporte: un Excel de más no se abre cómodo y la generación
# se vuelve un DoS accidental.
REPORTES_MAX_FILAS = 10_000


def _aplicar_estilos_ws(ws, money_cols: set[str] | None = None):
    """Aplica encabezado azul + zebra striping + freeze panes + auto-width.

    `money_cols`: nombres de columna (exactos) que reciben formato '$#,##0.00'.
    El resto de los numéricos quedan sin formato (importante: stocks, días,
    cantidades enteras o decimales sin moneda).
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    money_cols = money_cols or set()

    headers = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
    money_idx = {i + 1 for i, h in enumerate(headers) if h in money_cols}

    if ws.max_row > 0:
        ws.freeze_panes = 'A2'

    for row_idx in range(1, ws.max_row + 1):
        is_header = (row_idx == 1)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if is_header:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
            else:
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill
                if col_idx in money_idx and isinstance(cell.value, (int, float)):
                    cell.number_format = '"$"#,##0.00'

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    n = len(str(cell.value))
                    if n > max_length:
                        max_length = n
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)


def _stream_excel(sheets: dict, filename: str, money_cols: set[str] | None = None):
    """Genera un .xlsx multi-hoja en memoria y lo devuelve con send_file.

    `sheets`: dict[nombre_hoja, list[dict]]. Los keys de la primera fila se
    usan como encabezados (orden estable de Python 3.7+).
    Aplica saneo anti CSV-injection con `safe_excel_value`.
    """
    from openpyxl import Workbook
    from app.utils import safe_excel_value

    wb = Workbook()
    # Workbook trae una hoja "Sheet" por default; la quitamos para que solo
    # queden las hojas que pasamos en `sheets`.
    default = wb.active
    wb.remove(default)

    for raw_name, rows in sheets.items():
        # Excel limita nombre de hoja a 31 chars y prohíbe / \ ? * [ ]
        safe = (raw_name or 'Hoja')[:31]
        for ch in r'/\?*[]:':
            safe = safe.replace(ch, '_')
        ws = wb.create_sheet(title=safe)
        if not rows:
            ws['A1'] = 'Sin datos para los filtros seleccionados'
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([safe_excel_value(row.get(h)) for h in headers])
        _aplicar_estilos_ws(ws, money_cols=money_cols)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


def _parse_fecha_arg(name: str, default: datetime.date | None):
    """Lee un query param YYYY-MM-DD o devuelve `default` si está vacío.
    Devuelve (date, error_response). Si error_response es None, todo bien.
    """
    raw = request.args.get(name)
    if not raw:
        return default, None
    try:
        return datetime.date.fromisoformat(raw), None
    except ValueError:
        return None, (jsonify({'detail': f"Parámetro '{name}' debe ser YYYY-MM-DD"}), 422)
