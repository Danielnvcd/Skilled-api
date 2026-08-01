"""Plantilla Excel del catálogo: definición de columnas, normalización de
celdas y construcción del workbook.

La misma plantilla sirve para importar (vacía, con instrucciones y validaciones)
y para exportar (con los productos actuales), así lo que se descarga se puede
editar y volver a subir sin tocar nada.
"""
import io

from app.models import Producto

from .._core import _es_categoria_cable


# Encabezados oficiales de la plantilla. El importador acepta también una
# variante en lowercase/sin tildes para no romper si el usuario edita los headers.
# Orden pensado para el usuario: las columnas de CABLE van JUNTO a Categoría
# (no al final). El importador empareja por NOMBRE de columna, no por posición,
# así que este orden se puede cambiar sin romper la importación.
PLANTILLA_HEADERS = [
    'Código (SKU)', 'Descripción', 'Marca', 'Categoría',
    # Solo para CABLE (categoría que contenga "cable"): obligatorios en cable,
    # se ignoran en cualquier otra categoría.
    'Tipo (cable)', 'Tamaño mm²/AWG (cable)',
    'Unidad', 'Stock Inicial',
    # Destino del stock inicial por fila (feature stock por proyecto). Opcionales:
    # si van vacías, el stock inicial cae en la bodega default y bucket general.
    # Solo aplican a productos NUEVOS (en existentes el stock no se toca).
    'Almacén', 'Proyecto',
    'Stock Mínimo', 'Precio Unitario', 'URL Imagen (opcional)',
    # Proveedor habitual del material: lo usa Compras Express para agrupar la
    # orden y mandarla por WhatsApp. Sin estas columnas, un catálogo importado
    # había que completarlo producto por producto para poder comprar.
    'Proveedor', 'Contacto proveedor',
]
# Columnas específicas de cable (para colorearlas distinto en la plantilla).
PLANTILLA_CABLE_COLS = {'Tipo (cable)', 'Tamaño mm²/AWG (cable)'}
# Tope de filas por importación. El guardia real contra archivos abusivos es el
# límite de 5 MB (un export de 8 000 productos pesa 0.29 MB), así que este número
# solo tiene que ir MUY por delante del catálogo: si se queda corto, el export
# del catálogo deja de poder reimportarse. Con 5 000 productos el export ya son
# ~5 100 filas contando las de sección.
PLANTILLA_MAX_FILAS = 15000

# Columnas del EXPORT del catálogo: subconjunto de PLANTILLA_HEADERS con SOLO lo
# que tiene sentido editar de un producto que YA existe.
#
# Quedan fuera a propósito:
#   - Stock Inicial / Almacén / Proyecto: el importador los ignora en productos
#     existentes (el stock real solo se mueve por movimientos), así que en el
#     export eran celdas que invitaban a editar algo que no iba a pasar.
#   - Stock Mínimo: se edita en la ficha del producto, no en masa por Excel.
# Como el importador empareja por NOMBRE de columna, un archivo sin ellas se
# sube igual: las ausentes se leen vacías y no pisan nada (ver `precio_provisto`
# y `stock_min_provisto` en importar.py). Las plantillas viejas con las 13
# columnas siguen importándose igual.
EXPORT_HEADERS = [
    'Código (SKU)', 'Descripción', 'Marca', 'Categoría',
    'Tipo (cable)', 'Tamaño mm²/AWG (cable)',
    'Unidad', 'Precio Unitario', 'URL Imagen (opcional)',
    'Proveedor', 'Contacto proveedor',
]

# Nombres de las columnas de destino. `prellenado` y `listas` se indexan por
# nombre de columna: una clave que no calce con el encabezado se ignora en
# silencio, así que quien las use debe pasar por estas constantes.
COL_ALMACEN = 'Almacén'
COL_PROYECTO = 'Proyecto'

# Cuántas filas se prellenan con el destino elegido al descargar la plantilla.
# Cuesta ~17 KB (6 000 filas serían 52 KB), así que conviene que sobre: una fila
# capturada MÁS ABAJO del prellenado se iría a la bodega predeterminada sin
# avisar, que es justo lo que este prellenado viene a evitar. Si alguien captura
# más de esto, copia las dos celdas hacia abajo.
PLANTILLA_FILAS_PRELLENADAS = 2000
# Tope de elementos por lista desplegable (hoja oculta `Listas`).
PLANTILLA_LISTA_MAX = 500


def _norm_header(s: str) -> str:
    """Normaliza un header para comparación: lower, sin acentos, sin espacios extra."""
    import unicodedata
    s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode('ascii')
    return s.strip().lower().replace('  ', ' ')


def _cell_str(value, maxlen=None) -> str:
    """Convierte valor de pandas a str limpio (NaN→'', strip). Trunca si maxlen."""
    import math
    if value is None:
        return ''
    try:
        if isinstance(value, float) and math.isnan(value):
            return ''
    except Exception:
        pass
    s = str(value).strip()
    if s.lower() == 'nan':
        return ''
    if maxlen and len(s) > maxlen:
        s = s[:maxlen]
    return s


def _norm_categoria(s: str) -> str:
    """Clave de comparación case/acento-insensitiva para nombres de categoría.
    Sirve solo como índice: el nombre original (con tildes y mayúsculas) se
    conserva como 'nombre canónico' en la primera ocurrencia."""
    import unicodedata
    s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode('ascii')
    return ' '.join(s.strip().lower().split())


def _categoria_similar(nueva: str, existentes) -> str | None:
    """Devuelve el nombre de categoría EXISTENTE más parecido a `nueva`, o None.

    Heurística: comparten raíz si el texto normalizado de una es subcadena del
    de la otra (con guarda de longitud). Atrapa variantes accidentales como
    'Cables' o 'Cable azul' frente a 'Cable' — para preguntarle al usuario si
    quiere crear una nueva categoría o agregar a la existente."""
    na = _norm_categoria(nueva)
    if len(na) < 3:
        return None
    mejor = None
    for ex in existentes:
        ne = _norm_categoria(ex)
        if len(ne) < 3 or ne == na:
            continue
        if ne in na or na in ne:
            # Sugerir la más corta/genérica (p. ej. 'Cable' antes que 'Cable rojo').
            if mejor is None or len(ne) < len(_norm_categoria(mejor)):
                mejor = ex
    return mejor


def _cell_number(value, default=0.0):
    """Convierte celda a float. Acepta '1,234.56' (formato MX), devuelve (valor, error_str_o_None)."""
    import math
    if value is None or value == '':
        return default, None
    try:
        if isinstance(value, float) and math.isnan(value):
            return default, None
    except Exception:
        pass
    if isinstance(value, (int, float)):
        return float(value), None
    s = str(value).strip().replace(',', '').replace('$', '').replace(' ', '')
    if not s:
        return default, None
    try:
        return float(s), None
    except ValueError:
        return default, f'no es un número válido ({value!r})'


# Marcador de fila de sección en el export: su celda Código empieza con esto,
# así el importador la reconoce y la salta (un SKU real nunca empieza con '#',
# lo prohíbe CODIGO_REGEX).
PLANTILLA_SECTION_PREFIX = '#'

# Tooltips por NOMBRE de columna (independiente del orden).
PLANTILLA_TOOLTIPS = {
    'Código (SKU)': 'Código único del producto (SKU). Acepta letras, números, -_./',
    'Descripción': 'Descripción corta del producto (máx 250 caracteres)',
    'Marca': 'OPCIONAL — Marca / fabricante del producto (ej: Cooper, Truper, Condumex). Máx 100 caracteres.',
    'Categoría': 'Categoría (ej: Tornillería, Eléctrico, Cable). Si contiene "cable", llena Tipo y Tamaño.',
    'Tipo (cable)': 'SOLO CABLE — Tipo de cable (THHN, THW, desnudo…). Llénalo SOLO si la Categoría es de cable; en otras categorías déjalo vacío.',
    'Tamaño mm²/AWG (cable)': 'SOLO CABLE — Tamaño en mm²/AWG (12, 2/0, 500 kcmil…). Llénalo SOLO si la Categoría es de cable; en otras categorías déjalo vacío.',
    'Unidad': 'Unidad de medida: Pza, Kg, Mts, Lts, Caja, Bote… (en cable se pone M sola).',
    'Stock Inicial': 'Cantidad inicial en almacén (número >= 0). En productos EXISTENTES este valor NO cambia el stock.',
    'Almacén': 'Bodega donde llega el stock inicial. Si elegiste una al descargar, ya viene llena; para cambiarla usa la lista de la celda. Vacío = bodega predeterminada. Solo aplica a productos NUEVOS.',
    'Proyecto': 'Proyecto al que se aparta el stock inicial. Si elegiste uno al descargar, ya viene lleno; para cambiarlo usa la lista de la celda. Vacío = General (libre). Solo aplica a productos NUEVOS.',
    'Stock Mínimo': 'Cantidad mínima antes de alertar de bajo stock (número >= 0)',
    'Precio Unitario': 'Precio unitario del material (número >= 0). Se usa para los costos por proyecto.',
    'URL Imagen (opcional)': 'OPCIONAL — URL HTTPS de la imagen del producto. Ej: https://cdn.miempresa.com/tornillo.jpg',
    'Proveedor': 'OPCIONAL — Proveedor habitual del material (ej: Ferretería López). Se usa en Compras Express para agrupar la orden.',
    'Contacto proveedor': 'OPCIONAL — Teléfono o correo del proveedor. Con un teléfono, Compras Express arma el mensaje de WhatsApp.',
}

# Anchos por NOMBRE de columna.
PLANTILLA_WIDTHS = {
    'Código (SKU)': 18, 'Descripción': 40, 'Marca': 18, 'Categoría': 18,
    'Tipo (cable)': 16, 'Tamaño mm²/AWG (cable)': 20, 'Unidad': 10,
    'Stock Inicial': 13, 'Almacén': 18, 'Proyecto': 18,
    'Stock Mínimo': 13, 'Precio Unitario': 14,
    'URL Imagen (opcional)': 42,
    'Proveedor': 24, 'Contacto proveedor': 22,
}


def _construir_plantilla_workbook(rows=None, instruccion=None, headers=None,
                                  prellenado=None, listas=None):
    """Construye el .xlsx de la plantilla y devuelve un BytesIO listo para
    `send_file`.

    `rows` (opcional) permite exportar el catálogo ya lleno. Cada elemento es:
      - una lista de valores alineada a `headers` (un producto), o
      - un dict {'__section__': nombre, 'count': n} → fila de sección resaltada
        (para agrupar por categoría; el importador la salta), o
      - None → fila en blanco separadora.

    `headers` (opcional) permite armar un SUBCONJUNTO de columnas — el export
    del catálogo usa EXPORT_HEADERS. Todo se referencia por NOMBRE de columna
    (no por letra fija), así que anchos, tooltips y validaciones se aplican solo
    a las columnas presentes y el orden se puede cambiar sin romper nada.

    `prellenado` {columna: valor} escribe ese valor en las primeras
    PLANTILLA_FILAS_PRELLENADAS filas vacías: sirve para el destino del stock
    inicial (Almacén / Proyecto) elegido al descargar la plantilla.

    `listas` {columna: [valores]} agrega listas desplegables leídas de una hoja
    oculta `Listas`. Se usa una hoja en vez de la lista escrita en línea porque
    esta última tiene un tope de 255 caracteres y se rompe con nombres que
    llevan coma.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.comments import Comment

    rows = rows or []
    headers = headers or PLANTILLA_HEADERS
    ncols = len(headers)
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    # Índice 0-based de Categoría en la fila de datos (None si no se exporta).
    cat_pos = col_idx['Categoría'] - 1 if 'Categoría' in col_idx else None

    def _letter(nombre):
        return get_column_letter(col_idx[nombre])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Materiales'

    # Estilos
    header_fill = PatternFill('solid', fgColor='1E40AF')        # azul (columnas normales)
    cable_header_fill = PatternFill('solid', fgColor='B45309')  # ámbar/cobre (columnas de cable)
    cable_cell_fill = PatternFill('solid', fgColor='FEF3C7')    # ámbar claro (celdas de cable)
    section_fill = PatternFill('solid', fgColor='E0E7FF')       # índigo suave (fila de sección)
    prefill_fill = PatternFill('solid', fgColor='F1F5F9')       # gris (celda ya resuelta)
    header_font = Font(bold=True, color='FFFFFF', size=11)
    title_font = Font(bold=True, color='111827', size=14)
    instr_font = Font(italic=True, color='6B7280', size=10)
    section_font = Font(bold=True, color='3730A3', size=10)
    thin = Side(border_style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_align = Alignment(vertical='top', wrap_text=True)  # evita que el texto largo invada la celda de al lado

    # Fila 1: título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value='Plantilla de Importación de Materiales — SKILLED')
    c.font = title_font
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Fila 2: instrucciones
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=instruccion or (
        '⚠ No alteres los encabezados de la fila 4. Precio y URL Imagen son opcionales '
        '(URL solo HTTPS o /static/...png). Si un SKU ya existe, se ACTUALIZA con los datos '
        'de esta fila (el stock NO se toca). Las columnas de CABLE (en ámbar) van junto a '
        'Categoría: llénalas SOLO si la categoría contiene "cable" — ahí la unidad se pone en '
        'M sola. En el resto de categorías déjalas vacías.'))
    c.font = instr_font
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 34

    # Fila 3 vacía como separador
    ws.row_dimensions[3].height = 6

    # Fila 4: encabezados oficiales (cable en ámbar) + tooltips + anchos
    for col, header in enumerate(headers, 1):
        es_cable_col = header in PLANTILLA_CABLE_COLS
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = cable_header_fill if es_cable_col else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = PLANTILLA_WIDTHS.get(header, 18)
        if header in PLANTILLA_TOOLTIPS:
            cell.comment = Comment(PLANTILLA_TOOLTIPS[header], 'Plantilla SKILLED')

    ws.row_dimensions[4].height = 42
    ws.freeze_panes = 'A5'

    # Filas de datos / secciones (solo en exportación).
    r = 5
    for item in rows:
        if item is None:
            # Fila en blanco separadora (el importador la ignora).
            r += 1
            continue
        if isinstance(item, dict) and '__section__' in item:
            # Fila de sección: combinada, resaltada y con el marcador '#' en A
            # para que el importador la salte.
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            cell = ws.cell(row=r, column=1,
                           value=f"{PLANTILLA_SECTION_PREFIX}  ▸  {item['__section__']}  ({item.get('count', 0)})")
            cell.fill = section_fill
            cell.font = section_font
            cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[r].height = 20
            r += 1
            continue
        # Producto: escribir celdas con wrap + borde; tinte ámbar en cable si aplica.
        es_cable = _es_categoria_cable(
            item[cat_pos] if cat_pos is not None and len(item) > cat_pos else ''
        )
        for col, valor in enumerate(item, 1):
            cell = ws.cell(row=r, column=col, value=valor)
            cell.alignment = data_align
            cell.border = border
            if headers[col - 1] in PLANTILLA_CABLE_COLS and es_cable:
                cell.fill = cable_cell_fill
        r += 1

    # Prellenado del destino (Almacén / Proyecto) elegido al descargar. Se pinta
    # en gris para que se lea como "ya viene resuelto"; el usuario solo teclea el
    # material. OJO: una fila con SOLO estas celdas llenas sigue siendo una fila
    # vacía para el importador (mira código/descripción/categoría/unidad), así
    # que las filas prellenadas de más no generan basura ni errores.
    destino = {c: v for c, v in (prellenado or {}).items() if c in col_idx and v}
    for i in range(PLANTILLA_FILAS_PRELLENADAS if destino else 0):
        for nombre, valor in destino.items():
            cell = ws.cell(row=r + i, column=col_idx[nombre], value=valor)
            cell.alignment = data_align
            cell.border = border
            cell.fill = prefill_fill

    # Data validations (por nombre → letra). Cubrir al menos 1000 filas o todas.
    last_row = 4 + max(1000, len(rows) + (PLANTILLA_FILAS_PRELLENADAS if destino else 0))

    num_dv = DataValidation(type='decimal', operator='greaterThanOrEqual', formula1=0,
                             showErrorMessage=True, errorTitle='Stock inválido',
                             error='El stock debe ser un número mayor o igual a 0.')
    num_cols = [n for n in ('Stock Inicial', 'Stock Mínimo', 'Precio Unitario') if n in col_idx]
    for nombre in num_cols:
        L = _letter(nombre)
        num_dv.add(f'{L}5:{L}{last_row}')
    if num_cols:
        ws.add_data_validation(num_dv)

    if 'Descripción' in col_idx:
        desc_dv = DataValidation(type='textLength', operator='between', formula1=1, formula2=250,
                                  showErrorMessage=True, errorTitle='Descripción inválida',
                                  error='Entre 1 y 250 caracteres.')
        Ld = _letter('Descripción')
        desc_dv.add(f'{Ld}5:{Ld}{last_row}')
        ws.add_data_validation(desc_dv)

    if 'Código (SKU)' in col_idx:
        sku_dv = DataValidation(type='textLength', operator='between', formula1=1, formula2=50,
                                  showErrorMessage=True, errorTitle='SKU inválido',
                                  error='Entre 1 y 50 caracteres.')
        La = _letter('Código (SKU)')
        sku_dv.add(f'{La}5:{La}{last_row}')
        ws.add_data_validation(sku_dv)

    # Listas desplegables desde una hoja oculta: el usuario elige de valores que
    # EXISTEN en el sistema, así no hay errores de dedo al importar.
    if listas:
        activas = {
            c: [str(v) for v in vals if v][:PLANTILLA_LISTA_MAX]
            for c, vals in listas.items() if c in col_idx and vals
        }
        activas = {c: v for c, v in activas.items() if v}
        if activas:
            ws_listas = wb.create_sheet('Listas')
            for i, (nombre, valores) in enumerate(activas.items(), 1):
                letra_lista = get_column_letter(i)
                ws_listas.cell(row=1, column=i, value=nombre)
                for j, valor in enumerate(valores, start=2):
                    ws_listas.cell(row=j, column=i, value=valor)
                rango = f'Listas!${letra_lista}$2:${letra_lista}${1 + len(valores)}'
                dv = DataValidation(type='list', formula1=rango, allow_blank=True,
                                    showErrorMessage=True)
                dv.errorTitle = f'{nombre} inválido'
                dv.error = (f'Elige un {nombre.lower()} de la lista: debe existir en el '
                            f'sistema tal cual. Déjalo vacío para usar el predeterminado.')
                ws.add_data_validation(dv)
                L = _letter(nombre)
                dv.add(f'{L}5:{L}{last_row}')
            ws_listas.sheet_state = 'hidden'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _producto_export_row(p: Producto, headers=None) -> list:
    """Fila de exportación alineada a `headers` (por defecto EXPORT_HEADERS).

    Se arma por NOMBRE de columna y luego se proyecta al orden pedido: así
    quitar, agregar o mover columnas no puede desalinear los valores.
    """
    valores = {
        'Código (SKU)': p.codigo or '',
        'Descripción': p.descripcion or '',
        'Marca': p.marca or '',
        'Categoría': p.categoria or '',
        'Tipo (cable)': p.cable_tipo or '',
        'Tamaño mm²/AWG (cable)': p.cable_calibre or '',
        'Unidad': p.unidad or '',
        'Stock Inicial': float(p.stock_actual or 0),
        # Almacén / Proyecto: vacíos si alguien los pide en el export — el stock
        # de un producto existente puede estar repartido en varios buckets, así
        # que no hay un único destino que exportar, y al reimportar NO se reubica.
        'Almacén': '',
        'Proyecto': '',
        'Stock Mínimo': float(p.stock_minimo or 0),
        'Precio Unitario': float(p.precio_unitario or 0),
        'URL Imagen (opcional)': p.imagen_url or '',
        'Proveedor': p.proveedor_default_nombre or '',
        'Contacto proveedor': p.proveedor_default_contacto or '',
    }
    return [valores.get(h, '') for h in (headers or EXPORT_HEADERS)]
