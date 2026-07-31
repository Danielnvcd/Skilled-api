"""Carga masiva por Excel: plantilla de tres columnas e importación.

La plantilla de catálogo pide trece columnas porque da de alta materiales; aquí
el material ya existe y lo único que varía es cuánto y dónde.
"""
import io

from flask import jsonify, request

from app.extensions import db
from app.models import Almacen, Producto, Proyecto, StockAlmacenProyecto

from .._core import bp, _require_inventario_admin
from .validacion import _resolver_lineas, _resumen


# ── Plantilla de Excel para la carga masiva ─────────────────────────────────

def _construir_plantilla(proyecto: Proyecto) -> "io.BytesIO":  # noqa: F821
    """Excel de TRES columnas: SKU, Cantidad, Bodega.

    La plantilla de catálogo pide trece columnas porque sirve para dar de alta
    materiales. Aquí el material ya existe: lo único que varía es cuánto y
    dónde. Pedir descripción, marca, unidad o precio otra vez sería hacer
    teclear datos que el sistema ya tiene.

    El PROYECTO tampoco es columna: la plantilla se descarga desde dentro del
    proyecto, así que ya se sabe cuál es. Ponerlo sería una oportunidad más de
    equivocarse.

    Se pre-llena con lo que el proyecto YA tiene, para que ajustar cantidades no
    obligue a teclear los SKU.
    """
    import io

    import openpyxl
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    COLS = ['SKU', 'Cantidad', 'Bodega']
    ANCHOS = {'SKU': 22, 'Cantidad': 12, 'Bodega': 24}
    AYUDAS = {
        'SKU': 'Código del material tal como está en el catálogo. Debe existir.',
        'Cantidad': 'Cuánto se asigna a este proyecto. Número mayor que cero.',
        'Bodega': 'Nombre de la bodega. Si se deja vacío se usa la predeterminada.',
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Asignación'

    azul = PatternFill('solid', fgColor='1E40AF')
    gris = PatternFill('solid', fgColor='F1F5F9')
    thin = Side(border_style='thin', color='CBD5E1')
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Fila 1: a qué proyecto va todo esto. Es contexto, no un dato a llenar.
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    c = ws.cell(row=1, column=1,
                value=f'Asignar material a: {proyecto.numero_proyecto}'
                      f'{" — " + proyecto.nombre if proyecto.nombre else ""}')
    c.font = Font(bold=True, color='111827', size=14)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLS))
    c = ws.cell(row=2, column=1, value=(
        'Llena una fila por material. El proyecto ya está definido — no hace falta '
        'repetirlo. Al subir el archivo verás una vista previa antes de aplicar nada.'
    ))
    c.font = Font(italic=True, color='6B7280', size=10)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[2].height = 28

    # Fila 3: encabezados.
    for i, col in enumerate(COLS, 1):
        cel = ws.cell(row=3, column=i, value=col)
        cel.fill = azul
        cel.font = Font(bold=True, color='FFFFFF', size=11)
        cel.alignment = Alignment(horizontal='center', vertical='center')
        cel.border = borde
        cel.comment = Comment(AYUDAS[col], 'SKILLED')
        ws.column_dimensions[get_column_letter(i)].width = ANCHOS[col]
    ws.freeze_panes = 'A4'

    # Pre-llenado con lo que el proyecto ya tiene.
    existentes = (
        db.session.query(Producto.codigo, StockAlmacenProyecto.cantidad, Almacen.nombre)
        .join(Producto, Producto.id == StockAlmacenProyecto.producto_id)
        .join(Almacen, Almacen.id == StockAlmacenProyecto.almacen_id)
        .filter(
            StockAlmacenProyecto.proyecto_id == proyecto.id,
            StockAlmacenProyecto.cantidad > 0,
            Producto.activo == True,   # noqa: E712
            Almacen.activo == True,    # noqa: E712
        )
        .order_by(Producto.codigo)
        .all()
    )
    fila = 4
    for codigo, cantidad, bodega in existentes:
        ws.cell(row=fila, column=1, value=codigo).border = borde
        # Se deja la CANTIDAD VACÍA a propósito: pre-llenarla con lo que ya
        # tiene invitaría a subir el archivo sin tocarlo y duplicar todo el
        # material del proyecto. El SKU y la bodega sí se dan hechos, que es
        # lo tedioso de teclear.
        ws.cell(row=fila, column=2).border = borde
        ws.cell(row=fila, column=3, value=bodega).border = borde
        for col in range(1, len(COLS) + 1):
            ws.cell(row=fila, column=col).fill = gris
        fila += 1

    ultima = max(fila + 200, 210)

    # Validación: cantidad numérica y positiva.
    dv = DataValidation(type='decimal', operator='greaterThan', formula1='0',
                        allow_blank=True, showErrorMessage=True)
    dv.error = 'La cantidad debe ser un número mayor que cero.'
    dv.errorTitle = 'Cantidad inválida'
    ws.add_data_validation(dv)
    dv.add(f'B4:B{ultima}')

    # Lista de bodegas válidas, para no depender de que se escriban bien.
    nombres = [a.nombre for a in Almacen.query.filter(Almacen.activo == True).all()]  # noqa: E712
    if nombres and sum(len(n) + 1 for n in nombres) < 250:
        dv_b = DataValidation(type='list', formula1='"' + ','.join(nombres) + '"',
                              allow_blank=True, showErrorMessage=True)
        dv_b.error = 'Elige una bodega de la lista.'
        dv_b.errorTitle = 'Bodega inválida'
        ws.add_data_validation(dv_b)
        dv_b.add(f'C4:C{ultima}')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@bp.route('/proyectos-materiales/<int:proyecto_id>/plantilla-asignacion', methods=['GET'])
@_require_inventario_admin
def plantilla_asignacion(proyecto_id: int):
    """Descarga la plantilla de asignación de ESTE proyecto."""
    from flask import send_file

    proyecto = Proyecto.query.get(proyecto_id)
    if not proyecto:
        return jsonify({'detail': 'Proyecto no encontrado'}), 404

    try:
        buf = _construir_plantilla(proyecto)
    except ImportError:
        return jsonify({'detail': 'openpyxl no instalado en el servidor'}), 500

    seguro = ''.join(ch for ch in (proyecto.numero_proyecto or 'proyecto')
                     if ch.isalnum() or ch in '-_')
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'asignacion_{seguro or "proyecto"}.xlsx',
    )


# ── Importar la plantilla llena ─────────────────────────────────────────────

@bp.route('/proyectos-materiales/<int:proyecto_id>/asignar/importar', methods=['POST'])
@_require_inventario_admin
def importar_asignacion(proyecto_id: int):
    """Lee el Excel y devuelve la MISMA previsualización que `previsualizar`.

    No escribe nada. El archivo se traduce a líneas y se pasa por
    `_resolver_lineas`, igual que la captura a mano: así el usuario ve exactamente
    lo que va a pasar, y confirmar dispara el mismo `/asignar` de siempre. Subir
    un archivo no es un camino distinto al de capturar — es solo otra forma de
    llenar las líneas.
    """
    import io as _io

    proyecto = Proyecto.query.get(proyecto_id)
    if not proyecto:
        return jsonify({'detail': 'Proyecto no encontrado'}), 404

    try:
        import pandas as pd
    except ImportError:
        return jsonify({'detail': 'pandas no instalado en el servidor'}), 500

    file = request.files.get('archivo')
    if not file or not file.filename:
        return jsonify({'detail': 'No se envió archivo'}), 400
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'detail': 'Formato no válido. Debe ser .xlsx o .xls'}), 400

    file.stream.seek(0, _io.SEEK_END)
    size = file.stream.tell()
    file.stream.seek(0)
    if size > 5 * 1024 * 1024:
        return jsonify({'detail': 'Archivo demasiado grande. Máximo 5 MB.'}), 413
    if size < 100:
        return jsonify({'detail': 'Archivo vacío o corrupto.'}), 400

    try:
        raw = pd.read_excel(file, header=None, nrows=2100)
    except Exception as e:
        return jsonify({
            'detail': 'No se pudo leer el Excel. Usa la plantilla y guárdalo como .xlsx.',
            'tecnico': str(e)[:200],
        }), 400

    def _norm(v):
        import unicodedata
        s = unicodedata.normalize('NFKD', str(v or ''))
        return ''.join(c for c in s if not unicodedata.combining(c)).strip().lower()

    # Los encabezados van en la fila 3 de la plantilla, pero se buscan por
    # contenido: si alguien inserta o borra una fila arriba, el archivo sigue
    # sirviendo. Solo SKU y Cantidad son obligatorios — sin bodega se usa la
    # predeterminada, que es justo lo que hace la captura a mano.
    ALIAS = {
        'sku': 'sku', 'codigo': 'sku', 'codigo (sku)': 'sku', 'clave': 'sku',
        'cantidad': 'cantidad', 'cant': 'cantidad', 'cantidad a asignar': 'cantidad',
        'bodega': 'almacen', 'almacen': 'almacen',
    }
    fila_encabezado, columnas = None, {}
    for ridx in range(min(12, len(raw))):
        encontrado = {}
        for i, val in enumerate(raw.iloc[ridx].tolist()):
            campo = ALIAS.get(_norm(val))
            if campo and campo not in encontrado:
                encontrado[campo] = i
        if 'sku' in encontrado and 'cantidad' in encontrado:
            fila_encabezado, columnas = ridx, encontrado
            break

    if fila_encabezado is None:
        return jsonify({
            'detail': 'No se encontraron las columnas SKU y Cantidad. '
                      'Descarga la plantilla del proyecto y llénala sin borrar los encabezados.',
        }), 400

    datos = raw.iloc[fila_encabezado + 1:]
    if len(datos) > 2000:
        return jsonify({'detail': f'Demasiadas filas ({len(datos)}). Máximo 2000.'}), 400

    lineas, vacias = [], 0
    for _, fila in datos.iterrows():
        def _celda(campo):
            i = columnas.get(campo)
            if i is None or i >= len(fila):
                return ''
            v = fila.iloc[i]
            return '' if pd.isna(v) else str(v).strip()

        sku, cantidad = _celda('sku'), _celda('cantidad')
        # Fila totalmente vacía = final del llenado, no un error. La plantilla
        # trae 200 filas en blanco a propósito; reportarlas como errores llenaría
        # la vista previa de ruido y escondería los problemas de verdad.
        if not sku and not cantidad:
            vacias += 1
            continue
        # SKU prellenado al que no le pusieron cantidad: tampoco es un error,
        # es material del proyecto que esta vez no se toca.
        if sku and not cantidad:
            vacias += 1
            continue
        lineas.append({'sku': sku, 'cantidad': cantidad, 'almacen': _celda('almacen')})

    if not lineas:
        return jsonify({
            'detail': 'El archivo no tiene ninguna fila con cantidad. '
                      'Escribe cuánto asignar en la columna Cantidad.',
        }), 422

    origen = (request.form.get('origen') or 'general').strip().lower()
    if origen not in ('general', 'entrada'):
        origen = 'general'
    modo = (request.form.get('modo') or 'sumar').strip().lower()
    if modo not in ('sumar', 'reemplazar'):
        modo = 'sumar'

    plan = _resolver_lineas(proyecto.id, lineas, origen, modo)
    return jsonify({
        'proyecto': {'id': proyecto.id, 'numero_proyecto': proyecto.numero_proyecto,
                     'nombre': proyecto.nombre or ''},
        'origen': origen,
        'modo': modo,
        'lineas': plan,
        'resumen': _resumen(plan),
        'filas_ignoradas': vacias,
    })
