"""Proyectos + Categorías + CategoriaConfig + Importación masiva desde Excel."""
import io
import re
from decimal import Decimal

from flask import jsonify, request, send_file, Response, current_app
from sqlalchemy import distinct as sql_distinct

from app.extensions import db, limiter
from app.models import (
    Producto, Proyecto, CategoriaConfig, StockPorAlmacen, StockAlmacenProyecto,
    SolicitudMaterial, SolicitudMaterialDetalle,
)

from ._core import (
    bp,
    _require_login, _require_inventario, _require_inventario_admin,
    _parse_or_422,
    CategoriaConfigUpsertSchema,
    CODIGO_REGEX, _IMAGEN_URL_REGEX,
    _audit, _almacen_default_id,
    _INV_ROLES,
)
from app.realtime import emit_to_role
# Reglas de cable compartidas (productos ya está cargado cuando se importa este
# módulo — ver orden en inventario_api/__init__.py).
from .productos import _es_categoria_cable, _CABLE_UNIDAD


# ─── Proyectos ────────────────────────────────────────────────────────────────

@bp.route('/proyectos/', methods=['GET'])
@_require_login
def get_proyectos():
    proyectos = (
        Proyecto.query
        .filter(Proyecto.activo == True)
        .order_by(Proyecto.numero_proyecto)
        .all()
    )
    return jsonify([
        {'id': p.id, 'numero_proyecto': p.numero_proyecto, 'nombre': p.nombre or ''}
        for p in proyectos
    ])


# ─── Categorías ───────────────────────────────────────────────────────────────

@bp.route('/categorias/', methods=['GET'])
@_require_inventario
def get_categorias():
    """Devuelve la unión de categorías presentes en el catálogo de productos
    y las registradas en `categorias_config` (admin pudo crear categorías sin
    haber capturado aún ningún producto)."""
    prod_rows = (
        db.session.query(sql_distinct(Producto.categoria))
        .filter(Producto.activo == True, Producto.categoria != None, Producto.categoria != '')
        .all()
    )
    cfg_rows = db.session.query(CategoriaConfig.nombre).all()
    nombres = {r[0] for r in prod_rows} | {r[0] for r in cfg_rows}
    return jsonify(sorted(nombres))


@bp.route('/categorias/resumen', methods=['GET'])
@_require_inventario
def get_categorias_resumen():
    """Conteo de productos por categoría (total + cuántos bajo mínimo) para las
    tarjetas del catálogo. Server-side: evita descargar miles de productos al
    cliente solo para contarlos. Incluye categorías de `categorias_config` que
    aún no tienen productos (total 0)."""
    rows = (
        db.session.query(
            Producto.categoria,
            db.func.count(Producto.id),
            db.func.coalesce(
                db.func.sum(
                    db.case((Producto.stock_actual <= Producto.stock_minimo, 1), else_=0)
                ), 0,
            ),
        )
        .filter(Producto.activo == True, Producto.categoria != None, Producto.categoria != '')  # noqa: E711,E712
        .group_by(Producto.categoria)
        .all()
    )
    resumen = {
        nombre: {'nombre': nombre, 'total': int(total or 0), 'bajo_minimo': int(bajos or 0)}
        for nombre, total, bajos in rows
    }
    # Categorías registradas en config pero sin productos todavía.
    for (nombre,) in db.session.query(CategoriaConfig.nombre).all():
        if nombre and nombre not in resumen:
            resumen[nombre] = {'nombre': nombre, 'total': 0, 'bajo_minimo': 0}

    return jsonify(sorted(resumen.values(), key=lambda r: r['nombre'].lower()))


# ─── CategoriaConfig (metadatos visuales por categoría) ──────────────────────

def _categoria_config_to_dict(c: CategoriaConfig) -> dict:
    return {
        'nombre': c.nombre,
        'imagen_url': c.imagen_url,
        'imagen_estado': c.imagen_estado,
        'updated_at': c.updated_at.isoformat() if c.updated_at else None,
    }


@bp.route('/categorias-config/', methods=['GET'])
@_require_login
def get_categorias_config():
    """Lista todas las configuraciones (imagen, etc.) por nombre de categoría.
    Lectura abierta a cualquier usuario autenticado: el dashboard de inventario
    también lo consume desde el rol solicitante_material."""
    rows = CategoriaConfig.query.order_by(CategoriaConfig.nombre).all()
    return jsonify([_categoria_config_to_dict(c) for c in rows])


@bp.route('/categorias-config/<string:nombre>', methods=['PUT'])
@_require_inventario_admin
def upsert_categoria_config(nombre: str):
    """Crea o actualiza la config de la categoría con `nombre`. Si imagen_url
    viene null o vacío, persiste null (UI lo trata como "quitar imagen")."""
    nombre = (nombre or '').strip()
    if not nombre or len(nombre) > 100:
        return jsonify({'detail': "Nombre de categoría inválido (1..100 caracteres)"}), 422

    data, err = _parse_or_422(CategoriaConfigUpsertSchema(), request.get_json(silent=True))
    if err: return err

    imagen = (data.get('imagen_url') or '').strip() or None

    cfg = CategoriaConfig.query.filter(CategoriaConfig.nombre == nombre).first()
    if cfg is None:
        cfg = CategoriaConfig(
            nombre=nombre,
            imagen_url=imagen,
            created_by_id=request.current_user.id,
        )
        db.session.add(cfg)
        _audit(request.current_user, f"Categoría '{nombre}' creada/actualizada")
    else:
        cfg.imagen_url = imagen
        _audit(request.current_user, f"Categoría '{nombre}' actualizada")

    db.session.commit()
    db.session.refresh(cfg)
    # Pipeline de imágenes → R2: si la imagen de categoría es URL externa, se
    # marca y se encola su descarga a WebP+R2 (no-op salvo producción con R2).
    from .imagenes import marcar_para_sync, encolar_sync
    if marcar_para_sync(cfg, cfg.imagen_url):
        db.session.commit()
        encolar_sync(request.current_user.id, [('categoria', cfg.id)])
    return jsonify(_categoria_config_to_dict(cfg))


@bp.route('/categorias-config/<string:nombre>', methods=['DELETE'])
@_require_inventario_admin
def delete_categoria_config(nombre: str):
    """Elimina una categoría.

    Por defecto solo borra la fila de config (metadatos visuales) y NO afecta
    productos. Para eliminar también todos los productos de la categoría —el
    flujo destructivo que el SPA debe confirmar con una advertencia— pasar
    `?con_productos=1`. En ese caso:
      - Se hace soft-delete (activo=False) de cada producto, igual que
        `delete_producto`, para preservar el histórico de movimientos y
        solicitudes (no se rompe el kardex ni los folios).
      - Se liberan las reservas pendientes de stock de esos productos para que
        no queden unidades apartadas por solicitudes sobre un producto muerto.
      - Se borra también la config de la categoría.
    """
    nombre = (nombre or '').strip()
    con_productos = (request.args.get('con_productos') or '').lower() in ('1', 'true', 'yes')

    cfg = CategoriaConfig.query.filter(CategoriaConfig.nombre == nombre).first()

    if not con_productos:
        if not cfg:
            return jsonify({'detail': 'Categoría no encontrada en config'}), 404
        db.session.delete(cfg)
        _audit(request.current_user, f"Categoría '{nombre}' config eliminada")
        db.session.commit()
        return Response(status=204)

    # Borrado en cascada: productos + config.
    productos = (
        Producto.query
        .filter(Producto.categoria == nombre, Producto.activo == True)  # noqa: E712
        .all()
    )
    # Si no hay ni productos ni config, la categoría no existe.
    if not productos and not cfg:
        return jsonify({'detail': 'Categoría no encontrada'}), 404

    # Guardia: no borrar productos que tengan entregas pendientes. Una solicitud
    # APROBADA y no entregada apartó stock y el solicitante la espera; si
    # desactiváramos el producto, la entrega posterior generaría una SALIDA
    # sobre un producto muerto y dejaría la solicitud colgada. Bloqueamos con
    # 409 y devolvemos qué solicitudes resolver primero (opción conservadora:
    # el almacenista decide, no cancelamos pedidos ajenos automáticamente).
    prod_ids = [p.id for p in productos]
    if prod_ids:
        # pendiente por línea = base − entregada, con base = aprobada (o
        # solicitada si aún no se tocó la aprobación, caso pre-8b). Mismo
        # criterio que `_reservas_de_solicitud`.
        base = db.func.coalesce(
            db.func.nullif(SolicitudMaterialDetalle.cantidad_aprobada, 0),
            SolicitudMaterialDetalle.cantidad_solicitada,
        )
        pendiente = base - db.func.coalesce(SolicitudMaterialDetalle.cantidad_entregada, 0)
        filas_bloqueo = (
            db.session.query(
                SolicitudMaterial.id,
                Producto.codigo,
                Producto.descripcion,
            )
            .join(SolicitudMaterialDetalle, SolicitudMaterialDetalle.solicitud_id == SolicitudMaterial.id)
            .join(Producto, Producto.id == SolicitudMaterialDetalle.producto_id)
            .filter(
                SolicitudMaterial.estatus == 'APROBADA',
                SolicitudMaterialDetalle.producto_id.in_(prod_ids),
                pendiente > 0,
            )
            .distinct()
            .all()
        )
        if filas_bloqueo:
            solicitudes_ids = sorted({r.id for r in filas_bloqueo})
            return jsonify({
                'detail': (
                    'No se puede eliminar la categoría: tiene productos con entregas '
                    f'pendientes en {len(solicitudes_ids)} solicitud(es) aprobada(s). '
                    'Entrega o rechaza esas solicitudes antes de borrar.'
                ),
                'codigo': 'ENTREGAS_PENDIENTES',
                'solicitudes': [f'SOL-{sid:06d}' for sid in solicitudes_ids],
                'productos': sorted({f'{r.codigo} — {r.descripcion}' for r in filas_bloqueo}),
            }), 409

    eliminados = 0
    for prod in productos:
        prod.activo = False  # Soft delete: conserva histórico (igual que delete_producto)
        # Libera la reserva que esté apartando para que no quede stock fantasma
        # apartado por solicitudes sobre un producto desactivado.
        if prod.stock_reservado:
            prod.stock_reservado = Decimal('0')
        eliminados += 1

    if cfg:
        db.session.delete(cfg)

    _audit(
        request.current_user,
        f"Categoría '{nombre}' eliminada con sus productos ({eliminados} desactivados)",
    )
    db.session.commit()

    # Websockets-first: refresca catálogos abiertos en otras sesiones. El front
    # invalida el namespace completo de productos con un solo emit (igual que la
    # importación masiva) y refresca las tarjetas de categorías.
    emit_to_role(_INV_ROLES, 'producto:changed', {
        'action': 'bulk_delete', 'categoria': nombre, 'count': eliminados,
    })

    return jsonify({
        'categoria': nombre,
        'productos_eliminados': eliminados,
        'config_eliminada': cfg is not None,
    })


# ─── Importar materiales desde Excel ─────────────────────────────────────────

# Encabezados oficiales de la plantilla. El importador acepta también una
# variante en lowercase/sin tildes para no romper si el usuario edita los headers.
# Orden pensado para el usuario: las columnas de CABLE van JUNTO a Categoría
# (no al final). El importador empareja por NOMBRE de columna, no por posición,
# así que este orden se puede cambiar sin romper la importación.
PLANTILLA_HEADERS = [
    'Código (SKU)', 'Descripción', 'Marca', 'Categoría',
    # Solo para CABLE (categoría que contenga "cable"): obligatorios en cable,
    # se ignoran en cualquier otra categoría.
    'Tipo (cable)', 'Tamaño mm²/AWG (cable)',
    'Unidad', 'Stock Inicial',
    # Destino del stock inicial por fila (feature stock por proyecto). Opcionales:
    # si van vacías, el stock inicial cae en la bodega default y bucket general.
    # Solo aplican a productos NUEVOS (en existentes el stock no se toca).
    'Almacén', 'Proyecto',
    'Stock Mínimo', 'Precio Unitario', 'URL Imagen (opcional)',
]
# Columnas específicas de cable (para colorearlas distinto en la plantilla).
PLANTILLA_CABLE_COLS = {'Tipo (cable)', 'Tamaño mm²/AWG (cable)'}
PLANTILLA_MAX_FILAS = 6000  # Tope anti-DoS; holgado para export con filas de sección


def _norm_header(s: str) -> str:
    """Normaliza un header para comparación: lower, sin acentos, sin espacios extra."""
    import unicodedata
    s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode('ascii')
    return s.strip().lower().replace('  ', ' ')


def _cell_str(value, maxlen=None) -> str:
    """Convierte valor de pandas a str limpio (NaN→'', strip). Trunca si maxlen."""
    import math
    if value is None:
        return ''
    try:
        if isinstance(value, float) and math.isnan(value):
            return ''
    except Exception:
        pass
    s = str(value).strip()
    if s.lower() == 'nan':
        return ''
    if maxlen and len(s) > maxlen:
        s = s[:maxlen]
    return s


def _norm_categoria(s: str) -> str:
    """Clave de comparación case/acento-insensitiva para nombres de categoría.
    Sirve solo como índice: el nombre original (con tildes y mayúsculas) se
    conserva como 'nombre canónico' en la primera ocurrencia."""
    import unicodedata
    s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii', 'ignore').decode('ascii')
    return ' '.join(s.strip().lower().split())


def _categoria_similar(nueva: str, existentes) -> str | None:
    """Devuelve el nombre de categoría EXISTENTE más parecido a `nueva`, o None.

    Heurística: comparten raíz si el texto normalizado de una es subcadena del
    de la otra (con guarda de longitud). Atrapa variantes accidentales como
    'Cables' o 'Cable azul' frente a 'Cable' — para preguntarle al usuario si
    quiere crear una nueva categoría o agregar a la existente."""
    na = _norm_categoria(nueva)
    if len(na) < 3:
        return None
    mejor = None
    for ex in existentes:
        ne = _norm_categoria(ex)
        if len(ne) < 3 or ne == na:
            continue
        if ne in na or na in ne:
            # Sugerir la más corta/genérica (p. ej. 'Cable' antes que 'Cable rojo').
            if mejor is None or len(ne) < len(_norm_categoria(mejor)):
                mejor = ex
    return mejor


def _cell_number(value, default=0.0):
    """Convierte celda a float. Acepta '1,234.56' (formato MX), devuelve (valor, error_str_o_None)."""
    import math
    if value is None or value == '':
        return default, None
    try:
        if isinstance(value, float) and math.isnan(value):
            return default, None
    except Exception:
        pass
    if isinstance(value, (int, float)):
        return float(value), None
    s = str(value).strip().replace(',', '').replace('$', '').replace(' ', '')
    if not s:
        return default, None
    try:
        return float(s), None
    except ValueError:
        return default, f'no es un número válido ({value!r})'


# Marcador de fila de sección en el export: su celda Código empieza con esto,
# así el importador la reconoce y la salta (un SKU real nunca empieza con '#',
# lo prohíbe CODIGO_REGEX).
PLANTILLA_SECTION_PREFIX = '#'

# Tooltips por NOMBRE de columna (independiente del orden).
PLANTILLA_TOOLTIPS = {
    'Código (SKU)': 'Código único del producto (SKU). Acepta letras, números, -_./',
    'Descripción': 'Descripción corta del producto (máx 250 caracteres)',
    'Marca': 'OPCIONAL — Marca / fabricante del producto (ej: Cooper, Truper, Condumex). Máx 100 caracteres.',
    'Categoría': 'Categoría (ej: Tornillería, Eléctrico, Cable). Si contiene "cable", llena Tipo y Tamaño.',
    'Tipo (cable)': 'SOLO CABLE — Tipo de cable (THHN, THW, desnudo…). Llénalo SOLO si la Categoría es de cable; en otras categorías déjalo vacío.',
    'Tamaño mm²/AWG (cable)': 'SOLO CABLE — Tamaño en mm²/AWG (12, 2/0, 500 kcmil…). Llénalo SOLO si la Categoría es de cable; en otras categorías déjalo vacío.',
    'Unidad': 'Unidad de medida: Pza, Kg, Mts, Lts, Caja, Bote… (en cable se pone M sola).',
    'Stock Inicial': 'Cantidad inicial en almacén (número >= 0). En productos EXISTENTES este valor NO cambia el stock.',
    'Almacén': 'OPCIONAL — Nombre de la bodega donde llega el stock inicial (ej: CDMX). Vacío = bodega default. Solo aplica a productos NUEVOS.',
    'Proyecto': 'OPCIONAL — Número/nombre del proyecto al que se aparta el stock inicial. Vacío = General (libre). Solo aplica a productos NUEVOS.',
    'Stock Mínimo': 'Cantidad mínima antes de alertar de bajo stock (número >= 0)',
    'Precio Unitario': 'Precio unitario del material (número >= 0). Se usa para los costos por proyecto.',
    'URL Imagen (opcional)': 'OPCIONAL — URL HTTPS de la imagen del producto. Ej: https://cdn.miempresa.com/tornillo.jpg',
}

# Anchos por NOMBRE de columna.
PLANTILLA_WIDTHS = {
    'Código (SKU)': 18, 'Descripción': 40, 'Marca': 18, 'Categoría': 18,
    'Tipo (cable)': 16, 'Tamaño mm²/AWG (cable)': 20, 'Unidad': 10,
    'Stock Inicial': 13, 'Almacén': 18, 'Proyecto': 18,
    'Stock Mínimo': 13, 'Precio Unitario': 14,
    'URL Imagen (opcional)': 42,
}


def _construir_plantilla_workbook(rows=None, instruccion=None):
    """Construye el .xlsx de la plantilla y devuelve un BytesIO listo para
    `send_file`.

    `rows` (opcional) permite exportar el catálogo ya lleno. Cada elemento es:
      - una lista de valores alineada a PLANTILLA_HEADERS (un producto), o
      - un dict {'__section__': nombre, 'count': n} → fila de sección resaltada
        (para agrupar por categoría; el importador la salta), o
      - None → fila en blanco separadora.

    Todo se referencia por NOMBRE de columna (no por letra fija), así que el
    orden de PLANTILLA_HEADERS se puede cambiar sin tocar validaciones/anchos.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.comments import Comment

    rows = rows or []
    ncols = len(PLANTILLA_HEADERS)
    col_idx = {h: i + 1 for i, h in enumerate(PLANTILLA_HEADERS)}
    cat_pos = col_idx['Categoría'] - 1  # índice 0-based en la fila de datos

    def _letter(nombre):
        return get_column_letter(col_idx[nombre])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Materiales'

    # Estilos
    header_fill = PatternFill('solid', fgColor='1E40AF')        # azul (columnas normales)
    cable_header_fill = PatternFill('solid', fgColor='B45309')  # ámbar/cobre (columnas de cable)
    cable_cell_fill = PatternFill('solid', fgColor='FEF3C7')    # ámbar claro (celdas de cable)
    section_fill = PatternFill('solid', fgColor='E0E7FF')       # índigo suave (fila de sección)
    header_font = Font(bold=True, color='FFFFFF', size=11)
    title_font = Font(bold=True, color='111827', size=14)
    instr_font = Font(italic=True, color='6B7280', size=10)
    section_font = Font(bold=True, color='3730A3', size=10)
    thin = Side(border_style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_align = Alignment(vertical='top', wrap_text=True)  # evita que el texto largo invada la celda de al lado

    # Fila 1: título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value='Plantilla de Importación de Materiales — SKILLED')
    c.font = title_font
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Fila 2: instrucciones
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=instruccion or (
        '⚠ No alteres los encabezados de la fila 4. Precio y URL Imagen son opcionales '
        '(URL solo HTTPS o /static/...png). Si un SKU ya existe, se ACTUALIZA con los datos '
        'de esta fila (el stock NO se toca). Las columnas de CABLE (en ámbar) van junto a '
        'Categoría: llénalas SOLO si la categoría contiene "cable" — ahí la unidad se pone en '
        'M sola. En el resto de categorías déjalas vacías.'))
    c.font = instr_font
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 34

    # Fila 3 vacía como separador
    ws.row_dimensions[3].height = 6

    # Fila 4: encabezados oficiales (cable en ámbar) + tooltips + anchos
    for col, header in enumerate(PLANTILLA_HEADERS, 1):
        es_cable_col = header in PLANTILLA_CABLE_COLS
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = cable_header_fill if es_cable_col else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = PLANTILLA_WIDTHS.get(header, 18)
        if header in PLANTILLA_TOOLTIPS:
            cell.comment = Comment(PLANTILLA_TOOLTIPS[header], 'Plantilla SKILLED')

    ws.row_dimensions[4].height = 42
    ws.freeze_panes = 'A5'

    # Filas de datos / secciones (solo en exportación).
    r = 5
    for item in rows:
        if item is None:
            # Fila en blanco separadora (el importador la ignora).
            r += 1
            continue
        if isinstance(item, dict) and '__section__' in item:
            # Fila de sección: combinada, resaltada y con el marcador '#' en A
            # para que el importador la salte.
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            cell = ws.cell(row=r, column=1,
                           value=f"{PLANTILLA_SECTION_PREFIX}  ▸  {item['__section__']}  ({item.get('count', 0)})")
            cell.fill = section_fill
            cell.font = section_font
            cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[r].height = 20
            r += 1
            continue
        # Producto: escribir celdas con wrap + borde; tinte ámbar en cable si aplica.
        es_cable = _es_categoria_cable(item[cat_pos] if len(item) > cat_pos else '')
        for col, valor in enumerate(item, 1):
            cell = ws.cell(row=r, column=col, value=valor)
            cell.alignment = data_align
            cell.border = border
            if PLANTILLA_HEADERS[col - 1] in PLANTILLA_CABLE_COLS and es_cable:
                cell.fill = cable_cell_fill
        r += 1

    # Data validations (por nombre → letra). Cubrir al menos 1000 filas o todas.
    last_row = 4 + max(1000, len(rows))

    num_dv = DataValidation(type='decimal', operator='greaterThanOrEqual', formula1=0,
                             showErrorMessage=True, errorTitle='Stock inválido',
                             error='El stock debe ser un número mayor o igual a 0.')
    for nombre in ('Stock Inicial', 'Stock Mínimo', 'Precio Unitario'):
        L = _letter(nombre)
        num_dv.add(f'{L}5:{L}{last_row}')
    ws.add_data_validation(num_dv)

    desc_dv = DataValidation(type='textLength', operator='between', formula1=1, formula2=250,
                              showErrorMessage=True, errorTitle='Descripción inválida',
                              error='Entre 1 y 250 caracteres.')
    Ld = _letter('Descripción')
    desc_dv.add(f'{Ld}5:{Ld}{last_row}')
    ws.add_data_validation(desc_dv)

    sku_dv = DataValidation(type='textLength', operator='between', formula1=1, formula2=50,
                              showErrorMessage=True, errorTitle='SKU inválido',
                              error='Entre 1 y 50 caracteres.')
    La = _letter('Código (SKU)')
    sku_dv.add(f'{La}5:{La}{last_row}')
    ws.add_data_validation(sku_dv)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _producto_export_row(p: Producto) -> list:
    """Fila de exportación alineada a PLANTILLA_HEADERS (mismo orden)."""
    return [
        p.codigo or '',
        p.descripcion or '',
        p.marca or '',
        p.categoria or '',
        p.cable_tipo or '',
        p.cable_calibre or '',
        p.unidad or '',
        float(p.stock_actual or 0),
        # Almacén / Proyecto: vacíos en el export a propósito — el stock de un
        # producto existente puede estar repartido en varios buckets, así que no
        # hay un único destino que exportar, y al reimportar NO se reubica stock.
        '',
        '',
        float(p.stock_minimo or 0),
        float(p.precio_unitario or 0),
        p.imagen_url or '',
    ]


@bp.route('/productos/plantilla-importar', methods=['GET'])
@_require_inventario
def get_plantilla_materiales():
    """Genera y sirve un Excel de plantilla VACÍA para carga masiva de productos.
    Incluye instrucciones, validaciones de celda y tooltips en cada columna."""
    try:
        buf = _construir_plantilla_workbook()
    except ImportError:
        return jsonify({'detail': 'openpyxl no instalado en el servidor'}), 500
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='plantilla_materiales.xlsx',
    )


@bp.route('/productos/exportar', methods=['GET'])
@_require_inventario_admin
def exportar_productos():
    """Exporta TODO el catálogo activo en el mismo formato de la plantilla, con
    los productos ya llenados. Sirve para editar en Excel y reimportar: el
    importador detecta y aplica SOLO los campos que cambiaron (el stock actual
    nunca se toca al reimportar productos existentes)."""
    from itertools import groupby
    productos = (
        Producto.query
        .filter(Producto.activo == True)  # noqa: E712
        .order_by(Producto.categoria.asc(), Producto.codigo.asc())
        .all()
    )

    # Agrupar por categoría: fila de sección resaltada + productos, con una fila
    # en blanco entre secciones. El importador salta secciones/blancos.
    rows = []
    for cat, grupo in groupby(productos, key=lambda p: p.categoria or 'Sin categoría'):
        grupo = list(grupo)
        if rows:
            rows.append(None)  # separador visual entre secciones
        rows.append({'__section__': cat, 'count': len(grupo)})
        rows.extend(_producto_export_row(p) for p in grupo)

    instruccion = (
        '⚠ Edita los valores y vuelve a subir este archivo en "Importar". El sistema aplica '
        'SOLO lo que cambió; las filas sin cambios se ignoran. El Stock Inicial NO modifica el '
        'stock real de productos existentes (usa Movimientos para eso). No borres el SKU (columna A) '
        'ni las filas de sección grises. Columnas de CABLE en ámbar: llénalas solo en productos de cable.'
    )
    try:
        buf = _construir_plantilla_workbook(rows=rows, instruccion=instruccion)
    except ImportError:
        return jsonify({'detail': 'openpyxl no instalado en el servidor'}), 500

    _audit(request.current_user, f'Exportó catálogo de materiales ({len(productos)} productos)')
    db.session.commit()
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='catalogo_materiales.xlsx',
    )


@bp.route('/productos/importar', methods=['POST'])
@_require_inventario_admin
@limiter.limit('5 per minute')
def importar_materiales():
    """Importa productos en masa desde un archivo Excel.

    Validaciones a prueba de tontos:
      - Tamaño max 5 MB.
      - Extensión .xlsx/.xls.
      - Headers tolerantes (acepta variantes sin tildes/lowercase, encuentra la
        fila de headers automáticamente).
      - Trim, normalización y truncado por seguridad.
      - SKU debe pasar CODIGO_REGEX (mismo que crear manual).
      - Descripción, categoría, unidad: validación de longitud.
      - Stocks: parseo tolerante ('1,234.56', '$50', etc.) con error claro.
      - URL imagen pasa por _IMAGEN_URL_REGEX (anti-XSS/SSRF).
      - Duplicados dentro del mismo archivo y contra DB.
      - Filas vacías se ignoran silenciosamente.
      - Tope de PLANTILLA_MAX_FILAS para evitar DoS.
    """
    try:
        import pandas as pd
    except ImportError:
        return jsonify({'detail': 'pandas no instalado en el servidor'}), 500

    file = request.files.get('archivo') or request.files.get('archivo_excel')
    if not file or not file.filename:
        return jsonify({'detail': 'No se envió archivo'}), 400
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'detail': 'Formato no válido. Debe ser .xlsx o .xls'}), 400

    # Validar tamaño antes de cargar a pandas (evita DoS).
    file.stream.seek(0, io.SEEK_END)
    size = file.stream.tell()
    file.stream.seek(0)
    MAX_BYTES = 5 * 1024 * 1024
    if size > MAX_BYTES:
        return jsonify({'detail': f'Archivo demasiado grande. Máximo {MAX_BYTES // (1024*1024)} MB.'}), 413
    if size < 100:
        return jsonify({'detail': 'Archivo vacío o corrupto.'}), 400

    # Buscar la fila de encabezados (acepta hasta fila 10 — la plantilla los pone en fila 4).
    # Si los encabezados no se encuentran, fallamos con mensaje claro.
    try:
        raw = pd.read_excel(file, header=None, nrows=PLANTILLA_MAX_FILAS + 20)
    except Exception as e:
        return jsonify({
            'detail': 'No se pudo leer el Excel. Asegúrate de usar la plantilla y guardar como .xlsx.',
            'tecnico': str(e)[:200],
        }), 400

    expected_norm = {_norm_header(h): h for h in PLANTILLA_HEADERS}
    header_row_idx = None
    column_map = {}  # idx_columna_excel -> nombre_oficial
    for ridx in range(min(10, len(raw))):
        row_vals = [_norm_header(v) for v in raw.iloc[ridx].tolist()]
        matches = {expected_norm[h]: i for i, h in enumerate(row_vals) if h in expected_norm}
        # Necesitamos al menos los 4 obligatorios para considerar esa fila como header
        obligatorios = {'Código (SKU)', 'Descripción', 'Categoría', 'Unidad'}
        if obligatorios.issubset(matches.keys()):
            header_row_idx = ridx
            column_map = {matches[h]: h for h in matches}
            break

    if header_row_idx is None:
        return jsonify({
            'detail': 'No se encontraron los encabezados esperados. '
                      'Descarga la plantilla nueva (debe incluir Código (SKU), Descripción, Categoría, Unidad).',
        }), 400

    # Datos = todo lo que está debajo del header
    data_df = raw.iloc[header_row_idx + 1:].reset_index(drop=True)
    if len(data_df) > PLANTILLA_MAX_FILAS:
        return jsonify({
            'detail': f'Demasiadas filas ({len(data_df)}). Máximo {PLANTILLA_MAX_FILAS} por importación.',
        }), 400

    user = request.current_user
    exitosos = 0       # productos nuevos creados
    actualizados = 0   # productos existentes con al menos un campo cambiado
    sin_cambios = 0    # productos existentes cuya fila era idéntica (se ignoran)
    cambios_detalle = []  # [{codigo, cambios: [str, ...]}] para el reporte al SPA
    errores = []
    skus_en_archivo = set()  # detectar duplicados intra-archivo
    # Pipeline de imágenes → R2: ids de productos cuya imagen es URL externa y
    # hay que descargar/convertir/subir. Se encola al final (no-op si R2 apagado).
    from .imagenes import marcar_para_sync, encolar_sync
    sync_ids = []

    codigo_re = re.compile(CODIGO_REGEX)
    imagen_re = re.compile(_IMAGEN_URL_REGEX)

    # Pausa 2: bodega default donde caerá el stock inicial de productos nuevos.
    # Si no hay ninguna bodega activa, el stock se importa pero queda
    # huérfano (cache lleno, stock_por_almacen vacío) hasta que se cree una.
    bodega_default_id = _almacen_default_id()

    # Feature stock por proyecto: caches para resolver las columnas opcionales
    # 'Almacén' y 'Proyecto' por fila (case/acento-insensitivo, una query cada
    # uno). El proyecto se resuelve por número o por nombre.
    from app.models import Almacen as _Almacen
    alm_by_name: dict[str, int] = {}
    for _a in _Almacen.query.filter(_Almacen.activo == True).all():  # noqa: E712
        alm_by_name.setdefault(_norm_categoria(_a.nombre), _a.id)
    proy_by_key: dict[str, int] = {}
    for _p in Proyecto.query.all():
        if _p.numero_proyecto:
            proy_by_key.setdefault(_norm_categoria(_p.numero_proyecto), _p.id)
        if _p.nombre:
            proy_by_key.setdefault(_norm_categoria(_p.nombre), _p.id)

    # Cache de categorías existentes para resolver case/acento-insensitivamente.
    # Unimos Producto.categoria (catálogo real) + CategoriaConfig (metadatos
    # visuales). El usuario puede capturar "tornilleria" en el Excel y, si ya
    # existe "Tornillería", reutilizamos el nombre canónico en vez de crear un
    # duplicado por dedazo. Cargamos una sola vez (evita N+1 consultas).
    cat_canon: dict[str, str] = {}  # clave normalizada -> nombre canónico
    for (nombre,) in db.session.query(sql_distinct(Producto.categoria)).filter(
        Producto.categoria != None, Producto.categoria != ''
    ).all():
        if nombre:
            cat_canon.setdefault(_norm_categoria(nombre), nombre)
    for (nombre,) in db.session.query(CategoriaConfig.nombre).all():
        if nombre:
            cat_canon.setdefault(_norm_categoria(nombre), nombre)

    categorias_creadas: list[str] = []  # nombres de las categorías nuevas (para reportar al SPA)

    def _g(row, header_oficial):
        """Lee la celda por header oficial, no por posición."""
        for col_idx, oficial in column_map.items():
            if oficial == header_oficial:
                return row.iloc[col_idx] if col_idx < len(row) else None
        return None

    # ── Confirmación de categorías nuevas parecidas a existentes ──────────────
    # Si el usuario ya decidió qué hacer con cada categoría ambigua, llega el
    # mapeo {nombre_en_archivo: nombre_existente}. Valor vacío/ausente = crear
    # como nueva con ese mismo nombre.
    import json as _json
    mapeo_raw = request.form.get('categoria_mapeo')
    categoria_mapeo: dict[str, str] = {}
    if mapeo_raw:
        try:
            _m = _json.loads(mapeo_raw) or {}
            categoria_mapeo = {str(k): str(v or '') for k, v in _m.items()}
        except (ValueError, TypeError):
            categoria_mapeo = {}

    # Primer intento (sin mapeo): si hay categorías NUEVAS parecidas a alguna
    # existente, no creamos nada — devolvemos la lista para que el SPA pregunte.
    if not categoria_mapeo:
        file_cats: dict[str, int] = {}
        for _, row in data_df.iterrows():
            cat = _cell_str(_g(row, 'Categoría'), maxlen=100)
            cod = _cell_str(_g(row, 'Código (SKU)'), maxlen=50)
            if not cat or cod.startswith(PLANTILLA_SECTION_PREFIX):
                continue
            file_cats[cat] = file_cats.get(cat, 0) + 1
        ambiguas = []
        for cat, n in file_cats.items():
            if _norm_categoria(cat) in cat_canon:
                continue  # ya existe exactamente → no preguntar
            sug = _categoria_similar(cat, cat_canon.values())
            if sug:
                ambiguas.append({'nombre': cat, 'sugerencia': sug, 'productos': n})
        if ambiguas:
            ambiguas.sort(key=lambda a: a['nombre'].lower())
            return jsonify({
                'necesita_confirmacion': True,
                'categorias_ambiguas': ambiguas,
                'categorias_existentes': sorted(set(cat_canon.values()), key=str.lower),
            }), 200

    for offset, row in data_df.iterrows():
        fila_excel = header_row_idx + offset + 2  # +1 por header + 1 porque Excel es 1-indexed
        try:
            codigo      = _cell_str(_g(row, 'Código (SKU)'), maxlen=50)
            descripcion = _cell_str(_g(row, 'Descripción'), maxlen=250)
            marca       = _cell_str(_g(row, 'Marca'), maxlen=100)
            categoria   = _cell_str(_g(row, 'Categoría'), maxlen=100)
            unidad      = _cell_str(_g(row, 'Unidad'), maxlen=50)
            imagen_url  = _cell_str(_g(row, 'URL Imagen (opcional)'), maxlen=500)
            cable_tipo_in    = _cell_str(_g(row, 'Tipo (cable)'), maxlen=60)
            cable_calibre_in = _cell_str(_g(row, 'Tamaño mm²/AWG (cable)'), maxlen=40)
            # Destino del stock inicial (feature stock por proyecto). Solo se usan
            # al crear un producto nuevo con stock > 0; en updates se ignoran.
            almacen_in  = _cell_str(_g(row, 'Almacén'), maxlen=100)
            proyecto_in = _cell_str(_g(row, 'Proyecto'), maxlen=100)

            # Aplicar la decisión del usuario: si esta categoría del archivo se
            # mapeó a una existente, la sustituimos antes de procesar.
            if categoria and categoria in categoria_mapeo:
                destino = categoria_mapeo[categoria].strip()
                if destino:
                    categoria = destino[:100]

            # Fila completamente vacía: ignorar sin reportar
            if not (codigo or descripcion or categoria or unidad):
                continue

            # Fila de sección del export (código empieza con '#'): saltar sin
            # error. Un SKU real nunca empieza con '#' (lo prohíbe CODIGO_REGEX).
            if codigo.startswith(PLANTILLA_SECTION_PREFIX):
                continue

            problemas = []
            if not codigo:
                problemas.append('falta SKU')
            elif not codigo_re.match(codigo):
                problemas.append('SKU contiene caracteres no permitidos (usa A-Z 0-9 - _ . /)')
            if not descripcion:
                problemas.append('falta descripción')
            if not categoria:
                problemas.append('falta categoría')
            if not unidad:
                unidad = 'Pza'  # default suave

            stock_actual, err_si = _cell_number(_g(row, 'Stock Inicial'), default=0.0)
            if err_si:
                problemas.append(f'stock inicial {err_si}')
            elif stock_actual < 0:
                problemas.append('stock inicial debe ser >= 0')
            elif stock_actual > 1_000_000:
                problemas.append('stock inicial fuera de rango (máx 1M)')

            stock_minimo, err_sm = _cell_number(_g(row, 'Stock Mínimo'), default=0.0)
            if err_sm:
                problemas.append(f'stock mínimo {err_sm}')
            elif stock_minimo < 0:
                problemas.append('stock mínimo debe ser >= 0')
            elif stock_minimo > 1_000_000:
                problemas.append('stock mínimo fuera de rango (máx 1M)')

            precio, err_pr = _cell_number(_g(row, 'Precio Unitario'), default=0.0)
            if err_pr:
                problemas.append(f'precio {err_pr}')
            elif precio < 0:
                problemas.append('precio debe ser >= 0')
            elif precio > 100_000_000:
                problemas.append('precio fuera de rango (máx 100M)')

            # Validar URL imagen (opcional). Si viene, debe ser HTTPS o /path.png.
            imagen_final = None
            if imagen_url:
                if not imagen_re.match(imagen_url):
                    problemas.append('URL imagen inválida (solo HTTPS o /static/...png)')
                else:
                    imagen_final = imagen_url

            if problemas:
                errores.append(f'Fila {fila_excel}: ' + '; '.join(problemas))
                continue

            # Duplicado intra-archivo
            sku_lower = codigo.lower()
            if sku_lower in skus_en_archivo:
                errores.append(f'Fila {fila_excel}: SKU "{codigo}" duplicado en este archivo')
                continue
            skus_en_archivo.add(sku_lower)

            # Producto existente (upsert por SKU). Se resuelve aquí para poder
            # reutilizarlo tanto en la validación de cable (heredar Tipo/Tamaño
            # ya guardados) como en el bloque de actualización más abajo.
            existente = Producto.query.filter(Producto.codigo == codigo).first()

            # ── Cable a prueba de tontos ──────────────────────────────────────
            # Si la categoría es de cable: Tipo y Tamaño son obligatorios (pueden
            # heredarse de un producto ya existente) y la unidad se fuerza a 'M'.
            # Si NO es cable: se ignoran esas columnas aunque las hayan llenado
            # (no se guardan). Detectamos por el texto crudo de la categoría
            # (mismo criterio, sin importar tildes/mayúsculas).
            es_cable = _es_categoria_cable(categoria)
            cable_tipo_final = (cable_tipo_in or '').strip() or (existente.cable_tipo if existente else None)
            cable_calibre_final = (cable_calibre_in or '').strip() or (existente.cable_calibre if existente else None)
            if es_cable:
                if not cable_tipo_final or not cable_calibre_final:
                    errores.append(
                        f'Fila {fila_excel}: la categoría es de cable; captura Tipo y Tamaño mm²/AWG'
                    )
                    continue
                unidad = _CABLE_UNIDAD  # metros, sin importar lo que traiga la celda
            else:
                # No es cable → no arrastrar datos de cable (foolproof).
                cable_tipo_final = None
                cable_calibre_final = None

            # Resolver categoría a prueba de tontos (sirve para alta y update):
            #  - Si ya existe una equivalente (case/acento-insensitiva), usar el
            #    nombre canónico para no romper el agrupado del dashboard.
            #  - Si es nueva, registrarla en CategoriaConfig (sin imagen) y
            #    agregarla al cache para que el resto del archivo la reutilice.
            cat_key = _norm_categoria(categoria)
            categoria_canonica = cat_canon.get(cat_key)
            if categoria_canonica is None:
                categoria_canonica = categoria  # primera ocurrencia → este es el canónico
                cat_canon[cat_key] = categoria_canonica
                db.session.add(CategoriaConfig(
                    nombre=categoria_canonica,
                    imagen_url=None,
                    created_by_id=user.id,
                ))
                categorias_creadas.append(categoria_canonica)

            precio_dec = Decimal(str(precio))
            stock_min_dec = Decimal(str(stock_minimo))
            # ¿Vinieron en blanco las celdas opcionales? En un update NO debemos
            # pisar un precio/mínimo existente con 0 solo porque la celda quedó
            # vacía. (En alta sí caen a 0, que es el default correcto.)
            precio_provisto = _cell_str(_g(row, 'Precio Unitario')) != ''
            stock_min_provisto = _cell_str(_g(row, 'Stock Mínimo')) != ''

            # ── UPSERT con DETECCIÓN DE CAMBIOS ──────────────────────────────
            # Solo aplicamos (y reportamos) los campos que REALMENTE cambiaron;
            # una fila idéntica a lo guardado se cuenta como "sin cambios" y no
            # se toca. Así el flujo exportar → editar → reimportar aplica solo lo
            # editado. NUNCA tocamos stock_actual/stock_por_almacen: el inventario
            # real se mueve solo por movimientos/ajustes. (`existente` ya se
            # resolvió arriba, en la validación de cable.)
            if existente:
                cambios_fila = []

                if descripcion != (existente.descripcion or ''):
                    cambios_fila.append('descripción')
                    existente.descripcion = descripcion
                if (marca or None) != (existente.marca or None):
                    cambios_fila.append(f'marca: {existente.marca or "—"} → {marca or "—"}')
                    existente.marca = marca or None
                if categoria_canonica != (existente.categoria or ''):
                    cambios_fila.append(f'categoría: {existente.categoria or "—"} → {categoria_canonica}')
                    existente.categoria = categoria_canonica
                if unidad != (existente.unidad or ''):
                    cambios_fila.append(f'unidad: {existente.unidad or "—"} → {unidad}')
                    existente.unidad = unidad
                # Cable (None si dejó de ser cable).
                if (existente.cable_tipo or None) != (cable_tipo_final or None):
                    cambios_fila.append(f'tipo cable: {existente.cable_tipo or "—"} → {cable_tipo_final or "—"}')
                    existente.cable_tipo = cable_tipo_final
                if (existente.cable_calibre or None) != (cable_calibre_final or None):
                    cambios_fila.append(f'tamaño cable: {existente.cable_calibre or "—"} → {cable_calibre_final or "—"}')
                    existente.cable_calibre = cable_calibre_final
                # Precio y stock mínimo: solo si la celda venía con valor.
                if precio_provisto and Decimal(str(existente.precio_unitario or 0)) != precio_dec:
                    cambios_fila.append(f'precio: {float(existente.precio_unitario or 0)} → {float(precio_dec)}')
                    existente.precio_unitario = precio_dec
                if stock_min_provisto and Decimal(str(existente.stock_minimo or 0)) != stock_min_dec:
                    cambios_fila.append(f'stock mín: {float(existente.stock_minimo or 0)} → {float(stock_min_dec)}')
                    existente.stock_minimo = stock_min_dec
                # Imagen: solo si vino una URL válida y es distinta.
                if imagen_final and (existente.imagen_url or '') != imagen_final:
                    cambios_fila.append('imagen')
                    existente.imagen_url = imagen_final
                    if marcar_para_sync(existente, imagen_final):
                        sync_ids.append(existente.id)
                # Reactivar si estaba soft-deleted.
                if not existente.activo:
                    cambios_fila.append('reactivado')
                    existente.activo = True

                if cambios_fila:
                    actualizados += 1
                    if len(cambios_detalle) < 500:  # cap para no inflar la respuesta
                        cambios_detalle.append({'codigo': existente.codigo, 'cambios': cambios_fila})
                else:
                    sin_cambios += 1
                continue

            stock_inicial_dec = Decimal(str(stock_actual))

            # Resolver destino del stock inicial (columnas opcionales Almacén /
            # Proyecto). Vacío → bodega default / bucket general (compat). Si viene
            # un nombre que no existe, error de fila claro para no colocar el stock
            # en el bucket equivocado en silencio.
            almacen_destino_id = bodega_default_id
            proyecto_destino_id = None
            if almacen_in:
                aid = alm_by_name.get(_norm_categoria(almacen_in))
                if not aid:
                    errores.append(f'Fila {fila_excel}: almacén "{almacen_in}" no existe (créalo o deja la celda vacía)')
                    continue
                almacen_destino_id = aid
            if proyecto_in:
                pid = proy_by_key.get(_norm_categoria(proyecto_in))
                if not pid:
                    errores.append(f'Fila {fila_excel}: proyecto "{proyecto_in}" no existe (usa su número o nombre, o deja la celda vacía)')
                    continue
                proyecto_destino_id = pid

            nuevo = Producto(
                codigo=codigo,
                descripcion=descripcion,
                marca=(marca or None),
                categoria=categoria_canonica,
                unidad=unidad,
                cable_tipo=cable_tipo_final,
                cable_calibre=cable_calibre_final,
                stock_actual=stock_inicial_dec,
                stock_minimo=stock_min_dec,
                precio_unitario=precio_dec,
                imagen_url=imagen_final,
                created_by_id=user.id,
            )
            db.session.add(nuevo)
            db.session.flush()  # asegura nuevo.id antes de crear StockPorAlmacen

            # Depositar stock inicial en el bucket resuelto (almacén, proyecto|
            # general) si hay bodega y >0. Feature stock por proyecto: la fuente de
            # verdad es stock_almacen_proyecto; se crea también el cache
            # stock_por_almacen consistente (producto nuevo = un solo bucket).
            if stock_inicial_dec > 0 and almacen_destino_id:
                db.session.add(StockAlmacenProyecto(
                    producto_id=nuevo.id,
                    almacen_id=almacen_destino_id,
                    proyecto_id=proyecto_destino_id,
                    cantidad=stock_inicial_dec,
                ))
                db.session.add(StockPorAlmacen(
                    producto_id=nuevo.id,
                    almacen_id=almacen_destino_id,
                    cantidad=stock_inicial_dec,
                ))

            exitosos += 1
            # Si la imagen importada es una URL externa, encolar su sync a R2.
            if marcar_para_sync(nuevo, imagen_final):
                sync_ids.append(nuevo.id)

        except Exception as e:
            errores.append(f'Fila {fila_excel}: error inesperado — {str(e)[:80]}')

    try:
        msg = (f'Importación masiva: {exitosos} creados, {actualizados} actualizados, '
               f'{sin_cambios} sin cambios, {len(errores)} errores')
        if categorias_creadas:
            msg += f', {len(categorias_creadas)} categorías nuevas'
        _audit(user, msg)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error commit importar_materiales')
        return jsonify({'detail': f'Error al guardar en base de datos: {str(e)[:100]}'}), 500

    # Refresca catálogos abiertos en otras sesiones tras la importación masiva.
    # Mandamos un solo emit aunque hayan sido N productos — el front invalida
    # el namespace completo de productos vía `useResource.invalidateOn`.
    if exitosos > 0 or actualizados > 0:
        emit_to_role(_INV_ROLES, 'producto:changed', {
            'action': 'bulk_import', 'count': exitosos + actualizados,
        })

    # Encolar la descarga/conversión/subida a R2 de las imágenes externas de
    # esta importación. No-op si R2 está apagado (sync_ids queda vacío).
    imagenes_job = encolar_sync(user.id, sync_ids) if sync_ids else None

    respuesta = {
        'exitosos': exitosos,
        'actualizados': actualizados,
        'sin_cambios': sin_cambios,
        'errores': errores,
        'total_procesadas': exitosos + actualizados + sin_cambios + len(errores),
        'categorias_creadas': categorias_creadas,
        'cambios_detalle': cambios_detalle,
    }
    # Clave `imagenes` solo si hay algo que sincronizar → la respuesta queda
    # idéntica a la de antes cuando R2 está apagado (no rompe nada).
    if sync_ids:
        respuesta['imagenes'] = {'pendientes': len(sync_ids), 'job_id': imagenes_job}
    return jsonify(respuesta)
