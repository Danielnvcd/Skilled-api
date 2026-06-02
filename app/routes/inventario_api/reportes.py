"""Reportes Excel del módulo de Inventario (Pausa 6).

Cinco reportes:
  - inventario_actual: stock vigente con filtros opcionales.
  - movimientos: histórico de entradas/salidas/ajustes/traspasos.
  - kardex: kardex de un producto en Excel (espejo del JSON).
  - consumo_proyecto: agregado de consumo por proyecto.
  - solicitudes: listado de solicitudes con totales.
"""
import io
import datetime
from decimal import Decimal

from flask import jsonify, request, send_file
from sqlalchemy.orm import joinedload, selectinload

from app.extensions import db, limiter, get_real_client_ip_flask
from app.models import (
    Producto, MovimientoInventario, SolicitudMaterial, SolicitudMaterialDetalle,
)

from ._core import (
    bp,
    _require_inventario_admin,
    _int_arg,
    _audit,
)


REPORTES_MAX_FILAS = 10_000


def _aplicar_estilos_ws(ws, money_cols: set[str] | None = None):
    """Aplica encabezado azul + zebra striping + freeze panes + auto-width.

    `money_cols`: nombres de columna (exactos) que reciben formato '$#,##0.00'.
    El resto de los numéricos quedan sin formato (importante: stocks, días,
    cantidades enteras o decimales sin moneda).
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    money_cols = money_cols or set()

    headers = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
    money_idx = {i + 1 for i, h in enumerate(headers) if h in money_cols}

    if ws.max_row > 0:
        ws.freeze_panes = 'A2'

    for row_idx in range(1, ws.max_row + 1):
        is_header = (row_idx == 1)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if is_header:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
            else:
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill
                if col_idx in money_idx and isinstance(cell.value, (int, float)):
                    cell.number_format = '"$"#,##0.00'

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    n = len(str(cell.value))
                    if n > max_length:
                        max_length = n
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)


def _stream_excel(sheets: dict, filename: str, money_cols: set[str] | None = None):
    """Genera un .xlsx multi-hoja en memoria y lo devuelve con send_file.

    `sheets`: dict[nombre_hoja, list[dict]]. Los keys de la primera fila se
    usan como encabezados (orden estable de Python 3.7+).
    Aplica saneo anti CSV-injection con `safe_excel_value`.
    """
    from openpyxl import Workbook
    from app.utils import safe_excel_value

    wb = Workbook()
    # Workbook trae una hoja "Sheet" por default; la quitamos para que solo
    # queden las hojas que pasamos en `sheets`.
    default = wb.active
    wb.remove(default)

    for raw_name, rows in sheets.items():
        # Excel limita nombre de hoja a 31 chars y prohíbe / \ ? * [ ]
        safe = (raw_name or 'Hoja')[:31]
        for ch in r'/\?*[]:':
            safe = safe.replace(ch, '_')
        ws = wb.create_sheet(title=safe)
        if not rows:
            ws['A1'] = 'Sin datos para los filtros seleccionados'
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([safe_excel_value(row.get(h)) for h in headers])
        _aplicar_estilos_ws(ws, money_cols=money_cols)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


def _parse_fecha_arg(name: str, default: datetime.date | None):
    """Lee un query param YYYY-MM-DD o devuelve `default` si está vacío.
    Devuelve (date, error_response). Si error_response es None, todo bien.
    """
    raw = request.args.get(name)
    if not raw:
        return default, None
    try:
        return datetime.date.fromisoformat(raw), None
    except ValueError:
        return None, (jsonify({'detail': f"Parámetro '{name}' debe ser YYYY-MM-DD"}), 422)


@bp.route('/reportes/inventario-actual.xlsx', methods=['GET'])
@limiter.limit(
    "10/minute",
    key_func=lambda: f"ip:{get_real_client_ip_flask()}",
)
@_require_inventario_admin
def reporte_inventario_actual():
    """Reporte de stock actual de todos los productos activos.

    Query params:
      - categoria: filtra por categoría exacta (opcional).
      - solo_bajo_minimo: 1 → solo productos con stock_actual ≤ stock_minimo.
    """
    q = Producto.query.filter(Producto.activo == True)  # noqa: E712
    categoria = (request.args.get('categoria') or '').strip()
    if categoria:
        q = q.filter(Producto.categoria == categoria)
    if request.args.get('solo_bajo_minimo') in ('1', 'true', 'True'):
        q = q.filter(Producto.stock_actual <= Producto.stock_minimo)
    productos = q.order_by(Producto.categoria, Producto.codigo).limit(REPORTES_MAX_FILAS).all()

    rows = []
    for p in productos:
        actual = float(p.stock_actual or 0)
        reservado = float(p.stock_reservado or 0)
        minimo = float(p.stock_minimo or 0)
        rows.append({
            'Código': p.codigo,
            'Descripción': p.descripcion,
            'Categoría': p.categoria,
            'Unidad': p.unidad,
            'Stock actual': actual,
            'Reservado': reservado,
            'Disponible': actual - reservado,
            'Mínimo': minimo,
            'Diferencia vs mínimo': actual - minimo,
            'Estado': 'BAJO' if actual <= minimo else 'OK',
        })

    _audit(request.current_user, f"Reporte inventario actual ({len(rows)} filas)")
    db.session.commit()

    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    return _stream_excel(
        {'Inventario': rows},
        f'inventario_actual_{ts}.xlsx',
    )


@bp.route('/reportes/movimientos.xlsx', methods=['GET'])
@limiter.limit(
    "10/minute",
    key_func=lambda: f"ip:{get_real_client_ip_flask()}",
)
@_require_inventario_admin
def reporte_movimientos():
    """Reporte de movimientos de inventario con filtros.

    Query params:
      - desde, hasta (YYYY-MM-DD): default últimos 30 días.
      - tipo: ENTRADA / SALIDA / AJUSTE / TRASPASO (opcional).
      - producto_id (opcional).
      - usuario_id (opcional).
    """
    hoy = datetime.date.today()
    desde, err = _parse_fecha_arg('desde', hoy - datetime.timedelta(days=30))
    if err: return err
    hasta, err = _parse_fecha_arg('hasta', hoy)
    if err: return err
    if desde > hasta:
        return jsonify({'detail': "'desde' no puede ser mayor que 'hasta'"}), 422

    tipo = request.args.get('tipo')
    if tipo and tipo not in ('ENTRADA', 'SALIDA', 'AJUSTE', 'TRASPASO'):
        return jsonify({'detail': "Parámetro 'tipo' inválido"}), 422

    prod_id, err = _int_arg('producto_id', 0, 0, 1_000_000_000)
    if err: return err
    usr_id, err = _int_arg('usuario_id', 0, 0, 1_000_000_000)
    if err: return err

    desde_dt = datetime.datetime.combine(desde, datetime.time.min)
    hasta_dt = datetime.datetime.combine(hasta, datetime.time.max)

    q = (
        MovimientoInventario.query
        .options(
            joinedload(MovimientoInventario.producto),
            joinedload(MovimientoInventario.almacen_origen),
            joinedload(MovimientoInventario.almacen_destino),
            joinedload(MovimientoInventario.usuario),
        )
        .filter(MovimientoInventario.fecha >= desde_dt, MovimientoInventario.fecha <= hasta_dt)
    )
    if tipo:
        q = q.filter(MovimientoInventario.tipo == tipo)
    if prod_id:
        q = q.filter(MovimientoInventario.producto_id == prod_id)
    if usr_id:
        q = q.filter(MovimientoInventario.usuario_id == usr_id)

    movs = q.order_by(MovimientoInventario.fecha.desc()).limit(REPORTES_MAX_FILAS).all()

    rows = []
    for m in movs:
        rows.append({
            'Fecha': m.fecha.strftime('%Y-%m-%d %H:%M') if m.fecha else '',
            'Tipo': m.tipo,
            'Código': m.producto.codigo if m.producto else '',
            'Descripción': m.producto.descripcion if m.producto else '',
            'Cantidad': float(m.cantidad or 0),
            'Unidad': m.producto.unidad if m.producto else '',
            'Almacén origen': m.almacen_origen.nombre if m.almacen_origen else '',
            'Almacén destino': m.almacen_destino.nombre if m.almacen_destino else '',
            'Usuario': m.usuario.username if m.usuario else '',
            'Motivo': m.motivo or '',
        })

    _audit(
        request.current_user,
        f"Reporte movimientos {desde} a {hasta} ({len(rows)} filas)",
    )
    db.session.commit()

    return _stream_excel(
        {'Movimientos': rows},
        f'movimientos_{desde}_{hasta}.xlsx',
    )


@bp.route('/reportes/kardex.xlsx', methods=['GET'])
@limiter.limit(
    "10/minute",
    key_func=lambda: f"ip:{get_real_client_ip_flask()}",
)
@_require_inventario_admin
def reporte_kardex_xlsx():
    """Kardex de un producto exportado a Excel.

    Query params:
      - producto_id (requerido).
      - desde, hasta (YYYY-MM-DD): default últimos 30 días.
    """
    prod_id, err = _int_arg('producto_id', 0, 0, 1_000_000_000)
    if err: return err
    if not prod_id:
        return jsonify({'detail': "Parámetro 'producto_id' es requerido"}), 422

    producto = Producto.query.filter(Producto.id == prod_id).first()
    if not producto:
        return jsonify({'detail': 'Producto no encontrado'}), 404

    hoy = datetime.date.today()
    desde, err = _parse_fecha_arg('desde', hoy - datetime.timedelta(days=30))
    if err: return err
    hasta, err = _parse_fecha_arg('hasta', hoy)
    if err: return err
    if desde > hasta:
        return jsonify({'detail': "'desde' no puede ser mayor que 'hasta'"}), 422

    # Reutiliza la misma fórmula de saldo corrido del endpoint JSON.
    def _delta(m: MovimientoInventario) -> Decimal:
        cant = m.cantidad or Decimal('0')
        if m.tipo == 'ENTRADA':
            return cant
        if m.tipo == 'SALIDA':
            return -cant
        if m.tipo == 'AJUSTE':
            return cant
        return Decimal('0')  # TRASPASO no altera total

    desde_dt = datetime.datetime.combine(desde, datetime.time.min)
    hasta_dt = datetime.datetime.combine(hasta, datetime.time.max)

    # Saldo inicial = stock_actual − Σ deltas posteriores a 'desde'.
    movs_post = (
        MovimientoInventario.query
        .filter(
            MovimientoInventario.producto_id == prod_id,
            MovimientoInventario.fecha >= desde_dt,
        )
        .all()
    )
    delta_post = sum((_delta(m) for m in movs_post), Decimal('0'))
    saldo_inicial = (producto.stock_actual or Decimal('0')) - delta_post

    movs = (
        MovimientoInventario.query
        .options(
            joinedload(MovimientoInventario.usuario),
            joinedload(MovimientoInventario.almacen_origen),
            joinedload(MovimientoInventario.almacen_destino),
        )
        .filter(
            MovimientoInventario.producto_id == prod_id,
            MovimientoInventario.fecha >= desde_dt,
            MovimientoInventario.fecha <= hasta_dt,
        )
        .order_by(MovimientoInventario.fecha.asc(), MovimientoInventario.id.asc())
        .limit(REPORTES_MAX_FILAS)
        .all()
    )

    rows = [{
        'Fecha': '',
        'Tipo': '— Saldo inicial —',
        'Cantidad': '',
        'Delta': '',
        'Saldo': float(saldo_inicial),
        'Almacén origen': '',
        'Almacén destino': '',
        'Usuario': '',
        'Motivo': '',
    }]
    saldo = saldo_inicial
    for m in movs:
        d = _delta(m)
        saldo = saldo + d
        rows.append({
            'Fecha': m.fecha.strftime('%Y-%m-%d %H:%M') if m.fecha else '',
            'Tipo': m.tipo,
            'Cantidad': float(m.cantidad or 0),
            'Delta': float(d),
            'Saldo': float(saldo),
            'Almacén origen': m.almacen_origen.nombre if m.almacen_origen else '',
            'Almacén destino': m.almacen_destino.nombre if m.almacen_destino else '',
            'Usuario': m.usuario.username if m.usuario else '',
            'Motivo': m.motivo or '',
        })

    info = [
        {'Campo': 'Producto', 'Valor': f"{producto.codigo} — {producto.descripcion}"},
        {'Campo': 'Unidad', 'Valor': producto.unidad},
        {'Campo': 'Stock actual', 'Valor': float(producto.stock_actual or 0)},
        {'Campo': 'Periodo', 'Valor': f"{desde} a {hasta}"},
        {'Campo': 'Movimientos', 'Valor': len(movs)},
    ]

    _audit(request.current_user, f"Reporte kardex prod #{prod_id} {desde} a {hasta}")
    db.session.commit()

    return _stream_excel(
        {'Resumen': info, 'Kardex': rows},
        f'kardex_{producto.codigo}_{desde}_{hasta}.xlsx',
    )


@bp.route('/reportes/consumo-proyecto.xlsx', methods=['GET'])
@limiter.limit(
    "10/minute",
    key_func=lambda: f"ip:{get_real_client_ip_flask()}",
)
@_require_inventario_admin
def reporte_consumo_proyecto():
    """Consumo por proyecto agrupado, basado en solicitudes entregadas o con
    entrega parcial dentro del rango.

    Query params:
      - desde, hasta (YYYY-MM-DD): default últimos 90 días.
      - estatus: por defecto ['ENTREGADA','APROBADA'] (incluye parciales).
    """
    hoy = datetime.date.today()
    desde, err = _parse_fecha_arg('desde', hoy - datetime.timedelta(days=90))
    if err: return err
    hasta, err = _parse_fecha_arg('hasta', hoy)
    if err: return err
    if desde > hasta:
        return jsonify({'detail': "'desde' no puede ser mayor que 'hasta'"}), 422

    desde_dt = datetime.datetime.combine(desde, datetime.time.min)
    hasta_dt = datetime.datetime.combine(hasta, datetime.time.max)

    estatus_arg = (request.args.get('estatus') or '').strip()
    if estatus_arg:
        estatus_list = [e.strip().upper() for e in estatus_arg.split(',') if e.strip()]
        if any(e not in ('PENDIENTE', 'APROBADA', 'RECHAZADA', 'ENTREGADA') for e in estatus_list):
            return jsonify({'detail': "Parámetro 'estatus' inválido"}), 422
    else:
        estatus_list = ['ENTREGADA', 'APROBADA']

    sols = (
        SolicitudMaterial.query
        .options(
            selectinload(SolicitudMaterial.detalles).joinedload(SolicitudMaterialDetalle.producto),
        )
        .filter(
            SolicitudMaterial.fecha_creacion >= desde_dt,
            SolicitudMaterial.fecha_creacion <= hasta_dt,
            SolicitudMaterial.estatus.in_(estatus_list),
        )
        .all()
    )

    # Agregamos por (proyecto, código). Usamos cantidad_entregada (real); para
    # ENTREGADAs pre-8b sin cantidad_entregada, caemos a cantidad_solicitada.
    agg: dict[tuple, dict] = {}
    for s in sols:
        proyecto = (s.proyecto or 'Sin proyecto').strip() or 'Sin proyecto'
        for d in (s.detalles or []):
            if (d.tipo_item or 'MATERIAL').upper() != 'MATERIAL' or not d.producto_id:
                continue
            cant_ent = Decimal(str(d.cantidad_entregada or 0))
            if cant_ent <= 0 and s.estatus == 'ENTREGADA':
                # Legacy: solicitudes pre-8b marcadas como ENTREGADA sin descontar.
                cant_ent = Decimal(str(d.cantidad_solicitada or 0))
            if cant_ent <= 0:
                continue
            codigo = d.producto.codigo if d.producto else f'(#{d.producto_id})'
            descripcion = d.producto.descripcion if d.producto else 'Producto eliminado'
            unidad = d.producto.unidad if d.producto else ''
            key = (proyecto, codigo)
            if key not in agg:
                agg[key] = {
                    'Proyecto': proyecto,
                    'Código': codigo,
                    'Descripción': descripcion,
                    'Unidad': unidad,
                    'Cantidad entregada': 0.0,
                    'Solicitudes': set(),
                }
            agg[key]['Cantidad entregada'] += float(cant_ent)
            agg[key]['Solicitudes'].add(s.id)

    rows = []
    for v in sorted(agg.values(), key=lambda r: (r['Proyecto'], r['Código'])):
        rows.append({
            'Proyecto': v['Proyecto'],
            'Código': v['Código'],
            'Descripción': v['Descripción'],
            'Unidad': v['Unidad'],
            'Cantidad entregada': round(v['Cantidad entregada'], 4),
            '# Solicitudes': len(v['Solicitudes']),
        })

    _audit(
        request.current_user,
        f"Reporte consumo por proyecto {desde} a {hasta} ({len(rows)} filas)",
    )
    db.session.commit()

    return _stream_excel(
        {'Consumo por proyecto': rows},
        f'consumo_proyecto_{desde}_{hasta}.xlsx',
    )


@bp.route('/reportes/solicitudes.xlsx', methods=['GET'])
@limiter.limit(
    "10/minute",
    key_func=lambda: f"ip:{get_real_client_ip_flask()}",
)
@_require_inventario_admin
def reporte_solicitudes():
    """Reporte de solicitudes del periodo con totales por solicitud.

    Query params:
      - desde, hasta (YYYY-MM-DD): default últimos 30 días.
      - estatus: PENDIENTE / APROBADA / RECHAZADA / ENTREGADA (opcional).
    """
    hoy = datetime.date.today()
    desde, err = _parse_fecha_arg('desde', hoy - datetime.timedelta(days=30))
    if err: return err
    hasta, err = _parse_fecha_arg('hasta', hoy)
    if err: return err
    if desde > hasta:
        return jsonify({'detail': "'desde' no puede ser mayor que 'hasta'"}), 422

    estatus = (request.args.get('estatus') or '').strip().upper()
    if estatus and estatus not in ('PENDIENTE', 'APROBADA', 'RECHAZADA', 'ENTREGADA'):
        return jsonify({'detail': "Parámetro 'estatus' inválido"}), 422

    desde_dt = datetime.datetime.combine(desde, datetime.time.min)
    hasta_dt = datetime.datetime.combine(hasta, datetime.time.max)

    q = (
        SolicitudMaterial.query
        .options(
            joinedload(SolicitudMaterial.solicitante),
            selectinload(SolicitudMaterial.detalles),
        )
        .filter(
            SolicitudMaterial.fecha_creacion >= desde_dt,
            SolicitudMaterial.fecha_creacion <= hasta_dt,
        )
    )
    if estatus:
        q = q.filter(SolicitudMaterial.estatus == estatus)
    sols = q.order_by(SolicitudMaterial.fecha_creacion.desc()).limit(REPORTES_MAX_FILAS).all()

    rows = []
    for s in sols:
        total_sol = sum(float(d.cantidad_solicitada or 0) for d in (s.detalles or []))
        total_aprob = sum(float(d.cantidad_aprobada or 0) for d in (s.detalles or []))
        total_ent = sum(float(d.cantidad_entregada or 0) for d in (s.detalles or []))
        rows.append({
            'ID': s.id,
            'Fecha': s.fecha_creacion.strftime('%Y-%m-%d %H:%M') if s.fecha_creacion else '',
            'Cierre': s.fecha_cierre.strftime('%Y-%m-%d %H:%M') if s.fecha_cierre else '',
            'Estatus': s.estatus,
            'Solicitante': s.solicitante.username if s.solicitante else '',
            'Proyecto': s.proyecto or '',
            'Líneas': len(s.detalles or []),
            'Total solicitado': round(total_sol, 4),
            'Total aprobado': round(total_aprob, 4),
            'Total entregado': round(total_ent, 4),
        })

    _audit(
        request.current_user,
        f"Reporte solicitudes {desde} a {hasta} ({len(rows)} filas)",
    )
    db.session.commit()

    return _stream_excel(
        {'Solicitudes': rows},
        f'solicitudes_{desde}_{hasta}.xlsx',
    )
