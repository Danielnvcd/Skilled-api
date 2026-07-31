"""Asignación de material a proyectos: previsualizar, aplicar y devolver.

Pensado para la sección «Material por proyecto», donde el proyecto es el
contexto y se trabaja con VARIAS líneas a la vez (a mano o desde Excel), en
lugar de un formulario de movimiento por material.

  POST /proyectos-materiales/<id>/asignar/previsualizar   valida, NO escribe
  POST /proyectos-materiales/<id>/asignar                 aplica en una transacción
  POST /proyectos-materiales/<id>/devolver                a General o a otro proyecto

── Principio que NO se rompe ────────────────────────────────────────────────
Todo pasa por los mismos helpers de stock que usan los movimientos
(`_depositar`, `_consumir_bucket_exacto`, `_recalcular_caches`) y genera un
`MovimientoInventario` por línea. No hay una vía paralela de tocar existencias:
si algo modifica el stock sin quedar en el kardex, el inventario deja de ser
auditable y se pierde la confianza en él.

── Decisiones de comportamiento ─────────────────────────────────────────────
* **Faltante ⇒ se asigna lo disponible, con aviso.** No se bloquea el lote
  entero por una línea corta (castiga todo el trabajo) ni se omite en silencio.
  Es seguro porque la previsualización lo muestra ANTES de aplicar.
* **Suma por defecto; reemplazar es explícito.** Asignar es acumulativo (llegó
  más material). Reemplazar sirve para cuadrar tras un conteo: más raro y más
  peligroso, así que nunca por defecto.
* **Origen `general`** mueve stock ya existente (REASIGNACION). **Origen
  `entrada`** crea stock nuevo que llegó directo para la obra (ENTRADA). Son
  cosas distintas y quien captura debe elegir.
"""
from decimal import Decimal, InvalidOperation

from flask import jsonify, request

from app.extensions import db
from app.models import (
    Almacen, MovimientoInventario, Producto, Proyecto, StockAlmacenProyecto,
)
from app.realtime import emit_to_role

from ._core import (
    bp,
    _require_inventario_admin,
    _almacen_default_id,
    _audit,
    _INV_ROLES,
    _consumir_bucket_exacto,
    _depositar,
    _es_error_de_lock,
    _recalcular_caches,
)


def _dec(v) -> Decimal:
    try:
        return Decimal(str(v))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0')


def _f(v) -> float:
    return float(Decimal(str(v or 0)))


# ── Validación compartida ───────────────────────────────────────────────────
# La usan TANTO la previsualización como la aplicación. Es deliberado: si cada
# una validara por su cuenta, la vista previa podría prometer un resultado y la
# aplicación hacer otro — y entonces la previsualización deja de servir para
# lo único que sirve, que es confiar en ella.

def _resolver_lineas(proyecto_id: int, lineas: list, origen: str, modo: str) -> list:
    """Convierte las líneas crudas en un plan evaluado, sin escribir nada.

    Devuelve una lista de dicts con el estado de cada línea:
      estado='ok'      se aplicará tal cual
      estado='aviso'   se aplicará AJUSTADA (no había suficiente)
      estado='error'   se omitirá, con el motivo
    """
    almacen_default = _almacen_default_id()
    plan = []

    for cruda in lineas:
        sku = str(cruda.get('sku') or cruda.get('codigo') or '').strip()
        producto_id = cruda.get('producto_id')
        cantidad = _dec(cruda.get('cantidad'))
        almacen_nombre = str(cruda.get('almacen') or '').strip()
        almacen_id = cruda.get('almacen_id')

        fila = {
            'sku': sku, 'cantidad_pedida': _f(cantidad),
            'estado': 'error', 'motivo': None,
            'producto_id': None, 'descripcion': None, 'unidad': None,
            'almacen_id': None, 'almacen_nombre': None,
            'cantidad_aplicada': 0.0, 'actual': 0.0, 'resultado': 0.0,
            # Cuánto hay en General de este material en esa bodega. La interfaz
            # lo muestra ANTES de que se escriba la cantidad: sin ese dato se
            # captura a ciegas. En origen 'entrada' no aplica (el material llega
            # de fuera, no hay tope) y queda en null.
            'disponible': None,
        }

        # ── Producto ────────────────────────────────────────────────────────
        producto = None
        if producto_id:
            producto = Producto.query.filter(
                Producto.id == producto_id, Producto.activo == True,  # noqa: E712
            ).first()
        elif sku:
            producto = Producto.query.filter(
                Producto.codigo == sku, Producto.activo == True,  # noqa: E712
            ).first()
        if not producto:
            fila['motivo'] = f"El código «{sku or producto_id}» no existe en el catálogo"
            plan.append(fila)
            continue
        fila.update({
            'producto_id': producto.id, 'sku': producto.codigo,
            'descripcion': producto.descripcion, 'unidad': producto.unidad,
        })

        # ── Cantidad ────────────────────────────────────────────────────────
        if cantidad <= 0:
            fila['motivo'] = 'La cantidad debe ser mayor que cero'
            plan.append(fila)
            continue

        # ── Bodega ──────────────────────────────────────────────────────────
        almacen = None
        if almacen_id:
            almacen = Almacen.query.filter(
                Almacen.id == almacen_id, Almacen.activo == True,  # noqa: E712
            ).first()
        elif almacen_nombre:
            almacen = Almacen.query.filter(
                db.func.lower(Almacen.nombre) == almacen_nombre.lower(),
                Almacen.activo == True,  # noqa: E712
            ).first()
            if not almacen:
                # Sugerencia por parecido: un typo en el nombre de la bodega es
                # el error más común al llenar el Excel a mano.
                parecidas = [
                    a.nombre for a in Almacen.query.filter(Almacen.activo == True).all()  # noqa: E712
                    if a.nombre.lower().startswith(almacen_nombre.lower()[:3])
                ]
                sugerencia = f' ¿Quisiste decir «{parecidas[0]}»?' if parecidas else ''
                fila['motivo'] = f'La bodega «{almacen_nombre}» no existe.{sugerencia}'
                plan.append(fila)
                continue
        else:
            almacen = Almacen.query.get(almacen_default) if almacen_default else None
        if not almacen:
            fila['motivo'] = 'No se indicó bodega y no hay una bodega predeterminada'
            plan.append(fila)
            continue
        fila.update({'almacen_id': almacen.id, 'almacen_nombre': almacen.nombre})

        # ── Existencias actuales ────────────────────────────────────────────
        actual = _bucket(producto.id, almacen.id, proyecto_id)
        fila['actual'] = _f(actual)

        disponible = _bucket(producto.id, almacen.id, None) if origen == 'general' else None
        if disponible is not None:
            fila['disponible'] = _f(disponible)

        # En modo reemplazar, la cantidad de la línea es el OBJETIVO, no el
        # incremento: se calcula el delta necesario para llegar a él.
        objetivo = cantidad if modo == 'reemplazar' else actual + cantidad
        delta = objetivo - actual

        if delta == 0:
            fila.update({'estado': 'ok', 'cantidad_aplicada': 0.0,
                         'resultado': _f(actual),
                         'motivo': 'Ya tiene esa cantidad; no hay cambio'})
            plan.append(fila)
            continue

        if delta > 0:
            # Hay que meter material al proyecto.
            if origen == 'general':
                if disponible <= 0:
                    fila['motivo'] = (
                        f'No hay stock general de este material en {almacen.nombre}'
                    )
                    plan.append(fila)
                    continue
                if disponible < delta:
                    # Decisión: aplicar lo disponible y avisar.
                    fila.update({
                        'estado': 'aviso',
                        'motivo': (
                            f'Solo hay {_f(disponible)} en General ({almacen.nombre}); '
                            f'se asignará esa cantidad'
                        ),
                    })
                    delta = disponible
                else:
                    fila['estado'] = 'ok'
            else:
                # Entrada nueva: el material llega de fuera, no hay tope.
                fila['estado'] = 'ok'
        else:
            # Modo reemplazar con objetivo MENOR: sobra material y se devuelve
            # a General. Nunca puede dejar el bucket en negativo porque el
            # objetivo se valida como >= 0 más arriba (cantidad > 0).
            fila['estado'] = 'ok'
            fila['motivo'] = 'Sobra material; se devolverá a General'

        fila['cantidad_aplicada'] = _f(delta)
        fila['resultado'] = _f(actual + delta)
        plan.append(fila)

    return plan


def _bucket(producto_id: int, almacen_id: int, proyecto_id: int | None) -> Decimal:
    """Existencia actual de un bucket, sin bloquear (solo lectura)."""
    q = db.session.query(StockAlmacenProyecto.cantidad).filter(
        StockAlmacenProyecto.producto_id == producto_id,
        StockAlmacenProyecto.almacen_id == almacen_id,
    )
    q = q.filter(StockAlmacenProyecto.proyecto_id.is_(None)) if proyecto_id is None \
        else q.filter(StockAlmacenProyecto.proyecto_id == proyecto_id)
    return Decimal(str(q.scalar() or 0))


def _resumen(plan: list) -> dict:
    return {
        'total': len(plan),
        'ok': sum(1 for f in plan if f['estado'] == 'ok'),
        'avisos': sum(1 for f in plan if f['estado'] == 'aviso'),
        'errores': sum(1 for f in plan if f['estado'] == 'error'),
        'unidades': _f(sum(_dec(f['cantidad_aplicada']) for f in plan
                           if f['estado'] != 'error')),
    }


def _leer_peticion(proyecto_id: int):
    """Valida proyecto, cuerpo y opciones comunes a los tres endpoints."""
    proyecto = Proyecto.query.get(proyecto_id)
    if not proyecto:
        return None, (jsonify({'detail': 'Proyecto no encontrado'}), 404)

    data = request.get_json(silent=True) or {}
    lineas = data.get('lineas')
    if not isinstance(lineas, list) or not lineas:
        return None, (jsonify({'detail': 'Se requiere al menos una línea'}), 422)
    if len(lineas) > 2000:
        return None, (jsonify({'detail': 'Máximo 2000 líneas por operación'}), 422)

    origen = (data.get('origen') or 'general').strip().lower()
    if origen not in ('general', 'entrada'):
        return None, (jsonify({'detail': "origen debe ser 'general' o 'entrada'"}), 422)

    modo = (data.get('modo') or 'sumar').strip().lower()
    if modo not in ('sumar', 'reemplazar'):
        return None, (jsonify({'detail': "modo debe ser 'sumar' o 'reemplazar'"}), 422)

    return (proyecto, lineas, origen, modo, data), None


@bp.route('/proyectos-materiales/<int:proyecto_id>/asignar/previsualizar', methods=['POST'])
@_require_inventario_admin
def previsualizar_asignacion(proyecto_id: int):
    """Simula la asignación. NO escribe absolutamente nada.

    Es el paso que hace segura la decisión de «asignar lo disponible cuando no
    alcanza»: el usuario ve el ajuste antes de aplicarlo, así que nunca es una
    sorpresa.
    """
    ctx, err = _leer_peticion(proyecto_id)
    if err:
        return err
    proyecto, lineas, origen, modo, _ = ctx

    plan = _resolver_lineas(proyecto.id, lineas, origen, modo)
    return jsonify({
        'proyecto': {'id': proyecto.id, 'numero_proyecto': proyecto.numero_proyecto,
                     'nombre': proyecto.nombre or ''},
        'origen': origen,
        'modo': modo,
        'lineas': plan,
        'resumen': _resumen(plan),
    })


@bp.route('/proyectos-materiales/<int:proyecto_id>/asignar', methods=['POST'])
@_require_inventario_admin
def aplicar_asignacion(proyecto_id: int):
    """Aplica la asignación. Todo o nada: una sola transacción.

    Las líneas con error se OMITEN (no abortan el lote); las de aviso se aplican
    ajustadas. Es el mismo resultado que mostró la previsualización porque ambas
    usan `_resolver_lineas`.
    """
    ctx, err = _leer_peticion(proyecto_id)
    if err:
        return err
    proyecto, lineas, origen, modo, data = ctx

    plan = _resolver_lineas(proyecto.id, lineas, origen, modo)
    aplicables = [f for f in plan if f['estado'] != 'error' and f['cantidad_aplicada'] != 0]
    if not aplicables:
        return jsonify({
            'ok': True, 'aplicadas': 0, 'lineas': plan, 'resumen': _resumen(plan),
            'detail': 'No hubo nada que aplicar',
        })

    from app.routes._api_helpers import current_user
    user = current_user()
    motivo = (data.get('motivo') or '').strip() or None

    try:
        for f in aplicables:
            delta = _dec(f['cantidad_aplicada'])
            pid, aid = f['producto_id'], f['almacen_id']
            producto = Producto.query.get(pid)

            if delta > 0:
                if origen == 'general':
                    # Mover de General al proyecto: REASIGNACION.
                    e = _consumir_bucket_exacto(pid, aid, None, delta)
                    if e:
                        # La previsualización dijo que alcanzaba; si aquí no,
                        # alguien movió stock en medio. Se aborta entero.
                        db.session.rollback()
                        return jsonify({
                            'detail': f'El stock cambió mientras se aplicaba ({f["sku"]}): {e}. '
                                      f'Vuelve a previsualizar.',
                        }), 409
                    _depositar(pid, aid, proyecto.id, delta)
                    tipo, p_org, p_dst = 'REASIGNACION', None, proyecto.id
                else:
                    # Material que llega de fuera directo al proyecto: ENTRADA.
                    _depositar(pid, aid, proyecto.id, delta)
                    tipo, p_org, p_dst = 'ENTRADA', None, proyecto.id
            else:
                # Sobrante en modo reemplazar: vuelve a General.
                e = _consumir_bucket_exacto(pid, aid, proyecto.id, -delta)
                if e:
                    db.session.rollback()
                    return jsonify({
                        'detail': f'El stock cambió mientras se aplicaba ({f["sku"]}): {e}. '
                                  f'Vuelve a previsualizar.',
                    }), 409
                _depositar(pid, aid, None, -delta)
                tipo, p_org, p_dst = 'REASIGNACION', proyecto.id, None

            _recalcular_caches(producto, aid)
            db.session.add(MovimientoInventario(
                tipo=tipo, producto_id=pid, cantidad=abs(delta),
                almacen_origen_id=aid if tipo == 'REASIGNACION' else None,
                almacen_destino_id=aid,
                proyecto_origen_id=p_org, proyecto_destino_id=p_dst,
                motivo=motivo or f'Asignación a {proyecto.numero_proyecto}',
                usuario_id=user.id,
            ))

        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        if _es_error_de_lock(exc):
            return jsonify({'detail': 'Stock bloqueado por otra operación, reintenta'}), 409
        raise

    _audit(user, (
        f'Asignó {len(aplicables)} materiales al proyecto '
        f'{proyecto.numero_proyecto} (origen: {origen}, modo: {modo})'
    ))
    db.session.commit()
    emit_to_role(_INV_ROLES, 'producto:changed', {'action': 'asignacion_proyecto',
                                                  'proyecto_id': proyecto.id})

    return jsonify({
        'ok': True,
        'aplicadas': len(aplicables),
        'lineas': plan,
        'resumen': _resumen(plan),
    })


@bp.route('/proyectos-materiales/<int:proyecto_id>/devolver', methods=['POST'])
@_require_inventario_admin
def devolver_material(proyecto_id: int):
    """Saca material del proyecto: a General o a otro proyecto.

    `destino_proyecto_id` ausente o null = General. Siempre es una REASIGNACION,
    nunca una salida: el material no se va del almacén, solo cambia de etiqueta.
    """
    proyecto = Proyecto.query.get(proyecto_id)
    if not proyecto:
        return jsonify({'detail': 'Proyecto no encontrado'}), 404

    data = request.get_json(silent=True) or {}
    lineas = data.get('lineas')
    if not isinstance(lineas, list) or not lineas:
        return jsonify({'detail': 'Se requiere al menos una línea'}), 422

    destino_id = data.get('destino_proyecto_id')
    if destino_id:
        destino = Proyecto.query.get(destino_id)
        if not destino:
            return jsonify({'detail': 'Proyecto destino no encontrado'}), 404
        if destino.id == proyecto.id:
            return jsonify({'detail': 'El destino debe ser distinto del origen'}), 422
        etiqueta_destino = destino.numero_proyecto
    else:
        destino_id, etiqueta_destino = None, 'General'

    from app.routes._api_helpers import current_user
    user = current_user()
    aplicadas, problemas = 0, []

    try:
        for cruda in lineas:
            pid = cruda.get('producto_id')
            aid = cruda.get('almacen_id')
            cantidad = _dec(cruda.get('cantidad'))
            if not pid or not aid or cantidad <= 0:
                problemas.append({'producto_id': pid, 'motivo': 'Línea incompleta'})
                continue
            producto = Producto.query.get(pid)
            if not producto:
                problemas.append({'producto_id': pid, 'motivo': 'Material no encontrado'})
                continue

            e = _consumir_bucket_exacto(pid, aid, proyecto.id, cantidad)
            if e:
                problemas.append({'producto_id': pid, 'sku': producto.codigo, 'motivo': e})
                continue
            _depositar(pid, aid, destino_id, cantidad)
            _recalcular_caches(producto, aid)
            db.session.add(MovimientoInventario(
                tipo='REASIGNACION', producto_id=pid, cantidad=cantidad,
                almacen_origen_id=aid, almacen_destino_id=aid,
                proyecto_origen_id=proyecto.id, proyecto_destino_id=destino_id,
                motivo=(data.get('motivo') or '').strip()
                       or f'Devolución de {proyecto.numero_proyecto} a {etiqueta_destino}',
                usuario_id=user.id,
            ))
            aplicadas += 1

        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        if _es_error_de_lock(exc):
            return jsonify({'detail': 'Stock bloqueado por otra operación, reintenta'}), 409
        raise

    _audit(user, (
        f'Devolvió {aplicadas} materiales de {proyecto.numero_proyecto} '
        f'a {etiqueta_destino}'
    ))
    db.session.commit()
    emit_to_role(_INV_ROLES, 'producto:changed', {'action': 'devolucion_proyecto',
                                                  'proyecto_id': proyecto.id})

    return jsonify({'ok': True, 'aplicadas': aplicadas, 'problemas': problemas,
                    'destino': etiqueta_destino})


# ── Resumen para las tarjetas de la pantalla principal ──────────────────────

@bp.route('/proyectos-materiales/resumen-asignacion', methods=['GET'])
@_require_inventario_admin
def resumen_asignacion():
    """Cuánto material tiene apartado cada proyecto, y cuánto queda libre.

    Alimenta las tarjetas de la sección «Material por proyecto». General va
    SIEMPRE primero: no es un proyecto más, es el stock libre del que sale casi
    toda asignación y el punto de referencia contra el que se leen los demás.

    Toda la agregación ocurre en SQL — no se baja el catálogo al cliente para
    contarlo.
    """
    filas = (
        db.session.query(
            StockAlmacenProyecto.proyecto_id,
            db.func.count(db.distinct(StockAlmacenProyecto.producto_id)),
            db.func.sum(StockAlmacenProyecto.cantidad),
            db.func.sum(StockAlmacenProyecto.cantidad * Producto.precio_unitario),
        )
        .join(Producto, Producto.id == StockAlmacenProyecto.producto_id)
        .join(Almacen, Almacen.id == StockAlmacenProyecto.almacen_id)
        .filter(
            StockAlmacenProyecto.cantidad > 0,
            Producto.activo == True,   # noqa: E712
            Almacen.activo == True,    # noqa: E712
        )
        .group_by(StockAlmacenProyecto.proyecto_id)
        .all()
    )
    por_proyecto = {
        pid: {'materiales': int(n or 0), 'unidades': _f(u), 'valor': _f(v)}
        for pid, n, u, v in filas
    }

    vacio = {'materiales': 0, 'unidades': 0.0, 'valor': 0.0}
    tarjetas = [{
        'proyecto_id': None,
        'numero_proyecto': 'General',
        'nombre': 'Stock libre, sin apartar',
        'es_general': True,
        **por_proyecto.get(None, vacio),
    }]

    # Se listan TODOS los proyectos activos, incluso sin material: poder ver
    # que una obra no tiene nada apartado es información, y además es el punto
    # de partida natural para asignarle.
    for p in (Proyecto.query
              .filter(Proyecto.activo == True)  # noqa: E712
              .order_by(Proyecto.numero_proyecto)
              .all()):
        tarjetas.append({
            'proyecto_id': p.id,
            'numero_proyecto': p.numero_proyecto,
            'nombre': p.nombre or '',
            'es_general': False,
            **por_proyecto.get(p.id, vacio),
        })

    return jsonify({
        'tarjetas': tarjetas,
        'total_apartado': _f(sum(
            d['unidades'] for pid, d in por_proyecto.items() if pid is not None
        )),
    })


# ── Plantilla de Excel para la carga masiva ─────────────────────────────────

def _construir_plantilla(proyecto: Proyecto) -> "io.BytesIO":  # noqa: F821
    """Excel de TRES columnas: SKU, Cantidad, Bodega.

    La plantilla de catálogo pide trece columnas porque sirve para dar de alta
    materiales. Aquí el material ya existe: lo único que varía es cuánto y
    dónde. Pedir descripción, marca, unidad o precio otra vez sería hacer
    teclear datos que el sistema ya tiene.

    El PROYECTO tampoco es columna: la plantilla se descarga desde dentro del
    proyecto, así que ya se sabe cuál es. Ponerlo sería una oportunidad más de
    equivocarse.

    Se pre-llena con lo que el proyecto YA tiene, para que ajustar cantidades no
    obligue a teclear los SKU.
    """
    import io

    import openpyxl
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    COLS = ['SKU', 'Cantidad', 'Bodega']
    ANCHOS = {'SKU': 22, 'Cantidad': 12, 'Bodega': 24}
    AYUDAS = {
        'SKU': 'Código del material tal como está en el catálogo. Debe existir.',
        'Cantidad': 'Cuánto se asigna a este proyecto. Número mayor que cero.',
        'Bodega': 'Nombre de la bodega. Si se deja vacío se usa la predeterminada.',
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Asignación'

    azul = PatternFill('solid', fgColor='1E40AF')
    gris = PatternFill('solid', fgColor='F1F5F9')
    thin = Side(border_style='thin', color='CBD5E1')
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Fila 1: a qué proyecto va todo esto. Es contexto, no un dato a llenar.
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    c = ws.cell(row=1, column=1,
                value=f'Asignar material a: {proyecto.numero_proyecto}'
                      f'{" — " + proyecto.nombre if proyecto.nombre else ""}')
    c.font = Font(bold=True, color='111827', size=14)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLS))
    c = ws.cell(row=2, column=1, value=(
        'Llena una fila por material. El proyecto ya está definido — no hace falta '
        'repetirlo. Al subir el archivo verás una vista previa antes de aplicar nada.'
    ))
    c.font = Font(italic=True, color='6B7280', size=10)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[2].height = 28

    # Fila 3: encabezados.
    for i, col in enumerate(COLS, 1):
        cel = ws.cell(row=3, column=i, value=col)
        cel.fill = azul
        cel.font = Font(bold=True, color='FFFFFF', size=11)
        cel.alignment = Alignment(horizontal='center', vertical='center')
        cel.border = borde
        cel.comment = Comment(AYUDAS[col], 'SKILLED')
        ws.column_dimensions[get_column_letter(i)].width = ANCHOS[col]
    ws.freeze_panes = 'A4'

    # Pre-llenado con lo que el proyecto ya tiene.
    existentes = (
        db.session.query(Producto.codigo, StockAlmacenProyecto.cantidad, Almacen.nombre)
        .join(Producto, Producto.id == StockAlmacenProyecto.producto_id)
        .join(Almacen, Almacen.id == StockAlmacenProyecto.almacen_id)
        .filter(
            StockAlmacenProyecto.proyecto_id == proyecto.id,
            StockAlmacenProyecto.cantidad > 0,
            Producto.activo == True,   # noqa: E712
            Almacen.activo == True,    # noqa: E712
        )
        .order_by(Producto.codigo)
        .all()
    )
    fila = 4
    for codigo, cantidad, bodega in existentes:
        ws.cell(row=fila, column=1, value=codigo).border = borde
        # Se deja la CANTIDAD VACÍA a propósito: pre-llenarla con lo que ya
        # tiene invitaría a subir el archivo sin tocarlo y duplicar todo el
        # material del proyecto. El SKU y la bodega sí se dan hechos, que es
        # lo tedioso de teclear.
        ws.cell(row=fila, column=2).border = borde
        ws.cell(row=fila, column=3, value=bodega).border = borde
        for col in range(1, len(COLS) + 1):
            ws.cell(row=fila, column=col).fill = gris
        fila += 1

    ultima = max(fila + 200, 210)

    # Validación: cantidad numérica y positiva.
    dv = DataValidation(type='decimal', operator='greaterThan', formula1='0',
                        allow_blank=True, showErrorMessage=True)
    dv.error = 'La cantidad debe ser un número mayor que cero.'
    dv.errorTitle = 'Cantidad inválida'
    ws.add_data_validation(dv)
    dv.add(f'B4:B{ultima}')

    # Lista de bodegas válidas, para no depender de que se escriban bien.
    nombres = [a.nombre for a in Almacen.query.filter(Almacen.activo == True).all()]  # noqa: E712
    if nombres and sum(len(n) + 1 for n in nombres) < 250:
        dv_b = DataValidation(type='list', formula1='"' + ','.join(nombres) + '"',
                              allow_blank=True, showErrorMessage=True)
        dv_b.error = 'Elige una bodega de la lista.'
        dv_b.errorTitle = 'Bodega inválida'
        ws.add_data_validation(dv_b)
        dv_b.add(f'C4:C{ultima}')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@bp.route('/proyectos-materiales/<int:proyecto_id>/plantilla-asignacion', methods=['GET'])
@_require_inventario_admin
def plantilla_asignacion(proyecto_id: int):
    """Descarga la plantilla de asignación de ESTE proyecto."""
    from flask import send_file

    proyecto = Proyecto.query.get(proyecto_id)
    if not proyecto:
        return jsonify({'detail': 'Proyecto no encontrado'}), 404

    try:
        buf = _construir_plantilla(proyecto)
    except ImportError:
        return jsonify({'detail': 'openpyxl no instalado en el servidor'}), 500

    seguro = ''.join(ch for ch in (proyecto.numero_proyecto or 'proyecto')
                     if ch.isalnum() or ch in '-_')
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'asignacion_{seguro or "proyecto"}.xlsx',
    )


# ── Importar la plantilla llena ─────────────────────────────────────────────

@bp.route('/proyectos-materiales/<int:proyecto_id>/asignar/importar', methods=['POST'])
@_require_inventario_admin
def importar_asignacion(proyecto_id: int):
    """Lee el Excel y devuelve la MISMA previsualización que `previsualizar`.

    No escribe nada. El archivo se traduce a líneas y se pasa por
    `_resolver_lineas`, igual que la captura a mano: así el usuario ve exactamente
    lo que va a pasar, y confirmar dispara el mismo `/asignar` de siempre. Subir
    un archivo no es un camino distinto al de capturar — es solo otra forma de
    llenar las líneas.
    """
    import io as _io

    proyecto = Proyecto.query.get(proyecto_id)
    if not proyecto:
        return jsonify({'detail': 'Proyecto no encontrado'}), 404

    try:
        import pandas as pd
    except ImportError:
        return jsonify({'detail': 'pandas no instalado en el servidor'}), 500

    file = request.files.get('archivo')
    if not file or not file.filename:
        return jsonify({'detail': 'No se envió archivo'}), 400
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'detail': 'Formato no válido. Debe ser .xlsx o .xls'}), 400

    file.stream.seek(0, _io.SEEK_END)
    size = file.stream.tell()
    file.stream.seek(0)
    if size > 5 * 1024 * 1024:
        return jsonify({'detail': 'Archivo demasiado grande. Máximo 5 MB.'}), 413
    if size < 100:
        return jsonify({'detail': 'Archivo vacío o corrupto.'}), 400

    try:
        raw = pd.read_excel(file, header=None, nrows=2100)
    except Exception as e:
        return jsonify({
            'detail': 'No se pudo leer el Excel. Usa la plantilla y guárdalo como .xlsx.',
            'tecnico': str(e)[:200],
        }), 400

    def _norm(v):
        import unicodedata
        s = unicodedata.normalize('NFKD', str(v or ''))
        return ''.join(c for c in s if not unicodedata.combining(c)).strip().lower()

    # Los encabezados van en la fila 3 de la plantilla, pero se buscan por
    # contenido: si alguien inserta o borra una fila arriba, el archivo sigue
    # sirviendo. Solo SKU y Cantidad son obligatorios — sin bodega se usa la
    # predeterminada, que es justo lo que hace la captura a mano.
    ALIAS = {
        'sku': 'sku', 'codigo': 'sku', 'codigo (sku)': 'sku', 'clave': 'sku',
        'cantidad': 'cantidad', 'cant': 'cantidad', 'cantidad a asignar': 'cantidad',
        'bodega': 'almacen', 'almacen': 'almacen',
    }
    fila_encabezado, columnas = None, {}
    for ridx in range(min(12, len(raw))):
        encontrado = {}
        for i, val in enumerate(raw.iloc[ridx].tolist()):
            campo = ALIAS.get(_norm(val))
            if campo and campo not in encontrado:
                encontrado[campo] = i
        if 'sku' in encontrado and 'cantidad' in encontrado:
            fila_encabezado, columnas = ridx, encontrado
            break

    if fila_encabezado is None:
        return jsonify({
            'detail': 'No se encontraron las columnas SKU y Cantidad. '
                      'Descarga la plantilla del proyecto y llénala sin borrar los encabezados.',
        }), 400

    datos = raw.iloc[fila_encabezado + 1:]
    if len(datos) > 2000:
        return jsonify({'detail': f'Demasiadas filas ({len(datos)}). Máximo 2000.'}), 400

    lineas, vacias = [], 0
    for _, fila in datos.iterrows():
        def _celda(campo):
            i = columnas.get(campo)
            if i is None or i >= len(fila):
                return ''
            v = fila.iloc[i]
            return '' if pd.isna(v) else str(v).strip()

        sku, cantidad = _celda('sku'), _celda('cantidad')
        # Fila totalmente vacía = final del llenado, no un error. La plantilla
        # trae 200 filas en blanco a propósito; reportarlas como errores llenaría
        # la vista previa de ruido y escondería los problemas de verdad.
        if not sku and not cantidad:
            vacias += 1
            continue
        # SKU prellenado al que no le pusieron cantidad: tampoco es un error,
        # es material del proyecto que esta vez no se toca.
        if sku and not cantidad:
            vacias += 1
            continue
        lineas.append({'sku': sku, 'cantidad': cantidad, 'almacen': _celda('almacen')})

    if not lineas:
        return jsonify({
            'detail': 'El archivo no tiene ninguna fila con cantidad. '
                      'Escribe cuánto asignar en la columna Cantidad.',
        }), 422

    origen = (request.form.get('origen') or 'general').strip().lower()
    if origen not in ('general', 'entrada'):
        origen = 'general'
    modo = (request.form.get('modo') or 'sumar').strip().lower()
    if modo not in ('sumar', 'reemplazar'):
        modo = 'sumar'

    plan = _resolver_lineas(proyecto.id, lineas, origen, modo)
    return jsonify({
        'proyecto': {'id': proyecto.id, 'numero_proyecto': proyecto.numero_proyecto,
                     'nombre': proyecto.nombre or ''},
        'origen': origen,
        'modo': modo,
        'lineas': plan,
        'resumen': _resumen(plan),
        'filas_ignoradas': vacias,
    })


# ── Stock libre (General) ───────────────────────────────────────────────────

@bp.route('/proyectos-materiales/general/existencias', methods=['GET'])
@_require_inventario_admin
def existencias_general():
    """Material libre —sin apartar a ninguna obra—, desglosado por bodega.

    Es el espejo de `/<id>/existencias`, pero para el bucket sin proyecto. Hace
    falta porque el sentido natural del flujo es General → obra: quien tiene
    material libre quiere mandarlo a un proyecto, y hasta ahora el único camino
    era entrar primero al proyecto y buscar de vuelta el material.

    A diferencia del de un proyecto, este SÍ pagina: General suele tener el
    catálogo casi entero, y bajarlo completo al navegador sería regalar unos
    cuantos megabytes por cada visita a la pantalla.
    """
    q = (request.args.get('q') or '').strip()
    try:
        pagina = max(1, int(request.args.get('page', 1)))
    except (TypeError, ValueError):
        pagina = 1
    try:
        por_pagina = min(200, max(1, int(request.args.get('per_page', 50))))
    except (TypeError, ValueError):
        por_pagina = 50

    base = (
        db.session.query(StockAlmacenProyecto.producto_id)
        .join(Producto, Producto.id == StockAlmacenProyecto.producto_id)
        .join(Almacen, Almacen.id == StockAlmacenProyecto.almacen_id)
        .filter(
            StockAlmacenProyecto.proyecto_id.is_(None),
            StockAlmacenProyecto.cantidad > 0,
            Producto.activo == True,   # noqa: E712
            Almacen.activo == True,    # noqa: E712
        )
    )
    if q:
        patron = f'%{q}%'
        base = base.filter(db.or_(Producto.codigo.ilike(patron),
                                  Producto.descripcion.ilike(patron)))

    # Se pagina por PRODUCTO, no por bucket: una fila de la tabla es un material
    # con sus bodegas al lado. Paginar por bucket partiría un material a la
    # mitad entre dos páginas.
    ids_q = base.group_by(StockAlmacenProyecto.producto_id).subquery()
    total = db.session.query(db.func.count()).select_from(ids_q).scalar() or 0

    ids = [
        pid for (pid,) in
        base.group_by(StockAlmacenProyecto.producto_id)
        .order_by(db.func.sum(StockAlmacenProyecto.cantidad).desc())
        .limit(por_pagina).offset((pagina - 1) * por_pagina).all()
    ]

    materiales, bodegas = [], {}
    if ids:
        filas = (
            db.session.query(
                Producto.id, Producto.codigo, Producto.descripcion, Producto.unidad,
                Producto.precio_unitario, Almacen.id, Almacen.nombre,
                StockAlmacenProyecto.cantidad,
            )
            .join(Producto, Producto.id == StockAlmacenProyecto.producto_id)
            .join(Almacen, Almacen.id == StockAlmacenProyecto.almacen_id)
            .filter(
                StockAlmacenProyecto.proyecto_id.is_(None),
                StockAlmacenProyecto.cantidad > 0,
                StockAlmacenProyecto.producto_id.in_(ids),
                Almacen.activo == True,  # noqa: E712
            )
            .all()
        )
        acumulado = {}
        for pid, codigo, desc, unidad, precio, aid, anombre, cant in filas:
            bodegas.setdefault(aid, anombre)
            m = acumulado.setdefault(pid, {
                'producto_id': pid, 'codigo': codigo, 'descripcion': desc,
                'unidad': unidad, 'precio_unitario': _f(precio or 0),
                'total': Decimal('0'), 'por_almacen': {},
            })
            c = _dec(cant)
            m['total'] += c
            m['por_almacen'][aid] = _f(_dec(m['por_almacen'].get(aid, 0)) + c)

        orden = {pid: i for i, pid in enumerate(ids)}
        for m in sorted(acumulado.values(), key=lambda r: orden[r['producto_id']]):
            total_m = m.pop('total')
            materiales.append({
                **m,
                'total': _f(total_m),
                'valor': _f(total_m * _dec(m['precio_unitario'])),
            })

    return jsonify({
        'almacenes': [{'id': aid, 'nombre': n}
                      for aid, n in sorted(bodegas.items(), key=lambda kv: kv[1])],
        'materiales': materiales,
        'total': int(total),
        'page': pagina,
        'per_page': por_pagina,
        'pages': max(1, (int(total) + por_pagina - 1) // por_pagina),
    })
