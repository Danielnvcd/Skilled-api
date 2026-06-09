"""Exportes a Excel: individual, bulk por IDs y exportar-todos.

Registra:
  GET  /<int:id>/exportar
  POST /bulk-exportar
  GET  /exportar-todos
"""
import io
from datetime import date

from flask import jsonify, request, send_file
from sqlalchemy import func
from sqlalchemy.orm import selectinload

from app.models import Proyecto, Trabajador
from app.routes._api_helpers import current_user, is_admin
from app.routes.api_auth import jwt_required
from app.utils import log_action, safe_excel_value

from ._core import bp, _authorized


def _build_export_styles():
    """Estilos compartidos para los dos exports XLSX."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    return {
        'AZUL_OSC': PatternFill('solid', fgColor='1E3A5F'),
        'AZUL_HDR': PatternFill('solid', fgColor='2563EB'),
        'BLANCO': PatternFill('solid', fgColor='FFFFFF'),
        'GRIS_ALT': PatternFill('solid', fgColor='F1F5F9'),
        'BORDER': Border(*(Side(style='thin', color='CBD5E1') for _ in range(4))),
        'BORDER_HDR': Border(*(Side(style='medium', color='1E3A5F') for _ in range(4))),
        'CENTER': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'LEFT': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'FONT_WHITE_BOLD': Font(name='Calibri', color='FFFFFF', bold=True, size=10),
        'FONT_BODY': Font(name='Calibri', color='374151', size=10),
        'FONT_TITLE': Font(name='Calibri', color='FFFFFF', bold=True, size=13),
    }


_HEADERS_EXPORT = [
    'No. Empleado', 'Nombre(s)', 'Apellidos', 'RFC', 'CURP', 'NSS',
    'Fecha Nacimiento', 'Sexo', 'Estado Civil', 'Nacionalidad', 'Edad',
    'Correo', 'Celular', 'Domicilio',
    'Área', 'Puesto', 'Tipo Movimiento', 'Tipo Contrato',
    'Fecha Ingreso', 'Tipo Jornada', 'Tipo Nómina', 'Tipo Pago',
    'Folio IDSE', 'Desc. Servicio',
    'Tipo Sangre', 'Lentes', 'Alergias', 'Estatura', 'Enf. Crónicas', 'Licencia Conducir',
    'Contacto Emergencia', 'Parentesco', 'Tel. Emergencia',
    'Salario Pactado/Sem', 'Salario Base', 'SDI', 'Letra/Categoría',
    'Hrs Extra', 'Infonavit', 'Caja Ahorro', 'Viáticos', 'Día Festivo', 'Pagos Efectivo',
    'Ubicación Actual', 'Estado Ubic.', 'No. Proyecto', 'Coord. a Cargo',
]


def _row_values(t: Trabajador, mask_pii: bool):
    def fmt_d(d):
        return d.strftime('%d/%m/%Y') if d else ''

    def fmt_m(v):
        return float(v) if v else ''

    def mask(val):
        if not val or not mask_pii:
            return val or ''
        return val[:3] + '***' + val[-2:] if len(val) > 5 else '***'

    return [
        t.no_empleado, t.nombre, t.nombre_apellidos,
        mask(t.rfc), mask(t.curp), mask(t.nss),
        fmt_d(t.fecha_nacimiento), t.sexo, t.estado_civil, t.nacionalidad,
        t.edad, t.correo, t.celular, t.domicilio,
        t.area, t.puesto, t.tipo_mov, t.tipo_cont,
        fmt_d(t.fecha_ingreso), t.tipo_jornada, t.tipo_nomina, t.tipo_pago,
        t.folio_mov_idse, t.descripcion_servicio,
        t.tipo_sangre, t.lentes, t.alergias, t.estatura,
        t.enfermedades_cronicas, t.licencia_conducir,
        t.contacto_emergencia, t.parentesco_contacto, t.numero_contacto_emerg,
        fmt_m(t.salario_real_pactado_x_sem), fmt_m(t.sb), fmt_m(t.sdi),
        t.letra, fmt_m(t.hr_extra), fmt_m(t.infonavit),
        fmt_m(t.caja_ahorro), fmt_m(t.viaticos),
        fmt_m(t.pago_dia_festivo), fmt_m(t.pagos_efectivo),
        t.ubicacion_actual, t.ubicacion_estado, t.no_proyecto, t.coord_a_cargo,
    ]


@bp.route('/<int:id>/exportar', methods=['GET'])
@jwt_required
def exportar_uno(id):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    t = (
        Trabajador.query.options(selectinload(Trabajador.credenciales))
        .filter_by(id=id)
        .first()
    )
    if not t:
        return jsonify({'error': 'No encontrado'}), 404
    if not _authorized(t):
        return jsonify({'error': 'Acceso denegado'}), 403

    S = _build_export_styles()
    headers = _HEADERS_EXPORT + ['Observaciones']
    mask_pii = not is_admin()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Empleado'
    ws.row_dimensions[1].height = 36
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    c = ws['A1']
    c.value = (
        f"{t.nombre} {t.nombre_apellidos}".upper()
        + f"  —  No. {t.no_empleado}  —  Generado: {date.today().strftime('%d/%m/%Y')}"
    )
    c.fill = S['AZUL_OSC']; c.font = S['FONT_TITLE']; c.alignment = S['CENTER']; c.border = S['BORDER_HDR']

    ws.row_dimensions[2].height = 22
    for i, header in enumerate(headers, 1):
        cc = ws.cell(row=2, column=i, value=header)
        cc.fill = S['AZUL_HDR']; cc.font = S['FONT_WHITE_BOLD']; cc.alignment = S['CENTER']; cc.border = S['BORDER_HDR']

    ws.row_dimensions[3].height = 18
    values = _row_values(t, mask_pii) + [t.observaciones]
    for i, val in enumerate(values, 1):
        cc = ws.cell(row=3, column=i, value=safe_excel_value(val) if val is not None else '')
        cc.fill = S['BLANCO']; cc.font = S['FONT_BODY']; cc.alignment = S['LEFT']; cc.border = S['BORDER']

    if t.credenciales:
        from openpyxl.styles import Font, PatternFill
        ws2 = wb.create_sheet('Credenciales')
        ws2.row_dimensions[1].height = 22
        for i, label in enumerate(['Planta', 'ID Credencial', 'Caducidad', 'Estado'], 1):
            cc = ws2.cell(row=1, column=i, value=label)
            cc.fill = S['AZUL_HDR']; cc.font = S['FONT_WHITE_BOLD']; cc.alignment = S['CENTER']; cc.border = S['BORDER_HDR']
        today = date.today()
        verde = PatternFill('solid', fgColor='D1FAE5')
        rojo = PatternFill('solid', fgColor='FEE2E2')
        for i, cred in enumerate(t.credenciales, 2):
            vencida = cred.fecha_caducidad and cred.fecha_caducidad < today
            cad_str = cred.fecha_caducidad.strftime('%d/%m/%Y') if cred.fecha_caducidad else ''
            estado = 'CADUCADA' if vencida else 'VIGENTE'
            row_fill = rojo if vencida else verde
            for j, val in enumerate([cred.planta, cred.credencial_id, cad_str, estado], 1):
                cc = ws2.cell(row=i, column=j, value=val or '')
                cc.fill = row_fill
                cc.font = Font(name='Calibri', color='7F1D1D' if vencida else '065F46', size=10)
                cc.alignment = S['LEFT']; cc.border = S['BORDER']
        for i, w in enumerate([20, 20, 14, 12], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

    widths = [14, 18, 20, 14, 20, 14, 14, 8, 14, 14, 6, 28, 14, 32,
              16, 20, 16, 16, 14, 14, 14, 12, 12, 20,
              12, 8, 16, 10, 16, 16,
              22, 14, 14,
              16, 14, 12, 14, 12, 12, 12, 12, 12, 14,
              20, 16, 14, 20, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    safe_name = f"{t.no_empleado}_{(t.nombre_apellidos or '').replace(' ', '_')}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=safe_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@bp.route('/bulk-exportar', methods=['POST'])
@jwt_required
def bulk_exportar():
    """Exporta a Excel solo los empleados de la selección recibida.

    Body: { "ids": [int, ...] }  (1..200 ids).

    Reusa el mismo layout y estilos del export-todos pero filtrado por IDs.
    El filtro de coordinador sigue aplicando: solo verá empleados de sus
    proyectos aunque haya pasado IDs ajenos en el body — anti IDOR.
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    payload = request.get_json(silent=True) or {}
    raw_ids = payload.get('ids') or []
    if not isinstance(raw_ids, list) or not raw_ids:
        return jsonify({'error': 'Lista de IDs vacía'}), 422
    if len(raw_ids) > 200:
        return jsonify({'error': 'Máximo 200 IDs por operación'}), 422
    try:
        ids = sorted({int(i) for i in raw_ids})
    except (TypeError, ValueError):
        return jsonify({'error': 'IDs deben ser enteros'}), 422

    base = Trabajador.query.filter(Trabajador.id.in_(ids))
    # Mismo gate que /exportar-todos: el coordinador solo ve a los suyos.
    # Si pasa IDs fuera de su scope se filtran silenciosamente.
    if current_user().role == 'coordinador':
        mis = Proyecto.query.filter_by(activo=True, coordinador_id=current_user().id).all()
        permitidos = {t.id for p in mis for t in p.participantes}
        base = base.filter(Trabajador.id.in_(permitidos))
    elif not is_admin():
        return jsonify({'error': 'Acceso denegado'}), 403

    trabajadores = base.order_by(func.lower(Trabajador.nombre)).all()
    if not trabajadores:
        return jsonify({'error': 'Ninguno de los IDs es accesible'}), 404

    S = _build_export_styles()
    mask_pii = not is_admin()

    wb = Workbook(); ws = wb.active; ws.title = 'Empleados'
    ws.row_dimensions[1].height = 38
    ws.merge_cells(f'A1:{get_column_letter(len(_HEADERS_EXPORT))}1')
    c = ws['A1']
    c.value = (
        f"LISTADO DE EMPLEADOS (SELECCIÓN) — "
        f"{len(trabajadores)} de {len(ids)} solicitados — "
        f"Generado: {date.today().strftime('%d/%m/%Y')}"
    )
    c.fill = S['AZUL_OSC']; c.font = S['FONT_TITLE']; c.alignment = S['CENTER']; c.border = S['BORDER_HDR']

    ws.row_dimensions[2].height = 22
    for i, header in enumerate(_HEADERS_EXPORT, 1):
        cc = ws.cell(row=2, column=i, value=header)
        cc.fill = S['AZUL_HDR']; cc.font = S['FONT_WHITE_BOLD']; cc.alignment = S['CENTER']; cc.border = S['BORDER_HDR']

    for row_idx, t in enumerate(trabajadores, start=3):
        fill = S['GRIS_ALT'] if (row_idx % 2 == 0) else S['BLANCO']
        ws.row_dimensions[row_idx].height = 16
        for i, val in enumerate(_row_values(t, mask_pii), 1):
            cc = ws.cell(row=row_idx, column=i, value=safe_excel_value(val) if val is not None else '')
            cc.fill = fill; cc.font = S['FONT_BODY']; cc.alignment = S['LEFT']; cc.border = S['BORDER']

    widths = [14, 18, 20, 14, 20, 14, 14, 8, 14, 14, 6, 28, 14, 32,
              16, 20, 16, 16, 14, 14, 14, 12, 12, 20,
              12, 8, 16, 10, 16, 16,
              22, 14, 14,
              16, 14, 12, 14, 12, 12, 12, 12, 12, 14,
              20, 16, 14, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"empleados_seleccion_{date.today().strftime('%Y%m%d')}.xlsx"
    log_action(f'Bulk export trabajadores: {len(trabajadores)} de {len(ids)} solicitados')
    return send_file(
        buf, as_attachment=True, download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@bp.route('/exportar-todos', methods=['GET'])
@jwt_required
def exportar_todos():
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    base = Trabajador.query.filter(Trabajador.activo == True, Trabajador.fecha_baja == None)  # noqa: E712
    if current_user().role == 'coordinador':
        mis = Proyecto.query.filter_by(activo=True, coordinador_id=current_user().id).all()
        ids = {t.id for p in mis for t in p.participantes}
        base = base.filter(Trabajador.id.in_(ids))

    trabajadores = base.order_by(func.lower(Trabajador.nombre)).all()
    S = _build_export_styles()
    mask_pii = not is_admin()

    wb = Workbook(); ws = wb.active; ws.title = 'Empleados'
    ws.row_dimensions[1].height = 38
    ws.merge_cells(f'A1:{get_column_letter(len(_HEADERS_EXPORT))}1')
    c = ws['A1']
    c.value = f"LISTADO DE EMPLEADOS — Generado: {date.today().strftime('%d/%m/%Y')}"
    c.fill = S['AZUL_OSC']; c.font = S['FONT_TITLE']; c.alignment = S['CENTER']; c.border = S['BORDER_HDR']

    ws.row_dimensions[2].height = 22
    for i, header in enumerate(_HEADERS_EXPORT, 1):
        cc = ws.cell(row=2, column=i, value=header)
        cc.fill = S['AZUL_HDR']; cc.font = S['FONT_WHITE_BOLD']; cc.alignment = S['CENTER']; cc.border = S['BORDER_HDR']

    for row_idx, t in enumerate(trabajadores, start=3):
        fill = S['GRIS_ALT'] if (row_idx % 2 == 0) else S['BLANCO']
        ws.row_dimensions[row_idx].height = 16
        for i, val in enumerate(_row_values(t, mask_pii), 1):
            cc = ws.cell(row=row_idx, column=i, value=safe_excel_value(val) if val is not None else '')
            cc.fill = fill; cc.font = S['FONT_BODY']; cc.alignment = S['LEFT']; cc.border = S['BORDER']

    widths = [14, 18, 20, 14, 20, 14, 14, 8, 14, 14, 6, 28, 14, 32,
              16, 20, 16, 16, 14, 14, 14, 12, 12, 20,
              12, 8, 16, 10, 16, 16,
              22, 14, 14,
              16, 14, 12, 14, 12, 12, 12, 12, 12, 14,
              20, 16, 14, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"empleados_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
