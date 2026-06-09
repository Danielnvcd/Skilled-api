"""Importación masiva desde Excel + plantilla descargable.

Registra:
  GET  /plantilla-importar
  POST /importar

`procesar_excel_trabajadores` se re-exporta desde `__init__.py` porque la
ruta legacy `trabajadores.procesar_importacion` la importa para reusar la
misma lógica de validación.
"""
import io
import os
import re
import traceback

from flask import current_app, jsonify, request, send_file

from app.extensions import db, limiter
from app.models import Trabajador
from app.realtime import emit_to_role
from app.routes._api_helpers import is_admin
from app.routes.api_auth import jwt_required
from ._core import _CURP_RE, _RFC_RE
from app.utils import log_action
from app.utils.excel import (
    ESTADO_CIVIL_ALIASES,
    LENTES_ALIASES,
    SEXO_ALIASES,
    TIPO_NOMINA_ALIASES,
    TIPO_PAGO_ALIASES,
    cell_date,
    cell_number,
    cell_str,
    find_header_row,
    normalize_choice,
    normalize_curp,
    normalize_rfc,
    read_cell,
)

from ._core import bp


# Encabezados oficiales de la plantilla de importación de empleados.
# OBLIGATORIOS para que la fila se considere válida: No. Empleado, Nombre(s), Apellidos.
# El resto es opcional — si la columna no viene, simplemente no se asigna.
PLANTILLA_EMP_HEADERS = [
    # IDENTIFICADORES
    'No. Empleado', 'Nombre(s)', 'Apellidos',
    # LABORALES
    'Area', 'Puesto', 'Tipo de Nomina (Semanal/Por hora/Cuadrado)',
    'Salario Real Pactado por Semana',
    'Sueldo Base (SB)', 'Salario Diario Integrado (SDI)', 'Letra',
    'Horas Extra', 'Infonavit', 'Caja de Ahorro', 'Viaticos', 'Pago Dia Festivo',
    'Tipo de Movimiento', 'Tipo de Contrato', 'Tipo de Jornada',
    'Descripcion de Servicio',
    'Fecha Ingreso (YYYY-MM-DD)', 'Fecha Inicio (YYYY-MM-DD)',
    'Termino de Prueba (YYYY-MM-DD)',
    # DATOS PERSONALES
    'CURP', 'RFC', 'NSS',
    'Fecha Nacimiento (YYYY-MM-DD)', 'Edad', 'Sexo (M/F)',
    'Estado Civil', 'Nacionalidad', 'Domicilio', 'Correo', 'Celular',
    # DATOS MEDICOS / EMERGENCIA
    'Tipo de Sangre', 'Alergias', 'Enfermedades Cronicas',
    'Contacto de Emergencia', 'Parentesco del Contacto',
    'Numero Contacto Emergencia',
    'Usa Lentes (Si/No)', 'Licencia de Conducir (Tipo)', 'Estatura',
    # FINANCIERO
    'Folio Mov IDSE', 'Tipo Pago',
    # UBICACION Y OPERACION
    'Ubicacion Estado', 'Observaciones',
]

PLANTILLA_EMP_REQUIRED = {'No. Empleado', 'Nombre(s)', 'Apellidos'}
PLANTILLA_EMP_MAX_FILAS = 5000  # Tope para evitar DoS por archivo gigante

_EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')


@bp.route('/plantilla-importar', methods=['GET'])
@jwt_required
def descargar_plantilla_importar():
    """Genera y sirve un Excel de plantilla para carga masiva de empleados.

    Generada en memoria (sin depender de archivos en disco). Incluye:
    - Título e instrucciones en filas 1-2.
    - Headers oficiales en fila 4 (estilo azul, freeze panes en A5).
    - Tooltips/comentarios en cada header indicando formato esperado.
    - Dropdowns para Sexo, Tipo de Nómina, Tipo de Pago, Tipo de Sangre,
      Usa Lentes y Estado Civil — sin permitir valores fuera del catálogo.
    - Validaciones de longitud para CURP (18), RFC (13), NSS (20).
    - Validaciones numéricas (>= 0) para todos los campos monetarios.
    """
    try:
        import openpyxl
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        return jsonify({'error': 'openpyxl no instalado en el servidor'}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Empleados'

    # Estilos (mismo lenguaje visual que la plantilla de inventario)
    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    title_font = Font(bold=True, color='111827', size=14)
    instr_font = Font(italic=True, color='6B7280', size=10)
    thin = Side(border_style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ncols = len(PLANTILLA_EMP_HEADERS)

    # Fila 1: título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value='Plantilla de Importación de Empleados — SKILLED')
    c.font = title_font
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Fila 2: instrucciones
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(
        row=2, column=1,
        value=(
            '⚠ No alteres los encabezados de la fila 4. Campos obligatorios: '
            'No. Empleado, Nombre(s), Apellidos. Fechas en formato YYYY-MM-DD '
            '(también se aceptan DD/MM/YYYY). Los duplicados de No. Empleado se ignoran.'
        ),
    )
    c.font = instr_font
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 38

    # Fila 3: separador
    ws.row_dimensions[3].height = 6

    # Fila 4: headers
    for col, header in enumerate(PLANTILLA_EMP_HEADERS, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    ws.row_dimensions[4].height = 38
    ws.freeze_panes = 'A5'

    # Tooltips por header — usa el índice (1-based) que corresponde a cada nombre
    # en PLANTILLA_EMP_HEADERS. Si el orden cambia, los tooltips siguen el header
    # gracias a este lookup por nombre.
    tooltips_por_header = {
        'No. Empleado': 'OBLIGATORIO — identificador único (máx 50). Duplicados se ignoran.',
        'Nombre(s)': 'OBLIGATORIO — nombre(s) del empleado (máx 250).',
        'Apellidos': 'OBLIGATORIO — apellidos completos (máx 250).',
        'CURP': '18 caracteres en MAYÚSCULAS. Ej: GARC850315HDFRRL09',
        'RFC': '13 caracteres (personas físicas) o 12 (morales). Ej: GARC850315ABC',
        'NSS': 'Número de Seguridad Social (máx 20). Solo dígitos.',
        'Sexo (M/F)': 'M, F, Masculino, Femenino, Hombre o Mujer — se normaliza a M/F.',
        'Tipo de Nomina (Semanal/Por hora/Cuadrado)': 'Semanal | Por hora | Cuadrado',
        'Estado Civil': 'Soltero(a), Casado(a), Unión Libre, Divorciado(a), Viudo(a).',
        'Tipo de Sangre': 'A+, A-, B+, B-, AB+, AB-, O+, O-.',
        'Usa Lentes (Si/No)': 'Sí o No.',
        'Tipo Pago': 'EFECTIVO o TRANSFERENCIA (SPEI, Depósito y Banco se mapean a TRANSFERENCIA).',
        'Fecha Ingreso (YYYY-MM-DD)': 'Fecha en formato ISO 2026-01-31 (también acepta 31/01/2026).',
        'Fecha Inicio (YYYY-MM-DD)': 'Misma regla de fecha.',
        'Termino de Prueba (YYYY-MM-DD)': 'Misma regla de fecha.',
        'Fecha Nacimiento (YYYY-MM-DD)': 'Misma regla de fecha.',
        'Salario Real Pactado por Semana': 'Número >= 0. Acepta "$1,234.56".',
        'Sueldo Base (SB)': 'Número >= 0.',
        'Salario Diario Integrado (SDI)': 'Número >= 0.',
        'Horas Extra': 'Número >= 0 (monto por hora extra).',
        'Infonavit': 'Número >= 0.',
        'Caja de Ahorro': 'Número >= 0.',
        'Viaticos': 'Número >= 0.',
        'Pago Dia Festivo': 'Número >= 0.',
        'Edad': 'Entero >= 0.',
        'Correo': 'Email (máx 150). Validación básica de formato.',
        'Celular': 'Número (máx 20). Sin espacios ni guiones.',
        'Contacto de Emergencia': 'Nombre del contacto (máx 200).',
        'Numero Contacto Emergencia': 'Tel del contacto (máx 20).',
    }
    for col, header in enumerate(PLANTILLA_EMP_HEADERS, 1):
        tip = tooltips_por_header.get(header)
        if tip:
            ws.cell(row=4, column=col).comment = Comment(tip, 'Plantilla SKILLED')

    # Anchos diferenciados por tipo de columna
    anchos_por_header = {
        'No. Empleado': 14, 'Nombre(s)': 22, 'Apellidos': 24,
        'Area': 16, 'Puesto': 20, 'Tipo de Nomina (Semanal/Por hora/Cuadrado)': 26,
        'Salario Real Pactado por Semana': 20, 'Sueldo Base (SB)': 14,
        'Salario Diario Integrado (SDI)': 18, 'Letra': 10,
        'Horas Extra': 12, 'Infonavit': 12, 'Caja de Ahorro': 14,
        'Viaticos': 12, 'Pago Dia Festivo': 14,
        'Tipo de Movimiento': 18, 'Tipo de Contrato': 16, 'Tipo de Jornada': 16,
        'Descripcion de Servicio': 24,
        'Fecha Ingreso (YYYY-MM-DD)': 18, 'Fecha Inicio (YYYY-MM-DD)': 18,
        'Termino de Prueba (YYYY-MM-DD)': 20,
        'CURP': 20, 'RFC': 16, 'NSS': 14,
        'Fecha Nacimiento (YYYY-MM-DD)': 20, 'Edad': 7, 'Sexo (M/F)': 12,
        'Estado Civil': 16, 'Nacionalidad': 14, 'Domicilio': 30,
        'Correo': 26, 'Celular': 14,
        'Tipo de Sangre': 12, 'Alergias': 20, 'Enfermedades Cronicas': 22,
        'Contacto de Emergencia': 22, 'Parentesco del Contacto': 18,
        'Numero Contacto Emergencia': 18,
        'Usa Lentes (Si/No)': 12, 'Licencia de Conducir (Tipo)': 18, 'Estatura': 10,
        'Folio Mov IDSE': 16, 'Tipo Pago': 14,
        'Ubicacion Estado': 16, 'Observaciones': 30,
    }
    for col, header in enumerate(PLANTILLA_EMP_HEADERS, 1):
        ws.column_dimensions[get_column_letter(col)].width = anchos_por_header.get(header, 16)

    # ── Data validations (dropdowns y rangos) ──
    # Aplican a filas 5..1004 para mantener el archivo ligero.
    FIRST_DATA_ROW = 5
    LAST_DATA_ROW = 1004

    def _col_letter_for(header_name: str) -> str | None:
        try:
            idx = PLANTILLA_EMP_HEADERS.index(header_name) + 1
            return get_column_letter(idx)
        except ValueError:
            return None

    def _add_list_dv(header_name: str, options: list[str], error_title: str, error_msg: str) -> None:
        col = _col_letter_for(header_name)
        if not col:
            return
        # openpyxl exige la fórmula entre comillas dobles. Si los valores tienen
        # tildes/espacios, Excel los acepta sin problema (string literal).
        formula = '"' + ','.join(options) + '"'
        dv = DataValidation(
            type='list', formula1=formula, allow_blank=True,
            showErrorMessage=True, errorTitle=error_title, error=error_msg,
        )
        dv.add(f'{col}{FIRST_DATA_ROW}:{col}{LAST_DATA_ROW}')
        ws.add_data_validation(dv)

    _add_list_dv('Sexo (M/F)', ['M', 'F'],
                 'Sexo inválido', 'Usa M o F.')
    _add_list_dv('Tipo de Nomina (Semanal/Por hora/Cuadrado)',
                 ['Semanal', 'Por hora', 'Cuadrado'],
                 'Tipo de nómina inválido', 'Usa Semanal, Por hora o Cuadrado.')
    _add_list_dv('Tipo Pago', ['EFECTIVO', 'TRANSFERENCIA'],
                 'Tipo de pago inválido', 'Usa EFECTIVO o TRANSFERENCIA.')
    _add_list_dv('Tipo de Sangre',
                 ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-'],
                 'Tipo de sangre inválido', 'Usa A+, A-, B+, B-, AB+, AB-, O+ u O-.')
    _add_list_dv('Usa Lentes (Si/No)', ['Sí', 'No'],
                 'Valor inválido', 'Usa Sí o No.')
    _add_list_dv('Estado Civil',
                 ['Soltero(a)', 'Casado(a)', 'Unión Libre', 'Divorciado(a)', 'Viudo(a)'],
                 'Estado civil inválido',
                 'Usa Soltero(a), Casado(a), Unión Libre, Divorciado(a) o Viudo(a).')

    # Longitudes para los identificadores fiscales
    def _add_len_dv(header_name: str, min_len: int, max_len: int, title: str, msg: str) -> None:
        col = _col_letter_for(header_name)
        if not col:
            return
        dv = DataValidation(
            type='textLength', operator='between',
            formula1=min_len, formula2=max_len,
            allow_blank=True,
            showErrorMessage=True, errorTitle=title, error=msg,
        )
        dv.add(f'{col}{FIRST_DATA_ROW}:{col}{LAST_DATA_ROW}')
        ws.add_data_validation(dv)

    _add_len_dv('CURP', 18, 18, 'CURP inválido', 'CURP debe tener exactamente 18 caracteres.')
    _add_len_dv('RFC', 12, 13, 'RFC inválido', 'RFC debe tener 12 (moral) o 13 (física) caracteres.')
    _add_len_dv('NSS', 1, 20, 'NSS inválido', 'NSS no debe exceder 20 caracteres.')

    # Números >= 0 para los campos monetarios y edad
    num_headers = [
        'Salario Real Pactado por Semana', 'Sueldo Base (SB)',
        'Salario Diario Integrado (SDI)', 'Horas Extra', 'Infonavit',
        'Caja de Ahorro', 'Viaticos', 'Pago Dia Festivo', 'Edad',
    ]
    num_dv = DataValidation(
        type='decimal', operator='greaterThanOrEqual', formula1=0,
        allow_blank=True, showErrorMessage=True,
        errorTitle='Número inválido',
        error='Debe ser un número mayor o igual a 0.',
    )
    for h in num_headers:
        col = _col_letter_for(h)
        if col:
            num_dv.add(f'{col}{FIRST_DATA_ROW}:{col}{LAST_DATA_ROW}')
    ws.add_data_validation(num_dv)

    # Servir en memoria
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='plantilla_empleados.xlsx',
    )


def procesar_excel_trabajadores(file) -> tuple[dict, int]:
    """Procesa un FileStorage de Excel y crea trabajadores en BD.

    Función pura (no toca request/response). La invocan tanto el endpoint API
    (`/api/trabajadores/importar`) como el endpoint UI legacy
    (`trabajadores.procesar_importacion`) para no duplicar lógica.

    Características "a prueba de tontos":
    - Tamaño máximo 5 MB, magic-bytes verificados (anti DoS / fake xlsx).
    - Detección automática de la fila de encabezados (hasta fila 10) — la
      plantilla los pone en fila 4, pero acepta otros layouts.
    - Headers tolerantes (lower, sin acentos, espacios colapsados) — el
      usuario puede capturar "no. empleado" o "No.   Empleado".
    - Mapeo por nombre de header, no por posición — si reordena columnas, sigue.
    - Tope de PLANTILLA_EMP_MAX_FILAS para evitar Excel gigantes.
    - Validación CURP/RFC con regex oficiales (los mismos del alta manual).
    - Normalización de vocabularios cerrados: sexo, tipo_nomina, tipo_pago,
      lentes (Sí/No/yes/no/1/0), estado civil.
    - Parseo tolerante de números formato MX ('$1,234.56', '50 ').
    - Parseo tolerante de fechas (ISO, MX dd/mm/yyyy, etc.).
    - Anti-fórmula injection: prefija ' a celdas que empiezan con = + - @.
    - Dedup intra-archivo y contra BD del No. Empleado.
    - Commit único al final: si falla, rollback total — nunca queda parcial.

    Returns:
        (payload_dict, http_status_int) — el caller decide cómo serializarlo.
        Cuando falla la lectura/validación temprana, payload tiene 'error' y
        status >= 400. Si todo OK, status=200 y payload tiene 'exitosos',
        'errores' y 'total_filas_procesadas'.
    """
    if not file or not getattr(file, 'filename', ''):
        return {'error': 'No se envió archivo'}, 400
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return {'error': 'Formato no válido. Debe ser .xlsx o .xls'}, 400

    # Magic bytes + tamaño antes de cargar a pandas (evita DoS).
    import filetype
    header_bytes = file.read(2048)
    file.seek(0)
    kind = filetype.guess(header_bytes)
    if not kind or kind.extension not in ('xlsx', 'xls', 'zip', 'cfb'):
        return {'error': 'El archivo no parece ser Excel válido'}, 400

    file.seek(0, os.SEEK_END)
    size = file.tell()
    file.seek(0)
    MAX_BYTES = 5 * 1024 * 1024
    if size > MAX_BYTES:
        return {'error': f'El archivo excede {MAX_BYTES // (1024*1024)}MB'}, 413
    if size < 100:
        return {'error': 'Archivo vacío o corrupto'}, 400

    try:
        import pandas as pd
    except ImportError:
        return {'error': 'pandas no instalado en el servidor'}, 500

    try:
        raw = pd.read_excel(file, header=None, nrows=PLANTILLA_EMP_MAX_FILAS + 20)
    except Exception as e:
        current_app.logger.error('Error leyendo Excel: %s', e)
        return {'error': 'No se pudo leer el Excel. Usa la plantilla y guarda como .xlsx.'}, 400

    header_row_idx, column_map = find_header_row(
        raw, PLANTILLA_EMP_HEADERS, PLANTILLA_EMP_REQUIRED, max_scan=10,
    )
    if header_row_idx is None:
        return {
            'error': 'No se encontraron los encabezados esperados. Descarga la '
                     'plantilla nueva (debe incluir No. Empleado, Nombre(s) y Apellidos).',
        }, 400

    data_df = raw.iloc[header_row_idx + 1:].reset_index(drop=True)
    if len(data_df) > PLANTILLA_EMP_MAX_FILAS:
        return {
            'error': f'Demasiadas filas ({len(data_df)}). Máximo {PLANTILLA_EMP_MAX_FILAS}.',
        }, 400

    errores: list[str] = []
    no_emps_en_archivo: set[str] = set()  # dedup intra-archivo
    a_insertar: list[Trabajador] = []

    def G(row, header_oficial):
        """Atajo: lee la celda por nombre oficial de header."""
        return read_cell(row, column_map, header_oficial)

    for offset, row in data_df.iterrows():
        fila_excel = header_row_idx + offset + 2  # +1 header + 1 (Excel es 1-indexed)
        problemas: list[str] = []

        # Obligatorios
        no_emp = cell_str(G(row, 'No. Empleado'), maxlen=50)
        nombre = cell_str(G(row, 'Nombre(s)'), maxlen=250)
        apellidos = cell_str(G(row, 'Apellidos'), maxlen=250)

        # Detectar fila completamente vacía y saltarla sin reportar — usuarios
        # arrastran filas extra al copiar y pegar; no queremos ruido.
        if not (no_emp or nombre or apellidos):
            continue

        if not no_emp:
            problemas.append('falta No. Empleado')
        if not nombre:
            problemas.append('falta Nombre(s)')
        if not apellidos:
            problemas.append('falta Apellidos')

        # Dedup intra-archivo (case sensitive — la BD también lo es).
        if no_emp and no_emp in no_emps_en_archivo:
            problemas.append(f'No. Empleado {no_emp!r} duplicado en este archivo')
        # Dedup contra BD.
        elif no_emp and Trabajador.query.filter_by(no_empleado=no_emp).first():
            problemas.append(f'el empleado {no_emp!r} ya existe en el sistema')

        # CURP / RFC: normalizar + validar regex si no están vacíos.
        curp_raw = cell_str(G(row, 'CURP'), maxlen=64)
        curp = normalize_curp(curp_raw) if curp_raw else ''
        if curp and not _CURP_RE.match(curp):
            problemas.append(f'CURP inválido ({curp!r}) — formato XXXX######HXXXXXX##')

        rfc_raw = cell_str(G(row, 'RFC'), maxlen=64)
        rfc = normalize_rfc(rfc_raw) if rfc_raw else ''
        if rfc and not _RFC_RE.match(rfc):
            problemas.append(f'RFC inválido ({rfc!r})')

        nss = cell_str(G(row, 'NSS'), maxlen=20)

        # Fechas (parseo tolerante: ISO, MX, US)
        fechas = {}
        for header_excel, attr in (
            ('Fecha Ingreso (YYYY-MM-DD)', 'fecha_ingreso'),
            ('Fecha Inicio (YYYY-MM-DD)', 'inicio'),
            ('Termino de Prueba (YYYY-MM-DD)', 'termino_prueba'),
            ('Fecha Nacimiento (YYYY-MM-DD)', 'fecha_nacimiento'),
        ):
            v, err = cell_date(G(row, header_excel))
            if err:
                problemas.append(f'{header_excel}: {err}')
            fechas[attr] = v

        # Números monetarios y edad
        nums: dict[str, float] = {}
        for header_excel, attr in (
            ('Salario Real Pactado por Semana', 'salario_real_pactado_x_sem'),
            ('Sueldo Base (SB)', 'sb'),
            ('Salario Diario Integrado (SDI)', 'sdi'),
            ('Horas Extra', 'hr_extra'),
            ('Infonavit', 'infonavit'),
            ('Caja de Ahorro', 'caja_ahorro'),
            ('Viaticos', 'viaticos'),
            ('Pago Dia Festivo', 'pago_dia_festivo'),
        ):
            v, err = cell_number(G(row, header_excel), default=0.0)
            if err:
                problemas.append(f'{header_excel}: {err}')
            elif v < 0:
                problemas.append(f'{header_excel}: debe ser >= 0')
            nums[attr] = v

        edad_val, err_edad = cell_number(G(row, 'Edad'), default=0.0)
        if err_edad:
            problemas.append(f'Edad: {err_edad}')
        elif edad_val < 0 or edad_val > 120:
            problemas.append('Edad: fuera de rango (0-120)')
        edad_int = int(edad_val) if edad_val > 0 else None

        # Vocabularios cerrados — normalización aceptando variantes comunes
        sexo = normalize_choice(G(row, 'Sexo (M/F)'), SEXO_ALIASES)
        if sexo and sexo not in ('M', 'F'):
            problemas.append(f'Sexo: {sexo!r} (usa M o F)')

        tipo_nomina = normalize_choice(
            G(row, 'Tipo de Nomina (Semanal/Por hora/Cuadrado)'),
            TIPO_NOMINA_ALIASES,
        )
        if tipo_nomina and tipo_nomina not in ('Semanal', 'Por hora', 'Cuadrado'):
            problemas.append(f'Tipo de Nómina: {tipo_nomina!r} (usa Semanal/Por hora/Cuadrado)')

        tipo_pago = normalize_choice(G(row, 'Tipo Pago'), TIPO_PAGO_ALIASES)
        if tipo_pago and tipo_pago not in ('EFECTIVO', 'TRANSFERENCIA'):
            problemas.append(f'Tipo Pago: {tipo_pago!r} (usa EFECTIVO o TRANSFERENCIA)')

        lentes = normalize_choice(G(row, 'Usa Lentes (Si/No)'), LENTES_ALIASES)
        # No validamos contra catálogo cerrado — usuarios pueden capturar 'Bifocales'.

        estado_civil = normalize_choice(G(row, 'Estado Civil'), ESTADO_CIVIL_ALIASES)

        correo = cell_str(G(row, 'Correo'), maxlen=150)
        if correo and not _EMAIL_RE.match(correo):
            problemas.append(f'Correo inválido ({correo!r})')

        # Si hay problemas, reportar y NO insertar.
        if problemas:
            ident = no_emp or '(sin No. Empleado)'
            errores.append(f'Fila {fila_excel} ({ident}): ' + '; '.join(problemas))
            continue

        # Marcar el no_emp como visto SOLO si pasó las validaciones (así una
        # fila con el mismo no_emp pero rechazada no bloquea reintentos).
        no_emps_en_archivo.add(no_emp)

        t = Trabajador(
            no_empleado=no_emp,
            nombre=nombre,
            nombre_apellidos=apellidos,
            tipo_mov=cell_str(G(row, 'Tipo de Movimiento'), maxlen=100),
            tipo_cont=cell_str(G(row, 'Tipo de Contrato'), maxlen=100),
            area=cell_str(G(row, 'Area'), maxlen=150),
            puesto=cell_str(G(row, 'Puesto'), maxlen=150),
            tipo_jornada=cell_str(G(row, 'Tipo de Jornada'), maxlen=100),
            descripcion_servicio=cell_str(G(row, 'Descripcion de Servicio')),
            fecha_ingreso=fechas['fecha_ingreso'],
            inicio=fechas['inicio'],
            termino_prueba=fechas['termino_prueba'],
            fecha_nacimiento=fechas['fecha_nacimiento'],
            curp=curp or None,
            rfc=rfc or None,
            nss=nss or None,
            sexo=sexo or None,
            estado_civil=cell_str(estado_civil, maxlen=50) or None,
            nacionalidad=cell_str(G(row, 'Nacionalidad'), maxlen=100),
            edad=edad_int,
            domicilio=cell_str(G(row, 'Domicilio')),
            correo=correo or None,
            celular=cell_str(G(row, 'Celular'), maxlen=20),
            tipo_sangre=cell_str(G(row, 'Tipo de Sangre'), maxlen=10),
            alergias=cell_str(G(row, 'Alergias')),
            enfermedades_cronicas=cell_str(G(row, 'Enfermedades Cronicas')),
            contacto_emergencia=cell_str(G(row, 'Contacto de Emergencia'), maxlen=200),
            parentesco_contacto=cell_str(G(row, 'Parentesco del Contacto'), maxlen=100),
            numero_contacto_emerg=cell_str(G(row, 'Numero Contacto Emergencia'), maxlen=20),
            lentes=cell_str(lentes, maxlen=20),
            licencia_conducir=cell_str(G(row, 'Licencia de Conducir (Tipo)'), maxlen=50),
            estatura=cell_str(G(row, 'Estatura'), maxlen=20),
            salario_real_pactado_x_sem=nums['salario_real_pactado_x_sem'],
            tipo_pago=tipo_pago or None,
            tipo_nomina=tipo_nomina or None,
            sb=nums['sb'],
            sdi=nums['sdi'],
            letra=cell_str(G(row, 'Letra'), maxlen=100),
            hr_extra=nums['hr_extra'],
            infonavit=nums['infonavit'],
            caja_ahorro=nums['caja_ahorro'],
            viaticos=nums['viaticos'],
            pago_dia_festivo=nums['pago_dia_festivo'],
            folio_mov_idse=cell_str(G(row, 'Folio Mov IDSE'), maxlen=100),
            ubicacion_estado=cell_str(G(row, 'Ubicacion Estado'), maxlen=100),
            observaciones=cell_str(G(row, 'Observaciones')),
            activo=True,
        )
        a_insertar.append(t)

    # Commit único de los exitosos. Si falla, rollback total — la transacción
    # atómica evita estado parcial en BD.
    exitosos = 0
    if a_insertar:
        try:
            db.session.add_all(a_insertar)
            db.session.commit()
            exitosos = len(a_insertar)
            log_action(f'Importó masivamente {exitosos} trabajadores desde Excel')
            emit_to_role(['admin', 'super_admin'], 'empleado:changed', {
                'action': 'imported', 'count': exitosos,
            })
        except Exception as e:
            db.session.rollback()
            current_app.logger.error('Error commit import trabajadores: %s\n%s', e, traceback.format_exc())
            return {
                'error': 'Error guardando los registros. Ningún empleado fue importado.',
                'tecnico': str(e)[:200],
            }, 500

    return {
        'exitosos': exitosos,
        'errores': errores,
        'total_filas_procesadas': exitosos + len(errores),
    }, 200


@bp.route('/importar', methods=['POST'])
@jwt_required
@limiter.limit('5 per minute')
def importar_excel():
    """Wrapper JSON del importador compartido. Ver `procesar_excel_trabajadores`
    para la lógica completa (validaciones, normalización, dedup, etc.).
    """
    if not is_admin():
        return jsonify({'error': 'Solo admin puede importar'}), 403
    file = request.files.get('archivo') or request.files.get('archivo_excel')
    payload, status = procesar_excel_trabajadores(file)
    return jsonify(payload), status
