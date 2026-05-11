import io
from flask import Blueprint, send_file, flash, redirect, current_app, request, url_for
from sqlalchemy import func
from datetime import datetime
import pandas as pd

from app.extensions import db, limiter
from app.models import (
    Trabajador, Prenomina, ReporteSemanal, RegistroDiarioHoras,
    Proyecto, Prestamo, AjustePeriodo, AjusteTrabajadorPeriodo, AjusteDescuento,
    AbonoPrestamo
)
from app.utils import login_required, admin_required, log_action
from sqlalchemy.orm import selectinload

bp = Blueprint('reportes', __name__, url_prefix='/reportes')

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def _aplicar_estilos_y_retornar(writer, output, filename):
    """Aplica un diseño premium y retorna la respuesta de Flask para descargar el Excel."""
    # Estilos Premium
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
    header_font = Font(name="Calibri", bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Very light slate
    
    total_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid") # Light Blue
    total_font = Font(name="Calibri", bold=True, color="1E3A8A")
    top_border = Border(top=Side(style='medium', color="1E3A8A"))

    workbook = writer.book
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Congelar la primera fila
        worksheet.freeze_panes = 'A2'
        
        # Encontrar la última fila
        max_row = worksheet.max_row
        
        # Iterar sobre las filas para aplicar estilos
        for row in range(1, max_row + 1):
            is_header = (row == 1)
            is_total = False
            
            # Verificar si es la fila de TOTAL
            first_cell = worksheet.cell(row=row, column=1)
            if first_cell.value and str(first_cell.value).upper() == 'TOTAL':
                is_total = True

            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                
                if is_header:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                else:
                    # Zebra striping (excepto la fila de total)
                    if not is_total and row % 2 == 0:
                        cell.fill = zebra_fill
                    
                    # Estilo fila TOTAL
                    if is_total:
                        cell.fill = total_fill
                        cell.font = total_font
                        cell.border = top_border
                    
                    # Formato de Moneda para valores numéricos grandes (heurística simple)
                    if isinstance(cell.value, (int, float)):
                        # Formato: $ 1,234.56
                        cell.number_format = '"$"#,##0.00'

        # Auto-Ajuste de columnas
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = min(adjusted_width, 40)
    
    writer.close()
    
    output.seek(0)
    
    log_action(f"Exportó Reporte Excel: {filename}")
    
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@bp.route('/excel_prenomina/<fecha_str>', methods=['GET'])
@login_required
@admin_required
@limiter.limit("10 per minute")
def excel_prenomina(fecha_str):
    try:
        try:
            fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            flash("Formato de fecha inválido.", "danger")
            return redirect(request.referrer or url_for('prenomina.index'))

        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        
        prenominas = Prenomina.query.options(
            selectinload(Prenomina.trabajador)
        ).filter_by(fecha_inicio=fecha_obj).all()
        if not prenominas:
            # Need to get review calculation like the UI does.
            from app.routes.prenomina import calcular_preview_prenomina
            reportes = ReporteSemanal.query.filter(
                ReporteSemanal.fecha_inicio_semana == fecha_obj,
                ReporteSemanal.estado.in_(['TERMINADO', 'PRENOMINA_CERRADA'])
            ).all()
            prenominas = calcular_preview_prenomina(fecha_obj, reportes)

        if not prenominas:
             flash("No hay datos calculables para la semana.", "warning")
             return redirect(request.referrer or url_for('prenomina.index'))

        data = []
        for p in prenominas:
            data.append({
                'Semana Inicio': p.fecha_inicio.strftime('%Y-%m-%d') if p.fecha_inicio else '',
                'Semana Fin': p.fecha_fin.strftime('%Y-%m-%d') if p.fecha_fin else '',
                'No. Empleado': p.trabajador.no_empleado if p.trabajador else '',
                'Nombre del Empleado': f"{p.trabajador.nombre} {p.trabajador.nombre_apellidos}" if p.trabajador else '',
                'Estado': p.estado if p.reporte_semanal_id is None else 'PREVIEW',
                'Salario Base': float(p.salario_base or 0),
                'Pago Horas Extras': float(p.pago_horas_extras or 0),
                'Pago Viáticos': float(p.pago_viaticos or 0),
                'Pago Festivos': float(p.pago_festivos or 0),
                'Otros Depósitos': float(p.depositos_otros or 0),
                'Depósitos Préstamos': float(p.depositos_prestamos or 0),
                'Total Percepciones': float(p.total_percepciones or 0),
                'Descuento Infonavit': float(p.descuento_infonavit or 0),
                'Ajuste Inbursa': float(p.ajuste_inbursa or 0),
                'Otros Descuentos': float(p.descuentos_otros or 0),
                'Abono Préstamos': float(p.descuento_prestamos or 0),
                'Descuento Incidencias': float(p.descuento_incidencias or 0),
                'Total Deducciones': float(p.total_deducciones or 0),
                'TOTAL A PAGAR': float(p.total_a_pagar or 0),
                'Método de Pago': p.tipo_pago or ''
            })
            
        if data:
            total_row = {
                'Semana Inicio': 'TOTAL',
                'Semana Fin': '',
                'No. Empleado': '',
                'Nombre del Empleado': '',
                'Estado': '',
                'Salario Base': sum(d['Salario Base'] for d in data),
                'Pago Horas Extras': sum(d['Pago Horas Extras'] for d in data),
                'Pago Viáticos': sum(d['Pago Viáticos'] for d in data),
                'Pago Festivos': sum(d['Pago Festivos'] for d in data),
                'Otros Depósitos': sum(d['Otros Depósitos'] for d in data),
                'Depósitos Préstamos': sum(d['Depósitos Préstamos'] for d in data),
                'Total Percepciones': sum(d['Total Percepciones'] for d in data),
                'Descuento Infonavit': sum(d['Descuento Infonavit'] for d in data),
                'Ajuste Inbursa': sum(d['Ajuste Inbursa'] for d in data),
                'Otros Descuentos': sum(d['Otros Descuentos'] for d in data),
                'Abono Préstamos': sum(d['Abono Préstamos'] for d in data),
                'Descuento Incidencias': sum(d['Descuento Incidencias'] for d in data),
                'Total Deducciones': sum(d['Total Deducciones'] for d in data),
                'TOTAL A PAGAR': sum(d['TOTAL A PAGAR'] for d in data),
                'Método de Pago': ''
            }
            data.append(total_row)
            
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='Prenómina', index=False)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        return _aplicar_estilos_y_retornar(writer, output, f"Reporte_Prenomina_{fecha_str}_{timestamp}.xlsx")

    except Exception as e:
        current_app.logger.error(f"Error generando Excel Prenómina: {e}")
        flash("Error exportando reporte.", "danger")
        return redirect(request.referrer or '/')



@bp.route('/excel_historico/<fecha_str>', methods=['GET'])
@login_required
@admin_required
@limiter.limit("10 per minute")
def excel_historico(fecha_str):
    try:
        try:
            fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            flash("Formato de fecha inválido.", "danger")
            return redirect(request.referrer or url_for('historico.index'))

        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        
        prenominas = Prenomina.query.filter(
            Prenomina.fecha_inicio == fecha_obj, 
            Prenomina.estado == 'APROBADO'
        ).order_by(Prenomina.trabajador_id).all()

        if not prenominas:
             flash("No se encontraron prenominas aprobadas para esta semana.", "warning")
             return redirect(request.referrer or url_for('historico.index'))

        data = []
        for p in prenominas:
            data.append({
                'Semana Inicio': p.fecha_inicio.strftime('%Y-%m-%d') if p.fecha_inicio else '',
                'Semana Fin': p.fecha_fin.strftime('%Y-%m-%d') if p.fecha_fin else '',
                'No. Empleado': p.trabajador.no_empleado if p.trabajador else '',
                'Nombre del Empleado': f"{p.trabajador.nombre} {p.trabajador.nombre_apellidos}" if p.trabajador else '',
                'Total Percepciones': float(p.total_percepciones or 0),
                'Total Deducciones': float(p.total_deducciones or 0),
                'TOTAL PAGADO': float(p.total_a_pagar or 0),
                'Método de Pago': p.tipo_pago or ''
            })
            
        if data:
            total_row = {
                'Semana Inicio': 'TOTAL',
                'Semana Fin': '',
                'No. Empleado': '',
                'Nombre del Empleado': '',
                'Total Percepciones': sum(d['Total Percepciones'] for d in data),
                'Total Deducciones': sum(d['Total Deducciones'] for d in data),
                'TOTAL PAGADO': sum(d['TOTAL PAGADO'] for d in data),
                'Método de Pago': ''
            }
            data.append(total_row)
            
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='Histórico', index=False)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        return _aplicar_estilos_y_retornar(writer, output, f"Reporte_Historico_{fecha_str}_{timestamp}.xlsx")

    except Exception as e:
        current_app.logger.error(f"Error generando Excel Histórico: {e}")
        flash("Error exportando reporte.", "danger")
        return redirect(request.referrer or '/')


@bp.route('/excel_proyecto_total/<int:proyecto_id>', methods=['GET'])
@login_required
@admin_required
@limiter.limit("10 per minute")
def excel_proyecto_total(proyecto_id):
    try:
        proyecto = Proyecto.query.get_or_404(proyecto_id)
        
        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        
        reportes = ReporteSemanal.query.filter_by(proyecto_id=proyecto.id, estado='PRENOMINA_CERRADA').order_by(ReporteSemanal.fecha_inicio_semana).all()
        
        if not reportes:
             flash("No hay semanas cerradas para este proyecto.", "warning")
             return redirect(request.referrer or url_for('proyecto_total.index'))

        data = []
        for rep in reportes:
            # Find workers that participated in this project this week
            trabajadores_in_project = db.session.query(RegistroDiarioHoras.trabajador_id).filter(
                RegistroDiarioHoras.reporte_id == rep.id
            ).distinct().all()
            
            t_ids = [t[0] for t in trabajadores_in_project]
            
            prenominas_rep = Prenomina.query.filter(
                Prenomina.fecha_inicio == rep.fecha_inicio_semana,
                Prenomina.estado == 'APROBADO',
                Prenomina.trabajador_id.in_(t_ids)
            ).all() if t_ids else []

            for p in prenominas_rep:
                data.append({
                    'Semana Inicio': rep.fecha_inicio_semana.strftime('%Y-%m-%d'),
                    'Semana Fin': rep.fecha_fin_semana.strftime('%Y-%m-%d'),
                    'No. Empleado': p.trabajador.no_empleado if p.trabajador else '',
                    'Nombre del Empleado': f"{p.trabajador.nombre} {p.trabajador.nombre_apellidos}" if p.trabajador else '',
                    'Salario Base': float(p.salario_base or 0),
                    'Pago Horas Extras': float(p.pago_horas_extras or 0),
                    'Pago Viáticos': float(p.pago_viaticos or 0),
                    'Pago Festivos': float(p.pago_festivos or 0),
                    'Otros Depósitos': float(p.depositos_otros or 0),
                    'Depósitos Préstamos': float(p.depositos_prestamos or 0),
                    'Total Percepciones': float(p.total_percepciones or 0),
                    'Descuento Infonavit': float(p.descuento_infonavit or 0),
                    'Ajuste Inbursa': float(p.ajuste_inbursa or 0),
                    'Otros Descuentos': float(p.descuentos_otros or 0),
                    'Abono Préstamos': float(p.descuento_prestamos or 0),
                    'Descuento Incidencias': float(p.descuento_incidencias or 0),
                    'Total Deducciones': float(p.total_deducciones or 0),
                    'TOTAL A PAGAR': float(p.total_a_pagar or 0)
                })
        
        if not data:
             flash("No hay datos de nómina para este proyecto.", "warning")
             return redirect(request.referrer or url_for('proyecto_total.index'))

        total_row = {
            'Semana Inicio': 'TOTAL',
            'Semana Fin': '',
            'No. Empleado': '',
            'Nombre del Empleado': '',
            'Salario Base': sum(d['Salario Base'] for d in data),
            'Pago Horas Extras': sum(d['Pago Horas Extras'] for d in data),
            'Pago Viáticos': sum(d['Pago Viáticos'] for d in data),
            'Pago Festivos': sum(d['Pago Festivos'] for d in data),
            'Otros Depósitos': sum(d['Otros Depósitos'] for d in data),
            'Depósitos Préstamos': sum(d['Depósitos Préstamos'] for d in data),
            'Total Percepciones': sum(d['Total Percepciones'] for d in data),
            'Descuento Infonavit': sum(d['Descuento Infonavit'] for d in data),
            'Ajuste Inbursa': sum(d['Ajuste Inbursa'] for d in data),
            'Otros Descuentos': sum(d['Otros Descuentos'] for d in data),
            'Abono Préstamos': sum(d['Abono Préstamos'] for d in data),
            'Descuento Incidencias': sum(d['Descuento Incidencias'] for d in data),
            'Total Deducciones': sum(d['Total Deducciones'] for d in data),
            'TOTAL A PAGAR': sum(d['TOTAL A PAGAR'] for d in data)
        }
        data.append(total_row)

        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='Proyecto Total', index=False)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        return _aplicar_estilos_y_retornar(writer, output, f"Reporte_ProyectoTotal_{proyecto.numero_proyecto}_{timestamp}.xlsx")

    except Exception as e:
        current_app.logger.error(f"Error generando Excel Proyecto: {e}")
        flash("Error exportando reporte.", "danger")
        return redirect(request.referrer or '/')


@bp.route('/excel_prestamos/<int:trabajador_id>', methods=['GET'])
@login_required
@admin_required
@limiter.limit("10 per minute")
def excel_prestamos(trabajador_id):
    try:
        trabajador = Trabajador.query.get_or_404(trabajador_id)
        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        
        prestamos = Prestamo.query.filter_by(trabajador_id=trabajador.id).order_by(Prestamo.creado_en.desc()).all()
        if not prestamos:
             flash("Este trabajador no tiene préstamos registrados.", "warning")
             return redirect(request.referrer or url_for('prestamos.index'))

        data = []
        for pr in prestamos:
            total_abonado = sum(float(a.monto or 0) for a in pr.abonos)
            saldo = float(pr.monto_total or 0) - total_abonado
            data.append({
                'ID Préstamo': pr.id,
                'Fecha Registro': pr.creado_en.strftime('%Y-%m-%d'),
                'Monto Original': float(pr.monto_total or 0),
                'Total Abonado': total_abonado,
                'Saldo Restante': saldo,
                'Descuento Semanal': float(pr.descuento_semanal or 0),
                'Estado': pr.estado,
                'Motivo': pr.motivo or ''
            })
            
        if data:
            total_row = {
                'ID Préstamo': 'TOTAL',
                'Fecha Registro': '',
                'Monto Original': sum(d['Monto Original'] for d in data),
                'Total Abonado': sum(d['Total Abonado'] for d in data),
                'Saldo Restante': sum(d['Saldo Restante'] for d in data),
                'Descuento Semanal': sum(d['Descuento Semanal'] for d in data),
                'Estado': '',
                'Motivo': ''
            }
            data.append(total_row)
            
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='Préstamos', index=False)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        nombre_file = f"Prestamos_{trabajador.no_empleado}_{timestamp}.xlsx"
        return _aplicar_estilos_y_retornar(writer, output, nombre_file)

    except Exception as e:
        current_app.logger.error(f"Error generando Excel Préstamos: {e}")
        flash("Error exportando reporte.", "danger")
        return redirect(request.referrer or '/')


@bp.route('/excel_ajustes/<int:periodo_id>', methods=['GET'])
@login_required
@admin_required
@limiter.limit("10 per minute")
def excel_ajustes(periodo_id):
    try:
        from app.models import AjustePeriodo, AjusteDescuento, AjusteTrabajadorPeriodo
        periodo = AjustePeriodo.query.get_or_404(periodo_id)
        
        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        
        # 1. Obtenemos los trabajadores y su progreso para la hoja Resumen
        trabajadores_periodo = AjusteTrabajadorPeriodo.query.filter_by(periodo_id=periodo.id).all()
        ajustes = AjusteDescuento.query.filter_by(periodo_id=periodo.id).order_by(AjusteDescuento.fecha_descuento).all()
        
        if not trabajadores_periodo and not ajustes:
             flash("Este periodo aún no tiene información.", "warning")
             return redirect(request.referrer or url_for('ajustes.index'))

        # Preparar data de Resumen por Trabajador
        descuentos_por_trabajador = {}
        for d in ajustes:
            descuentos_por_trabajador.setdefault(d.trabajador_id, []).append(d)

        data_resumen = []
        for tp in trabajadores_periodo:
            t_id = tp.trabajador_id
            total_desc = sum(float(d.monto or 0) for d in descuentos_por_trabajador.get(t_id, []))
            data_resumen.append({
                'No. Empleado': tp.trabajador.no_empleado if tp.trabajador else '',
                'Nombre del Empleado': tp.trabajador.nombre_completo if tp.trabajador else '',
                'Meta (Depósito)': float(tp.monto_meta),
                'Total Descontado': total_desc,
                'Saldo Restante': float(tp.monto_meta) - total_desc
            })
            
        if data_resumen:
            total_row_res = {
                'No. Empleado': 'TOTAL',
                'Nombre del Empleado': '',
                'Meta (Depósito)': sum(d['Meta (Depósito)'] for d in data_resumen),
                'Total Descontado': sum(d['Total Descontado'] for d in data_resumen),
                'Saldo Restante': sum(d['Saldo Restante'] for d in data_resumen)
            }
            data_resumen.append(total_row_res)
            
        df_resumen = pd.DataFrame(data_resumen)
        if not df_resumen.empty:
            df_resumen.to_excel(writer, sheet_name='Resumen Trabajadores', index=False)

        # Preparar data de Detalle de Descuentos
        data_detalle = []
        for aj in ajustes:
            data_detalle.append({
                'No. Empleado': aj.trabajador.no_empleado if aj.trabajador else '',
                'Nombre del Empleado': aj.trabajador.nombre_completo if aj.trabajador else '',
                'Fecha Aplicación': aj.fecha_descuento.strftime('%Y-%m-%d'),
                'Monto Descontado': float(aj.monto or 0),
                'Notas': aj.notas or ''
            })
            
        if data_detalle:
            total_row_det = {
                'No. Empleado': 'TOTAL',
                'Nombre del Empleado': '',
                'Fecha Aplicación': '',
                'Monto Descontado': sum(d['Monto Descontado'] for d in data_detalle),
                'Notas': ''
            }
            data_detalle.append(total_row_det)
            
        df_detalle = pd.DataFrame(data_detalle)
        if not df_detalle.empty:
            df_detalle.to_excel(writer, sheet_name='Detalle Descuentos', index=False)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        nombre_file = f"Ajuste_Inbursa_{periodo.id}_{timestamp}.xlsx"
        return _aplicar_estilos_y_retornar(writer, output, nombre_file)

    except Exception as e:
        current_app.logger.error(f"Error generando Excel Ajustes: {e}")
        flash("Error exportando reporte.", "danger")
        return redirect(request.referrer or '/')
